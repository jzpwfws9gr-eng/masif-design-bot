import os
import re
import json
import shutil
import asyncio
from datetime import datetime, timedelta
from collections import Counter, defaultdict
import difflib

try:
    import requests
except Exception:
    requests = None

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, MessageHandler, ContextTypes, CallbackQueryHandler, filters

TOKEN = (os.getenv("BOT_TOKEN") or os.getenv("TELEGRAM_BOT_TOKEN") or os.getenv("TELEGRAM_TOKEN") or os.getenv("TOKEN"))

PARTICIPANTS = [
    "عادل عيد", "فهد فارس", "نواف فارس", "خالد عبدالرحمن",
    "محمد عبدالرحمن", "سلطان رباح", "فارس سالم", "عبدالرحمن سالم",
    "ممدوح غزاي", "محمد محسن", "طلال عبدالله", "مشعل غزاي",
]

HEADERS = [
    "المشارك", "الحارس", "اللاعب 1", "اللاعب 2", "اللاعب 3", "الكابتن",
    "نقاط الحارس", "نقاط لاعب 1", "نقاط لاعب 2", "نقاط لاعب 3",
    "نقاط الكابتن", "مجموع اليوم"
]

GOAL_POINTS = {1: 5, 2: 10, 3: 15, 4: 20, 5: 25, 6: 30}
LOCKED_FILE = "locked_days.json"
BACKUP_PREFIX = "backup_"
CUP_FILE = "cup_state.json"

# -------------------- أدوات عامة --------------------

def excel_file(day):
    return f"fantasy_day_{day}.xlsx"


def get_numbers(text):
    return [int(x) for x in re.findall(r"\d+", text or "")]


def get_day(text, default=5):
    nums = get_numbers(text)
    return str(nums[0]) if nums else str(default)


def normalize_name(name):
    """بدون توحيد أسماء تلقائي: فقط تنظيف مسافات."""
    if name is None:
        return ""
    name = str(name).strip()
    name = re.sub(r"\s+", " ", name)
    return name


def is_no_participation(value):
    value = normalize_name(value)
    return value in ("", "لم يشارك")


def has_participated(row_values):
    return any(not is_no_participation(v) for v in row_values[:5])


def score_to_int(value):
    try:
        return int(value or 0)
    except Exception:
        return 0


def ordinal_day(day):
    names = {
        1: "الأول", 2: "الثاني", 3: "الثالث", 4: "الرابع", 5: "الخامس",
        6: "السادس", 7: "السابع", 8: "الثامن", 9: "التاسع", 10: "العاشر",
        11: "الحادي عشر", 12: "الثاني عشر", 13: "الثالث عشر", 14: "الرابع عشر", 15: "الخامس عشر",
    }
    try:
        return names.get(int(day), str(day))
    except Exception:
        return str(day)


def get_existing_days(start_day=1, end_day=31):
    days = []
    for filename in os.listdir("."):
        m = re.match(r"fantasy_day_(\d+)\.xlsx$", filename)
        if not m:
            continue
        day = int(m.group(1))
        if start_day <= day <= end_day:
            days.append(day)
    return sorted(days)


def load_locked_days():
    if not os.path.exists(LOCKED_FILE):
        return set()
    try:
        with open(LOCKED_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {str(x) for x in data.get("locked_days", [])}
    except Exception:
        return set()


def save_locked_days(days):
    def sort_key(x):
        try:
            return int(x)
        except Exception:
            return 0
    with open(LOCKED_FILE, "w", encoding="utf-8") as f:
        json.dump({"locked_days": sorted([str(x) for x in days], key=sort_key)}, f, ensure_ascii=False, indent=2)


def is_locked(day):
    return str(day) in load_locked_days()


def current_data_files():
    files = []
    for filename in os.listdir("."):
        if (
            filename.startswith("fantasy_day_") and filename.endswith(".xlsx")
        ) or filename in ("overall_ranking.xlsx", "fantasy_dashboard.xlsx", LOCKED_FILE, CUP_FILE):
            files.append(filename)
    return sorted(files)


def backup_files(reason="auto", move=False, files=None):
    files = files if files is not None else current_data_files()
    files = [f for f in files if os.path.exists(f)]
    if not files:
        return None, []

    safe_reason = re.sub(r"[^A-Za-z0-9_\-\u0600-\u06FF]+", "_", reason).strip("_") or "auto"
    folder = f"{BACKUP_PREFIX}{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_reason}"
    os.makedirs(folder, exist_ok=True)

    for filename in files:
        target = os.path.join(folder, filename)
        if move:
            shutil.move(filename, target)
        else:
            shutil.copy2(filename, target)

    return folder, files


def latest_backup_folder():
    folders = [f for f in os.listdir(".") if os.path.isdir(f) and f.startswith(BACKUP_PREFIX)]
    if not folders:
        return None
    return sorted(folders)[-1]

# -------------------- تنسيق الإكسل --------------------

def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    light_blue = PatternFill("solid", fgColor="D9EAF7")
    green_fill = PatternFill("solid", fgColor="C6EFCE")

    white_font = Font(color="FFFFFF", bold=True, size=12)
    normal_font = Font(size=12)
    gray_font = Font(color="808080", size=12)
    bold_font = Font(bold=True, size=12)

    thin = Side(style="thin", color="5B9BD5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True

    if ws.max_row >= 1 and ws.max_column >= 1:
        last_col = ws.cell(row=1, column=ws.max_column).column_letter
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    for col in range(1, ws.max_column + 1):
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.font = normal_font

            if cell.row == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
                cell.font = white_font
            else:
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                if cell.row % 2 == 0:
                    cell.fill = light_blue

                if cell.value == "لم يشارك":
                    cell.font = gray_font

                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.fill = green_fill
                    cell.font = bold_font

    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 24


def style_title_cell(cell):
    cell.font = Font(bold=True, size=16, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_dashboard_sheet(ws):
    from openpyxl.utils import get_column_letter

    ws.sheet_view.rightToLeft = True

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

# -------------------- ملفات الأيام --------------------

def create_blank_workbook(day):
    wb = Workbook()
    ws = wb.active
    ws.title = f"اليوم{day}"
    ws.append(HEADERS)
    for name in PARTICIPANTS:
        ws.append([name] + ["لم يشارك"] * 5 + [0, 0, 0, 0, 0, 0])
    style_sheet(ws)
    return wb


def get_or_create_day_workbook(day):
    file_name = excel_file(day)
    sheet_name = f"اليوم{day}"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            ws.title = sheet_name
    else:
        wb = create_blank_workbook(day)
        ws = wb[sheet_name]

    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=idx).value = header

    existing = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            existing[normalize_name(name)] = row

    for name in PARTICIPANTS:
        if name not in existing:
            ws.append([name] + ["لم يشارك"] * 5 + [0, 0, 0, 0, 0, 0])

    return wb, ws, file_name


def participant_row(ws, participant):
    participant = normalize_name(participant)
    for row in range(2, ws.max_row + 1):
        if normalize_name(ws.cell(row=row, column=1).value) == participant:
            return row
    return None


def update_day_data(day, data):
    wb, ws, file_name = get_or_create_day_workbook(day)

    warnings = []
    for participant, values in data.items():
        row = participant_row(ws, participant)
        if not row:
            warnings.append(participant)
            continue

        for col, value in zip(range(2, 7), values):
            ws.cell(row=row, column=col).value = normalize_name(value)

        for col in range(7, 13):
            ws.cell(row=row, column=col).value = 0

    style_sheet(ws)
    wb.save(file_name)
    return file_name, warnings


def read_day_rows(day, data_only=True):
    file_name = excel_file(day)
    sheet_name = f"اليوم{day}"
    if not os.path.exists(file_name):
        return []

    wb = load_workbook(file_name, data_only=data_only)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = []
    for row in range(2, ws.max_row + 1):
        item = {
            "participant": normalize_name(ws.cell(row=row, column=1).value),
            "keeper": normalize_name(ws.cell(row=row, column=2).value),
            "p1": normalize_name(ws.cell(row=row, column=3).value),
            "p2": normalize_name(ws.cell(row=row, column=4).value),
            "p3": normalize_name(ws.cell(row=row, column=5).value),
            "captain": normalize_name(ws.cell(row=row, column=6).value),
            "keeper_points": score_to_int(ws.cell(row=row, column=7).value),
            "p1_points": score_to_int(ws.cell(row=row, column=8).value),
            "p2_points": score_to_int(ws.cell(row=row, column=9).value),
            "p3_points": score_to_int(ws.cell(row=row, column=10).value),
            "captain_points": score_to_int(ws.cell(row=row, column=11).value),
            "total": score_to_int(ws.cell(row=row, column=12).value),
        }
        item["participated"] = has_participated([item["keeper"], item["p1"], item["p2"], item["p3"], item["captain"]])
        rows.append(item)
    return rows

# -------------------- النتائج والحساب --------------------

def parse_results(text):
    goals_count = defaultdict(int)
    clean_sheets = []
    mode = None

    for line in text.splitlines()[1:]:
        line = line.strip()
        if not line:
            continue

        if "الأهداف" in line or "اهداف" in line:
            mode = "goals"
            continue

        if "كلين" in line or "شيت" in line or "الكلين" in line:
            mode = "clean"
            continue

        if mode == "goals":
            if "|" in line:
                player, count = line.split("|", 1)
                player = normalize_name(player)
                try:
                    count = int(str(count).strip())
                except Exception:
                    count = 1
                goals_count[player] += count
            else:
                goals_count[normalize_name(line)] += 1

        elif mode == "clean":
            clean_sheets.append(normalize_name(line))

    clean_sheets = list(dict.fromkeys([x for x in clean_sheets if x]))
    goals_count = {k: v for k, v in goals_count.items() if k}
    return goals_count, clean_sheets


def goals_to_points(goals_count):
    return {player: GOAL_POINTS.get(count, count * 5) for player, count in goals_count.items()}


def lineup_names_for_day(day):
    names = set()
    for row in read_day_rows(day):
        if not row["participated"]:
            continue
        for key in ("keeper", "p1", "p2", "p3", "captain"):
            val = row[key]
            if val and val != "لم يشارك":
                names.add(val)
    return names


def validate_results_names(day, goals_count, clean_sheets):
    existing = lineup_names_for_day(day)
    goal_missing = [name for name in goals_count.keys() if name not in existing]
    clean_missing = [name for name in clean_sheets if name not in existing]
    return goal_missing, clean_missing


def calculate_points(day, goals_count, clean_sheets):
    file_name = excel_file(day)
    sheet_name = f"اليوم{day}"

    if not os.path.exists(file_name):
        return None

    wb = load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    goals_points = goals_to_points(goals_count)

    for row in range(2, ws.max_row + 1):
        keeper = normalize_name(ws.cell(row=row, column=2).value)
        p1 = normalize_name(ws.cell(row=row, column=3).value)
        p2 = normalize_name(ws.cell(row=row, column=4).value)
        p3 = normalize_name(ws.cell(row=row, column=5).value)
        captain = normalize_name(ws.cell(row=row, column=6).value)

        keeper_points = 5 if keeper in clean_sheets else 0
        p1_points = goals_points.get(p1, 0)
        p2_points = goals_points.get(p2, 0)
        p3_points = goals_points.get(p3, 0)

        captain_points = 0
        if captain == p1:
            captain_points = p1_points
        elif captain == p2:
            captain_points = p2_points
        elif captain == p3:
            captain_points = p3_points
        elif captain == keeper:
            captain_points = keeper_points

        total = keeper_points + p1_points + p2_points + p3_points + captain_points

        for col, val in zip(range(7, 13), [keeper_points, p1_points, p2_points, p3_points, captain_points, total]):
            ws.cell(row=row, column=col).value = val

    style_sheet(ws)
    wb.save(file_name)
    return file_name


def clear_day_points(day):
    file_name = excel_file(day)
    sheet_name = f"اليوم{day}"
    if not os.path.exists(file_name):
        return None

    wb = load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        for col in range(7, 13):
            ws.cell(row=row, column=col).value = 0

    style_sheet(ws)
    wb.save(file_name)
    return file_name

# -------------------- تحليل البيانات --------------------

def collect_stats(start_day=1, end_day=31):
    days = get_existing_days(start_day, end_day)
    per_day = {}
    totals = {name: 0 for name in PARTICIPANTS}
    daily_wins = {name: 0 for name in PARTICIPANTS}
    participation_count = {name: 0 for name in PARTICIPANTS}
    zero_days = {name: 0 for name in PARTICIPANTS}
    scores_by_day = {name: {} for name in PARTICIPANTS}
    cumulative_by_day = {name: {} for name in PARTICIPANTS}

    captain_choice_count = Counter()
    captain_points_by_player = Counter()
    captain_points_by_participant = Counter()

    keeper_choice_count = Counter()
    keeper_points_by_keeper = Counter()
    keeper_points_by_participant = Counter()
    keeper_success_by_participant = Counter()

    player_choice_count = Counter()
    player_points = Counter()
    player_zero_selections = Counter()
    captain_selections_by_player = Counter()

    for day in days:
        rows = read_day_rows(day)
        day_scores = {}
        participants_today = 0
        zero_today = 0

        for row in rows:
            name = row["participant"]
            if name not in totals:
                continue

            total = row["total"]
            totals[name] += total
            scores_by_day[name][day] = total
            day_scores[name] = total

            if row["participated"]:
                participants_today += 1
                participation_count[name] += 1
                if total == 0:
                    zero_days[name] += 1
                    zero_today += 1

                keeper = row["keeper"]
                if keeper and keeper != "لم يشارك":
                    keeper_choice_count[keeper] += 1
                    keeper_points_by_keeper[keeper] += row["keeper_points"]
                    keeper_points_by_participant[name] += row["keeper_points"]
                    if row["keeper_points"] > 0:
                        keeper_success_by_participant[name] += 1

                captain = row["captain"]
                if captain and captain != "لم يشارك":
                    captain_choice_count[captain] += 1
                    captain_selections_by_player[captain] += 1
                    captain_points_by_player[captain] += row["captain_points"]
                    captain_points_by_participant[name] += row["captain_points"]

                for p_key, pts_key in (("p1", "p1_points"), ("p2", "p2_points"), ("p3", "p3_points")):
                    player = row[p_key]
                    pts = row[pts_key]
                    if player and player != "لم يشارك":
                        player_choice_count[player] += 1
                        player_points[player] += pts
                        if pts == 0:
                            player_zero_selections[player] += 1

        if day_scores:
            max_score = max(day_scores.values())
            winners = [n for n, s in day_scores.items() if s == max_score and max_score > 0]
            for winner in winners:
                daily_wins[winner] += 1

            scores_values = list(day_scores.values())
            per_day[day] = {
                "scores": day_scores,
                "max_score": max_score,
                "winners": winners,
                "participants": participants_today,
                "zeros": zero_today,
                "avg": round(sum(scores_values) / len(scores_values), 2) if scores_values else 0,
                "min_score": min(scores_values) if scores_values else 0,
            }

        for name in PARTICIPANTS:
            cumulative_by_day[name][day] = totals[name]

    ranking = sorted(PARTICIPANTS, key=lambda n: (totals[n], daily_wins[n]), reverse=True)

    return {
        "days": days,
        "per_day": per_day,
        "totals": totals,
        "daily_wins": daily_wins,
        "participation_count": participation_count,
        "zero_days": zero_days,
        "scores_by_day": scores_by_day,
        "cumulative_by_day": cumulative_by_day,
        "ranking": ranking,
        "captain_choice_count": captain_choice_count,
        "captain_points_by_player": captain_points_by_player,
        "captain_points_by_participant": captain_points_by_participant,
        "keeper_choice_count": keeper_choice_count,
        "keeper_points_by_keeper": keeper_points_by_keeper,
        "keeper_points_by_participant": keeper_points_by_participant,
        "keeper_success_by_participant": keeper_success_by_participant,
        "player_choice_count": player_choice_count,
        "player_points": player_points,
        "player_zero_selections": player_zero_selections,
        "captain_selections_by_player": captain_selections_by_player,
    }


def create_overall_ranking(start_day=1, end_day=31):
    stats = collect_stats(start_day, end_day)
    days_found = stats["days"]

    wb = Workbook()
    ws = wb.active
    ws.title = "الترتيب العام"
    ws.append(["المركز", "المشارك", "المجموع", "أسطورة اليوم"])

    current_rank = 0
    last_score = None
    real_index = 0

    for name in stats["ranking"]:
        real_index += 1
        score = stats["totals"][name]
        if score != last_score:
            current_rank = real_index
            last_score = score
        ws.append([current_rank, name, score, stats["daily_wins"][name]])

    style_sheet(ws)
    file_name = "overall_ranking.xlsx"
    wb.save(file_name)
    return file_name, days_found, stats


def build_ranking_text(stats, start_day=1, end_day=31):
    days = stats["days"]
    if not days:
        return "ما فيه أيام محسوبة."

    medals = ["🥇", "🥈", "🥉"]
    title = f"🏆 الترتيب العام لفانتزي المصيف 2026 بعد اليوم {max(days)} 🏆"
    lines = [title, ""]

    for idx, name in enumerate(stats["ranking"], start=1):
        score = stats["totals"][name]
        legends = stats["daily_wins"][name]
        icon = medals[idx - 1] if idx <= 3 else f"{idx}."
        if legends:
            lines.append(f"{icon} {name} — {score} نقطة | أسطورة اليوم: {legends}")
        else:
            lines.append(f"{icon} {name} — {score} نقطة")

    return "\n".join(lines)

# -------------------- الداشبورد --------------------

def create_dashboard(start_day=1, end_day=31):
    from collections import defaultdict, Counter
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    try:
        from openpyxl.chart import DoughnutChart
    except Exception:
        DoughnutChart = PieChart

    stats = collect_stats(start_day, end_day)
    days = stats["days"]
    file_name = "fantasy_dashboard.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # ألوان Dashboard فاتح مثل الصورة الثانية
    BG = "F5F7FB"
    WHITE = "FFFFFF"
    NAVY = "243B53"
    BLUE = "4F7DF3"
    PURPLE = "7C3AED"
    L_PURPLE = "EFE7FF"
    L_BLUE = "EAF1FF"
    GREEN = "12B886"
    L_GREEN = "E7F8F1"
    GOLD = "F2B705"
    L_GOLD = "FFF6D6"
    RED = "F06565"
    L_RED = "FFECEC"
    TEXT = "1F2937"
    MUTED = "6B7280"
    GRID = "D9E2EF"
    HEADER = "3B4A6B"

    thin = Side(style="thin", color=GRID)
    no_side = Side(style=None)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    INVALID_SELECTIONS = {"", "لم يشارك", "تم حجب المشاركة / تأخير"}

    def is_real_selection(value):
        return normalize_name(value) not in INVALID_SELECTIONS

    def points_source(base_pts, cap_extra):
        parts = []
        try:
            base_pts = int(base_pts or 0)
        except Exception:
            base_pts = 0
        try:
            cap_extra = int(cap_extra or 0)
        except Exception:
            cap_extra = 0

        if base_pts > 0:
            goals = base_pts // 5
            if goals <= 1:
                parts.append("هدف")
            elif goals == 2:
                parts.append("هدفين")
            elif goals <= 10:
                parts.append(f"{goals} أهداف")
            else:
                parts.append(f"{base_pts} نقطة")
        if cap_extra > 0:
            parts.append("كابتن")
        return " + ".join(parts) if parts else "بدون نقاط"

    def safe_pct(part, total):
        return f"{round((part / total) * 100, 1)}%" if total else "0%"

    def safe_avg(total, count):
        return round(total / count, 2) if count else 0

    def sheet_setup(ws, title=None):
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        for col in range(1, 22):
            ws.column_dimensions[get_column_letter(col)].width = 16
        for row in range(1, 90):
            ws.row_dimensions[row].height = 22
        if title:
            ws["A1"] = title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            c = ws["A1"]
            c.font = Font(bold=True, size=18, color=WHITE)
            c.fill = PatternFill("solid", fgColor=HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 34

    def style_table(ws, min_row, max_row, min_col, max_col, header_row=None, autofilter=True):
        header_row = header_row or min_row
        if max_row < min_row or max_col < min_col:
            return
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                if r == header_row:
                    cell.fill = PatternFill("solid", fgColor=HEADER)
                    cell.font = Font(bold=True, size=11, color=WHITE)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = Font(size=11, color=TEXT)
                    cell.fill = PatternFill("solid", fgColor=("F8FAFC" if r % 2 == 0 else WHITE))
        ws.row_dimensions[header_row].height = 30
        if autofilter:
            ws.auto_filter.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{max_row}"

    def section_title(ws, row, col, title, width=4, color=HEADER):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
        c = ws.cell(row=row, column=col)
        c.value = title
        c.font = Font(bold=True, size=13, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.row_dimensions[row].height = 30

    def card(ws, cell_range, title, value, color=BLUE, sub_text=None):
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        text = f"{title}\n{value}"
        if sub_text:
            text += f"\n{sub_text}"
        c.value = text
        c.font = Font(bold=True, size=13, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def light_card(ws, cell_range, title, value, color_fill=L_BLUE, color_text=TEXT):
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value = f"{title}\n{value}"
        c.font = Font(bold=True, size=12, color=color_text)
        c.fill = PatternFill("solid", fgColor=color_fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    def add_bar_chart(ws, title, data_col, header_row, min_row, max_row, cats_col, anchor, width=12, height=7, chart_type="bar"):
        if max_row < min_row:
            return
        chart = BarChart()
        chart.type = chart_type
        chart.style = 10
        chart.title = title
        chart.y_axis.title = ""
        chart.x_axis.title = ""
        data = Reference(ws, min_col=data_col, min_row=header_row, max_row=max_row)
        cats = Reference(ws, min_col=cats_col, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = width
        chart.height = height
        ws.add_chart(chart, anchor)

    def add_line_chart_local(ws, title, min_col, max_col, header_row, min_row, max_row, cats_col, anchor, width=18, height=8):
        if max_row < min_row or max_col < min_col:
            return
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = "النقاط"
        chart.x_axis.title = "الجولة"
        data = Reference(ws, min_col=min_col, max_col=max_col, min_row=header_row, max_row=max_row)
        cats = Reference(ws, min_col=cats_col, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = width
        chart.height = height
        ws.add_chart(chart, anchor)

    def add_donut_chart(ws, title, labels_col, values_col, min_row, max_row, anchor, width=9, height=7):
        if max_row < min_row:
            return
        chart = DoughnutChart()
        chart.title = title
        data = Reference(ws, min_col=values_col, min_row=min_row, max_row=max_row)
        labels = Reference(ws, min_col=labels_col, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        chart.width = width
        chart.height = height
        if hasattr(chart, "holeSize"):
            chart.holeSize = 55
        ws.add_chart(chart, anchor)

    # -------------------- حسابات إضافية للصفحات الجديدة --------------------
    participant_count = len(PARTICIPANTS)
    total_slots = participant_count * len(days) if days else 0
    total_active_slots = 0

    day_summary = {}
    active_scores_by_participant = {name: {} for name in PARTICIPANTS}

    keeper_participants = defaultdict(set)
    keeper_rounds = defaultdict(set)
    keeper_clean_rounds = defaultdict(set)
    keeper_round_detail = defaultdict(lambda: {"selected": 0, "people": set(), "clean": False, "total_points": 0})

    captain_by_participant = {name: {"count": 0, "success": 0, "points": 0, "best_player": "-", "best_day": "-", "best_points": -1} for name in PARTICIPANTS}
    captain_player_count = Counter()
    captain_player_points = Counter()

    player_participants = defaultdict(set)
    player_round_detail = defaultdict(lambda: {"selected": 0, "captains": 0, "base_points": 0, "distributed": 0})
    player_actual_by_day = defaultdict(dict)
    player_total_distributed = Counter()
    player_actual_total = Counter()
    player_choice_count = Counter()
    player_captain_count = Counter()

    selection_detail_rows = []

    for day in days:
        rows = read_day_rows(day)
        active_totals = []
        total_day_points = 0

        for row in rows:
            name = row["participant"]
            total = row["total"]
            total_day_points += total

            if row["participated"]:
                total_active_slots += 1
                active_totals.append(total)
                active_scores_by_participant[name][day] = total

                # الحراس
                keeper = row["keeper"]
                if is_real_selection(keeper):
                    keeper_participants[keeper].add(name)
                    keeper_rounds[keeper].add(day)
                    kd = keeper_round_detail[(day, keeper)]
                    kd["selected"] += 1
                    kd["people"].add(name)
                    kd["total_points"] += row["keeper_points"]
                    if row["keeper_points"] > 0:
                        kd["clean"] = True
                        keeper_clean_rounds[keeper].add(day)


                # الكباتن - حسب المشارك
                captain = row["captain"]
                if is_real_selection(captain):
                    captain_by_participant[name]["count"] += 1
                    captain_by_participant[name]["points"] += row["captain_points"]
                    captain_player_count[captain] += 1
                    captain_player_points[captain] += row["captain_points"]
                    if row["captain_points"] > 0:
                        captain_by_participant[name]["success"] += 1
                    if row["captain_points"] > captain_by_participant[name]["best_points"]:
                        captain_by_participant[name]["best_points"] = row["captain_points"]
                        captain_by_participant[name]["best_player"] = captain
                        captain_by_participant[name]["best_day"] = day

                # اللاعبون
                for p_key, pts_key in (("p1", "p1_points"), ("p2", "p2_points"), ("p3", "p3_points")):
                    player = row[p_key]
                    base_pts = row[pts_key]
                    if is_real_selection(player):
                        player_participants[player].add(name)
                        player_choice_count[player] += 1
                        pd = player_round_detail[(day, player)]
                        pd["selected"] += 1
                        pd["base_points"] = max(pd["base_points"], base_pts)
                        player_actual_by_day[player][day] = max(player_actual_by_day[player].get(day, 0), base_pts)

                        cap_extra = row["captain_points"] if row["captain"] == player else 0
                        if row["captain"] == player:
                            pd["captains"] += 1
                            player_captain_count[player] += 1

                        selection_detail_rows.append([
                            day,
                            name,
                            player,
                            "نعم 👑" if row["captain"] == player else "لا",
                            base_pts,
                            cap_extra,
                            base_pts + cap_extra,
                            points_source(base_pts, cap_extra),
                        ])

        for (d, player), detail in list(player_round_detail.items()):
            if d == day:
                detail["distributed"] = detail["base_points"] * detail["selected"] + detail["base_points"] * detail["captains"]

        info = stats["per_day"].get(day, {})
        winners = info.get("winners", [])
        day_summary[day] = {
            "participants": info.get("participants", 0),
            "non_participants": max(participant_count - info.get("participants", 0), 0),
            "total_points": total_day_points,
            "avg_active": round(sum(active_totals) / len(active_totals), 2) if active_totals else 0,
            "max_score": info.get("max_score", 0),
            "winners": winners,
            "legends_count": len(winners),
        }

    for player, by_day in player_actual_by_day.items():
        player_actual_total[player] = sum(by_day.values())

    for (day, player), detail in player_round_detail.items():
        player_total_distributed[player] += detail["distributed"]

    ranking = stats["ranking"]
    leader = ranking[0] if ranking else "-"
    leader_points = stats["totals"].get(leader, 0)

    highest_daily = 0
    highest_daily_name = "-"
    highest_daily_day = "-"
    for day, info in day_summary.items():
        if info["max_score"] > highest_daily:
            highest_daily = info["max_score"]
            highest_daily_name = "، ".join(info["winners"]) or "-"
            highest_daily_day = day

    most_legend = "-"
    if PARTICIPANTS:
        most_legend = max(PARTICIPANTS, key=lambda n: stats["daily_wins"][n])
        if stats["daily_wins"].get(most_legend, 0) == 0:
            most_legend = "-"

    best_keeper = "-"
    if keeper_rounds:
        best_keeper = max(keeper_rounds.keys(), key=lambda k: stats["keeper_points_by_keeper"][k])

    best_captain_participant = "-"
    if ranking:
        best_captain_participant = max(ranking, key=lambda n: captain_by_participant[n]["points"])

    # أكثر لاعب وزع نقاط على المشاركين في جولة واحدة
    top_round_player = None
    if player_round_detail:
        top_round_player = max(player_round_detail.items(), key=lambda x: x[1]["distributed"])

    best_actual_player = "-"
    if player_actual_total:
        best_actual_player = max(player_actual_total.keys(), key=lambda p: player_actual_total[p])

    most_selected_player = "-"
    if player_choice_count:
        most_selected_player = player_choice_count.most_common(1)[0][0]

    most_captained_player = "-"
    if player_captain_count:
        most_captained_player = player_captain_count.most_common(1)[0][0]

    # اللاعب الذهبي: أكثر لاعب أفاد المشاركين في البطولة كاملة
    golden_player = "-"
    if player_total_distributed:
        golden_player = max(player_total_distributed.keys(), key=lambda p: player_total_distributed[p])

    def build_rank_map_from_scores(score_map):
        ordered = sorted(score_map.keys(), key=lambda n: score_map.get(n, 0), reverse=True)
        ranks = {}
        current_rank = 0
        last_score = None
        real_index = 0
        for n in ordered:
            real_index += 1
            s = score_map.get(n, 0)
            if s != last_score:
                current_rank = real_index
                last_score = s
            ranks[n] = current_rank
        return ranks

    def movement_text(name, current_rank, previous_rank):
        if previous_rank is None:
            return "جديد"
        diff = previous_rank - current_rank
        if diff > 0:
            return f"↑ صعد {diff} مركز" if diff == 1 else f"↑ صعد {diff} مراكز"
        if diff < 0:
            drop = abs(diff)
            return f"↓ نزل مركز" if drop == 1 else f"↓ نزل {drop} مراكز"
        return "— ثابت"

    current_rank_map = build_rank_map_from_scores(stats["totals"])
    previous_rank_map = {}
    if len(days) >= 2:
        prev_day = days[-2]
        prev_scores = {name: stats["cumulative_by_day"][name].get(prev_day, 0) for name in PARTICIPANTS}
        previous_rank_map = build_rank_map_from_scores(prev_scores)

    # -------------------- 1) لوحة عامة Dashboard --------------------
    ws = wb.create_sheet("لوحة عامة")
    sheet_setup(ws)
    ws.sheet_properties.tabColor = PURPLE

    # خلفية/عنوان
    for row in range(1, 50):
        for col in range(1, 14):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BG)

    ws.merge_cells("A1:M2")
    ws["A1"] = "لوحة فانتزي المصيف 2026"
    ws["A1"].font = Font(bold=True, size=22, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # كروت عليا
    card(ws, "A4:C6", "🏆 المتصدر", f"{leader}\n{leader_points} نقطة", PURPLE)
    card(ws, "D4:F6", "👥 المشاركون", f"{participant_count}", BLUE)
    card(ws, "G4:I6", "📅 الجولات", f"{len(days)}", GREEN)
    card(ws, "J4:M6", "🔥 أعلى نتيجة يومية", f"{highest_daily_name}\n{highest_daily} نقطة - اليوم {highest_daily_day}", GOLD)

    light_card(ws, "A8:C10", "⭐ أكثر أسطورة يوم", f"{most_legend}\n{stats['daily_wins'].get(most_legend, 0)} مرات", L_GOLD)
    light_card(ws, "D8:F10", "🧤 أفضل حارس", f"{best_keeper}\n{stats['keeper_points_by_keeper'].get(best_keeper, 0)} نقطة", L_BLUE)
    light_card(ws, "G8:I10", "👑 أفضل مشارك بالكباتن", f"{best_captain_participant}\n{captain_by_participant.get(best_captain_participant, {}).get('points', 0)} نقطة إضافية", L_PURPLE)
    if top_round_player:
        (top_day, top_player), top_info = top_round_player
        light_card(ws, "J8:M10", "🔥 أكثر لاعب وزّع نقاط على المشاركين في جولة واحدة", f"{top_player} - اليوم {top_day}\nوزّع {top_info['distributed']} نقطة", L_RED)
    else:
        light_card(ws, "J8:M10", "🔥 أكثر لاعب وزّع نقاط على المشاركين في جولة واحدة", "-", L_RED)

    # بيانات مخفية للرسوم
    data_col = 16  # P
    ws.cell(row=1, column=data_col).value = "بيانات"
    # نسبة المشاركة
    ws.cell(row=2, column=data_col).value = "شاركوا"
    ws.cell(row=2, column=data_col + 1).value = total_active_slots
    ws.cell(row=3, column=data_col).value = "لم يشاركوا"
    ws.cell(row=3, column=data_col + 1).value = max(total_slots - total_active_slots, 0)

    # توزيع الأساطير
    leg_start = 6
    ws.cell(row=leg_start, column=data_col).value = "المشارك"
    ws.cell(row=leg_start, column=data_col + 1).value = "مرات أسطورة اليوم"
    leg_rows = [(n, stats["daily_wins"][n]) for n in ranking if stats["daily_wins"][n] > 0][:8]
    if not leg_rows:
        leg_rows = [("-", 0)]
    for i, (name, val) in enumerate(leg_rows, start=leg_start + 1):
        ws.cell(row=i, column=data_col).value = name
        ws.cell(row=i, column=data_col + 1).value = val

    # أفضل 5
    top_start = 17
    ws.cell(row=top_start, column=data_col).value = "المشارك"
    ws.cell(row=top_start, column=data_col + 1).value = "النقاط"
    for i, name in enumerate(ranking[:5], start=top_start + 1):
        ws.cell(row=i, column=data_col).value = name
        ws.cell(row=i, column=data_col + 1).value = stats["totals"][name]

    # نقاط الجولات
    day_start = 26
    ws.cell(row=day_start, column=data_col).value = "الجولة"
    ws.cell(row=day_start, column=data_col + 1).value = "مجموع نقاط الجولة"
    for i, day in enumerate(days, start=day_start + 1):
        ws.cell(row=i, column=data_col).value = day
        ws.cell(row=i, column=data_col + 1).value = day_summary[day]["total_points"]

    # تطور أول 5
    evo_start = 35
    ws.cell(row=evo_start, column=data_col).value = "الجولة"
    for idx, name in enumerate(ranking[:5], start=data_col + 1):
        ws.cell(row=evo_start, column=idx).value = name
    for r, day in enumerate(days, start=evo_start + 1):
        ws.cell(row=r, column=data_col).value = day
        for c, name in enumerate(ranking[:5], start=data_col + 1):
            ws.cell(row=r, column=c).value = stats["cumulative_by_day"][name].get(day, 0)

    # الترتيب العام داخل لوحة عامة
    section_title(ws, 12, 1, "الترتيب العام", 6, HEADER)
    ranking_headers = ["المركز", "المشارك", "النقاط", "الفارق عن المتصدر", "حركة الترتيب", "أسطورة اليوم"]
    for col, h in enumerate(ranking_headers, start=1):
        ws.cell(row=13, column=col).value = h

    current_rank = 0
    last_score = None
    real_index = 0
    rank_row = 14
    for name in stats["ranking"]:
        real_index += 1
        score = stats["totals"][name]
        if score != last_score:
            current_rank = real_index
            last_score = score
        prev_rank = previous_rank_map.get(name) if previous_rank_map else None
        move = movement_text(name, current_rank, prev_rank)
        values = [current_rank, name, score, leader_points - score, move, stats["daily_wins"][name]]
        for col, val in enumerate(values, start=1):
            ws.cell(row=rank_row, column=col).value = val
        # تلوين حركة الترتيب فقط بدون إزعاج الجدول
        move_cell = ws.cell(row=rank_row, column=5)
        if isinstance(move, str) and move.startswith("↑"):
            move_cell.font = Font(bold=True, color=GREEN)
        elif isinstance(move, str) and move.startswith("↓"):
            move_cell.font = Font(bold=True, color=RED)
        elif move == "جديد":
            move_cell.font = Font(bold=True, color=BLUE)
        rank_row += 1

    style_table(ws, 13, max(13, rank_row - 1), 1, 6, header_row=13)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["E"].width = 20

    # الرسوم تحت جدول الترتيب
    chart_top = max(31, rank_row + 3)
    add_donut_chart(ws, "نسبة المشاركة", data_col, data_col + 1, 2, 3, f"A{chart_top}", width=8, height=7)
    add_donut_chart(ws, "توزيع أسطورة اليوم", data_col, data_col + 1, leg_start + 1, leg_start + len(leg_rows), f"D{chart_top}", width=8, height=7)
    add_bar_chart(ws, "أفضل 5 مشاركين", data_col + 1, top_start, top_start + 1, top_start + len(ranking[:5]), data_col, f"G{chart_top}", width=8, height=7, chart_type="bar")
    add_bar_chart(ws, "مجموع نقاط الجولات", data_col + 1, day_start, day_start + 1, day_start + len(days), data_col, f"J{chart_top}", width=8, height=7, chart_type="col")
    add_line_chart_local(ws, "تطور نقاط أفضل 5", data_col + 1, data_col + min(5, len(ranking)), evo_start, evo_start + 1, evo_start + len(days), data_col, f"A{chart_top + 14}", width=18, height=8)

    for col in range(data_col, data_col + 8):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # -------------------- 2) تطور النقاط ✅ كما هو --------------------
    ws = wb.create_sheet("تطور النقاط")
    ws.append(["اليوم"] + stats["ranking"])

    for day in days:
        ws.append([day] + [stats["cumulative_by_day"][name].get(day, 0) for name in stats["ranking"]])

    style_sheet(ws)

    if days:
        add_line_chart_local(ws, "تطور مجموع النقاط يومًا بعد يوم", 2, 1 + len(stats["ranking"]), 1, 2, ws.max_row, 1, "A16", width=20, height=10)

    # -------------------- 3) تحليل الأيام ✅ مع تنسيق أسطورة اليوم --------------------
    ws = wb.create_sheet("تحليل الأيام")
    sheet_setup(ws, "تحليل الأيام")
    headers = ["اليوم", "عدد المشاركين", "غير المشاركين", "مجموع نقاط اليوم", "متوسط نقاط المشاركين", "أعلى نقاط", "عدد الأساطير", "أسطورة اليوم"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = h

    row_i = 3
    for day in days:
        info = day_summary[day]
        winners_text = "\n".join([f"⭐ {w}" for w in info["winners"]]) if info["winners"] else "-"
        values = [
            day,
            info["participants"],
            info["non_participants"],
            info["total_points"],
            info["avg_active"],
            info["max_score"],
            info["legends_count"],
            winners_text,
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_i, column=col).value = val
        ws.row_dimensions[row_i].height = max(28, 18 * max(1, info["legends_count"]))
        row_i += 1

    ws.column_dimensions["H"].width = 34
    style_table(ws, 2, max(2, ws.max_row), 1, 8, header_row=2)
    if ws.max_row >= 3:
        add_bar_chart(ws, "مجموع نقاط كل يوم", 4, 2, 3, ws.max_row, 1, f"A{ws.max_row + 4}", width=12, height=7, chart_type="col")
        add_bar_chart(ws, "متوسط نقاط المشاركين", 5, 2, 3, ws.max_row, 1, f"H{ws.max_row + 4}", width=12, height=7, chart_type="col")

    # -------------------- 4) تحليل المشاركين ✅ كما هو --------------------
    ws = wb.create_sheet("تحليل المشاركين")
    ws.append(["المشارك", "المجموع", "عدد المشاركات", "نسبة المشاركة", "أسطورة اليوم", "أفضل يوم", "نقاط أفضل يوم", "أسوأ يوم", "نقاط أسوأ يوم", "أيام صفر"])

    total_days = len(days)

    for name in stats["ranking"]:
        day_scores = stats["scores_by_day"][name]
        if day_scores:
            best_day = max(day_scores, key=lambda d: day_scores[d])
            worst_day = min(day_scores, key=lambda d: day_scores[d])
            best_score = day_scores[best_day]
            worst_score = day_scores[worst_day]
        else:
            best_day = worst_day = "-"
            best_score = worst_score = 0

        pc = stats["participation_count"][name]
        pct = f"{round((pc / total_days) * 100, 1)}%" if total_days else "0%"
        ws.append([name, stats["totals"][name], pc, pct, stats["daily_wins"][name], best_day, best_score, worst_day, worst_score, stats["zero_days"][name]])

    style_sheet(ws)
    add_bar_chart(ws, "مجموع نقاط المشاركين", 2, 1, 2, ws.max_row, 1, f"A{ws.max_row + 4}", width=14, height=8, chart_type="bar")

    # -------------------- 5) تحليل الكباتن ✅ حسب المشاركين --------------------
    ws = wb.create_sheet("تحليل الكباتن")
    sheet_setup(ws, "تحليل الكباتن")

    card(ws, "A3:C5", "👑 أفضل مشارك في الكباتن", f"{best_captain_participant}\n{captain_by_participant.get(best_captain_participant, {}).get('points', 0)} نقطة إضافية", PURPLE)
    top_cap_player = captain_player_count.most_common(1)[0][0] if captain_player_count else "-"
    card(ws, "D3:F5", "🔥 أكثر لاعب تم اختياره كابتن", f"{top_cap_player}\n{captain_player_count.get(top_cap_player, 0)} مرات", BLUE)
    top_success = "-"
    if ranking:
        top_success = max(ranking, key=lambda n: (safe_avg(captain_by_participant[n]["success"], captain_by_participant[n]["count"]), captain_by_participant[n]["points"]))
    card(ws, "G3:I5", "🎯 أعلى نجاح كابتن", f"{top_success}\n{safe_pct(captain_by_participant[top_success]['success'], captain_by_participant[top_success]['count']) if top_success != '-' else '0%'}", GREEN)

    section_title(ws, 7, 1, "أداء المشاركين في اختيارات الكابتن", 8, HEADER)
    headers = ["المشارك", "مرات اختيار كابتن", "كباتن جابوا نقاط", "نقاط الكابتن الإضافية", "أفضل كابتن اختاره", "نسبة نجاح الكابتن"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=8, column=col).value = h

    row_i = 9
    for name in ranking:
        d = captain_by_participant[name]
        count = d["count"]
        best_caption = "-"
        if d["best_points"] > 0:
            best_caption = f"{d['best_player']} +{d['best_points']}"
        ws.cell(row=row_i, column=1).value = name
        ws.cell(row=row_i, column=2).value = count
        ws.cell(row=row_i, column=3).value = d["success"]
        ws.cell(row=row_i, column=4).value = d["points"]
        ws.cell(row=row_i, column=5).value = best_caption
        ws.cell(row=row_i, column=6).value = safe_pct(d["success"], count)
        row_i += 1

    style_table(ws, 8, row_i - 1, 1, 6, header_row=8)

    start2 = row_i + 3
    section_title(ws, start2, 1, "أكثر اللاعبين اختيارًا كابتن", 4, HEADER)
    for col, h in enumerate(["اللاعب", "مرات كابتن", "نقاط كابتن", "متوسط نقاط كابتن"], start=1):
        ws.cell(row=start2 + 1, column=col).value = h

    r = start2 + 2
    for player, count in captain_player_count.most_common():
        pts = captain_player_points[player]
        ws.cell(row=r, column=1).value = player
        ws.cell(row=r, column=2).value = count
        ws.cell(row=r, column=3).value = pts
        ws.cell(row=r, column=4).value = safe_avg(pts, count)
        r += 1

    style_table(ws, start2 + 1, max(start2 + 1, r - 1), 1, 4, header_row=start2 + 1)
    if row_i > 9:
        add_bar_chart(ws, "نقاط الكابتن الإضافية حسب المشارك", 4, 8, 9, row_i - 1, 1, f"A{row_i + 3}", width=13, height=8, chart_type="bar")
    if r > start2 + 2:
        add_bar_chart(ws, "مرات اختيار اللاعب كابتن", 2, start2 + 1, start2 + 2, min(r - 1, start2 + 16), 1, f"H{row_i + 3}", width=13, height=8, chart_type="col")

    # -------------------- 6) تحليل الحراس ✅ بدون تكرار الكلين شيت --------------------
    ws = wb.create_sheet("تحليل الحراس")
    sheet_setup(ws, "تحليل الحراس")

    top_keeper_points = best_keeper
    top_keeper_choice = stats["keeper_choice_count"].most_common(1)[0][0] if stats["keeper_choice_count"] else "-"
    best_keeper_rate = "-"
    if keeper_rounds:
        best_keeper_rate = max(keeper_rounds.keys(), key=lambda k: (len(keeper_clean_rounds[k]) / len(keeper_rounds[k]) if keeper_rounds[k] else 0, len(keeper_clean_rounds[k])))

    card(ws, "A3:C5", "🧤 أفضل حارس نقاط", f"{top_keeper_points}\n{stats['keeper_points_by_keeper'].get(top_keeper_points, 0)} نقطة للمشاركين", BLUE)
    card(ws, "D3:F5", "👥 أكثر حارس اختيارًا", f"{top_keeper_choice}\n{stats['keeper_choice_count'].get(top_keeper_choice, 0)} اختيار", PURPLE)
    card(ws, "G3:I5", "🎯 أعلى نسبة نجاح", f"{best_keeper_rate}\n{safe_pct(len(keeper_clean_rounds[best_keeper_rate]), len(keeper_rounds[best_keeper_rate])) if best_keeper_rate != '-' else '0%'}", GREEN)

    section_title(ws, 7, 1, "ملخص الحراس في البطولة", 8, HEADER)
    headers = ["الحارس", "إجمالي الاختيارات", "عدد المشاركين الذين اختاروه", "مرات الكلين شيت", "إجمالي نقاط المشاركين منه", "نسبة النجاح"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=8, column=col).value = h

    r = 9
    keeper_list = sorted(keeper_rounds.keys(), key=lambda k: (stats["keeper_points_by_keeper"][k], stats["keeper_choice_count"][k]), reverse=True)
    for keeper in keeper_list:
        rounds = sorted(keeper_rounds[keeper])
        clean_count = len(keeper_clean_rounds[keeper])
        ws.cell(row=r, column=1).value = keeper
        ws.cell(row=r, column=2).value = stats["keeper_choice_count"][keeper]
        ws.cell(row=r, column=3).value = len(keeper_participants[keeper])
        ws.cell(row=r, column=4).value = clean_count
        ws.cell(row=r, column=5).value = stats["keeper_points_by_keeper"][keeper]
        ws.cell(row=r, column=6).value = safe_pct(clean_count, len(rounds))
        r += 1

    style_table(ws, 8, max(8, r - 1), 1, 6, header_row=8)

    start2 = r + 3
    section_title(ws, start2, 1, "تفصيل الحراس حسب الجولة", 6, HEADER)
    headers = ["الجولة", "الحارس", "عدد من اختاروه", "كلين شيت؟", "نقاط الحارس الأساسية", "إجمالي نقاط المشاركين"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=start2 + 1, column=col).value = h

    rr = start2 + 2
    for (day, keeper), d in sorted(keeper_round_detail.items(), key=lambda x: (x[0][0], x[0][1])):
        ws.cell(row=rr, column=1).value = day
        ws.cell(row=rr, column=2).value = keeper
        ws.cell(row=rr, column=3).value = d["selected"]
        ws.cell(row=rr, column=4).value = "نعم" if d["clean"] else "لا"
        ws.cell(row=rr, column=5).value = 5 if d["clean"] else 0
        ws.cell(row=rr, column=6).value = d["total_points"]
        rr += 1

    style_table(ws, start2 + 1, max(start2 + 1, rr - 1), 1, 6, header_row=start2 + 1)
    if r > 9:
        add_bar_chart(ws, "إجمالي نقاط المشاركين من الحراس", 5, 8, 9, min(r - 1, 23), 1, f"A{rr + 3}", width=13, height=8, chart_type="bar")

    # -------------------- 7) تحليل اللاعبين ✅ تأثير اللاعبين --------------------
    ws = wb.create_sheet("تحليل اللاعبين")
    sheet_setup(ws, "تحليل اللاعبين")

    if top_round_player:
        (top_day, top_player), top_info = top_round_player
        card(ws, "A3:C6", "🔥 أكثر لاعب وزّع نقاط على المشاركين في جولة واحدة", f"{top_player} — اليوم {top_day}\nوزّع {top_info['distributed']} نقطة على المشاركين في هذه الجولة", RED)
    else:
        card(ws, "A3:C6", "🔥 أكثر لاعب وزّع نقاط على المشاركين في جولة واحدة", "-", RED)

    card(ws, "D3:F6", "⚽ أعلى لاعب نقاط فعلية في البطولة", f"{best_actual_player}\n{player_actual_total.get(best_actual_player, 0)} نقطة", BLUE)
    card(ws, "G3:I6", "👥 أكثر لاعب تم اختياره في البطولة", f"{most_selected_player}\n{player_choice_count.get(most_selected_player, 0)} اختيار", PURPLE)
    card(ws, "J3:L6", "👑 أكثر لاعب تم اختياره كابتن في البطولة", f"{most_captained_player}\n{player_captain_count.get(most_captained_player, 0)} مرة", GOLD)

    section_title(ws, 8, 1, "تفصيل تأثير اللاعبين على المشاركين في كل جولة", 6, HEADER)
    headers = ["الجولة", "اللاعب", "نقاط اللاعب في الجولة", "عدد اللي اختاروه", "عدد اللي كبتنوه", "النقاط التي وزعها على المشاركين في هذه الجولة"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=9, column=col).value = h

    r = 10
    round_rows = sorted(player_round_detail.items(), key=lambda x: (x[0][0], x[1]["distributed"]), reverse=False)
    round_rows = sorted(round_rows, key=lambda x: (x[0][0], -x[1]["distributed"], x[0][1]))
    for (day, player), d in round_rows:
        ws.cell(row=r, column=1).value = day
        ws.cell(row=r, column=2).value = player
        ws.cell(row=r, column=3).value = d["base_points"]
        ws.cell(row=r, column=4).value = d["selected"]
        ws.cell(row=r, column=5).value = d["captains"]
        ws.cell(row=r, column=6).value = d["distributed"]
        r += 1

    style_table(ws, 9, max(9, r - 1), 1, 6, header_row=9)

    start2 = r + 3
    section_title(ws, start2, 1, "ملخص اللاعبين في البطولة", 6, HEADER)
    headers = ["اللاعب", "إجمالي الاختيارات", "عدد المشاركين الذين اختاروه", "مرات كابتن", "نقاط اللاعب الفعلية في البطولة", "إجمالي النقاط التي أفاد بها المشاركين في البطولة"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=start2 + 1, column=col).value = h

    rr = start2 + 2
    player_list = sorted(player_choice_count.keys(), key=lambda p: (player_total_distributed[p], player_actual_total[p], player_choice_count[p]), reverse=True)
    for player in player_list:
        ws.cell(row=rr, column=1).value = player
        ws.cell(row=rr, column=2).value = player_choice_count[player]
        ws.cell(row=rr, column=3).value = len(player_participants[player])
        ws.cell(row=rr, column=4).value = player_captain_count[player]
        ws.cell(row=rr, column=5).value = player_actual_total[player]
        ws.cell(row=rr, column=6).value = player_total_distributed[player]
        rr += 1

    style_table(ws, start2 + 1, max(start2 + 1, rr - 1), 1, 6, header_row=start2 + 1)
    if r > 10:
        add_bar_chart(ws, "أعلى تأثير للاعب في جولة", 6, 9, 10, min(r - 1, 24), 2, f"A{rr + 4}", width=14, height=8, chart_type="bar")
    if rr > start2 + 2:
        add_bar_chart(ws, "إجمالي تأثير اللاعبين في البطولة", 6, start2 + 1, start2 + 2, min(rr - 1, start2 + 16), 1, f"H{rr + 4}", width=14, height=8, chart_type="bar")

    # -------------------- 8) تفصيل اختيارات اللاعبين ✅ حسب الجولة --------------------
    ws = wb.create_sheet("تفصيل اختيارات اللاعبين")
    sheet_setup(ws, "تفصيل اختيارات اللاعبين")

    headers = ["الجولة", "المشارك", "اللاعب", "كابتن؟", "نقاط اللاعب", "نقاط الكابتن", "الإجمالي", "مصدر النقاط"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = h

    round_fills = ["EAF1FF", "E7F8F1", "EFE7FF", "FFF6D6", "E8F7FF", "FFEAF4", "FFF0E0", "F3F4F6", "E6FFFA", "F5F0FF"]
    r = 3
    for row in sorted(selection_detail_rows, key=lambda x: (x[0], x[1], 0 if "نعم" in str(x[3]) else 1, x[2])):
        day = int(row[0]) if str(row[0]).isdigit() else 0
        fill = PatternFill("solid", fgColor=round_fills[(day - 1) % len(round_fills)] if day else "FFFFFF")
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col)
            cell.value = val
            cell.fill = fill
        if "نعم" in str(row[3]):
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=L_GOLD)
            ws.cell(row=r, column=4).font = Font(bold=True, color=TEXT)
        if row[6] and row[6] > 0:
            ws.cell(row=r, column=7).font = Font(bold=True, color=NAVY)
        r += 1

    style_table(ws, 2, max(2, r - 1), 1, 8, header_row=2)
    # نعيد تطبيق ألوان الجولات بعد تنسيق الجدول
    for rr2 in range(3, r):
        day = ws.cell(row=rr2, column=1).value
        try:
            day_int = int(day)
            fill = PatternFill("solid", fgColor=round_fills[(day_int - 1) % len(round_fills)])
            for c in range(1, 9):
                ws.cell(row=rr2, column=c).fill = fill
        except Exception:
            pass
        if "نعم" in str(ws.cell(row=rr2, column=4).value):
            ws.cell(row=rr2, column=4).fill = PatternFill("solid", fgColor=L_GOLD)
            ws.cell(row=rr2, column=4).font = Font(bold=True, color=TEXT)
        if score_to_int(ws.cell(row=rr2, column=7).value) > 0:
            ws.cell(row=rr2, column=7).font = Font(bold=True, color=NAVY)

    # -------------------- 9) سجل الأساطير ✅ --------------------
    ws = wb.create_sheet("سجل الأساطير")
    sheet_setup(ws, "سجل الأساطير")

    headers = ["الجولة", "أسطورة اليوم", "عدد الأساطير", "أعلى نقاط"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col).value = h

    r = 3
    for day in days:
        info = day_summary[day]
        winners = info["winners"]
        ws.cell(row=r, column=1).value = f"اليوم {day}"
        ws.cell(row=r, column=2).value = " + ".join(winners) if winners else "-"
        ws.cell(row=r, column=3).value = info["legends_count"]
        ws.cell(row=r, column=4).value = info["max_score"]
        r += 1

    style_table(ws, 2, max(2, r - 1), 1, 4, header_row=2)
    ws.column_dimensions["B"].width = 42
    if r > 3:
        add_bar_chart(ws, "أعلى نقاط أسطورة اليوم", 4, 2, 3, r - 1, 1, "F3", width=13, height=7, chart_type="col")

    # تنسيق عام نهائي
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            current_width = ws.column_dimensions[letter].width or 12
            ws.column_dimensions[letter].width = max(min(current_width, 42), 14)
        for row in range(1, ws.max_row + 1):
            if not ws.row_dimensions[row].height:
                ws.row_dimensions[row].height = 22

    # عروض خاصة
    if "تحليل اللاعبين" in wb.sheetnames:
        ws = wb["تحليل اللاعبين"]
        ws.column_dimensions["F"].width = 42
    if "تفصيل اختيارات اللاعبين" in wb.sheetnames:
        ws = wb["تفصيل اختيارات اللاعبين"]
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["H"].width = 26
    if "تحليل الأيام" in wb.sheetnames:
        ws = wb["تحليل الأيام"]
        ws.column_dimensions["H"].width = 36

    wb.save(file_name)
    return file_name, stats


# ============================================================
# أوامر إضافية لفانتزي المصيف
# أضف هذا القسم قبل: async def dashboard
# ============================================================

PENDING_IMPORTS = {}
LAST_UPLOADED_FILES = {}
INVALID_SELECTIONS_GLOBAL = {"", "لم يشارك", "تم حجب المشاركة / تأخير"}

# استبدل دالة is_no_participation القديمة بهذا التعريف
# أو اترك هذا التعريف بعد القديمة عشان يغطيها.
def is_no_participation(value):
    value = normalize_name(value)
    return value in INVALID_SELECTIONS_GLOBAL


def load_workbook_safely(path, data_only=True):
    """يفتح ملفات الإكسل حتى لو فيها مشكلة styles مثل family val > 14."""
    import zipfile
    import tempfile
    import re as _re

    try:
        return load_workbook(path, data_only=data_only)
    except Exception:
        fixed_path = os.path.join(tempfile.gettempdir(), f"fixed_{os.path.basename(path)}")
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(fixed_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    text = data.decode("utf-8", errors="ignore")
                    text = _re.sub(r'<family val="(?:1[5-9]|[2-9][0-9]+)"\s*/>', '<family val="2"/>', text)
                    data = text.encode("utf-8")
                zout.writestr(item, data)
        return load_workbook(fixed_path, data_only=data_only)


def cell_text(value):
    return normalize_name(value)


def cell_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def parse_import_excel(path):
    """
    يستورد ملف الإكسل كامل مرة وحدة.
    يدعم صفحات باسم: يوم 1، يوم 2، ...
    ويسحب:
    - جدول النتائج من الأعمدة A:B:C
    - جدول المشاركين من عند عمود "المشارك"
    ويتجاهل الأيام الفاضية بالكامل حتى لا ينشئ أيام وهمية.
    """
    wb = load_workbook_safely(path, data_only=True)
    imported = {}

    TRUE_VALUES = {"نعم", "yes", "Yes", "YES", "صح", "✓", "✅", "1", "true", "True"}

    for ws in wb.worksheets:
        m = re.search(r"يوم\s*(\d+)", str(ws.title))
        if not m:
            continue

        day = int(m.group(1))

        # نبحث عن خانة "المشارك" في أول الصفوف
        header_row = None
        participant_col = None
        for r in range(1, min(ws.max_row, 25) + 1):
            for c in range(1, min(ws.max_column, 25) + 1):
                if cell_text(ws.cell(r, c).value) == "المشارك":
                    header_row = r
                    participant_col = c
                    break
            if header_row:
                break

        if not header_row or not participant_col:
            continue

        lineups = {}
        active_found = False

        # المشاركون: المشارك، الحارس، اللاعب 1، اللاعب 2، اللاعب 3، الكابتن
        for r in range(header_row + 1, ws.max_row + 1):
            participant = cell_text(ws.cell(r, participant_col).value)
            if not participant or participant not in PARTICIPANTS:
                continue

            values = [cell_text(ws.cell(r, participant_col + offset).value) for offset in range(1, 6)]

            # لو الصف فاضي أو كله عدم مشاركة
            if all(is_no_participation(v) for v in values):
                values = ["لم يشارك"] * 5
            else:
                active_found = True

            lineups[participant] = values

        # النتائج: غالبًا في A:B:C = اسم اللاعب | عدد الأهداف | كلين شيت؟
        goals_count = defaultdict(int)
        clean_sheets = []

        for r in range(header_row + 1, ws.max_row + 1):
            name = cell_text(ws.cell(r, 1).value)
            goals_raw = ws.cell(r, 2).value
            clean_raw = cell_text(ws.cell(r, 3).value)

            if not name or is_no_participation(name):
                continue

            goals = cell_int(goals_raw, 0)
            if goals > 0:
                goals_count[name] += goals

            if clean_raw in TRUE_VALUES:
                if name not in clean_sheets:
                    clean_sheets.append(name)

        # لا نستورد الأيام الفاضية بالكامل، مثل يوم 7 وما بعده إذا ما انلعبت
        if not active_found and not goals_count and not clean_sheets:
            continue

        if lineups:
            imported[day] = {
                "lineups": lineups,
                "goals_count": dict(goals_count),
                "clean_sheets": clean_sheets,
            }

    return imported

def import_summary_text(imported):
    lines = ["تم قراءة الملف ✅", "", "الأيام الموجودة:"]
    if not imported:
        lines.append("ما لقيت أيام قابلة للاستيراد.")
        return "\n".join(lines)

    for day in sorted(imported):
        data = imported[day]
        participants = len(data["lineups"])
        has_results = bool(data["goals_count"] or data["clean_sheets"])
        status = "نتائج موجودة" if has_results else "بدون نتائج"
        exists = " — موجود مسبقًا وسيتم استبداله" if os.path.exists(excel_file(day)) else ""
        lines.append(f"اليوم {day}: {participants} مشارك + {status}{exists}")

    lines.append("")
    lines.append("للاعتماد اكتب:")
    lines.append("/اعتماد_استيراد")
    lines.append("")
    lines.append("للإلغاء اكتب:")
    lines.append("/إلغاء_استيراد")
    return "\n".join(lines)


async def _download_document_to_folder(update: Update, context: ContextTypes.DEFAULT_TYPE, folder="uploads"):
    """يحفظ أي ملف مرسل ويرجع المسار المحلي."""
    document = update.message.document
    if not document:
        return None, None

    filename = document.file_name or "uploaded_file"
    os.makedirs(folder, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_.\-\u0600-\u06FF]+", "_", filename)
    local_path = os.path.join(folder, f"{update.effective_chat.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}")

    tg_file = await context.bot.get_file(document.file_id)
    await tg_file.download_to_drive(local_path)
    return local_path, filename


async def remember_last_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """يحفظ آخر ملف Excel أو ZIP أرسله المستخدم، عشان نقدر نستخدمه بعدين بدون تعليق."""
    local_path, filename = await _download_document_to_folder(update, context, "uploads")
    if not local_path:
        return

    chat_id = update.effective_chat.id
    LAST_UPLOADED_FILES.setdefault(chat_id, {})
    lower = (filename or "").lower()
    caption = (update.message.caption or "").strip()

    if lower.endswith((".xlsx", ".xlsm")):
        LAST_UPLOADED_FILES[chat_id]["excel"] = local_path
        if caption.startswith("/استيراد_ملف"):
            await _run_import_from_path(update, context, local_path)
        else:
            await update.message.reply_text(
                "وصل ملف الإكسل ✅\n"
                "اكتب الآن:\n"
                "/استيراد_ملف"
            )
        return

    if lower.endswith(".zip"):
        LAST_UPLOADED_FILES[chat_id]["zip"] = local_path
        if caption.startswith("/استرجاع_نسخة"):
            await _run_restore_from_zip_path(update, context, local_path)
        else:
            await update.message.reply_text(
                "وصل ملف ZIP ✅\n"
                "للاسترجاع اكتب الآن:\n"
                "/استرجاع_نسخة"
            )
        return

    await update.message.reply_text("وصل الملف، لكن أحتاج Excel أو ZIP فقط.")


async def _run_import_from_path(update: Update, context: ContextTypes.DEFAULT_TYPE, local_path):
    try:
        imported = parse_import_excel(local_path)
    except Exception as e:
        await update.message.reply_text(f"صار خطأ أثناء قراءة الإكسل ❌\n{e}")
        return

    if not imported:
        await update.message.reply_text("ما قدرت أستخرج أيام من الملف. تأكد أن الصفحات باسم: يوم 1، يوم 2 ...")
        return

    chat_id = update.effective_chat.id
    PENDING_IMPORTS[chat_id] = {"path": local_path, "data": imported}
    await update.message.reply_text(import_summary_text(imported))


async def import_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    يدعم طريقتين:
    1) إرسال ملف الإكسل بتعليق /استيراد_ملف
    2) إرسال ملف الإكسل لحاله، ثم كتابة /استيراد_ملف
    """
    document = update.message.document
    chat_id = update.effective_chat.id

    if document:
        filename = document.file_name or "import.xlsx"
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            await update.message.reply_text("الملف لازم يكون Excel بصيغة .xlsx أو .xlsm")
            return
        local_path, _ = await _download_document_to_folder(update, context, "imports")
        LAST_UPLOADED_FILES.setdefault(chat_id, {})["excel"] = local_path
        await _run_import_from_path(update, context, local_path)
        return

    local_path = LAST_UPLOADED_FILES.get(chat_id, {}).get("excel")
    if not local_path or not os.path.exists(local_path):
        await update.message.reply_text(
            "ما لقيت ملف Excel محفوظ.\n"
            "أرسل ملف الإكسل لحاله أولًا، وبعدها اكتب:\n"
            "/استيراد_ملف"
        )
        return

    await _run_import_from_path(update, context, local_path)


async def approve_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    pending = PENDING_IMPORTS.get(chat_id)
    if not pending:
        await update.message.reply_text("ما فيه استيراد معلق. أرسل ملف الإكسل مع /استيراد_ملف أولًا.")
        return

    imported = pending["data"]
    files_to_backup = [excel_file(day) for day in imported.keys() if os.path.exists(excel_file(day))]
    if os.path.exists(LOCKED_FILE):
        files_to_backup.append(LOCKED_FILE)
    if files_to_backup:
        backup_files("before_import_excel", files=files_to_backup)

    locked = load_locked_days()
    saved_days = []
    warnings = []

    for day in sorted(imported):
        data = imported[day]
        file_name, missing = update_day_data(str(day), data["lineups"])
        if missing:
            warnings.extend([f"اليوم {day}: مشارك غير موجود بالقائمة: {m}" for m in missing])
        calculate_points(str(day), data["goals_count"], data["clean_sheets"])
        locked.add(str(day))
        saved_days.append(day)

    save_locked_days(locked)
    PENDING_IMPORTS.pop(chat_id, None)

    msg = [
        "تم اعتماد الاستيراد ✅",
        f"الأيام المحفوظة: {', '.join(map(str, saved_days))}",
        "تم حساب النتائج وقفل الأيام المستوردة.",
    ]
    if warnings:
        msg.append("\nتنبيهات:")
        msg.extend(warnings[:20])
    await update.message.reply_text("\n".join(msg))


async def cancel_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id in PENDING_IMPORTS:
        PENDING_IMPORTS.pop(chat_id, None)
        await update.message.reply_text("تم إلغاء الاستيراد ✅")
    else:
        await update.message.reply_text("ما فيه استيراد معلق.")


async def backup_zip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    import zipfile

    files = [f for f in os.listdir(".") if re.match(r"fantasy_day_\d+\.xlsx$", f)]
    if os.path.exists(LOCKED_FILE):
        files.append(LOCKED_FILE)

    if not files:
        await update.message.reply_text("ما فيه ملفات أيام عشان أسوي نسخة احتياطية.")
        return

    zip_name = f"fantasy_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    with zipfile.ZipFile(zip_name, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(files):
            z.write(f)

    with open(zip_name, "rb") as f:
        await update.message.reply_document(document=f, filename=zip_name, caption="نسخة احتياطية لأيام الفانتزي ✅")


async def _run_restore_from_zip_path(update: Update, context: ContextTypes.DEFAULT_TYPE, local_path):
    import zipfile

    backup_files("before_restore_zip")
    restored = []
    try:
        with zipfile.ZipFile(local_path, "r") as z:
            for name in z.namelist():
                base = os.path.basename(name)
                if re.match(r"fantasy_day_\d+\.xlsx$", base) or base == LOCKED_FILE:
                    with z.open(name) as src, open(base, "wb") as dst:
                        dst.write(src.read())
                    restored.append(base)
    except Exception as e:
        await update.message.reply_text(f"صار خطأ في استرجاع النسخة ❌\n{e}")
        return

    if not restored:
        await update.message.reply_text("ملف ZIP ما فيه ملفات أيام صالحة.")
        return

    await update.message.reply_text("تم استرجاع النسخة ✅\n" + "\n".join(sorted(restored)))


async def restore_backup_zip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """يدعم إرسال ZIP لحاله ثم /استرجاع_نسخة، أو ZIP بتعليق."""
    document = update.message.document
    chat_id = update.effective_chat.id

    if document:
        filename = document.file_name or "backup.zip"
        if not filename.lower().endswith(".zip"):
            await update.message.reply_text("الملف لازم يكون ZIP.")
            return
        local_path, _ = await _download_document_to_folder(update, context, "restore_uploads")
        LAST_UPLOADED_FILES.setdefault(chat_id, {})["zip"] = local_path
        await _run_restore_from_zip_path(update, context, local_path)
        return

    local_path = LAST_UPLOADED_FILES.get(chat_id, {}).get("zip")
    if not local_path or not os.path.exists(local_path):
        await update.message.reply_text(
            "ما لقيت ملف ZIP محفوظ.\n"
            "أرسل ملف ZIP لحاله أولًا، وبعدها اكتب:\n"
            "/استرجاع_نسخة"
        )
        return

    await _run_restore_from_zip_path(update, context, local_path)


async def clean_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """ينظف القفل من أيام ما لها ملف فعلي."""
    locked = load_locked_days()
    existing = {str(d) for d in get_existing_days(1, 99)}
    cleaned = {d for d in locked if d in existing}
    removed = sorted(locked - cleaned, key=lambda x: int(x))
    save_locked_days(cleaned)
    if removed:
        await update.message.reply_text("تم تنظيف الأيام الوهمية ✅\nأزيلت من القفل: " + "، ".join(removed))
    else:
        await update.message.reply_text("ما فيه أيام وهمية. كل شيء تمام ✅")



# -------------------- أوامر تيليجرام --------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n"
        "/نتائج 5\n"
        "/احصائيات\n"
        "/احصائيات 1 6\n"
        "/ترتيب_نص\n\n"
        "أوامر الفحص:\n"
        "/الأيام\n"
        "/فحص 5\n"
        "/مشاركين 5\n"
        "/اسطورة 5\n"
        "/مقارنة 4 5\n\n"
        "أوامر الاستيراد والنسخ:\n"
        "/استيراد_ملف — أرسل ملف الإكسل لحاله ثم اكتب الأمر\n"
        "/اعتماد_استيراد\n"
        "/إلغاء_استيراد\n"
        "/نسخة_احتياطية\n"
        "/استرجاع_نسخة — أرسل ملف ZIP لحاله ثم اكتب الأمر\n"
        "/تنظيف_الأيام\n\n"
        "أوامر الأمان:\n"
        "/مسح_نتائج 5\n"
        "/مسح_يوم 5\n"
        "/مسح_الكل تأكيد\n"
        "/استرجاع_آخر\n"
        "/قفل_يوم 5\n"
        "/فتح_يوم 5"
    )




# ============================================================
# V19 — كأس المصيف اليدوي + تصميم مواجهات الكأس
# ============================================================

CUP_ROUNDS = [
    ("r12", "دور 12"),
    ("qf", "ربع النهائي"),
    ("sf", "نصف النهائي"),
    ("final", "النهائي"),
]

def load_cup_state():
    if not os.path.exists(CUP_FILE):
        return {"active": False}
    try:
        with open(CUP_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {"active": False}
        return data
    except Exception:
        return {"active": False}

def save_cup_state(state):
    with open(CUP_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def cup_round_key_by_offset(offset):
    if 0 <= offset < len(CUP_ROUNDS):
        return CUP_ROUNDS[offset][0]
    return None

def cup_round_name(round_key):
    for key, name in CUP_ROUNDS:
        if key == round_key:
            return name
    return str(round_key or "")

def cup_next_round(round_key):
    keys = [x[0] for x in CUP_ROUNDS]
    if round_key not in keys:
        return None
    idx = keys.index(round_key)
    return keys[idx + 1] if idx + 1 < len(keys) else None

def cup_previous_ranking_before(start_day):
    try:
        prev_end = max(0, int(start_day) - 1)
    except Exception:
        prev_end = 0

    if prev_end >= 1:
        stats = collect_stats(1, prev_end)
        ranking = [n for n in stats.get("ranking", []) if n in PARTICIPANTS]
        missing = [n for n in PARTICIPANTS if n not in ranking]
        return ranking + missing
    return PARTICIPANTS[:]

def cup_seed_number(state, participant):
    try:
        return int(state.get("seed_by", {}).get(participant, 99))
    except Exception:
        return 99

def cup_make_match(a, b, match_id=None):
    return {
        "id": match_id or "",
        "a": a,
        "b": b,
        "score_a": None,
        "score_b": None,
        "active_a": None,
        "active_b": None,
        "winner": None,
        "note": "",
    }

def cup_initial_state(start_day):
    seeds = cup_previous_ranking_before(start_day)[:12]
    while len(seeds) < 12:
        seeds.append(f"مشارك {len(seeds)+1}")

    seed_by = {name: i + 1 for i, name in enumerate(seeds)}

    # دور 12: أول 4 عندهم راحة، والباقي يلعبون
    r12_matches = [
        cup_make_match(seeds[4], seeds[11], "R12-1"),  # 5 vs 12
        cup_make_match(seeds[7], seeds[8], "R12-2"),   # 8 vs 9
        cup_make_match(seeds[5], seeds[10], "R12-3"),  # 6 vs 11
        cup_make_match(seeds[6], seeds[9], "R12-4"),   # 7 vs 10
    ]

    return {
        "active": True,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "start_day": int(start_day),
        "end_day": int(start_day) + 3,
        "seeds": seeds,
        "seed_by": seed_by,
        "champion": None,
        "rounds": {
            "r12": {"day": int(start_day), "name": "دور 12", "matches": r12_matches, "completed": False},
            "qf": {"day": int(start_day) + 1, "name": "ربع النهائي", "matches": [], "completed": False},
            "sf": {"day": int(start_day) + 2, "name": "نصف النهائي", "matches": [], "completed": False},
            "final": {"day": int(start_day) + 3, "name": "النهائي", "matches": [], "completed": False},
        },
    }

def cup_scores_for_day(day):
    rows = read_day_rows(day)
    scores = {}
    active = {}
    for row in rows:
        name = row.get("participant")
        if not name:
            continue
        scores[name] = int(row.get("total", 0) or 0)
        active[name] = bool(row.get("participated"))
    for name in PARTICIPANTS:
        scores.setdefault(name, 0)
        active.setdefault(name, False)
    return scores, active

def cup_decide_winner(state, match, scores, active):
    a = match.get("a", "")
    b = match.get("b", "")
    sa = int(scores.get(a, 0) or 0)
    sb = int(scores.get(b, 0) or 0)
    aa = bool(active.get(a, False))
    ab = bool(active.get(b, False))

    # القواعد المعتمدة:
    # إذا واحد شارك والثاني لا: المشارك يتأهل حتى لو صفر
    # إذا الاثنين شاركوا: الأعلى نقاط يتأهل
    # إذا الاثنين لم يشاركوا: الأعلى بالتصنيف/البذر يتأهل
    if aa and not ab:
        winner = a
        note = "تأهل لأنه شارك"
    elif ab and not aa:
        winner = b
        note = "تأهل لأنه شارك"
    elif aa and ab:
        if sa > sb:
            winner = a
            note = "فاز بالنقاط"
        elif sb > sa:
            winner = b
            note = "فاز بالنقاط"
        else:
            # عند التعادل: الأعلى بذرًا يتأهل
            seed_a = cup_seed_number(state, a)
            seed_b = cup_seed_number(state, b)
            winner = a if seed_a <= seed_b else b
            note = "تعادل — تأهل الأعلى تصنيفًا"
    else:
        seed_a = cup_seed_number(state, a)
        seed_b = cup_seed_number(state, b)
        winner = a if seed_a <= seed_b else b
        note = "لم يشارك الطرفان — تأهل الأعلى تصنيفًا"

    match["score_a"] = sa
    match["score_b"] = sb
    match["active_a"] = aa
    match["active_b"] = ab
    match["winner"] = winner
    match["note"] = note
    return winner

def cup_build_next_round_matches(state, completed_round_key):
    seeds = state.get("seeds", [])
    rounds = state.get("rounds", {})

    if completed_round_key == "r12":
        r12 = rounds.get("r12", {}).get("matches", [])
        if len(r12) < 4 or len(seeds) < 12:
            return []
        # ترتيب ربع النهائي
        # Seed 1 vs Winner 8/9
        # Seed 4 vs Winner 5/12
        # Seed 2 vs Winner 7/10
        # Seed 3 vs Winner 6/11
        return [
            cup_make_match(seeds[0], r12[1].get("winner"), "QF-1"),
            cup_make_match(seeds[3], r12[0].get("winner"), "QF-2"),
            cup_make_match(seeds[1], r12[3].get("winner"), "QF-3"),
            cup_make_match(seeds[2], r12[2].get("winner"), "QF-4"),
        ]

    if completed_round_key == "qf":
        qf = rounds.get("qf", {}).get("matches", [])
        if len(qf) < 4:
            return []
        return [
            cup_make_match(qf[0].get("winner"), qf[1].get("winner"), "SF-1"),
            cup_make_match(qf[2].get("winner"), qf[3].get("winner"), "SF-2"),
        ]

    if completed_round_key == "sf":
        sf = rounds.get("sf", {}).get("matches", [])
        if len(sf) < 2:
            return []
        return [
            cup_make_match(sf[0].get("winner"), sf[1].get("winner"), "FINAL"),
        ]

    return []

def cup_process_day(day):
    state = load_cup_state()
    if not state.get("active"):
        return state, None, "لا يوجد كأس مفعّلة."

    try:
        day_i = int(day)
        start_day = int(state.get("start_day"))
    except Exception:
        return state, None, "رقم اليوم غير صحيح."

    if day_i < start_day or day_i > int(state.get("end_day", start_day + 3)):
        return state, None, "اليوم خارج فترة الكأس الحالية."

    round_key = cup_round_key_by_offset(day_i - start_day)
    if not round_key:
        return state, None, "اليوم خارج مراحل الكأس."

    round_data = state.get("rounds", {}).get(round_key, {})
    matches = round_data.get("matches", [])
    if not matches:
        return state, None, f"لا توجد مواجهات في {cup_round_name(round_key)}."

    if round_data.get("completed"):
        return state, round_key, "هذه المرحلة محسوبة مسبقًا."

    scores, active = cup_scores_for_day(day_i)
    for match in matches:
        cup_decide_winner(state, match, scores, active)

    round_data["completed"] = True
    state["rounds"][round_key] = round_data

    next_key = cup_next_round(round_key)
    if next_key:
        next_matches = cup_build_next_round_matches(state, round_key)
        state["rounds"][next_key]["matches"] = next_matches
    else:
        # النهائي انتهى
        winner = matches[0].get("winner") if matches else None
        state["champion"] = winner
        state["active"] = False

    save_cup_state(state)
    return state, round_key, "تم حساب مرحلة الكأس ✅"

def cup_pending_round_key(state):
    if not state.get("rounds"):
        return None
    for key, _name in CUP_ROUNDS:
        rd = state["rounds"].get(key, {})
        if rd.get("matches") and not rd.get("completed"):
            return key
    if state.get("champion"):
        return "champion"
    return None

def cup_results_text(state):
    if not state or not state.get("rounds"):
        return "لا توجد كأس محفوظة."
    lines = [
        "🏆 كأس المصيف",
        f"الفترة: اليوم {state.get('start_day')} إلى اليوم {state.get('end_day')}",
        "",
    ]

    for key, name in CUP_ROUNDS:
        rd = state.get("rounds", {}).get(key, {})
        matches = rd.get("matches", [])
        if not matches:
            continue
        lines.append(f"📌 {name} — اليوم {rd.get('day')}")
        for m in matches:
            a, b = m.get("a", ""), m.get("b", "")
            if m.get("winner"):
                lines.append(f"- {a} {m.get('score_a', 0)} × {m.get('score_b', 0)} {b} | المتأهل: {m.get('winner')}")
            else:
                lines.append(f"- {a} × {b}")
        lines.append("")

    if state.get("champion"):
        lines.append(f"🏆 بطل كأس المصيف: {state.get('champion')}")
    else:
        pending = cup_pending_round_key(state)
        if pending and pending != "champion":
            lines.append(f"القادم: {cup_round_name(pending)}")
    return "\n".join(lines).strip()

def create_cup_matches_image(state, round_key=None, title_suffix=None):
    ensure_generated_dir()
    if round_key is None:
        round_key = cup_pending_round_key(state)
    if not round_key:
        return None

    width = 1200
    if round_key == "champion":
        height = 850
        img, draw = design_canvas(None, width, height, "gold")
        draw_design_header(draw, width, "كأس المصيف 2026", "البطل", img)
        draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#F59E0B")
        champion = state.get("champion", "غير محدد")
        draw_text(draw, (width//2, 430), "🏆", get_font(82), fill="#FDE68A")
        draw_text(draw, (width//2, 535), champion, get_font(58), fill="#FFFFFF", max_width=850)
        draw_text(draw, (width//2, 625), "بطل كأس المصيف", get_font(36), fill="#FDE68A")
        footer_event(draw, width, height)
        path = os.path.join(GENERATED_DIR, "cup_champion.png")
        img.save(path, quality=95)
        return path

    rd = state.get("rounds", {}).get(round_key, {})
    matches = rd.get("matches", [])
    count = max(len(matches), 1)
    row_h, gap, name_size = v16_fit_row_metrics(count, "match")
    if count <= 2:
        row_h += 10
        name_size += 2

    content_h = count * row_h + max(0, count - 1) * gap
    height = max(780, 245 + content_h + 200)
    img, draw = design_canvas(None, width, height, "gold")
    subtitle = title_suffix or f"{cup_round_name(round_key)} — اليوم {rd.get('day', '')}"
    draw_design_header(draw, width, "كأس المصيف 2026", subtitle, img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#F59E0B")

    available_h = (fy2 - fy1) - 70
    y = fy1 + 38 + max(0, (available_h - content_h) // 2)
    for i, m in enumerate(matches, start=1):
        a = m.get("a") or "—"
        b = m.get("b") or "—"
        accent = "#F59E0B" if m.get("winner") else v16_accent(i)
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=28, fill="#0B1020", outline=accent, width=2)
        cy = y + row_h//2

        # بذر/تصنيف اللاعب
        seed_a = cup_seed_number(state, a)
        seed_b = cup_seed_number(state, b)
        rounded_rect(draw, (width-190, cy-34, width-132, cy+34), radius=16, fill="#05070D", outline="#FDE68A80", width=1)
        rounded_rect(draw, (132, cy-34, 190, cy+34), radius=16, fill="#05070D", outline="#FDE68A80", width=1)
        draw_text(draw, (width-161, cy), str(seed_a if seed_a != 99 else "-"), get_font(28), fill="#FDE68A")
        draw_text(draw, (161, cy), str(seed_b if seed_b != 99 else "-"), get_font(28), fill="#FDE68A")

        a_fill = "#FDE68A" if m.get("winner") == a else "#FFFFFF"
        b_fill = "#FDE68A" if m.get("winner") == b else "#FFFFFF"
        draw_text(draw, (width-390, cy), a, get_font(name_size), fill=a_fill, max_width=330)
        draw_text(draw, (390, cy), b, get_font(name_size), fill=b_fill, max_width=330)

        if m.get("winner"):
            score = f"{m.get('score_a', 0)} - {m.get('score_b', 0)}"
            rounded_rect(draw, (width//2-110, cy-34, width//2+110, cy+34), radius=20, fill="#05070D", outline="#FDE68A", width=2)
            draw_text(draw, (width//2, cy), score, get_font(max(30, name_size+4)), fill="#FDE68A")
        else:
            draw_text(draw, (width//2, cy), "×", get_font(max(42, name_size+18)), fill="#FDE68A")

        y += row_h + gap

    footer_event(draw, width, height)
    path = os.path.join(GENERATED_DIR, f"cup_{round_key}.png")
    img.save(path, quality=95)
    return path

def cup_started_caption(state):
    seeds = state.get("seeds", [])
    bye = "، ".join(seeds[:4]) if len(seeds) >= 4 else "لا يوجد"
    return (
        f"تم بدء كأس المصيف 🏆\n"
        f"الفترة: اليوم {state.get('start_day')} إلى اليوم {state.get('end_day')}\n\n"
        f"اليوم {state.get('start_day')}: دور 12\n"
        f"اليوم {int(state.get('start_day'))+1}: ربع النهائي\n"
        f"اليوم {int(state.get('start_day'))+2}: نصف النهائي\n"
        f"اليوم {int(state.get('start_day'))+3}: النهائي\n\n"
        f"راحة دور 12: {bye}"
    )

async def start_cup_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if not nums:
        await update.message.reply_text("اكتبها كذا:\n/بدء_الكاس 5")
        return
    start_day = int(nums[0])
    if start_day < 1:
        await update.message.reply_text("رقم اليوم غير صحيح.")
        return

    old = load_cup_state()
    if old.get("active"):
        await update.message.reply_text(
            "فيه كأس مفعّلة حاليًا.\n"
            "إذا تبي تلغيها اكتب: /الغاء_الكاس"
        )
        return

    backup_files(f"before_start_cup_{start_day}", files=[CUP_FILE] if os.path.exists(CUP_FILE) else [])
    state = cup_initial_state(start_day)
    save_cup_state(state)

    path = create_cup_matches_image(state, "r12", "دور 12")
    if path:
        await send_photo_path(update, path, cup_started_caption(state))
    else:
        await update.message.reply_text(cup_started_caption(state))

async def cup_status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    state = load_cup_state()
    if not state.get("rounds"):
        await update.message.reply_text("لا توجد كأس محفوظة حاليًا.\nلبدء كأس جديد: /بدء_الكاس 5")
        return
    pending = cup_pending_round_key(state)
    status = "مفعّلة ✅" if state.get("active") else "منتهية ✅"
    msg = (
        f"🏆 حالة كأس المصيف\n"
        f"الحالة: {status}\n"
        f"الفترة: اليوم {state.get('start_day')} إلى اليوم {state.get('end_day')}\n"
    )
    if state.get("champion"):
        msg += f"البطل: {state.get('champion')}"
    elif pending:
        msg += f"المرحلة الحالية: {cup_round_name(pending)}"
    await update.message.reply_text(msg)

async def cup_results_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    state = load_cup_state()
    await update.message.reply_text(cup_results_text(state))

async def cancel_cup_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "تأكيد" not in (update.message.text or ""):
        await update.message.reply_text(
            "لإلغاء كأس المصيف الحالية اكتب:\n"
            "/الغاء_الكاس تأكيد\n\n"
            "تنبيه: هذا يلغي البطولة الحالية فقط، ولا يغير نقاط الفانتزي."
        )
        return

    state = load_cup_state()
    if not state.get("rounds"):
        await update.message.reply_text("ما فيه كأس محفوظة أصلًا.")
        return

    backup_files("before_cancel_cup", files=[CUP_FILE] if os.path.exists(CUP_FILE) else [])
    state["active"] = False
    state["cancelled_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_cup_state(state)
    await update.message.reply_text("تم إلغاء كأس المصيف الحالية ✅\nتقدر تبدأ كأس جديد بأمر:\n/بدء_الكاس 7")

async def cup_matches_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    state = load_cup_state()
    if not state.get("rounds"):
        await update.message.reply_text("لا توجد كأس محفوظة حاليًا.\nلبدء كأس جديد: /بدء_الكاس 5")
        return

    pending = cup_pending_round_key(state)
    if not pending:
        await update.message.reply_text(cup_results_text(state))
        return

    path = create_cup_matches_image(state, pending)
    if path:
        caption = cup_results_text(state) if pending == "champion" else f"🏆 مواجهات كأس المصيف — {cup_round_name(pending)}"
        await send_photo_path(update, path, caption)
    else:
        await update.message.reply_text(cup_results_text(state))

async def cup_after_fantasy_results(update: Update, day):
    state = load_cup_state()
    if not state.get("active"):
        return
    try:
        day_i = int(day)
        if day_i < int(state.get("start_day")) or day_i > int(state.get("end_day")):
            return
    except Exception:
        return

    state, round_key, msg = cup_process_day(day_i)
    if not round_key:
        return

    # أرسل نتائج المرحلة النصية
    await update.message.reply_text(f"🏆 كأس المصيف\n{msg}\n\n{cup_results_text(state)}")

    # بعد حساب اليوم: أرسل تصميم المرحلة القادمة، أو البطل إذا انتهى النهائي
    next_key = cup_pending_round_key(state)
    if next_key:
        path = create_cup_matches_image(state, next_key)
        if path:
            if next_key == "champion":
                caption = f"🏆 بطل كأس المصيف: {state.get('champion')}"
            else:
                caption = f"🏆 مواجهات {cup_round_name(next_key)} في كأس المصيف"
            await send_photo_path(update, path, caption)


# ============================================================
# V20 — اعتماد النتائج النهائي + إعادة الكأس من يوم
# ============================================================

def day_has_result_lines(message_text):
    lines = [l.strip() for l in (message_text or "").splitlines()[1:] if l.strip()]
    return bool(lines)

def clear_cup_match_result(match):
    match["score_a"] = None
    match["score_b"] = None
    match["active_a"] = None
    match["active_b"] = None
    match["winner"] = None
    match["note"] = ""
    return match

def reset_cup_from_day(state, from_day):
    if not state.get("rounds"):
        return state, "لا توجد بطولة كأس محفوظة."

    start_day = int(state.get("start_day", 0))
    end_day = int(state.get("end_day", start_day + 3))
    from_day = int(from_day)

    if from_day < start_day or from_day > end_day:
        return state, f"اليوم {from_day} خارج فترة الكأس الحالية."

    offset = from_day - start_day
    keys = [x[0] for x in CUP_ROUNDS]
    reset_key = keys[offset]

    # إذا نعيد من ربع/نصف/نهائي، نحتاج نضمن أن المرحلة مبنية من نتائج المرحلة السابقة.
    if reset_key == "qf":
        state["rounds"]["qf"]["matches"] = cup_build_next_round_matches(state, "r12")
    elif reset_key == "sf":
        state["rounds"]["sf"]["matches"] = cup_build_next_round_matches(state, "qf")
    elif reset_key == "final":
        state["rounds"]["final"]["matches"] = cup_build_next_round_matches(state, "sf")

    # امسح نتائج المرحلة المطلوبة وما بعدها، وامسح مواجهات المراحل التي بعدها
    for idx, key in enumerate(keys):
        rd = state["rounds"].get(key, {})
        if idx < offset:
            continue

        if idx == offset:
            rd["matches"] = [clear_cup_match_result(m) for m in rd.get("matches", [])]
            rd["completed"] = False
        else:
            rd["matches"] = []
            rd["completed"] = False

        state["rounds"][key] = rd

    state["champion"] = None
    state["active"] = True
    state["reset_from_day"] = from_day
    state["reset_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return state, f"تمت إعادة الكأس من اليوم {from_day} ✅"

async def approve_results_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    day = get_day(text)

    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nإذا تبي تعيد اعتماده اكتب أولًا:\n/فتح_يوم {day}")
        return

    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}. أضف التشكيلات أولًا.")
        return

    # لو أرسل نتائج داخل أمر الاعتماد، نعيد الحساب قبل الاعتماد.
    if day_has_result_lines(text):
        goals_count, clean_sheets = parse_results(text)
        goal_missing, clean_missing = validate_results_names(day, goals_count, clean_sheets)
        backup_files(f"before_approve_results_day_{day}", files=[excel_file(day), LOCKED_FILE, CUP_FILE])
        file_name = calculate_points(day, goals_count, clean_sheets)
    else:
        goal_missing, clean_missing = [], []
        file_name = excel_file(day)

    rows = read_day_rows(day)
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    legends_text = "، ".join(winners) if winners else "لا يوجد"

    warnings = []
    if goal_missing:
        warnings.append("⚠️ هدافون غير موجودين في تشكيلات اليوم:\n" + "\n".join(goal_missing))
    if clean_missing:
        warnings.append("⚠️ حراس كلين شيت غير موجودين في تشكيلات اليوم:\n" + "\n".join(clean_missing))

    caption = (
        f"تم اعتماد نتائج اليوم {day} رسميًا ✅\n"
        f"تم قفل اليوم {day} 🔒\n\n"
        f"🏆 أسطورة اليوم: {legends_text} — {max_score} نقطة"
    )
    if warnings:
        caption += "\n\n" + "\n\n".join(warnings)

    # أرسل ملف اليوم المعتمد
    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=caption)

    # احسب الكأس فقط هنا، وليس في /نتائج
    try:
        await cup_after_fantasy_results(update, day)
    except Exception as e:
        await update.message.reply_text(f"تم اعتماد اليوم، لكن تعذر تحديث الكأس ❌\n{e}")

    # اقفل اليوم بعد الاعتماد
    days = load_locked_days()
    days.add(str(day))
    save_locked_days(days)

async def reset_cup_from_day_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if not nums:
        await update.message.reply_text("اكتبها كذا:\n/إعادة_الكاس_من 7")
        return

    from_day = int(nums[0])
    state = load_cup_state()
    if not state.get("rounds"):
        await update.message.reply_text("لا توجد كأس محفوظة حاليًا.\nلبدء كأس جديد: /بدء_الكاس 7")
        return

    backup_files(f"before_reset_cup_from_{from_day}", files=[CUP_FILE] if os.path.exists(CUP_FILE) else [])
    state, msg = reset_cup_from_day(state, from_day)
    save_cup_state(state)

    pending = cup_pending_round_key(state)
    if pending:
        path = create_cup_matches_image(state, pending)
        if path:
            await send_photo_path(update, path, msg + f"\n\nالمرحلة الحالية: {cup_round_name(pending)}")
            return

    await update.message.reply_text(msg)

# -------------------- V18: قراءة النموذج الرسمي الطويل في /اضافه --------------------

def clean_pick_value(value):
    """
    تنظيف بسيط للاسم بدون Alias:
    - يشيل الإيموجي والزخارف
    - يحافظ على النص كما كتبه المستخدم قدر الإمكان
    """
    value = normalize_name(value)
    value = value.replace("：", ":")
    # إزالة رموز شائعة تظهر بعد أسماء اللاعبين
    value = re.sub(r"[👑🐐🔥⚽️✅🧤🏆⭐🌟🥇🥈🥉📋\-–—]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value

def extract_after_colon(line):
    line = str(line or "").strip().replace("：", ":")
    if ":" in line:
        return clean_pick_value(line.split(":", 1)[1])
    # احتياط لو كتب بدون نقطتين: "الحارس نايلاند"
    for key in ["الحارس", "حارس", "الكابتن", "كابتن", "اللاعب", "لاعب"]:
        if key in line:
            return clean_pick_value(line.split(key, 1)[1])
    return ""

def participant_line_name(line):
    line = normalize_name(line)
    if not line:
        return ""
    cleaned = re.sub(r"[^\u0600-\u06FFa-zA-Z\s]", " ", line)
    cleaned = normalize_name(cleaned)
    for p in PARTICIPANTS:
        if cleaned == p:
            return p
    return ""

def parse_pipe_lineup_lines(lines):
    data = {}
    bad_lines = []
    for line in lines:
        raw = line.strip()
        if not raw:
            continue
        if "|" not in raw:
            continue
        parts = [clean_pick_value(p) for p in raw.split("|")]
        if len(parts) != 6:
            bad_lines.append(raw)
            continue
        participant = normalize_name(parts[0])
        data[participant] = [normalize_name(x) for x in parts[1:]]
    return data, bad_lines

def parse_official_lineup_blocks(lines):
    """
    يقرأ النموذج الرسمي مثل:
    فهد فارس
    🧤 الحارس: أوناي سيمون
    اللاعب 1: أولمو
    اللاعب 2: سالم الدوسري
    اللاعب 3: داروين نونيز
    👑 الكابتن : نونيز

    ويدعم:
    لاعب: ميسي
    لاعب ٢: هالاند
    لاعب ٣: ...
    """
    data = {}
    bad_blocks = []
    current = None
    values = None

    def finish_current():
        nonlocal current, values
        if not current:
            return
        keeper = values.get("keeper", "")
        p1 = values.get("p1", "")
        p2 = values.get("p2", "")
        p3 = values.get("p3", "")
        captain = values.get("captain", "")
        if any([keeper, p1, p2, p3, captain]):
            # إذا ناقص شيء نخليه فارغ بدل ما نرفض كامل المشاركة
            data[current] = [keeper, p1, p2, p3, captain]
        current = None
        values = None

    def put_player(value, preferred=None):
        if not value:
            return
        if preferred and not values.get(preferred):
            values[preferred] = value
            return
        for key in ["p1", "p2", "p3"]:
            if not values.get(key):
                values[key] = value
                return

    for raw in lines:
        line = normalize_name(raw)
        if not line:
            continue

        # إذا لقى اسم مشارك جديد، يقفل السابق
        p_name = participant_line_name(line)
        if p_name:
            finish_current()
            current = p_name
            values = {"keeper": "", "p1": "", "p2": "", "p3": "", "captain": ""}
            continue

        if not current:
            continue

        line_no_space = line.replace(" ", "")
        value = extract_after_colon(line)

        if "الحارس" in line or line.startswith("حارس") or "🧤" in raw:
            if value:
                values["keeper"] = value
            continue

        if "الكابتن" in line or "كابتن" in line or "👑" in raw:
            if value:
                values["captain"] = value
            continue

        if "لاعب" in line:
            if not value:
                continue
            preferred = None
            if re.search(r"(3|٣|الثالث|ثالث)", line_no_space):
                preferred = "p3"
            elif re.search(r"(2|٢|الثاني|ثاني)", line_no_space):
                preferred = "p2"
            elif re.search(r"(1|١|الأول|اول|الأول|الاول)", line_no_space):
                preferred = "p1"
            put_player(value, preferred)
            continue

    finish_current()
    return data, bad_blocks

def parse_add_day_text(message_text):
    """
    يقبل الصيغتين:
    1) فهد|حارس|لاعب1|لاعب2|لاعب3|كابتن
    2) النموذج الرسمي الطويل لكل مشارك
    """
    lines = (message_text or "").splitlines()[1:]

    pipe_data, pipe_bad = parse_pipe_lineup_lines(lines)
    official_data, official_bad = parse_official_lineup_blocks(lines)

    # لو نفس المشارك موجود بالصيغتين، آخر قراءة من النموذج الرسمي تغطي
    data = {}
    data.update(pipe_data)
    data.update(official_data)

    # الأسطر السيئة فقط للبايب؛ لا نحاسب النموذج الطويل على الزخارف والعناوين
    bad_lines = pipe_bad[:]

    return data, bad_lines

async def add_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    day = get_day(text)

    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return

    data, bad_lines = parse_add_day_text(text)

    if not data:
        await update.message.reply_text(
            "ما لقيت مشاركين بصيغة صحيحة.\n\n"
            "الصيغة المختصرة:\n"
            "/اضافه 5\n"
            "فهد فارس|الحارس|لاعب1|لاعب2|لاعب3|الكابتن\n\n"
            "أو النموذج الرسمي:\n"
            "فهد فارس\n"
            "🧤 الحارس: أوناي سيمون\n"
            "اللاعب 1: أولمو\n"
            "اللاعب 2: سالم الدوسري\n"
            "اللاعب 3: داروين نونيز\n"
            "👑 الكابتن : نونيز"
        )
        return

    backup_files(f"before_add_day_{day}", files=[excel_file(day), LOCKED_FILE])
    file_name, unknown = update_day_data(day, data)

    caption = f"تم إنشاء/تحديث ملف اليوم {day} ✅\nعدد المشاركين المرسلين: {len(data)}"
    caption += "\n\n✅ تم قبول الصيغتين: السطر الواحد + النموذج الرسمي"

    if unknown:
        caption += "\n\n⚠️ أسماء مشاركين غير موجودة بالقائمة:\n" + "\n".join(unknown)
    if bad_lines:
        caption += "\n\n⚠️ أسطر مختصرة لم أفهمها:\n" + "\n".join(bad_lines[:5])

    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=caption)

async def results_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    day = get_day(text)

    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return

    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}. أضف التشكيلات أولًا.")
        return

    goals_count, clean_sheets = parse_results(text)
    goal_missing, clean_missing = validate_results_names(day, goals_count, clean_sheets)

    backup_files(f"before_temp_results_day_{day}", files=[excel_file(day), LOCKED_FILE, CUP_FILE])
    file_name = calculate_points(day, goals_count, clean_sheets)

    goals_text = "\n".join([f"- {name}: {count} هدف = {GOAL_POINTS.get(count, count * 5)} نقطة" for name, count in goals_count.items()]) or "لا يوجد"
    clean_text = "\n".join([f"- {name}" for name in clean_sheets]) or "لا يوجد"

    rows = read_day_rows(day)
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    legends_text = "، ".join(winners) if winners else "لا يوجد"

    warnings = []
    if goal_missing:
        warnings.append("⚠️ هدافون غير موجودين في تشكيلات اليوم:\n" + "\n".join(goal_missing))
    if clean_missing:
        warnings.append("⚠️ حراس كلين شيت غير موجودين في تشكيلات اليوم:\n" + "\n".join(clean_missing))

    caption = (
        f"تم تحديث نقاط اليوم {day} مؤقتًا ✅\n"
        f"لن يتم حساب الكأس أو اعتماد اليوم إلا بعد الأمر:\n"
        f"/اعتماد_نتائج {day}\n\n"
        f"الأهداف:\n{goals_text}\n\n"
        f"الكلين شيت:\n{clean_text}\n\n"
        f"🏆 المتصدر/أسطورة اليوم حاليًا: {legends_text} — {max_score} نقطة"
    )
    if warnings:
        caption += "\n\n" + "\n\n".join(warnings)

    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=caption)

async def overall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if len(nums) >= 2:
        start_day, end_day = nums[0], nums[1]
        if start_day > end_day:
            start_day, end_day = end_day, start_day
    elif len(nums) == 1:
        start_day, end_day = 1, nums[0]
    else:
        start_day, end_day = 1, 31

    file_name, days_found, stats = create_overall_ranking(start_day, end_day)
    if not days_found:
        await update.message.reply_text("ما لقيت أي ملفات أيام محسوبة في النطاق المطلوب.")
        return

    with open(file_name, "rb") as file:
        await update.message.reply_document(
            document=file,
            filename=file_name,
            caption=(
                f"تم إنشاء الترتيب العام ✅\n"
                f"النطاق: من اليوم {start_day} إلى اليوم {end_day}\n"
                f"الأيام المحسوبة: {', '.join(map(str, days_found))}"
            )
        )


async def ranking_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if len(nums) >= 2:
        start_day, end_day = nums[0], nums[1]
        if start_day > end_day:
            start_day, end_day = end_day, start_day
    elif len(nums) == 1:
        start_day, end_day = 1, nums[0]
    else:
        start_day, end_day = 1, 31
    stats = collect_stats(start_day, end_day)
    await update.message.reply_text(build_ranking_text(stats, start_day, end_day))


async def list_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    days = get_existing_days(1, 99)
    if not days:
        await update.message.reply_text("ما فيه أيام محفوظة.")
        return

    locked = load_locked_days()
    lines = ["📅 الأيام الموجودة:"]
    for day in days:
        rows = read_day_rows(day)
        participants = sum(1 for r in rows if r["participated"])
        scored = any(r["total"] > 0 for r in rows)
        lock = "🔒" if str(day) in locked else "🔓"
        calc = "محسوب" if scored else "بدون نتائج"
        lines.append(f"اليوم {day} {lock} — {participants} مشارك — {calc}")
    await update.message.reply_text("\n".join(lines))


async def inspect_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    rows = read_day_rows(day)
    if not rows:
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return

    participants = [r for r in rows if r["participated"]]
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    zeros = [r["participant"] for r in participants if r["total"] == 0]
    avg = round(sum(r["total"] for r in rows) / len(rows), 2) if rows else 0
    locked = "مقفل 🔒" if is_locked(day) else "مفتوح 🔓"

    lines = [
        f"🔎 فحص اليوم {day}",
        f"الحالة: {locked}",
        f"عدد المشاركين: {len(participants)}",
        f"أعلى نقاط: {max_score}",
        f"أسطورة اليوم: {'، '.join(winners) if winners else 'لا يوجد'}",
        f"عدد اللي جابوا 0: {len(zeros)}",
        f"متوسط النقاط: {avg}",
    ]
    if zeros:
        lines.append("\nالأصفار:")
        lines.extend([f"- {name}" for name in zeros])
    await update.message.reply_text("\n".join(lines))


async def participants_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    rows = read_day_rows(day)
    if not rows:
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return

    yes = [r["participant"] for r in rows if r["participated"]]
    no = [r["participant"] for r in rows if not r["participated"]]
    lines = [f"👥 مشاركو اليوم {day}:"]
    lines.extend([f"✅ {name}" for name in yes] or ["لا يوجد"])
    lines.append("\nلم يشارك:")
    lines.extend([f"❌ {name}" for name in no] or ["لا يوجد"])
    await update.message.reply_text("\n".join(lines))


async def legend_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    rows = read_day_rows(day)
    if not rows:
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    if not winners:
        await update.message.reply_text(f"ما فيه أسطورة لليوم {day} حتى الآن.")
        return
    await update.message.reply_text(f"🏆 أسطورة اليوم {ordinal_day(day)}:\n" + "\n".join([f"- {w} — {max_score} نقطة" for w in winners]))


async def compare_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if len(nums) < 2:
        await update.message.reply_text("اكتب رقمين، مثال:\n/مقارنة 4 5\nيعني مقارنة الترتيب بعد اليوم 4 وبعد اليوم 5")
        return

    old_end, new_end = nums[0], nums[1]
    old_stats = collect_stats(1, old_end)
    new_stats = collect_stats(1, new_end)
    if not old_stats["days"] or not new_stats["days"]:
        await update.message.reply_text("ما لقيت أيام كافية للمقارنة.")
        return

    lines = [f"📊 مقارنة الترتيب بعد اليوم {old_end} وبعد اليوم {new_end}:", ""]
    old_pos = {name: i for i, name in enumerate(old_stats["ranking"], start=1)}
    new_pos = {name: i for i, name in enumerate(new_stats["ranking"], start=1)}

    for name in new_stats["ranking"]:
        diff_points = new_stats["totals"][name] - old_stats["totals"].get(name, 0)
        move = old_pos.get(name, new_pos[name]) - new_pos[name]
        move_text = "ثابت"
        if move > 0:
            move_text = f"صعد {move}"
        elif move < 0:
            move_text = f"نزل {abs(move)}"
        lines.append(f"{name}: +{diff_points} نقطة | {move_text}")

    await update.message.reply_text("\n".join(lines))


async def clear_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text != "/مسح_الكل تأكيد":
        await update.message.reply_text(
            "⚠️ أمر مسح الكل خطير.\n\n"
            "للتنفيذ اكتب بالضبط:\n"
            "/مسح_الكل تأكيد\n\n"
            "راح أنقل الملفات لنسخة احتياطية بدل حذفها."
        )
        return

    folder, files = backup_files("clear_all", move=True)
    if not files:
        await update.message.reply_text("ما لقيت ملفات أمسحها.")
        return

    await update.message.reply_text(f"✅ تم نقل الملفات للنسخة الاحتياطية.\n📁 {folder}\n📌 عدد الملفات: {len(files)}")


async def clear_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if not nums:
        await update.message.reply_text("اكتب رقم اليوم، مثال:\n/مسح_يوم 5")
        return
    day = str(nums[0])
    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return

    file_name = excel_file(day)
    if not os.path.exists(file_name):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return

    folder, _ = backup_files(f"before_clear_day_{day}", files=[file_name, LOCKED_FILE])
    os.remove(file_name)
    await update.message.reply_text(f"تم مسح ملف اليوم {day} ✅\nتم حفظ نسخة احتياطية: {folder}")


async def clear_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return

    backup_files(f"before_clear_results_{day}", files=[excel_file(day), LOCKED_FILE])
    file_name = clear_day_points(day)
    if not file_name:
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return

    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=f"تم مسح نتائج اليوم {day} مع بقاء المشاركين ✅")


async def restore_last(update: Update, context: ContextTypes.DEFAULT_TYPE):
    folder = latest_backup_folder()
    if not folder:
        await update.message.reply_text("ما فيه نسخة احتياطية أسترجعها.")
        return

    backup_files("before_restore")

    restored = []
    for filename in os.listdir(folder):
        if filename.endswith(".xlsx") or filename == LOCKED_FILE:
            shutil.copy2(os.path.join(folder, filename), filename)
            restored.append(filename)

    if not restored:
        await update.message.reply_text(f"النسخة {folder} ما فيها ملفات قابلة للاسترجاع.")
        return

    await update.message.reply_text(f"✅ تم استرجاع آخر نسخة احتياطية:\n📁 {folder}\n\nالملفات:\n" + "\n".join(restored))


async def lock_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    days = load_locked_days()
    days.add(str(day))
    save_locked_days(days)
    await update.message.reply_text(f"تم قفل اليوم {day} 🔒")


async def unlock_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    days = load_locked_days()
    days.discard(str(day))
    save_locked_days(days)
    await update.message.reply_text(f"تم فتح اليوم {day} 🔓")


async def dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        nums = get_numbers(update.message.text or update.message.caption or "")

        if len(nums) >= 2:
            start_day, end_day = nums[0], nums[1]
            if start_day > end_day:
                start_day, end_day = end_day, start_day
        elif len(nums) == 1:
            start_day, end_day = 1, nums[0]
        else:
            start_day, end_day = 1, 31

        await update.message.reply_text("جاري إنشاء ملف الإحصائيات... ⏳")
        file_name, stats = create_dashboard(start_day, end_day)

        if not stats.get("days"):
            await update.message.reply_text(
                "ما لقيت أيام لإحصائيات الداشبورد.\n"
                "تأكد أن ملفات الأيام موجودة مثل:\n"
                "fantasy_day_1.xlsx\n"
                "fantasy_day_2.xlsx"
            )
            return

        if not os.path.exists(file_name):
            await update.message.reply_text(
                "صار خطأ: ملف الإحصائيات ما انحفظ.\n"
                f"اسم الملف المتوقع: {file_name}"
            )
            return

        caption = (
            "تم إنشاء ملف الإحصائيات الكامل ✅\n"
            f"النطاق: من اليوم {start_day} إلى اليوم {end_day}\n"
            f"الأيام المحسوبة: {', '.join(map(str, stats['days']))}\n\n"
            "الصفحات:\n"
            "لوحة عامة\n"
            "تطور النقاط\n"
            "تحليل الأيام\n"
            "تحليل المشاركين\n"
            "تحليل الكباتن\n"
            "تحليل الحراس\n"
            "تحليل اللاعبين\n"
            "تفصيل اختيارات اللاعبين\n"
            "سجل الأساطير"
        )

        with open(file_name, "rb") as file:
            await update.message.reply_document(
                document=file,
                filename=file_name,
                caption=caption
            )

    except Exception as e:
        await update.message.reply_text(
            "صار خطأ أثناء إنشاء الإحصائيات ❌\n\n"
            f"السبب:\n{e}"
        )



# ============================================================
# V8 - الصور والتصاميم والإعلانات والأعلام
# ============================================================

SETTINGS_FILE = "bot_settings.json"
GENERATED_DIR = "generated_images"
FLAGS_ZIP = "worldcup_2026_flags_pack.zip"
FLAGS_JSON = "flags_map.json"
FLAGS_DIR = os.path.join("assets", "flags")

RESULT_ANNOUNCEMENT_TEMPLATE = """أساطير اليوم {day_name} لفانتزي المصيف :
\"          🏆( {legends} ) 🏆

📌 يرجى مراجعة النقاط، وفي حال وجود أي خطأ في الاحتساب أو تسجيل النقاط فإن مدة الاعتراض أو طلب المراجعة هي 24 ساعة من وقت إرسال هذه الرسالة.
⏳ بعد انتهاء المدة تُعتمد النتائج بشكل نهائي ولا يقبل أي اعتراض أو تعديل لاحقاً.

بالتوفيق للجميع ⚽️✅🧤🏆"""

MATCHES_ANNOUNCEMENT_TEMPLATE = """🏆 فانتزي المصيف 2026  🏆
ً           🔥🔥🔥 مباريات اليوم ( {day_name} )  🔥🔥🔥🔥  

{matches_text}

📋 نموذج المشاركة الرسمي المعتمد
🏆 تشكيلة الفانتزي - اليوم ( {day_name} )
🧤 الحارس:
 اللاعب 1:
 اللاعب 2:
 اللاعب 3:
👑 الكابتن :"""

# نحاول دعم تشكيل العربية داخل الصور، ولو المكتبات غير موجودة نستمر بدون كراش.
try:
    from PIL import Image, ImageDraw, ImageFont, ImageFilter, features as PIL_FEATURES
except Exception:
    Image = ImageDraw = ImageFont = ImageFilter = PIL_FEATURES = None

try:
    PIL_RAQM = bool(PIL_FEATURES and PIL_FEATURES.check("raqm"))
except Exception:
    PIL_RAQM = False

try:
    import arabic_reshaper
    from bidi.algorithm import get_display
except Exception:
    arabic_reshaper = None
    get_display = None


def ar_text(text):
    text = "" if text is None else str(text)
    if arabic_reshaper and get_display:
        try:
            return get_display(arabic_reshaper.reshape(text))
        except Exception:
            return text
    return text


def _safe_filename(name):
    name = normalize_name(name)
    name = re.sub(r"[^A-Za-z0-9_.\-\u0600-\u06FF]+", "_", name).strip("_")
    return name or "file"


def ensure_generated_dir():
    os.makedirs(GENERATED_DIR, exist_ok=True)


def load_settings():
    default = {"auto_images": True}
    if not os.path.exists(SETTINGS_FILE):
        return default
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        default.update(data if isinstance(data, dict) else {})
    except Exception:
        pass
    return default


def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def admin_id_set():
    raw = os.getenv("ADMIN_IDS") or os.getenv("ADMINS") or ""
    ids = set()
    for part in raw.replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            ids.add(int(part))
        except Exception:
            pass
    return ids


def is_admin_user(update):
    ids = admin_id_set()
    if not ids:
        # إذا ما ضبطت ADMIN_IDS في Railway نخلي البوت يشتغل عشان ما ينقفل عليك.
        return True
    user = getattr(update, "effective_user", None)
    return bool(user and user.id in ids)


def admin_only(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not is_admin_user(update):
            await update.message.reply_text("هذا الأمر للمشرفين فقط 🔒")
            return
        return await func(update, context)
    return wrapper


async def who_am_i(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    chat = update.effective_chat
    await update.message.reply_text(
        f"User ID: {user.id if user else '-'}\n"
        f"Chat ID: {chat.id if chat else '-'}\n\n"
        "حط User ID في Railway داخل ADMIN_IDS لو تبي تقفل أوامر الإدارة."
    )


def font_candidates():
    """
    ترتيب خطوط عربي محسّن للتصاميم.
    البوت يختار أول خط موجود من القائمة، لذلك ارفع هذه الملفات بجانب bot.py:
    Tajawal-ExtraBold.ttf / Tajawal-Black.ttf / Cairo-Bold.ttf / NotoNaskhArabic-Bold.ttf / Amiri-Bold.ttf
    """
    names = [
        "Tajawal-ExtraBold.ttf",
        "Tajawal-Black.ttf",
        "Tajawal-Bold.ttf",
        "Cairo-Bold.ttf",
        "Cairo-Bold-1.ttf",
        "NotoNaskhArabic-Bold.ttf",
        "NotoNaskhArabic-Regular.ttf",
        "Amiri-Bold.ttf",
        "DejaVuSans-Bold.ttf",
    ]
    bases = ["", ".", "/app", "/mnt/data", "/usr/share/fonts/truetype/noto", "/usr/share/fonts/truetype/dejavu"]
    paths = []
    for base in bases:
        for name in names:
            paths.append(os.path.join(base, name) if base else name)
    return paths


def get_font(size):
    if not ImageFont:
        return None
    for path in font_candidates():
        try:
            if os.path.exists(path):
                return ImageFont.truetype(path, size)
        except Exception:
            pass
    try:
        return ImageFont.truetype("DejaVuSans-Bold.ttf", size)
    except Exception:
        return ImageFont.load_default()


def font_size(font, default=24):
    try:
        return int(getattr(font, "size", default) or default)
    except Exception:
        return default


def get_font_from_names(size, names):
    if not ImageFont:
        return None
    bases = ["", ".", "/app", "/mnt/data", "/usr/share/fonts/truetype/noto", "/usr/share/fonts/truetype/dejavu"]
    for base in bases:
        for name in names:
            path = os.path.join(base, name) if base else name
            try:
                if os.path.exists(path):
                    return ImageFont.truetype(path, size)
            except Exception:
                pass
    return get_font(size)


def get_arabic_fallback_font(size):
    """
    يستخدم فقط إذا ما كان RAQM متوفرًا.
    Noto Naskh يدعم العربية مع reshaper/bidi أفضل من بعض خطوط العناوين.
    """
    return get_font_from_names(size, [
        "NotoNaskhArabic-Bold.ttf",
        "NotoNaskhArabic-Regular.ttf",
        "Cairo-Bold.ttf",
        "Cairo-Bold-1.ttf",
        "Amiri-Bold.ttf",
        "Tajawal-Bold.ttf",
        "DejaVuSans-Bold.ttf",
    ])

def has_arabic(text):
    return bool(re.search(r"[\u0600-\u06FF]", str(text or "")))



def clean_draw_text(text):
    """تنظيف نصوص الصور من الرموز التي تظهر كمربعات أو فواصل مزعجة داخل التقرير."""
    s = "" if text is None else str(text)
    replacements = {
        "🏆": "",
        "👑": "",
        "🧤": "",
        "👥": "",
        "⚽": "",
        "🔥": "",
        "😅": "",
        "⚔️": "",
        "⚔": "",
        "🥇": "1",
        "🥈": "2",
        "🥉": "3",
        "\ufe0f": "",
        "□": "",
        "■": "",
        "▪": "",
        "▫": "",
        "▢": "",
        "▣": "",
        "▤": "",
        "▥": "",
        "▦": "",
        "▧": "",
        "▨": "",
        "▩": "",
        "◻": "",
        "◼": "",
        "◽": "",
        "◾": "",
        "◻️": "",
        "◼️": "",
        "☐": "",
        "☑": "",
        "☒": "",
        "¦": "-",
        "|": "-",
        "\u200f": "",
        "\u200e": "",
        "\u2066": "",
        "\u2067": "",
        "\u2068": "",
        "\u2069": "",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = re.sub(r"\s+([\-—])\s+", r" \1 ", s)
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip(" -—")

def draw_text(draw, xy, text, font, fill="white", anchor="mm", align="center", max_width=None, spacing=8):
    """
    V12: رسم عربي آمن داخل الصور.
    - إذا RAQM متوفر: نرسم النص الأصلي direction=rtl حتى لا تظهر مربعات/فواصل داخل الكلمات.
    - إذا RAQM غير متوفر: نستخدم arabic_reshaper + bidi مع خط NotoNaskh كبديل.
    """
    text = clean_draw_text("" if text is None else str(text))
    if font is None:
        font = get_font(24)

    if max_width:
        lines = wrap_text(draw, text, font, max_width)
        line_h = int(font_size(font, 24) * 1.30)
        total_h = line_h * len(lines)
        x, y = xy
        start_y = y - total_h / 2 if anchor == "mm" else y
        for i, line in enumerate(lines):
            draw_text(draw, (x, start_y + i * line_h), line, font, fill=fill, anchor="ma", align=align)
        return

    try:
        if has_arabic(text) and PIL_RAQM:
            draw.text(xy, text, font=font, fill=fill, anchor=anchor, align=align, direction="rtl", language="ar")
        elif has_arabic(text):
            f = get_arabic_fallback_font(font_size(font, 24))
            draw.text(xy, ar_text(text), font=f, fill=fill, anchor=anchor, align=align)
        else:
            draw.text(xy, text, font=font, fill=fill, anchor=anchor, align=align)
    except Exception:
        try:
            f = get_arabic_fallback_font(font_size(font, 24)) if has_arabic(text) else font
            display = ar_text(text) if has_arabic(text) else text
            draw.text(xy, display, font=f, fill=fill, anchor=anchor, align=align)
        except Exception:
            draw.text(xy, text, font=font, fill=fill, anchor=anchor)

def text_width(draw, text, font):
    text = clean_draw_text("" if text is None else str(text))
    try:
        if has_arabic(text) and PIL_RAQM:
            bbox = draw.textbbox((0, 0), text, font=font, direction="rtl", language="ar")
        elif has_arabic(text):
            f = get_arabic_fallback_font(font_size(font, 24))
            bbox = draw.textbbox((0, 0), ar_text(text), font=f)
        else:
            bbox = draw.textbbox((0, 0), text, font=font)
        return bbox[2] - bbox[0]
    except Exception:
        return len(str(text)) * 12

def wrap_text(draw, text, font, max_width):
    text = clean_draw_text("" if text is None else str(text))
    final_lines = []
    for raw_line in text.splitlines() or [""]:
        words = raw_line.split()
        if not words:
            final_lines.append("")
            continue
        line = ""
        for word in words:
            test = word if not line else line + " " + word
            if text_width(draw, test, font) <= max_width:
                line = test
            else:
                if line:
                    final_lines.append(line)
                line = word
        if line:
            final_lines.append(line)
    return final_lines


def rounded_rect(draw, box, radius=24, fill=None, outline=None, width=1):
    try:
        draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)
    except Exception:
        draw.rectangle(box, fill=fill, outline=outline, width=width)


def make_canvas(width, height, theme="purple"):
    if not Image:
        raise RuntimeError("Pillow غير مثبت. أضف Pillow في requirements.txt")
    img = Image.new("RGB", (width, height), "#101827")
    draw = ImageDraw.Draw(img)
    # خلفية متدرجة بسيطة
    for y in range(height):
        r = int(18 + y / height * 12)
        g = int(24 + y / height * 10)
        b = int(45 + y / height * 35)
        draw.line([(0, y), (width, y)], fill=(r, g, b))
    # دوائر ديكور
    for i, (x, y, rr, col) in enumerate([
        (120, 80, 180, "#7C3AED"), (width-130, 120, 220, "#2563EB"),
        (width//2, height-80, 300, "#F2B705")
    ]):
        overlay = Image.new("RGBA", (width, height), (0,0,0,0))
        od = ImageDraw.Draw(overlay)
        od.ellipse((x-rr, y-rr, x+rr, y+rr), fill=col + "30")
        overlay = overlay.filter(ImageFilter.GaussianBlur(40))
        img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
        draw = ImageDraw.Draw(img)
    return img, draw


def ensure_flags_assets():
    if os.path.exists(FLAGS_DIR) and os.path.exists(FLAGS_JSON):
        return
    if os.path.exists(FLAGS_ZIP):
        try:
            import zipfile
            with zipfile.ZipFile(FLAGS_ZIP, "r") as z:
                z.extractall(".")
        except Exception:
            pass


def load_flags_map():
    ensure_flags_assets()
    data = {}
    if os.path.exists(FLAGS_JSON):
        try:
            with open(FLAGS_JSON, "r", encoding="utf-8") as f:
                obj = json.load(f)
            aliases = obj.get("aliases", {}) if isinstance(obj, dict) else {}
            for k, v in aliases.items():
                data[normalize_name(k)] = v
                data[normalize_name(k).lower()] = v
        except Exception:
            pass

    # Aliases إضافية عشان الأعلام تطلع حتى لو كتبنا الاسم بصيغة مختلفة.
    alias_groups = [
        ["الولايات المتحدة", "الولايات المتحدة الأمريكية", "امريكا", "أمريكا", "اميركا", "أميركا", "USA", "US", "United States", "United States of America"],
        ["باراغواي", "الباراغواي", "بارجواي", "Paraguay"],
        ["اسكتلندا", "إسكتلندا", "سكوتلندا", "Scotland"],
        ["ساحل العاج", "كوت ديفوار", "Ivory Coast", "Cote dIvoire", "Côte d’Ivoire"],
        ["كوريا الجنوبية", "كوريا الجنوبيه", "South Korea", "Korea Republic"],
    ]
    for group in alias_groups:
        filename = None
        for key in group:
            nk = normalize_name(key)
            filename = data.get(nk) or data.get(nk.lower())
            if filename:
                break
        if filename:
            for key in group:
                nk = normalize_name(key)
                data[nk] = filename
                data[nk.lower()] = filename

    return data


def flag_path_for(team_name):
    flags = load_flags_map()
    name = normalize_name(team_name)
    filename = flags.get(name) or flags.get(name.lower())
    if filename:
        path = os.path.join(FLAGS_DIR, filename)
        if os.path.exists(path):
            return path
    return None



def flag_of(team_name):
    """توافق للأمر الموسع /مباريات_الأيام10."""
    try:
        return flag_path_for(team_name)
    except Exception:
        return None

def paste_flag(base, team_name, box):
    if not Image:
        return
    path = flag_path_for(team_name)
    x1, y1, x2, y2 = box
    w, h = int(x2-x1), int(y2-y1)
    if path and os.path.exists(path):
        try:
            flag = Image.open(path).convert("RGBA")
            flag.thumbnail((w, h), Image.LANCZOS)
            px = int(x1 + (w - flag.width) / 2)
            py = int(y1 + (h - flag.height) / 2)
            base.paste(flag, (px, py), flag)
            return
        except Exception:
            pass
    # بديل إذا العلم ناقص
    d = ImageDraw.Draw(base)
    rounded_rect(d, box, radius=18, fill="#FFFFFF22", outline="#FFFFFF55", width=2)
    f = get_font(24)
    draw_text(d, ((x1+x2)//2, (y1+y2)//2), normalize_name(team_name)[:2], f, fill="#FFFFFF")


def build_result_announcement(day, winners):
    day_name = ordinal_day(day)
    legends = " + ".join(winners) if winners else "لا يوجد"
    return RESULT_ANNOUNCEMENT_TEMPLATE.format(day_name=day_name, legends=legends)


def build_day_summary(day):
    rows = read_day_rows(day)
    participants = [r for r in rows if r.get("participated")]
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    sorted_rows = sorted(rows, key=lambda r: r["total"], reverse=True)
    top_captain = max(rows, key=lambda r: r.get("captain_points", 0), default=None)
    top_keeper = max(rows, key=lambda r: r.get("keeper_points", 0), default=None)

    player_points = Counter()
    keeper_points = Counter()
    for r in rows:
        for k, pk in (("p1", "p1_points"), ("p2", "p2_points"), ("p3", "p3_points")):
            if r.get(pk, 0) > 0 and not is_no_participation(r.get(k)):
                player_points[r[k]] += r[pk]
        if r.get("keeper_points", 0) > 0 and not is_no_participation(r.get("keeper")):
            keeper_points[r["keeper"]] += r["keeper_points"]

    return {
        "rows": rows,
        "participants": participants,
        "max_score": max_score,
        "winners": winners,
        "sorted_rows": sorted_rows,
        "top_captain": top_captain,
        "top_keeper": top_keeper,
        "player_points": player_points,
        "keeper_points": keeper_points,
    }


def create_daily_result_image(day, goals_count=None, clean_sheets=None):
    ensure_generated_dir()
    data = build_day_summary(day)
    rows = data["sorted_rows"]
    h = max(1050, 470 + len(rows) * 58)
    img, draw = make_canvas(1400, h)
    title_f = get_font(58)
    sub_f = get_font(34)
    small_f = get_font(26)
    row_f = get_font(28)

    draw_text(draw, (700, 85), f"فانتزي المصيف 2026 — اليوم {ordinal_day(day)}", title_f, fill="#FFFFFF")
    draw_text(draw, (700, 145), "نتائج اليوم وترتيب المشاركين", sub_f, fill="#FDE68A")

    winners = data["winners"]
    legends = " + ".join(winners) if winners else "لا يوجد"
    rounded_rect(draw, (80, 190, 1320, 340), radius=38, fill="#7C3AEDDD", outline="#FFFFFF33", width=2)
    draw_text(draw, (700, 235), "🏆 أسطورة اليوم 🏆", sub_f, fill="#FFFFFF")
    draw_text(draw, (700, 295), f"{legends} — {data['max_score']} نقطة", get_font(44), fill="#FFF6D6")

    # كروت جانبية
    top_cap = data["top_captain"]
    top_keep = data["top_keeper"]
    cards = [
        (80, 370, 455, 500, "👑 أفضل كابتن", f"{top_cap['participant']} +{top_cap['captain_points']}" if top_cap else "-", "#2563EB"),
        (512, 370, 887, 500, "🧤 أفضل حارس", f"{top_keep['participant']} +{top_keep['keeper_points']}" if top_keep else "-", "#10B981"),
        (945, 370, 1320, 500, "👥 المشاركون", f"{len(data['participants'])} مشارك", "#F59E0B"),
    ]
    for x1,y1,x2,y2,t,v,c in cards:
        rounded_rect(draw, (x1,y1,x2,y2), radius=28, fill=c+"DD", outline="#FFFFFF33", width=2)
        draw_text(draw, ((x1+x2)//2, y1+38), t, small_f, fill="#FFFFFF")
        draw_text(draw, ((x1+x2)//2, y1+92), v, sub_f, fill="#FFFFFF")

    # جدول ترتيب اليوم
    y = 550
    rounded_rect(draw, (80, y, 1320, y+58), radius=18, fill="#FFFFFF22", outline="#FFFFFF30", width=1)
    headers = [(1180, "المشارك"), (850, "النقاط"), (640, "الحارس"), (410, "الكابتن"), (170, "المركز")]
    for x, t in headers:
        draw_text(draw, (x, y+30), t, small_f, fill="#FFFFFF")
    y += 70
    for idx, r in enumerate(rows, start=1):
        fill = "#FFFFFF16" if idx % 2 else "#FFFFFF0C"
        if idx == 1:
            fill = "#F2B70544"
        rounded_rect(draw, (80, y, 1320, y+48), radius=14, fill=fill, outline="#FFFFFF18", width=1)
        draw_text(draw, (1180, y+25), r["participant"], row_f, fill="#FFFFFF")
        draw_text(draw, (850, y+25), str(r["total"]), row_f, fill="#FDE68A")
        draw_text(draw, (640, y+25), f"+{r['keeper_points']}", row_f, fill="#A7F3D0")
        draw_text(draw, (410, y+25), f"+{r['captain_points']}", row_f, fill="#C4B5FD")
        medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else str(idx)
        draw_text(draw, (170, y+25), medal, row_f, fill="#FFFFFF")
        y += 58

    # الهدافين والكلين شيت
    y += 25
    box_h = 180
    rounded_rect(draw, (80, y, 670, y+box_h), radius=26, fill="#111827CC", outline="#FFFFFF22", width=2)
    rounded_rect(draw, (730, y, 1320, y+box_h), radius=26, fill="#111827CC", outline="#FFFFFF22", width=2)
    goals_lines = []
    if goals_count:
        goals_lines = [f"{p} — {c}" for p,c in goals_count.items()]
    else:
        goals_lines = [f"{p} — {pts} نقطة" for p,pts in data["player_points"].most_common(5)] or ["لا يوجد"]
    clean_lines = clean_sheets or list(data["keeper_points"].keys()) or ["لا يوجد"]
    draw_text(draw, (375, y+35), "⚽ الهدافون", sub_f, fill="#FFFFFF")
    draw_text(draw, (375, y+105), "\n".join(goals_lines[:4]), small_f, fill="#E5E7EB", max_width=520)
    draw_text(draw, (1025, y+35), "🧤 الكلين شيت", sub_f, fill="#FFFFFF")
    draw_text(draw, (1025, y+105), "\n".join(clean_lines[:4]), small_f, fill="#E5E7EB", max_width=520)

    path = os.path.join(GENERATED_DIR, f"daily_result_day_{day}.png")
    img.save(path, quality=95)
    return path


def build_daily_summary_text(day):
    data = build_day_summary(day)
    winners = " و ".join(data["winners"]) if data["winners"] else "لا يوجد"
    top_rows = data["sorted_rows"][:5]
    lines = [
        f"📊 ملخص اليوم {ordinal_day(day)}",
        f"👥 عدد المشاركين: {len(data['participants'])}",
        f"🏆 أسطورة اليوم: {winners}",
        f"🔥 أعلى نقاط: {data['max_score']}",
        "",
        "أول 5 في اليوم:",
    ]
    for i, r in enumerate(top_rows, start=1):
        lines.append(f"{i}. {r['participant']} — {r['total']} نقطة")
    return "\n".join(lines)


def create_overall_ranking_image(start_day=1, end_day=31):
    """
    V21: صورة ترتيب الفانتزي بنفس هوية التصاميم الجديدة
    بدال الجدول القديم اللي كانت بعض الخانات تطلع بيضاء/باهتة.
    """
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    ranking = stats.get("ranking", [])
    count = max(len(ranking), 1)
    width = 1200

    # تصميم مرن حسب عدد المشاركين
    if count <= 8:
        row_h, gap, name_size = 92, 12, 30
    elif count <= 12:
        row_h, gap, name_size = 78, 10, 26
    else:
        row_h, gap, name_size = 66, 8, 22

    header_h = 58
    content_h = header_h + 24 + count * row_h + max(0, count - 1) * gap
    height = max(900, 245 + content_h + 175)

    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "ترتيب فانتزي المصيف 2026", f"من اليوم {start_day} إلى اليوم {end_day}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#22C55E")

    leader = ranking[0] if ranking else None
    leader_points = int(stats.get("totals", {}).get(leader, 0) or 0) if leader else 0

    y = fy1 + 34

    # هيدر الجدول
    rounded_rect(draw, (92, y, width-92, y+header_h), radius=20, fill="#05070D", outline="#FFFFFF70", width=1)
    draw_text(draw, (1080, y+header_h//2), "المركز", get_font(22), fill="#FDE68A")
    draw_text(draw, (850, y+header_h//2), "المشارك", get_font(22), fill="#FDE68A")
    draw_text(draw, (570, y+header_h//2), "النقاط", get_font(22), fill="#FDE68A")
    draw_text(draw, (385, y+header_h//2), "الفارق", get_font(22), fill="#FDE68A")
    draw_text(draw, (190, y+header_h//2), "أسطورة", get_font(22), fill="#FDE68A")
    y += header_h + 24

    current_rank = 0
    last_score = None
    real_index = 0
    medals = {1: "🥇", 2: "🥈", 3: "🥉"}

    for name in ranking:
        real_index += 1
        score = int(stats.get("totals", {}).get(name, 0) or 0)
        if score != last_score:
            current_rank = real_index
            last_score = score

        accent = "#F59E0B" if current_rank == 1 else v16_accent(real_index)
        fill = "#1A1407" if current_rank == 1 else "#0B1020"
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=22, fill=fill, outline=accent, width=2)
        cy = y + row_h//2

        rank_label = medals.get(current_rank, str(current_rank))
        rank_color = "#FDE68A" if current_rank <= 3 else "#FFFFFF"
        draw_text(draw, (1080, cy), rank_label, get_font(max(24, name_size)), fill=rank_color)

        # الاسم أوضح وبمساحة كافية
        draw_text(draw, (835, cy), name, get_font(name_size), fill="#FFFFFF", max_width=330)

        # النقاط
        draw_text(draw, (570, cy), str(score), get_font(max(26, name_size+2)), fill="#FDE68A")

        # الفارق عن المتصدر
        diff = leader_points - score
        diff_text = "—" if diff == 0 else str(diff)
        draw_text(draw, (385, cy), diff_text, get_font(max(24, name_size)), fill="#E5E7EB")

        # مرات أسطورة اليوم
        legends = int(stats.get("daily_wins", {}).get(name, 0) or 0)
        legends_text = str(legends)
        draw_text(draw, (190, cy), legends_text, get_font(max(24, name_size)), fill="#C4B5FD")

        y += row_h + gap

    footer_event(draw, width, height)

    path = os.path.join(GENERATED_DIR, f"overall_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path

def create_legends_image(start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    days = stats["days"]
    h = max(800, 260 + len(days) * 72)
    img, draw = make_canvas(1400, h)
    draw_text(draw, (700, 90), "سجل أساطير الفانتزي", get_font(60), fill="#FFFFFF")
    draw_text(draw, (700, 150), f"من اليوم {start_day} إلى اليوم {end_day}", get_font(30), fill="#FDE68A")
    y = 230
    for day in days:
        info = stats["per_day"].get(day, {})
        winners = info.get("winners", [])
        names = " + ".join(winners) if winners else "لا يوجد"
        max_score = info.get("max_score", 0)
        rounded_rect(draw, (110, y, 1290, y+56), radius=18, fill="#FFFFFF14", outline="#FFFFFF22", width=1)
        draw_text(draw, (1130, y+29), f"اليوم {ordinal_day(day)}", get_font(30), fill="#FFFFFF")
        draw_text(draw, (700, y+29), names, get_font(30), fill="#FDE68A")
        draw_text(draw, (210, y+29), f"{max_score} نقطة", get_font(28), fill="#A7F3D0")
        y += 72
    path = os.path.join(GENERATED_DIR, f"legends_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path


def parse_range_from_text(text):
    nums = get_numbers(text or "")
    if len(nums) >= 2:
        a, b = nums[0], nums[1]
        return (min(a, b), max(a, b))
    if len(nums) == 1:
        return (1, nums[0])
    return (1, 31)


def page_name_from_command(text, default="لوحة عامة"):
    text = text or ""
    # نحذف الأمر والأرقام ونبقي اسم الصفحة
    parts = text.splitlines()[0].split()
    if not parts:
        return default
    rest = []
    for p in parts[1:]:
        if re.fullmatch(r"\d+", p):
            continue
        rest.append(p)
    name = " ".join(rest).strip()
    return name or default


def find_sheet_name(wb, requested):
    requested = normalize_name(requested)
    if requested in wb.sheetnames:
        return requested
    for s in wb.sheetnames:
        if requested and requested in s:
            return s
    for s in wb.sheetnames:
        if s in requested:
            return s
    return wb.sheetnames[0] if wb.sheetnames else None


def render_worksheet_to_images(xlsx_path, sheet_name, max_rows_per_image=32):
    ensure_generated_dir()
    wb = load_workbook(xlsx_path, data_only=False)
    sheet_name = find_sheet_name(wb, sheet_name)
    if not sheet_name:
        return []
    ws = wb[sheet_name]

    # نحدد حدود البيانات المفيدة بدون الأعمدة المخفية قدر الإمكان
    max_col = min(ws.max_column, 12)
    non_empty_rows = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v not in (None, "") for v in vals):
            non_empty_rows.append(r)
    if not non_empty_rows:
        non_empty_rows = [1]
    min_row, max_row = min(non_empty_rows), max(non_empty_rows)

    images = []
    col_w = 170
    row_h = 54
    title_h = 95
    page_index = 1
    for start in range(min_row, max_row + 1, max_rows_per_image):
        end = min(start + max_rows_per_image - 1, max_row)
        width = max(1100, max_col * col_w + 100)
        height = title_h + (end-start+1) * row_h + 80
        img, draw = make_canvas(width, height)
        draw_text(draw, (width//2, 52), f"{sheet_name}" + (f" - {page_index}" if max_row-min_row+1 > max_rows_per_image else ""), get_font(42), fill="#FFFFFF")
        y = title_h
        for r in range(start, end + 1):
            x = 50
            base_fill = "#FFFFFF14" if r % 2 else "#FFFFFF0A"
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                val = "" if cell.value is None else str(cell.value)
                fill = base_fill
                try:
                    fg = cell.fill.fgColor.rgb
                    if fg and fg != "00000000":
                        fill = "#" + fg[-6:] + "DD"
                except Exception:
                    pass
                rounded_rect(draw, (x, y, x+col_w-6, y+row_h-6), radius=10, fill=fill, outline="#FFFFFF20", width=1)
                font = get_font(22 if r != 1 else 24)
                draw_text(draw, (x + (col_w-6)//2, y + (row_h-6)//2), val[:42], font, fill="#FFFFFF", max_width=col_w-20)
                x += col_w
            y += row_h
        out = os.path.join(GENERATED_DIR, f"sheet_{_safe_filename(sheet_name)}_{page_index}.png")
        img.save(out, quality=92)
        images.append(out)
        page_index += 1
    return images


def parse_matches_text(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    if not lines:
        return "اليوم", []
    # أول سطر هو الأمر
    if len(lines) == 1:
        return "اليوم", []
    day_name = lines[1]
    matches = []
    for line in lines[2:]:
        # يدعم: فريق|فريق|وقت أو فريق × فريق | وقت
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 3:
                matches.append((parts[0], parts[1], parts[2]))
            elif len(parts) == 2 and "×" in parts[0]:
                a, b = [p.strip() for p in parts[0].split("×", 1)]
                matches.append((a, b, parts[1]))
        elif "×" in line:
            before, *after = re.split(r"[-—|]", line, maxsplit=1)
            a, b = [p.strip() for p in before.split("×", 1)]
            time = after[0].strip() if after else ""
            matches.append((a, b, time))
    return day_name, matches


def build_matches_announcement(day_name, matches):
    matches_text = "\n".join([f"{a} × {b} — {t}" for a,b,t in matches]) or ""
    return MATCHES_ANNOUNCEMENT_TEMPLATE.format(day_name=day_name, matches_text=matches_text)


def create_matches_image(day_name, matches):
    """تصميم مباريات للقروب: صورة أفقية واضحة بالأعلام، بدون قالب جامد.
    يضبط حجم الصفوف حسب عدد المباريات ويقلل الفراغ.
    """
    ensure_generated_dir()
    count = max(len(matches), 1)
    width = 1600
    # ارتفاع مناسب للقروب، ويزيد إذا المباريات كثيرة
    row_h = 170 if count <= 3 else 140
    gap = 22 if count <= 3 else 16
    header_h = 230
    footer_h = 125
    height = max(900, header_h + count * row_h + max(0, count-1) * gap + footer_h + 40)

    img, draw = make_canvas(width, height)

    # طبقة تظليل في المنتصف عشان النص يبان
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    od.rounded_rectangle((60, 45, width-60, height-45), radius=48, fill="#02061770", outline="#FFFFFF18", width=2)
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    # العنوان
    draw_text(draw, (width//2, 88), "مونديال المصيف 2026", get_font(58), fill="#FFFFFF")
    draw_text(draw, (width//2, 154), f"مباريات اليوم {day_name}", get_font(44), fill="#FDE68A")
    draw.line((210, 205, width-210, 205), fill="#FFFFFF35", width=2)

    colors = ["#7C3AED", "#2563EB", "#0891B2", "#059669", "#D97706", "#DC2626", "#4F46E5"]
    y = header_h
    for i, (a, b, t) in enumerate(matches, start=1):
        c = colors[(i-1) % len(colors)]
        # كرت المباراة
        rounded_rect(draw, (90, y, width-90, y+row_h), radius=36, fill=c+"D5", outline="#FFFFFF30", width=2)

        # مناطق الأعلام
        flag_w = 180 if count <= 3 else 145
        flag_h = 110 if count <= 3 else 90
        cy = y + row_h//2
        # يمين الفريق الأول
        paste_flag(img, a, (width-285, cy-flag_h//2, width-105, cy+flag_h//2))
        # يسار الفريق الثاني
        paste_flag(img, b, (105, cy-flag_h//2, 285, cy+flag_h//2))

        # أسماء المنتخبات
        name_font = get_font(46 if count <= 3 else 38)
        draw_text(draw, (width-470, cy), a, name_font, fill="#FFFFFF", max_width=360)
        draw_text(draw, (470, cy), b, name_font, fill="#FFFFFF", max_width=360)

        # علامة × والوقت
        draw_text(draw, (width//2, cy-20), "×", get_font(58 if count <= 3 else 48), fill="#FDE68A")
        if t:
            badge_w = 210 if count <= 3 else 180
            badge_h = 44 if count <= 3 else 38
            rounded_rect(draw, (width//2-badge_w//2, cy+32, width//2+badge_w//2, cy+32+badge_h), radius=18, fill="#020617E6", outline="#FFFFFF40", width=1)
            draw_text(draw, (width//2, cy+32+badge_h//2), t, get_font(26 if count <= 3 else 22), fill="#FFFFFF")
        y += row_h + gap

    # عبارة الختام
    footer_y = min(height-86, y+55)
    draw_text(draw, (width//2, footer_y), "حياكم في محلكم 🏆", get_font(42), fill="#FFFFFF")

    path = os.path.join(GENERATED_DIR, f"matches_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path


async def send_photo_path(update, path, caption=None):
    with open(path, "rb") as f:
        await update.message.reply_photo(photo=f, caption=caption or "")


# -------------------- أوامر الصور الجديدة --------------------

async def enable_auto_images(update: Update, context: ContextTypes.DEFAULT_TYPE):
    settings = load_settings()
    settings["auto_images"] = True
    save_settings(settings)
    await update.message.reply_text("تم تفعيل الصور التلقائية ✅")


async def disable_auto_images(update: Update, context: ContextTypes.DEFAULT_TYPE):
    settings = load_settings()
    settings["auto_images"] = False
    save_settings(settings)
    await update.message.reply_text("تم إيقاف الصور التلقائية ✅")


async def daily_image_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return
    try:
        path = create_daily_result_image(day)
        await send_photo_path(update, path, build_result_announcement(day, build_day_summary(day)["winners"]))
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء صورة اليوم ❌\n{e}")


async def overall_image_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    start_day, end_day = parse_range_from_text(update.message.text)
    try:
        path = create_overall_ranking_image(start_day, end_day)
        await send_photo_path(update, path, f"الترتيب العام من اليوم {start_day} إلى {end_day} ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء صورة الترتيب ❌\n{e}")


async def legends_image_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    start_day, end_day = parse_range_from_text(update.message.text)
    try:
        path = create_legends_image(start_day, end_day)
        await send_photo_path(update, path, f"سجل الأساطير من اليوم {start_day} إلى {end_day} ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء صورة الأساطير ❌\n{e}")


async def dashboard_sheet_image_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    start_day, end_day = parse_range_from_text(update.message.text)
    sheet = page_name_from_command(update.message.text, "لوحة عامة")
    await update.message.reply_text("جاري تجهيز صورة الصفحة... ⏳")
    try:
        xlsx, stats = create_dashboard(start_day, end_day)
        if not stats.get("days"):
            await update.message.reply_text("ما فيه أيام في النطاق المطلوب.")
            return
        images = render_worksheet_to_images(xlsx, sheet)
        if not images:
            await update.message.reply_text("ما قدرت أطلع صورة للصفحة.")
            return
        for i, path in enumerate(images, start=1):
            cap = f"{sheet}" if len(images) == 1 else f"{sheet} — جزء {i}"
            await send_photo_path(update, path, cap)
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء صورة الإحصائيات ❌\n{e}")


async def all_dashboard_images_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    start_day, end_day = parse_range_from_text(update.message.text)
    await update.message.reply_text("جاري تجهيز صور كل الصفحات... ⏳")
    try:
        xlsx, stats = create_dashboard(start_day, end_day)
        if not stats.get("days"):
            await update.message.reply_text("ما فيه أيام في النطاق المطلوب.")
            return
        wb = load_workbook(xlsx, read_only=True)
        for sheet in wb.sheetnames:
            images = render_worksheet_to_images(xlsx, sheet)
            for i, path in enumerate(images, start=1):
                cap = f"{sheet}" if len(images) == 1 else f"{sheet} — جزء {i}"
                await send_photo_path(update, path, cap)
        wb.close()
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء صور الإحصائيات ❌\n{e}")


async def announcement_day_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return
    await update.message.reply_text(build_result_announcement(day, build_day_summary(day)["winners"]))


async def summary_day_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day = get_day(update.message.text)
    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}.")
        return
    await update.message.reply_text(build_daily_summary_text(day))


async def matches_design_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, matches = parse_matches_text(update.message.text)
    if not matches:
        await update.message.reply_text(
            "اكتبها كذا:\n"
            "/تصميم_مباريات\n"
            "السادس\n"
            "فرنسا|البرازيل|10:00 م\n"
            "الأرجنتين|ألمانيا|12:00 ص"
        )
        return
    try:
        path = create_matches_image(day_name, matches)
        caption = build_matches_announcement(day_name, matches)
        await send_photo_path(update, path, caption)
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات ❌\n{e}")


async def clean_temp_files(update: Update, context: ContextTypes.DEFAULT_TYPE):
    removed = 0
    targets = [GENERATED_DIR, "imports", "uploads", "restore_uploads"]
    for folder in targets:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            path = os.path.join(folder, name)
            try:
                if os.path.isfile(path):
                    os.remove(path)
                    removed += 1
            except Exception:
                pass
    await update.message.reply_text(f"تم تنظيف الملفات المؤقتة ✅\nعدد الملفات المحذوفة: {removed}")


# -------------------- إعادة تعريف /start و /نتائج لإضافة الصور التلقائية --------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n"
        "/نتائج 5\n"
        "/احصائيات\n"
        "/احصائيات 1 6\n"
        "/ترتيب_نص\n\n"
        "أوامر الصور:\n"
        "/صورة_اليوم 6\n"
        "/صورة_الترتيب 1 6\n"
        "/صورة_الاساطير 1 6\n"
        "/صورة_احصائيات 1 6 لوحة عامة\n"
        "/صور_الاحصائيات 1 6\n"
        "/تفعيل_الصور_التلقائية\n"
        "/إيقاف_الصور_التلقائية\n\n"
        "أوامر المباريات:\n"
        "/تصميم_مباريات\n\n"
        "أوامر الفحص والنشر:\n"
        "/الأيام\n"
        "/فحص 5\n"
        "/مشاركين 5\n"
        "/اسطورة 5\n"
        "/مقارنة 4 5\n"
        "/اعلان_اليوم 5\n"
        "/ملخص_اليوم 5\n\n"
        "أوامر الاستيراد والنسخ:\n"
        "/استيراد_ملف — أرسل ملف الإكسل لحاله ثم اكتب الأمر\n"
        "/اعتماد_استيراد\n"
        "/إلغاء_استيراد\n"
        "/نسخة_احتياطية\n"
        "/استرجاع_نسخة — أرسل ملف ZIP لحاله ثم اكتب الأمر\n"
        "/تنظيف_الأيام\n"
        "/تنظيف_الملفات\n\n"
        "أوامر الأمان:\n"
        "/مسح_نتائج 5\n"
        "/مسح_يوم 5\n"
        "/مسح_الكل تأكيد\n"
        "/استرجاع_آخر\n"
        "/قفل_يوم 5\n"
        "/فتح_يوم 5\n"
        "/من_انا"
    )


async def results_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    day = get_day(text)

    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return

    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}. أضف التشكيلات أولًا.")
        return

    goals_count, clean_sheets = parse_results(text)
    goal_missing, clean_missing = validate_results_names(day, goals_count, clean_sheets)

    backup_files(f"before_results_day_{day}", files=[excel_file(day), LOCKED_FILE])
    file_name = calculate_points(day, goals_count, clean_sheets)

    goals_text = "\n".join([f"- {name}: {count} هدف = {GOAL_POINTS.get(count, count * 5)} نقطة" for name, count in goals_count.items()]) or "لا يوجد"
    clean_text = "\n".join([f"- {name}" for name in clean_sheets]) or "لا يوجد"

    rows = read_day_rows(day)
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    legends_text = "، ".join(winners) if winners else "لا يوجد"

    warnings = []
    if goal_missing:
        warnings.append("⚠️ هدافون غير موجودين في تشكيلات اليوم:\n" + "\n".join(goal_missing))
    if clean_missing:
        warnings.append("⚠️ حراس كلين شيت غير موجودين في تشكيلات اليوم:\n" + "\n".join(clean_missing))

    caption = (
        f"تم حساب نقاط اليوم {day} ✅\n\n"
        f"الأهداف:\n{goals_text}\n\n"
        f"الكلين شيت:\n{clean_text}\n\n"
        f"🏆 أسطورة اليوم: {legends_text} — {max_score} نقطة"
    )
    if warnings:
        caption += "\n\n" + "\n\n".join(warnings)

    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=caption)

    if load_settings().get("auto_images", True):
        try:
            path = create_daily_result_image(day, goals_count=goals_count, clean_sheets=clean_sheets)
            await send_photo_path(update, path, build_result_announcement(day, winners))
        except Exception as e:
            await update.message.reply_text(f"تم حساب النتائج، لكن تعذر إنشاء الصورة التلقائية ❌\n{e}")



# ============================================================
# V10 - إضافات نهائية: مواجهات بعد النتائج، كأس الفانتزي، بطاقات، تقارير، هدافين، نتائج مباريات
# ============================================================
import random

MATCHUPS_FILE = "matchups_history.json"
CUP_FILE = "fantasy_cup_history.json"


def font_candidates():
    """خطوط عربية مرتبة بالأولوية: عنوان ثم نص، يختار أول الموجود."""
    return [
        "Tajawal-Black.ttf",
        "Tajawal-ExtraBold.ttf",
        "Cairo-Bold.ttf",
        "Cairo-Bold-1.ttf",
        "NotoNaskhArabic-Bold.ttf",
        "Amiri-Bold.ttf",
        "NotoNaskhArabic-Regular.ttf",
        "/app/Tajawal-Black.ttf",
        "/app/Tajawal-ExtraBold.ttf",
        "/app/Cairo-Bold.ttf",
        "/app/Cairo-Bold-1.ttf",
        "/app/NotoNaskhArabic-Bold.ttf",
        "/app/Amiri-Bold.ttf",
        "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]


def load_json_file(path, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if data is not None else default
    except Exception:
        return default


def save_json_file(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def active_participants_for_day(day):
    return [r for r in read_day_rows(day) if r.get("participated")]


def daily_score_map(day):
    rows = read_day_rows(day)
    return {r["participant"]: r["total"] for r in rows}


def previous_ranking_before_day(day):
    try:
        d = int(day)
    except Exception:
        d = 1
    if d <= 1:
        return {name: i for i, name in enumerate(PARTICIPANTS, start=1)}
    stats = collect_stats(1, d - 1)
    if not stats["days"]:
        return {name: i for i, name in enumerate(PARTICIPANTS, start=1)}
    return {name: i for i, name in enumerate(stats["ranking"], start=1)}


def generate_matchups_for_day(day, force=False):
    """مواجهات اليوم تظهر بعد النتائج فقط. يدخل فقط المشاركون فعليًا في ذلك اليوم."""
    day_key = str(day)
    history = load_json_file(MATCHUPS_FILE, {})
    if (not force) and day_key in history:
        return history[day_key]

    active = [r["participant"] for r in active_participants_for_day(day)]
    if not active:
        result = {"day": day_key, "pairs": [], "bye": None, "note": "لا يوجد مشاركون"}
        history[day_key] = result
        save_json_file(MATCHUPS_FILE, history)
        return result

    # قرعة ثابتة حسب اليوم، غير مرتبطة بالنقاط.
    names = sorted(active)
    rnd = random.Random(int(day) * 2026 + 77)
    rnd.shuffle(names)

    bye = None
    if len(names) % 2 == 1:
        bye = names.pop()

    pairs = []
    scores = daily_score_map(day)
    for i in range(0, len(names), 2):
        a, b = names[i], names[i + 1]
        pa, pb = scores.get(a, 0), scores.get(b, 0)
        if pa > pb:
            winner, status = a, "فوز"
        elif pb > pa:
            winner, status = b, "فوز"
        else:
            # في التعادل يفوز الأعلى ترتيبًا قبل الجولة
            ranks = previous_ranking_before_day(day)
            winner = a if ranks.get(a, 999) <= ranks.get(b, 999) else b
            status = "تعادل — حُسم بالأفضلية"
        pairs.append({"a": a, "a_points": pa, "b": b, "b_points": pb, "winner": winner, "status": status})

    result = {"day": day_key, "pairs": pairs, "bye": bye, "note": ""}
    history[day_key] = result
    save_json_file(MATCHUPS_FILE, history)
    return result


def matchup_lines(day):
    data = generate_matchups_for_day(day)
    if not data.get("pairs") and not data.get("bye"):
        return ["⚔️ مواجهات اليوم:", "لا توجد مواجهات — لا يوجد مشاركون في هذا اليوم."]
    lines = [f"⚔️ مواجهات اليوم {ordinal_day(day)}"]
    for p in data.get("pairs", []):
        mark = "🤝" if "تعادل" in p.get("status", "") else "✅"
        lines.append(f"{p['a']} {p['a_points']} - {p['b_points']} {p['b']} {mark} الفائز: {p['winner']}")
    if data.get("bye"):
        lines.append(f"🟡 راحة الجولة: {data['bye']}")
    return lines


def matchup_wins_map(start_day=1, end_day=31):
    wins = Counter()
    for day in get_existing_days(start_day, end_day):
        data = generate_matchups_for_day(day)
        for p in data.get("pairs", []):
            if p.get("winner"):
                wins[p["winner"]] += 1
        if data.get("bye"):
            # الراحة لا تحتسب فوزًا.
            pass
    return wins


def add_matchups_sheet_to_dashboard(xlsx_path, start_day=1, end_day=31):
    wb = load_workbook(xlsx_path)
    if "سجل المواجهات" in wb.sheetnames:
        del wb["سجل المواجهات"]
    ws = wb.create_sheet("سجل المواجهات")
    ws.sheet_view.rightToLeft = True
    ws.append(["اليوم", "المشارك 1", "نقاطه", "المشارك 2", "نقاطه", "النتيجة", "الفائز"])
    for day in get_existing_days(start_day, end_day):
        data = generate_matchups_for_day(day)
        if not data.get("pairs") and not data.get("bye"):
            ws.append([day, "لا توجد مواجهات", "", "لا يوجد مشاركون", "", "-", "-"])
            continue
        for p in data.get("pairs", []):
            ws.append([day, p["a"], p["a_points"], p["b"], p["b_points"], p.get("status", ""), p.get("winner", "")])
        if data.get("bye"):
            ws.append([day, data["bye"], "راحة", "-", "-", "راحة الجولة", data["bye"]])
    style_sheet(ws)
    wb.save(xlsx_path)


def build_daily_summary_text(day):
    data = build_day_summary(day)
    winners = " و ".join(data["winners"]) if data["winners"] else "لا يوجد"
    top_rows = data["sorted_rows"][:5]
    lines = [
        f"📊 ملخص اليوم {ordinal_day(day)}",
        f"👥 عدد المشاركين: {len(data['participants'])}",
        f"🏆 أسطورة اليوم: {winners} — {data['max_score']} نقطة",
        "",
        "🔥 أفضل 5:",
    ]
    for idx, r in enumerate(top_rows, start=1):
        lines.append(f"{idx}. {r['participant']} — {r['total']} نقطة")
    top_cap = data.get("top_captain")
    top_keep = data.get("top_keeper")
    if top_cap:
        lines.append(f"\n👑 أفضل كابتن: {top_cap['participant']} — +{top_cap['captain_points']}")
    if top_keep:
        lines.append(f"🧤 أفضل حارس: {top_keep['participant']} — +{top_keep['keeper_points']}")
    lines.append("")
    lines.extend(matchup_lines(day))
    return "\n".join(lines)


def create_daily_result_image(day, goals_count=None, clean_sheets=None):
    ensure_generated_dir()
    data = build_day_summary(day)
    rows = data["sorted_rows"]
    matchups = generate_matchups_for_day(day)
    extra_match_rows = len(matchups.get("pairs", [])) + (1 if matchups.get("bye") else 0)
    h = max(1180, 520 + len(rows) * 54 + extra_match_rows * 44 + 260)
    img, draw = make_canvas(1400, h)
    title_f = get_font(58)
    sub_f = get_font(34)
    small_f = get_font(25)
    row_f = get_font(27)

    draw_text(draw, (700, 85), f"فانتزي المصيف 2026 — اليوم {ordinal_day(day)}", title_f, fill="#FFFFFF")
    draw_text(draw, (700, 145), "نتائج اليوم وترتيب المشاركين", sub_f, fill="#FDE68A")

    winners = data["winners"]
    legends = " + ".join(winners) if winners else "لا يوجد"
    rounded_rect(draw, (80, 190, 1320, 340), radius=38, fill="#7C3AEDDD", outline="#FFFFFF33", width=2)
    draw_text(draw, (700, 235), "🏆 أسطورة اليوم 🏆", sub_f, fill="#FFFFFF")
    draw_text(draw, (700, 295), f"{legends} — {data['max_score']} نقطة", get_font(44), fill="#FFF6D6")

    top_cap = data["top_captain"]
    top_keep = data["top_keeper"]
    cards = [
        (80, 370, 455, 500, "👑 أفضل كابتن", f"{top_cap['participant']} +{top_cap['captain_points']}" if top_cap else "-", "#2563EB"),
        (512, 370, 887, 500, "🧤 أفضل حارس", f"{top_keep['participant']} +{top_keep['keeper_points']}" if top_keep else "-", "#10B981"),
        (945, 370, 1320, 500, "👥 المشاركون", f"{len(data['participants'])} مشارك", "#F59E0B"),
    ]
    for x1,y1,x2,y2,t,v,c in cards:
        rounded_rect(draw, (x1,y1,x2,y2), radius=28, fill=c+"DD", outline="#FFFFFF33", width=2)
        draw_text(draw, ((x1+x2)//2, y1+38), t, small_f, fill="#FFFFFF")
        draw_text(draw, ((x1+x2)//2, y1+92), v, sub_f, fill="#FFFFFF", max_width=x2-x1-20)

    y = 545
    rounded_rect(draw, (80, y, 1320, y+58), radius=18, fill="#FFFFFF22", outline="#FFFFFF30", width=1)
    headers = [(1180, "المشارك"), (850, "النقاط"), (640, "الحارس"), (410, "الكابتن"), (170, "المركز")]
    for x, t in headers:
        draw_text(draw, (x, y+30), t, small_f, fill="#FFFFFF")
    y += 70
    for idx, r in enumerate(rows, start=1):
        fill = "#FFFFFF16" if idx % 2 else "#FFFFFF0C"
        if idx == 1:
            fill = "#F2B70544"
        rounded_rect(draw, (80, y, 1320, y+48), radius=14, fill=fill, outline="#FFFFFF18", width=1)
        draw_text(draw, (1180, y+25), r["participant"], row_f, fill="#FFFFFF")
        draw_text(draw, (850, y+25), str(r["total"]), row_f, fill="#FDE68A")
        draw_text(draw, (640, y+25), f"+{r['keeper_points']}", row_f, fill="#A7F3D0")
        draw_text(draw, (410, y+25), f"+{r['captain_points']}", row_f, fill="#C4B5FD")
        medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else str(idx)
        draw_text(draw, (170, y+25), medal, row_f, fill="#FFFFFF")
        y += 54

    y += 22
    rounded_rect(draw, (80, y, 1320, y+max(145, 58 + extra_match_rows*42)), radius=26, fill="#111827CC", outline="#FFFFFF22", width=2)
    draw_text(draw, (700, y+35), "⚔️ مواجهات اليوم", sub_f, fill="#FDE68A")
    yy = y + 82
    if not matchups.get("pairs") and not matchups.get("bye"):
        draw_text(draw, (700, yy), "لا توجد مواجهات — لا يوجد مشاركون", small_f, fill="#FFFFFF")
        yy += 42
    else:
        for p in matchups.get("pairs", []):
            line = f"{p['a']} {p['a_points']} - {p['b_points']} {p['b']}  |  الفائز: {p['winner']}"
            draw_text(draw, (700, yy), line, small_f, fill="#FFFFFF")
            yy += 42
        if matchups.get("bye"):
            draw_text(draw, (700, yy), f"🟡 راحة الجولة: {matchups['bye']}", small_f, fill="#FDE68A")
            yy += 42
    y = yy + 25

    box_h = 170
    rounded_rect(draw, (80, y, 670, y+box_h), radius=26, fill="#111827CC", outline="#FFFFFF22", width=2)
    rounded_rect(draw, (730, y, 1320, y+box_h), radius=26, fill="#111827CC", outline="#FFFFFF22", width=2)
    goals_lines = [f"{p} — {c}" for p,c in (goals_count or {}).items()] if goals_count else [f"{p} — {pts} نقطة" for p,pts in data["player_points"].most_common(5)] or ["لا يوجد"]
    clean_lines = clean_sheets or list(data["keeper_points"].keys()) or ["لا يوجد"]
    draw_text(draw, (375, y+35), "⚽ الهدافون", sub_f, fill="#FFFFFF")
    draw_text(draw, (375, y+105), "\n".join(goals_lines[:4]), small_f, fill="#E5E7EB", max_width=520)
    draw_text(draw, (1025, y+35), "🧤 الكلين شيت", sub_f, fill="#FFFFFF")
    draw_text(draw, (1025, y+105), "\n".join(clean_lines[:4]), small_f, fill="#E5E7EB", max_width=520)

    path = os.path.join(GENERATED_DIR, f"daily_result_day_{day}.png")
    img.save(path, quality=95)
    return path


def create_matches_image(day_name, matches):
    ensure_generated_dir()
    count = max(len(matches), 1)
    width = 1600
    row_h = 170 if count <= 3 else 140
    gap = 22 if count <= 3 else 16
    header_h = 230
    footer_h = 125
    height = max(900, header_h + count * row_h + max(0, count-1) * gap + footer_h + 40)
    img, draw = make_canvas(width, height)
    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)
    od.rounded_rectangle((60,45,width-60,height-45), radius=48, fill="#02061770", outline="#FFFFFF18", width=2)
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)
    draw_text(draw, (width//2, 88), "مونديال المصيف 2026", get_font(58), fill="#FFFFFF")
    draw_text(draw, (width//2, 154), f"مباريات اليوم {day_name}", get_font(44), fill="#FDE68A")
    draw.line((210,205,width-210,205), fill="#FFFFFF35", width=2)
    colors = ["#7C3AED", "#2563EB", "#0891B2", "#059669", "#D97706", "#DC2626", "#4F46E5"]
    y = header_h
    for i, (a,b,t) in enumerate(matches, start=1):
        c = colors[(i-1)%len(colors)]
        rounded_rect(draw, (90,y,width-90,y+row_h), radius=36, fill=c+"D5", outline="#FFFFFF30", width=2)
        flag_w = 180 if count <= 3 else 145
        flag_h = 110 if count <= 3 else 90
        cy = y + row_h//2
        paste_flag(img, a, (width-285, cy-flag_h//2, width-105, cy+flag_h//2))
        paste_flag(img, b, (105, cy-flag_h//2, 285, cy+flag_h//2))
        name_font = get_font(46 if count <= 3 else 38)
        draw_text(draw, (width-470, cy), a, name_font, fill="#FFFFFF", max_width=360)
        draw_text(draw, (470, cy), b, name_font, fill="#FFFFFF", max_width=360)
        draw_text(draw, (width//2, cy-20), "×", get_font(58 if count <= 3 else 48), fill="#FDE68A")
        if t:
            badge_w = 210 if count <= 3 else 180
            badge_h = 44 if count <= 3 else 38
            rounded_rect(draw, (width//2-badge_w//2, cy+32, width//2+badge_w//2, cy+32+badge_h), radius=18, fill="#020617E6", outline="#FFFFFF40", width=1)
            draw_text(draw, (width//2, cy+32+badge_h//2), t, get_font(26 if count <= 3 else 22), fill="#FFFFFF")
        y += row_h + gap
    footer_y = min(height-86, y+55)
    draw_text(draw, (width//2, footer_y), "المصيف ينقل لكم الحدث", get_font(42), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"matches_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path


def parse_match_results_text(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    results = []
    for line in lines[1:]:
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 4:
                try:
                    results.append((parts[0], int(parts[1]), int(parts[2]), parts[3]))
                    continue
                except Exception:
                    pass
            if len(parts) == 2:
                m1 = re.match(r"(.+?)(\d+)\s*$", parts[0])
                m2 = re.match(r"^(\d+)\s*(.+)$", parts[1])
                if m1 and m2:
                    results.append((normalize_name(m1.group(1)), int(m1.group(2)), int(m2.group(1)), normalize_name(m2.group(2))))
        m = re.match(r"(.+?)\s+(\d+)\s*[-–:]\s*(\d+)\s+(.+)$", line)
        if m:
            results.append((normalize_name(m.group(1)), int(m.group(2)), int(m.group(3)), normalize_name(m.group(4))))
    return results


def create_match_results_image(results):
    ensure_generated_dir()
    width = 1600
    row_h = 165 if len(results) <= 3 else 135
    height = max(820, 230 + len(results)*row_h + 180)
    img, draw = make_canvas(width, height)
    draw_text(draw, (width//2, 92), "مونديال المصيف 2026", get_font(64), fill="#FFFFFF")
    draw_text(draw, (width//2, 160), "نتائج مباريات اليوم", get_font(46), fill="#FDE68A")
    y = 240
    colors = ["#7C3AED", "#2563EB", "#0891B2", "#059669", "#D97706"]
    for i,(a,sa,sb,b) in enumerate(results, start=1):
        c = colors[(i-1)%len(colors)]
        rounded_rect(draw, (90,y,width-90,y+row_h), radius=36, fill=c+"D5", outline="#FFFFFF30", width=2)
        cy = y + row_h//2
        paste_flag(img, a, (width-280, cy-55, width-105, cy+55))
        paste_flag(img, b, (105, cy-55, 280, cy+55))
        draw_text(draw, (width-480, cy), a, get_font(44), fill="#FFFFFF", max_width=340)
        draw_text(draw, (480, cy), b, get_font(44), fill="#FFFFFF", max_width=340)
        rounded_rect(draw, (width//2-125, cy-44, width//2+125, cy+44), radius=26, fill="#020617E6", outline="#FDE68A", width=2)
        draw_text(draw, (width//2, cy), f"{sa} - {sb}", get_font(54), fill="#FDE68A")
        y += row_h + 24
    draw_text(draw, (width//2, height-70), "المصيف ينقل لكم الحدث", get_font(42), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, "match_results_today.png")
    img.save(path, quality=95)
    return path


def build_match_results_caption(results):
    lines = ["🏆 نتائج مباريات اليوم 🏆", ""]
    for a,sa,sb,b in results:
        lines.append(f"{a} {sa} - {sb} {b}")
    lines.append("\nالمصيف ينقل لكم الحدث")
    return "\n".join(lines)


async def match_results_today_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    results = parse_match_results_text(update.message.text)
    if not results:
        await update.message.reply_text("اكتبها كذا:\n/نتائج_مباريات_اليوم\nالسعودية|2|1|إسبانيا\nفرنسا|3|0|البرازيل")
        return
    try:
        path = create_match_results_image(results)
        await send_photo_path(update, path, build_match_results_caption(results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج المباريات ❌\n{e}")


def parse_scorers_text(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    items = []
    for line in lines[1:]:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 2:
            name = parts[0]
            goals = score_to_int(parts[1])
            team = parts[2] if len(parts) >= 3 else ""
            if name and goals > 0:
                items.append((name, goals, team))
    return sorted(items, key=lambda x: (-x[1], x[0]))


def create_top_scorers_image(items):
    ensure_generated_dir()
    width = 1400
    row_h = 110
    height = max(850, 230 + len(items[:10])*row_h + 150)
    img, draw = make_canvas(width, height)
    draw_text(draw, (width//2, 85), "مونديال المصيف 2026", get_font(60), fill="#FFFFFF")
    draw_text(draw, (width//2, 150), "هدافين البطولة حتى الآن", get_font(44), fill="#FDE68A")
    y = 235
    for i,(name,goals,team) in enumerate(items[:10], start=1):
        fill = "#F2B70555" if i == 1 else "#FFFFFF16"
        rounded_rect(draw, (90,y,width-90,y+88), radius=28, fill=fill, outline="#FFFFFF25", width=2)
        draw_text(draw, (1220, y+44), str(i), get_font(46), fill="#FDE68A" if i == 1 else "#FFFFFF")
        if team:
            paste_flag(img, team, (1040, y+12, 1130, y+76))
        draw_text(draw, (790, y+44), name, get_font(42), fill="#FFFFFF", max_width=430)
        draw_text(draw, (250, y+44), f"{goals} {'هدف' if goals == 1 else 'أهداف'}", get_font(38), fill="#FDE68A")
        y += row_h
    draw_text(draw, (width//2, height-70), "المصيف ينقل لكم الحدث", get_font(38), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, "top_scorers.png")
    img.save(path, quality=95)
    return path


def build_top_scorers_caption(items):
    lines = ["⚽ هدافين البطولة حتى الآن", ""]
    for i,(name,goals,team) in enumerate(items, start=1):
        team_txt = f" — {team}" if team else ""
        lines.append(f"{i}. {name}{team_txt} — {goals} {'هدف' if goals == 1 else 'أهداف'}")
    lines.append("\nالمصيف ينقل لكم الحدث")
    return "\n".join(lines)


async def top_scorers_design_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items = parse_scorers_text(update.message.text)
    if not items:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_هدافين\nراؤول خيمينيز|4|المكسيك\nفلوريان فيرتز|3|ألمانيا")
        return
    try:
        path = create_top_scorers_image(items)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين ❌\n{e}")


def find_participant_name(query):
    q = normalize_name(query)
    if not q:
        return None
    for name in PARTICIPANTS:
        if normalize_name(name) == q:
            return name
    for name in PARTICIPANTS:
        if q in normalize_name(name) or normalize_name(name) in q:
            return name
    return None


def participant_best_captain(name, start_day=1, end_day=31):
    best = ("-", 0, "-")
    for day in get_existing_days(start_day, end_day):
        for r in read_day_rows(day):
            if r["participant"] == name and r.get("captain_points", 0) > best[1]:
                best = (r.get("captain", "-"), r.get("captain_points", 0), day)
    return best


def create_participant_card_image(name, start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    if name not in PARTICIPANTS:
        raise ValueError("اسم المشارك غير موجود")

    rank = stats["ranking"].index(name) + 1 if name in stats["ranking"] else "-"
    total = stats["totals"].get(name, 0)
    day_scores = stats["scores_by_day"].get(name, {})
    best_day = max(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    worst_day = min(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    best_score = day_scores.get(best_day, 0) if best_day != "-" else 0
    worst_score = day_scores.get(worst_day, 0) if worst_day != "-" else 0
    best_cap, best_cap_pts, best_cap_day = participant_best_captain(name, start_day, end_day)
    wins = matchup_wins_map(start_day, end_day).get(name, 0)
    pc = stats["participation_count"].get(name, 0)
    days_count = len(stats["days"])
    pct = f"{round(pc / days_count * 100, 1)}%" if days_count else "0%"

    img, draw = design_canvas(None, 1400, 980, "purple")
    draw_design_header(draw, 1400, "بطاقة مشارك فانتزي المصيف", clean_draw_text(name), img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, 1400, 980, top=205, bottom_pad=94, accent="#A855F7")

    def card(box, title, value, value_size=42, fill="#0B1020", outline="#FFFFFF40"):
        x1, y1, x2, y2 = box
        rounded_rect(draw, box, radius=28, fill=fill, outline=outline, width=2)
        draw_text(draw, ((x1 + x2) // 2, y1 + 34), clean_draw_text(title), get_font(28), fill="#E5E7EB")
        draw_text(draw, ((x1 + x2) // 2, (y1 + y2) // 2 + 14), clean_draw_text(value), get_font(value_size), fill="#FFFFFF", max_width=x2 - x1 - 36)

    card((85, 255, 390, 410), "المركز", f"#{rank}", 52, "#08111F", "#FFFFFF50")
    card((548, 255, 852, 410), "النقاط", f"{total}", 52, "#08111F", "#FFFFFF50")
    card((1015, 255, 1320, 410), "أسطورة اليوم", f"{stats['daily_wins'].get(name, 0)}", 52, "#08111F", "#FFFFFF50")

    best_text = f"اليوم {best_day}\n{best_score} نقطة" if best_day != "-" else "-"
    worst_text = f"اليوم {worst_day}\n{worst_score} نقطة" if worst_day != "-" else "-"
    cap_label = best_cap if best_cap and best_cap != "-" else "-"
    cap_text = f"{cap_label}\n{best_cap_pts} نقطة" if cap_label != "-" else "-"
    if best_cap_day not in (None, "-", "") and cap_label != "-":
        cap_text = f"{cap_label}\nاليوم {best_cap_day} - {best_cap_pts} نقطة"

    card((85, 455, 390, 610), "أفضل يوم", best_text, 34, "#0C1628", "#10B98155")
    card((548, 455, 852, 610), "أسوأ يوم", worst_text, 34, "#0C1628", "#EF444455")
    card((1015, 455, 1320, 610), "نسبة المشاركة", pct, 46, "#0C1628", "#3B82F655")

    card((85, 655, 645, 840), "أفضل كابتن", cap_text, 33, "#0C1628", "#F59E0B55")
    card((760, 655, 1320, 840), "انتصارات المواجهات اليومية", f"{wins}", 56, "#0C1628", "#06B6D455")

    draw_text(draw, (700, 905), f"الفترة من اليوم {start_day} إلى {end_day}", get_font(28), fill="#FDE68A")
    path = os.path.join(GENERATED_DIR, f"participant_card_{_safe_filename(name)}.png")
    img.save(path, quality=95)
    return path

async def participant_card_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    query = re.sub(r"^/بطاقة\s*", "", text).strip()
    name = find_participant_name(query)
    if not name:
        await update.message.reply_text("اكتب اسم المشارك بعد الأمر، مثال:\n/بطاقة فارس سالم")
        return
    try:
        path = create_participant_card_image(name)
        await send_photo_path(update, path, f"بطاقة {name} ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء البطاقة ❌\n{e}")


def period_stats(start_day, end_day):
    days = get_existing_days(start_day, end_day)
    totals = Counter()
    cap_points = Counter()
    keeper_points = Counter()
    player_impact = Counter()
    player_zero_popularity = Counter()
    active_counts = Counter()
    for day in days:
        for r in read_day_rows(day):
            if r.get("participated"):
                totals[r["participant"]] += r["total"]
                active_counts[r["participant"]] += 1
                cap_points[r["participant"]] += r.get("captain_points", 0)
                if not is_no_participation(r.get("keeper")):
                    keeper_points[r["keeper"]] += r.get("keeper_points", 0)
                for pk, ptsk in (("p1","p1_points"),("p2","p2_points"),("p3","p3_points")):
                    p = r.get(pk)
                    if not is_no_participation(p):
                        pts = r.get(ptsk,0)
                        extra = r.get("captain_points",0) if r.get("captain") == p else 0
                        player_impact[p] += pts + extra
                        if pts == 0:
                            player_zero_popularity[p] += 1
    return days, totals, cap_points, keeper_points, player_impact, player_zero_popularity, active_counts


def resolve_cup_match(a, b, day, ranks_before):
    rows = {r["participant"]: r for r in read_day_rows(day)}
    ra, rb = rows.get(a), rows.get(b)
    a_play = bool(ra and ra.get("participated"))
    b_play = bool(rb and rb.get("participated"))
    pa = ra.get("total", 0) if ra else 0
    pb = rb.get("total", 0) if rb else 0
    if a_play and not b_play:
        return a, pa, pb, "انسحاب الخصم"
    if b_play and not a_play:
        return b, pa, pb, "انسحاب الخصم"
    if not a_play and not b_play:
        winner = a if ranks_before.get(a, 999) <= ranks_before.get(b, 999) else b
        return winner, pa, pb, "الاثنان لم يشاركا — حُسمت بالأفضلية"
    if pa > pb:
        return a, pa, pb, "فوز بالنقاط"
    if pb > pa:
        return b, pa, pb, "فوز بالنقاط"
    winner = a if ranks_before.get(a, 999) <= ranks_before.get(b, 999) else b
    return winner, pa, pb, "تعادل — حُسمت بالأفضلية"


def compute_fantasy_cup(start_day, end_day):
    """كأس كل 4 أيام: دور تمهيدي ثم ربع ثم نصف ثم نهائي. مناسب لـ12 مشارك."""
    ranks = previous_ranking_before_day(start_day)
    seeds = sorted(PARTICIPANTS, key=lambda n: ranks.get(n, 999))
    days = list(range(int(start_day), int(end_day)+1))
    if len(days) < 4:
        return {"champion": "-", "rounds": [], "note": "الفترة أقل من 4 أيام"}
    d1,d2,d3,d4 = days[:4]
    rounds = []
    # دور 12: أول 4 راحة، 5-12 يلعبون
    bye4 = seeds[:4]
    prelim_pairs = [(seeds[4], seeds[11]), (seeds[5], seeds[10]), (seeds[6], seeds[9]), (seeds[7], seeds[8])] if len(seeds) >= 12 else []
    prelim_winners = []
    for a,b in prelim_pairs:
        w,pa,pb,st = resolve_cup_match(a,b,d1,ranks)
        prelim_winners.append(w)
        rounds.append({"round":"دور 12", "day":d1, "a":a, "b":b, "pa":pa, "pb":pb, "winner":w, "status":st})
    q_players = [bye4[0], prelim_winners[0], bye4[3], prelim_winners[3], bye4[1], prelim_winners[1], bye4[2], prelim_winners[2]] if len(prelim_winners)==4 else seeds[:8]
    q_pairs = [(q_players[0],q_players[1]), (q_players[2],q_players[3]), (q_players[4],q_players[5]), (q_players[6],q_players[7])]
    q_winners = []
    for a,b in q_pairs:
        w,pa,pb,st = resolve_cup_match(a,b,d2,ranks)
        q_winners.append(w)
        rounds.append({"round":"ربع النهائي", "day":d2, "a":a, "b":b, "pa":pa, "pb":pb, "winner":w, "status":st})
    s_pairs = [(q_winners[0], q_winners[1]), (q_winners[2], q_winners[3])]
    s_winners = []
    for a,b in s_pairs:
        w,pa,pb,st = resolve_cup_match(a,b,d3,ranks)
        s_winners.append(w)
        rounds.append({"round":"نصف النهائي", "day":d3, "a":a, "b":b, "pa":pa, "pb":pb, "winner":w, "status":st})
    a,b = s_winners[0], s_winners[1]
    w,pa,pb,st = resolve_cup_match(a,b,d4,ranks)
    rounds.append({"round":"النهائي", "day":d4, "a":a, "b":b, "pa":pa, "pb":pb, "winner":w, "status":st})
    return {"champion": w, "rounds": rounds, "note": ""}


def create_period_report_image(start_day, end_day):
    ensure_generated_dir()
    days, totals, cap_points, keeper_points, player_impact, player_zero, active_counts = period_stats(start_day, end_day)
    if not days:
        raise ValueError("ما فيه أيام في الفترة")

    champion = totals.most_common(1)[0][0] if totals else "-"
    best_cap = cap_points.most_common(1)[0] if cap_points else ("-", 0)
    best_keeper = keeper_points.most_common(1)[0] if keeper_points else ("-", 0)
    best_player = player_impact.most_common(1)[0] if player_impact else ("-", 0)
    disappointment = player_zero.most_common(1)[0] if player_zero else ("-", 0)
    mw = matchup_wins_map(start_day, end_day).most_common(1)
    king_matchups = mw[0] if mw else ("-", 0)
    cup = compute_fantasy_cup(start_day, end_day)
    cup_champ = cup.get("champion", "-")

    img, draw = design_canvas(None, 1400, 1160, "purple")
    draw_design_header(draw, 1400, f"تقرير الفترة من اليوم {start_day} إلى {end_day}", "فانتزي المصيف 2026", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, 1400, 1160, top=205, bottom_pad=92, accent="#F59E0B")

    def info_card(box, title, value, accent="#FFFFFF35", value_size=33):
        x1, y1, x2, y2 = box
        rounded_rect(draw, box, radius=26, fill="#0B1020", outline=accent, width=2)
        draw_text(draw, ((x1 + x2) // 2, y1 + 32), clean_draw_text(title), get_font(27), fill="#E5E7EB")
        draw_text(draw, ((x1 + x2) // 2, (y1 + y2) // 2 + 10), clean_draw_text(value), get_font(value_size), fill="#FFFFFF", max_width=x2 - x1 - 36)

    top_row = [
        ((80, 255, 430, 405), "بطل الفترة", f"{champion}\n{totals.get(champion, 0)} نقطة", "#F59E0B66"),
        ((525, 255, 875, 405), "أكثر اللاعبين صعودا", _period_biggest_rise_text(start_day, end_day), "#10B98166"),
        ((970, 255, 1320, 405), "أكثر اللاعبين تراجعا", _period_biggest_drop_text(start_day, end_day), "#EF444466"),
    ]
    second_row = [
        ((80, 450, 565, 600), "أعلى نقاط يومية في الفترة", _period_best_day_text(start_day, end_day), "#06B6D466"),
        ((615, 450, 1320, 600), "أكثر من فاز بأسطورة اليوم", f"{_period_most_legend_name(start_day, end_day)}\n{_period_most_legend_count(start_day, end_day)} مرات", "#A855F766"),
    ]
    for box, title, value, accent in top_row + second_row:
        info_card(box, title, value, accent, 31)

    draw_text(draw, (700, 655), "أفضل 5 في الفترة", get_font(36), fill="#FDE68A")
    y = 705
    row_colors = ["#F59E0B", "#7C3AED", "#06B6D4", "#F59E0B", "#2563EB"]
    for i, (name, pts) in enumerate(totals.most_common(5), start=1):
        accent = row_colors[(i - 1) % len(row_colors)]
        rounded_rect(draw, (165, y, 1235, y + 56), radius=18, fill="#0B1020", outline=accent, width=2)
        draw_text(draw, (1110, y + 28), f"{i}. {name}", get_font(31), fill="#FFFFFF", max_width=540)
        draw_text(draw, (280, y + 28), f"{pts} نقطة", get_font(31), fill="#FDE68A")
        y += 68

    info_card((80, 1060, 670, 1130), "أفضل كابتن", f"{best_cap[0]}\n{best_cap[1]} نقطة", "#F59E0B66", 29)
    info_card((730, 1060, 1320, 1130), "أفضل حارس", f"{best_keeper[0]}\n{best_keeper[1]} نقطة", "#10B98166", 29)

    path = os.path.join(GENERATED_DIR, f"period_report_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path

async def period_report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if len(nums) >= 2:
        start_day, end_day = nums[0], nums[1]
    else:
        await update.message.reply_text("اكتب الفترة، مثال:\n/تقرير_الفترة 1 4")
        return
    if start_day > end_day:
        start_day, end_day = end_day, start_day
    try:
        path = create_period_report_image(start_day, end_day)
        await send_photo_path(update, path, f"تقرير الفترة {start_day} - {end_day} ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء تقرير الفترة ❌\n{e}")


async def dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        nums = get_numbers(update.message.text)
        if len(nums) >= 2:
            start_day, end_day = nums[0], nums[1]
            if start_day > end_day:
                start_day, end_day = end_day, start_day
        elif len(nums) == 1:
            start_day, end_day = 1, nums[0]
        else:
            start_day, end_day = 1, 31
        await update.message.reply_text("جاري إنشاء ملف الإحصائيات... ⏳")
        file_name, stats = create_dashboard(start_day, end_day)
        if not stats.get("days"):
            await update.message.reply_text("ما لقيت أيام لإحصائيات الداشبورد.")
            return
        add_matchups_sheet_to_dashboard(file_name, start_day, end_day)
        caption = (
            "تم إنشاء ملف الإحصائيات الكامل ✅\n"
            f"النطاق: من اليوم {start_day} إلى اليوم {end_day}\n"
            f"الأيام المحسوبة: {', '.join(map(str, stats['days']))}\n\n"
            "الصفحات تشمل سجل المواجهات ✅"
        )
        with open(file_name, "rb") as file:
            await update.message.reply_document(document=file, filename=file_name, caption=caption)
    except Exception as e:
        await update.message.reply_text(f"صار خطأ أثناء إنشاء الإحصائيات ❌\n\nالسبب:\n{e}")


async def results_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    day = get_day(text)
    if is_locked(day):
        await update.message.reply_text(f"اليوم {day} مقفل ✅\nلفتحه اكتب: /فتح_يوم {day}")
        return
    if not os.path.exists(excel_file(day)):
        await update.message.reply_text(f"ما لقيت ملف اليوم {day}. أضف التشكيلات أولًا.")
        return
    goals_count, clean_sheets = parse_results(text)
    goal_missing, clean_missing = validate_results_names(day, goals_count, clean_sheets)
    backup_files(f"before_results_day_{day}", files=[excel_file(day), LOCKED_FILE, MATCHUPS_FILE])
    file_name = calculate_points(day, goals_count, clean_sheets)
    # المواجهات بعد النتائج فقط
    generate_matchups_for_day(day, force=True)
    goals_text = "\n".join([f"- {name}: {count} هدف = {GOAL_POINTS.get(count, count * 5)} نقطة" for name, count in goals_count.items()]) or "لا يوجد"
    clean_text = "\n".join([f"- {name}" for name in clean_sheets]) or "لا يوجد"
    rows = read_day_rows(day)
    max_score = max([r["total"] for r in rows], default=0)
    winners = [r["participant"] for r in rows if r["total"] == max_score and max_score > 0]
    legends_text = "، ".join(winners) if winners else "لا يوجد"
    warnings = []
    if goal_missing:
        warnings.append("⚠️ هدافون غير موجودين في تشكيلات اليوم:\n" + "\n".join(goal_missing))
    if clean_missing:
        warnings.append("⚠️ حراس كلين شيت غير موجودين في تشكيلات اليوم:\n" + "\n".join(clean_missing))
    caption = (
        f"تم حساب نقاط اليوم {day} ✅\n\n"
        f"الأهداف:\n{goals_text}\n\n"
        f"الكلين شيت:\n{clean_text}\n\n"
        f"🏆 أسطورة اليوم: {legends_text} — {max_score} نقطة\n\n"
        + "\n".join(matchup_lines(day))
    )
    if warnings:
        caption += "\n\n" + "\n\n".join(warnings)
    with open(file_name, "rb") as file:
        await update.message.reply_document(document=file, filename=file_name, caption=caption)
    await update.message.reply_text(build_result_announcement(day, winners))
    if load_settings().get("auto_images", True):
        try:
            path = create_daily_result_image(day, goals_count=goals_count, clean_sheets=clean_sheets)
            await send_photo_path(update, path, build_result_announcement(day, winners))
        except Exception as e:
            await update.message.reply_text(f"تم حساب النتائج، لكن تعذر إنشاء الصورة التلقائية ❌\n{e}")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n/نتائج 5\n/احصائيات\n/احصائيات 1 6\n/ترتيب_نص\n\n"
        "أوامر الصور والتقارير:\n"
        "/صورة_اليوم 6\n/صورة_الترتيب 1 6\n/صورة_الاساطير 1 6\n/صورة_احصائيات 1 6 لوحة عامة\n/صور_الاحصائيات 1 6\n"
        "/بطاقة فارس سالم\n/تقرير_الفترة 1 4\n/تفعيل_الصور_التلقائية\n/إيقاف_الصور_التلقائية\n\n"
        "أوامر التصاميم:\n"
        "/تصميم_مباريات\n/تصميم_مباريات_تلقائي\n"
        "/تصميم_نتائج_مباريات\n/تصميم_نتائج_مباريات_تلقائي\n"
        "/تصميم_ترتيب_مجموعة\n/تصميم_ترتيب_مجموعة_تلقائي\n"
        "/تصميم_هدافين\n/تصميم_هدافين_تلقائي\n"
        "\nالتصميم الإضافي V24:\n"
        "/تصميم_مباريات_ستايل2\n/تصميم_نتائج_مباريات_ستايل2\n"
        "/تصميم_هدافين_ستايل2\n/تصميم_ترتيب_مجموعة_ستايل2\n"
        "/تصميم_مباريات_اطار\n/تصميم_نتائج_مباريات_اطار\n"
        "/تصميم_جميع_المجموعات\n\n"
        "أوامر الفحص والنشر:\n"
        "/الأيام\n/فحص 5\n/مشاركين 5\n/اسطورة 5\n/مقارنة 4 5\n/اعلان_اليوم 5\n/ملخص_اليوم 5\n\n"
        "أوامر الاستيراد والنسخ:\n"
        "/استيراد_ملف — أرسل ملف الإكسل لحاله ثم اكتب الأمر\n/اعتماد_استيراد\n/إلغاء_استيراد\n/نسخة_احتياطية\n/استرجاع_نسخة — أرسل ملف ZIP لحاله ثم اكتب الأمر\n/تنظيف_الأيام\n/تنظيف_الملفات\n\n"
        "أوامر الأمان:\n"
        "/مسح_نتائج 5\n/مسح_يوم 5\n/مسح_الكل تأكيد\n/استرجاع_آخر\n/قفل_يوم 5\n/فتح_يوم 5\n/معرفي"
    )




# ============================================================
# V13 — تصاميم القوالب + التصاميم التلقائية
# الأوامر:
# /تصميم_مباريات + /تصميم_مباريات_تلقائي
# /تصميم_نتائج_مباريات + /تصميم_نتائج_مباريات_تلقائي
# /تصميم_ترتيب_مجموعة + /تصميم_ترتيب_مجموعة_تلقائي
# /تصميم_هدافين + /تصميم_هدافين_تلقائي
# ============================================================

DESIGN_ZIP = "design_assets.zip"
TEMPLATE_DIR = os.path.join("assets", "templates")

def ensure_design_assets():
    logo_ok = os.path.exists(os.path.join("assets", "logos", "masif_logo_mark.png"))
    templates_ok = os.path.exists(TEMPLATE_DIR)
    if templates_ok and logo_ok:
        return
    if os.path.exists(DESIGN_ZIP):
        try:
            import zipfile
            with zipfile.ZipFile(DESIGN_ZIP, "r") as z:
                z.extractall(".")
        except Exception:
            pass

def template_path(name):
    ensure_design_assets()
    p = os.path.join(TEMPLATE_DIR, name)
    return p if os.path.exists(p) else None

def design_canvas(template_name=None, width=1200, height=1500, theme="purple"):
    """
    V16 broadcast-style canvas:
    - very dark background
    - soft green/red/purple edge glows
    - thin inner colored frame
    """
    if template_name:
        p = template_path(template_name)
        if p and os.path.exists(p):
            try:
                img = Image.open(p).convert("RGB").resize((width, height), Image.LANCZOS)
                return img, ImageDraw.Draw(img)
            except Exception:
                pass

    img = Image.new("RGB", (width, height), "#05070D")
    draw = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height, 1)
        val = int(5 + 11 * t)
        draw.line((0, y, width, y), fill=(val, val + 2, val + 8))

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    theme_colors = {
        "purple": ("#22C55E", "#EF4444", "#7C3AED"),
        "blue": ("#06B6D4", "#22C55E", "#EF4444"),
        "gold": ("#F59E0B", "#7C3AED", "#22C55E"),
    }
    left, right, bottom = theme_colors.get(theme, theme_colors["purple"])

    def hx(c):
        c = c.lstrip("#")
        return tuple(int(c[i:i+2], 16) for i in (0, 2, 4))

    od.ellipse((-250, 65, 285, 610), fill=(*hx(left), 70))
    od.ellipse((width-285, 85, width+250, 640), fill=(*hx(right), 65))
    od.ellipse((-170, height-500, 380, height+70), fill=(*hx(bottom), 42))
    od.ellipse((width-380, height-500, width+170, height+70), fill=(*hx("#0EA5E9"), 35))
    overlay = overlay.filter(ImageFilter.GaussianBlur(88))
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    return img, ImageDraw.Draw(img)


def event_logo_path():
    ensure_design_assets()
    for p in [
        os.path.join("assets", "logos", "masif_logo_mark.png"),
        os.path.join("assets", "logos", "masif_logo_full.png"),
    ]:
        if os.path.exists(p):
            return p
    return None

def paste_event_logo(img, width, y=48):
    """
    شعار استراحة المصيف بجانب العنوان.
    يستخدم نسخة mark لأنها أوضح بحجم صغير.
    """
    p = event_logo_path()
    if not p:
        return
    try:
        logo = Image.open(p).convert("RGBA")
        logo.thumbnail((130, 100), Image.LANCZOS)
        badge_w = max(108, logo.width + 24)
        badge_h = max(88, logo.height + 18)
        x = width - badge_w - 64

        # Badge خلف الشعار عشان يبان فوق الخلفية الداكنة
        overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
        od = ImageDraw.Draw(overlay)
        od.rounded_rectangle((x, y, x + badge_w, y + badge_h), radius=24, fill=(5, 7, 13, 135), outline=(253, 230, 138, 90), width=1)
        img_rgba = img.convert("RGBA")
        img_rgba = Image.alpha_composite(img_rgba, overlay)
        img.paste(img_rgba.convert("RGB"))

        lx = x + (badge_w - logo.width) // 2
        ly = y + (badge_h - logo.height) // 2
        img.paste(logo, (lx, ly), logo)
    except Exception:
        pass

def clean_group_title_for_design(title):
    title = str(title or "").strip()
    # تنظيف رموز/إيموجي تسبب مربعات داخل عنوان المجموعة، مع تحويل الحروف الإنجليزية لعربي واضح.
    title = re.sub(r"[🏆⭐🔥⚽️✅❌📌:：\-–—|]+", " ", title)
    title = re.sub(r"\s+", " ", title).strip()
    mapping = {
        "A": "أ", "B": "ب", "C": "ج", "D": "د", "E": "هـ", "F": "و",
        "G": "ز", "H": "ح", "I": "ط", "J": "ي", "K": "ك", "L": "ل",
    }
    def repl(m):
        letter = m.group(1).upper()
        return "المجموعة " + mapping.get(letter, letter)
    title = re.sub(r"المجموعة\s+([A-Za-z])\b", repl, title)
    title = re.sub(r"Group\s+([A-Za-z])\b", repl, title, flags=re.I)
    # لو السطر مجرد A/B/C نحوله قبل إضافة كلمة المجموعة في دوال الرسم.
    if re.fullmatch(r"[A-Za-z]", title):
        title = mapping.get(title.upper(), title.upper())
    return title

def draw_design_header(draw, width, title, subtitle, img=None):
    # V23: بدون شعار في جميع التصاميم
    draw_text(draw, (width//2, 96), title, get_font(54), fill="#FFFFFF")
    draw_text(draw, (width//2, 160), subtitle, get_font(34), fill="#FDE68A")
    draw.line((210, 213, width-210, 213), fill="#FFFFFF45", width=1)

def footer_event(draw, width, height):
    draw.line((300, height-92, width-300, height-92), fill="#FFFFFF35", width=1)
    draw_text(draw, (width//2, height-58), "المصيف ينقل لكم الحدث", get_font(30), fill="#FFFFFF")


def draw_broadcast_inner_frame(draw, width, height, top=250, bottom_pad=118, accent="#22C55E"):
    # إطار خفيف جدًا جدًا — المحتوى كله داخله
    left = 50
    right = width - 50
    bottom = height - bottom_pad
    try:
        draw.rounded_rectangle((left, top, right, bottom), radius=38, outline="#FFFFFF55", width=1)
        draw.rounded_rectangle((left+3, top+3, right-3, bottom-3), radius=35, outline=accent, width=1)
    except Exception:
        draw.rectangle((left, top, right, bottom), outline="#FFFFFF55", width=1)
    return (left, top, right, bottom)

def v16_row_style(index):
    # ألوان صفوف رياضية واضحة داخل الإطار
    return ["#151A26", "#0B1020", "#111827", "#101820"][index % 4]

def v16_accent(index):
    return ["#22C55E", "#EF4444", "#7C3AED", "#06B6D4", "#F59E0B", "#2563EB"][index % 6]

def v16_fit_row_metrics(count, kind="match"):
    # يعطي راحته إذا مباراة وحدة أو 10، ويصغّر تلقائيًا عند الكثرة
    if kind == "standings":
        if count <= 4:
            return 120, 18, 40
        if count <= 6:
            return 108, 14, 36
        return 92, 10, 30
    if kind == "scorers":
        if count <= 5:
            return 118, 18, 38
        if count <= 8:
            return 102, 13, 32
        return 88, 10, 28
    # match/results
    if count <= 2:
        return 145, 24, 40
    if count <= 5:
        return 126, 18, 34
    if count <= 8:
        return 104, 13, 29
    return 90, 10, 25

def parse_match_results_design_text(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    if len(lines) <= 1:
        return "اليوم", []
    idx = 1
    day_name = "اليوم"
    # إذا السطر الثاني ليس نتيجة، نعتبره اسم اليوم
    if idx < len(lines) and "|" not in lines[idx] and not re.search(r"\d+\s*[-–:]\s*\d+", lines[idx]):
        day_name = lines[idx]
        idx += 1
    results = []
    for line in lines[idx:]:
        if "|" in line:
            parts = [p.strip() for p in line.split("|")]
            if len(parts) >= 4:
                try:
                    results.append((parts[0], int(parts[1]), int(parts[2]), parts[3]))
                    continue
                except Exception:
                    pass
            if len(parts) == 2:
                m1 = re.match(r"(.+?)(\d+)\s*$", parts[0])
                m2 = re.match(r"^(\d+)\s*(.+)$", parts[1])
                if m1 and m2:
                    results.append((normalize_name(m1.group(1)), int(m1.group(2)), int(m2.group(1)), normalize_name(m2.group(2))))
                    continue
        m = re.match(r"(.+?)\s+(\d+)\s*[-–:]\s*(\d+)\s+(.+)$", line)
        if m:
            results.append((normalize_name(m.group(1)), int(m.group(2)), int(m.group(3)), normalize_name(m.group(4))))
    return day_name, results

def parse_group_standing_text(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    if len(lines) <= 1:
        return "المجموعة", []
    group_title = lines[1]
    rows = []
    for line in lines[2:]:
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 4:
            team = parts[0]
            played = cell_int(parts[1], 0)
            diff = cell_int(parts[2], 0)
            pts = cell_int(parts[3], 0)
            rows.append((team, played, diff, pts))
    rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
    return group_title, rows

def create_matches_template_image(day_name, matches, use_template=True):
    ensure_generated_dir()
    count = max(len(matches), 1)
    width = 1200
    row_h, gap, name_size = v16_fit_row_metrics(count, "match")
    # تكبير بسيط للصفوف إذا العدد قليل
    if count <= 2:
        row_h += 12
        name_size += 2
    frame_top = 245
    content_h = count * row_h + max(0, count - 1) * gap
    height = max(760, frame_top + content_h + 195)
    img, draw = design_canvas("matches_template.png" if use_template else None, width, height, "purple")
    draw_design_header(draw, width, "مونديال المصيف 2026", f"مباريات اليوم {day_name}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#22C55E")

    available_h = (fy2 - fy1) - 70
    y = fy1 + 38 + max(0, (available_h - content_h) // 2)
    for i, (a, b, t) in enumerate(matches, start=1):
        accent = v16_accent(i-1)
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=28, fill="#0B1020", outline=accent, width=2)
        cy = y + row_h//2

        # V17: تكبير الأعلام
        flag_w = min(150, max(94, row_h - 18))
        paste_flag(img, a, (width-238, cy-flag_w//2, width-238+flag_w, cy+flag_w//2))
        paste_flag(img, b, (238-flag_w, cy-flag_w//2, 238, cy+flag_w//2))

        draw_text(draw, (width-400, cy-4), a, get_font(name_size), fill="#FFFFFF", max_width=285)
        draw_text(draw, (400, cy-4), b, get_font(name_size), fill="#FFFFFF", max_width=285)

        draw_text(draw, (width//2, cy-20), "×", get_font(max(40, name_size+18)), fill="#FDE68A")
        if t:
            badge_w = 214 if row_h >= 104 else 178
            badge_h = 46 if row_h >= 104 else 36
            rounded_rect(draw, (width//2-badge_w//2, cy+24, width//2+badge_w//2, cy+24+badge_h), radius=16, fill="#05070D", outline="#FFFFFF99", width=1)
            draw_text(draw, (width//2, cy+24+badge_h//2), t, get_font(max(20, name_size-12)), fill="#FFFFFF")

        y += row_h + gap

    footer_event(draw, width, height)
    suffix = "template" if use_template else "auto"
    path = os.path.join(GENERATED_DIR, f"matches_{suffix}_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_match_results_template_image(day_name, results, use_template=True):
    ensure_generated_dir()
    count = max(len(results), 1)
    width = 1200
    row_h, gap, name_size = v16_fit_row_metrics(count, "match")
    if count <= 2:
        row_h += 10
        name_size += 2
    frame_top = 245
    content_h = count * row_h + max(0, count - 1) * gap
    height = max(760, frame_top + content_h + 195)
    img, draw = design_canvas("match_results_template.png" if use_template else None, width, height, "blue")
    draw_design_header(draw, width, "مونديال المصيف 2026", f"نتائج مباريات اليوم {day_name}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#06B6D4")

    available_h = (fy2 - fy1) - 70
    y = fy1 + 38 + max(0, (available_h - content_h) // 2)
    for i, (a, sa, sb, b) in enumerate(results, start=1):
        accent = v16_accent(i)
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=28, fill="#0B1020", outline=accent, width=2)
        cy = y + row_h//2

        flag_w = min(146, max(92, row_h - 18))
        paste_flag(img, a, (width-232, cy-flag_w//2, width-232+flag_w, cy+flag_w//2))
        paste_flag(img, b, (232-flag_w, cy-flag_w//2, 232, cy+flag_w//2))

        draw_text(draw, (width-395, cy), a, get_font(name_size), fill="#FFFFFF", max_width=280)
        draw_text(draw, (395, cy), b, get_font(name_size), fill="#FFFFFF", max_width=280)

        badge_w = 230 if row_h >= 104 else 190
        badge_h = 70 if row_h >= 104 else 56
        rounded_rect(draw, (width//2-badge_w//2, cy-badge_h//2, width//2+badge_w//2, cy+badge_h//2), radius=22, fill="#05070D", outline="#FDE68A", width=2)
        draw_text(draw, (width//2, cy), f"{sb} - {sa}", get_font(max(32, name_size+12)), fill="#FDE68A")
        y += row_h + gap

    footer_event(draw, width, height)
    suffix = "template" if use_template else "auto"
    path = os.path.join(GENERATED_DIR, f"match_results_{suffix}_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_group_standing_image(group_title, rows, use_template=True):
    ensure_generated_dir()
    count = max(len(rows), 1)
    width = 1200
    row_h, gap, name_size = v16_fit_row_metrics(count, "standings")
    header_h = 58
    content_h = header_h + 26 + count * row_h + max(0, count - 1) * gap
    height = max(880, 245 + content_h + 180)
    img, draw = design_canvas("group_standing_template.png" if use_template else None, width, height, "purple")
    draw_design_header(draw, width, "مونديال المصيف 2026", clean_group_title_for_design(group_title), img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#22C55E")

    y = fy1 + 36
    rounded_rect(draw, (92, y, width-92, y+header_h), radius=22, fill="#05070D", outline="#FFFFFF66", width=1)
    draw_text(draw, (960, y+header_h//2), "المنتخب", get_font(26), fill="#FFFFFF")
    draw_text(draw, (500, y+header_h//2), "لعب", get_font(24), fill="#FDE68A")
    draw_text(draw, (370, y+header_h//2), "+/-", get_font(24), fill="#FDE68A")
    draw_text(draw, (225, y+header_h//2), "نقاط", get_font(24), fill="#FDE68A")
    y += header_h + 26

    for i, (team, played, diff, pts) in enumerate(rows, start=1):
        accent = v16_accent(i-1)
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=26, fill="#0B1020", outline=accent, width=2)
        cy = y + row_h//2

        # V17: الرقم داخل أكثر من الحافة
        draw_text(draw, (1062, cy), str(i), get_font(max(28, name_size)), fill="#FDE68A" if i == 1 else "#FFFFFF")

        # V17: تكبير أعلام الترتيب
        flag_w = min(120, max(76, row_h-18))
        paste_flag(img, team, (900, cy-flag_w//2, 900+flag_w, cy+flag_w//2))

        draw_text(draw, (710, cy), team, get_font(name_size), fill="#FFFFFF", max_width=330)
        draw_text(draw, (500, cy), str(played), get_font(max(26, name_size-2)), fill="#FFFFFF")
        draw_text(draw, (370, cy), f"{diff:+d}", get_font(max(26, name_size-2)), fill="#FFFFFF")
        draw_text(draw, (225, cy), str(pts), get_font(max(30, name_size+3)), fill="#FDE68A")
        y += row_h + gap

    footer_event(draw, width, height)
    suffix = "template" if use_template else "auto"
    path = os.path.join(GENERATED_DIR, f"group_{suffix}_{_safe_filename(group_title)}.png")
    img.save(path, quality=95)
    return path

def create_top_scorers_template_image(items, use_template=True):
    ensure_generated_dir()
    items = sorted(items, key=lambda x: (-x[1], x[0]))
    count = max(len(items[:10]), 1)
    width = 1200
    row_h, gap, name_size = v16_fit_row_metrics(count, "scorers")
    content_h = count * row_h + max(0, count - 1) * gap
    height = max(830, 245 + content_h + 180)
    img, draw = design_canvas("scorers_template.png" if use_template else None, width, height, "gold")
    draw_design_header(draw, width, "مونديال المصيف 2026", "هدافين البطولة حتى الآن", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=112, accent="#F59E0B")

    available_h = (fy2 - fy1) - 70
    y = fy1 + 38 + max(0, (available_h - content_h) // 2)
    for i, (name, goals, team) in enumerate(items[:10], start=1):
        accent = "#F59E0B" if i == 1 else v16_accent(i)
        fill = "#1A1407" if i == 1 else "#0B1020"
        rounded_rect(draw, (92, y, width-92, y+row_h), radius=26, fill=fill, outline=accent, width=2)
        cy = y + row_h//2

        # V17: الرقم داخل أكثر من الحافة
        draw_text(draw, (1062, cy), str(i), get_font(max(30, name_size+4)), fill="#FDE68A")

        # V17: تكبير العلم
        if team:
            flag_w = min(110, max(68, row_h-18))
            paste_flag(img, team, (885, cy-flag_w//2, 885+flag_w, cy+flag_w//2))

        # مساحة أكبر للاسم
        draw_text(draw, (655, cy), name, get_font(name_size), fill="#FFFFFF", max_width=460)
        draw_text(draw, (245, cy), f"{goals} {'هدف' if goals == 1 else 'أهداف'}", get_font(max(24, name_size-2)), fill="#FDE68A")
        y += row_h + gap

    footer_event(draw, width, height)
    suffix = "template" if use_template else "auto"
    path = os.path.join(GENERATED_DIR, f"scorers_{suffix}.png")
    img.save(path, quality=95)
    return path

def build_group_standing_caption(group_title, rows):
    lines = [f"🏆 {group_title}", ""]
    for i, (team, played, diff, pts) in enumerate(rows, start=1):
        lines.append(f"{i}. {team} — لعب {played} | فارق {diff:+d} | نقاط {pts}")
    lines.append("\nالمصيف ينقل لكم الحدث")
    return "\n".join(lines)

async def design_matches_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, matches = parse_matches_text(update.message.text)
    if not matches:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات\nالسادس\nفرنسا|البرازيل|10:00 م\nالأرجنتين|ألمانيا|12:00 ص")
        return
    try:
        path = create_matches_template_image(day_name, matches, use_template=True)
        await send_photo_path(update, path, build_matches_today_v31_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات بالقالب ❌\n{e}")

async def design_matches_auto_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, matches = parse_matches_text(update.message.text)
    if not matches:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات_تلقائي\nالسادس\nفرنسا|البرازيل|10:00 م")
        return
    try:
        path = create_matches_template_image(day_name, matches, use_template=False)
        await send_photo_path(update, path, build_matches_today_v31_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات التلقائي ❌\n{e}")

async def design_match_results_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, results = parse_match_results_design_text(update.message.text)
    if not results:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_نتائج_مباريات\nالسادس\nالسعودية|2|1|إسبانيا\nفرنسا|3|0|البرازيل")
        return
    try:
        path = create_match_results_template_image(day_name, results, use_template=True)
        await send_photo_path(update, path, build_match_results_caption(results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج المباريات بالقالب ❌\n{e}")

async def design_match_results_auto_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, results = parse_match_results_design_text(update.message.text)
    if not results:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_نتائج_مباريات_تلقائي\nالسادس\nالسعودية|2|1|إسبانيا")
        return
    try:
        path = create_match_results_template_image(day_name, results, use_template=False)
        await send_photo_path(update, path, build_match_results_caption(results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج المباريات التلقائي ❌\n{e}")

async def design_group_standing_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    group_title, rows = parse_group_standing_text(update.message.text)
    if not rows:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_ترتيب_مجموعة\nالمجموعة C\nاسكتلندا|1|1|3\nالبرازيل|1|0|1\nالمغرب|1|0|1\nهايتي|1|-1|0")
        return
    try:
        path = create_group_standing_image(group_title, rows, use_template=True)
        await send_photo_path(update, path, build_group_standing_caption(group_title, rows))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم ترتيب المجموعة بالقالب ❌\n{e}")

async def design_group_standing_auto_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    group_title, rows = parse_group_standing_text(update.message.text)
    if not rows:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_ترتيب_مجموعة_تلقائي\nالمجموعة C\nاسكتلندا|1|1|3\nالبرازيل|1|0|1")
        return
    try:
        path = create_group_standing_image(group_title, rows, use_template=False)
        await send_photo_path(update, path, build_group_standing_caption(group_title, rows))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم ترتيب المجموعة التلقائي ❌\n{e}")

async def design_scorers_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items = parse_scorers_text(update.message.text)
    if not items:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_هدافين\nراؤول خيمينيز|4|المكسيك\nفلوريان فيرتز|3|ألمانيا")
        return
    try:
        path = create_top_scorers_template_image(items, use_template=True)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين بالقالب ❌\n{e}")

async def design_scorers_auto_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items = parse_scorers_text(update.message.text)
    if not items:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_هدافين_تلقائي\nراؤول خيمينيز|4|المكسيك")
        return
    try:
        path = create_top_scorers_template_image(items, use_template=False)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين التلقائي ❌\n{e}")


# ============================================================
# V23 إضافات التصميم الإضافي:
# ستايل 2 أزرق/ذهبي + ستايل 3 إطار + جميع المجموعات
# ============================================================

def _style2_canvas(width, height):
    img = Image.new("RGB", (width, height), "#061633")
    d = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height, 1)
        d.line((0, y, width, y), fill=(3, int(24 + 20*t), int(70 + 30*t)))

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    od.ellipse((80, -210, width-80, 560), fill=(15, 92, 190, 92))
    od.ellipse((width*0.55, 330, width*1.20, 1050), fill=(0, 120, 255, 46))
    od.rectangle((width*0.61, 430, width*0.90, 640), fill=(0, 115, 255, 36))
    od.ellipse((-270, 500, 300, height+120), fill=(0, 70, 180, 70))
    overlay = overlay.filter(ImageFilter.GaussianBlur(28))
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    d = ImageDraw.Draw(img)

    for i in range(0, 64, 7):
        d.arc((80+i*2, 20+i, width-80-i*2, 720+i), 190, 335, fill="#1D6FEA55", width=2)
    for i in range(18):
        x = 60 + i * 18
        y = height - 215 + (i % 3) * 8
        d.ellipse((x, y, x+5, y+5), fill="#FBBF2435")
    return img, d

def _ampm_from_time(t):
    t = normalize_name(t)
    if "ص" in t or "AM" in t.upper():
        period = "ص"
    elif "م" in t or "PM" in t.upper():
        period = "م"
    else:
        period = ""
    tm = re.sub(r"(ص|م|AM|PM|am|pm)", "", t).strip()
    return tm, period

def create_matches_style2_image(day_name, matches):
    ensure_generated_dir()
    count = max(len(matches), 1)
    width = 1200

    if count == 1:
        row_h, gap, header_h, footer_h, card_pad = 245, 0, 330, 95, 95
    elif count == 2:
        row_h, gap, header_h, footer_h, card_pad = 215, 28, 320, 95, 115
    elif count <= 4:
        row_h, gap, header_h, footer_h, card_pad = 175, 22, 310, 95, 160
    else:
        row_h, gap, header_h, footer_h, card_pad = 150, 15, 290, 90, 150

    height = max(820, header_h + count * row_h + max(0, count-1) * gap + footer_h)
    img, draw = _style2_canvas(width, height)

    draw_text(draw, (width//2, 62), "MONDIAL AL MASEEF 2026", get_font(30), fill="#FFFFFF")
    draw_text(draw, (width//2, 132), "مباريات اليوم", get_font(80 if count <= 2 else 72), fill="#FFFFFF")
    rounded_rect(draw, (width//2-235, 202, width//2+235, 260), radius=18, fill="#FBBF24", outline="#00000055", width=2)
    draw_text(draw, (width//2, 232), f"اليوم {day_name}", get_font(30), fill="#061633")

    y = header_h
    if count == 1:
        y = 350

    for a, b, t in matches:
        x1, x2 = card_pad, width-card_pad
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=32, fill="#0638A5", outline="#14B8F5", width=4)

        # نفس روح الأطراف الملونة
        draw.line((x1+14, y+row_h-4, x2-14, y+row_h-4), fill="#EF4444", width=4)
        draw.line((x1+170, y+6, x2-170, y+6), fill="#22C55E", width=3)
        draw.line((x1+390, y+6, x1+470, y+6), fill="#FBBF24", width=3)
        draw.arc((x1, y, x1+58, y+58), 180, 270, fill="#14B8F5", width=8)
        draw.arc((x2-58, y+row_h-58, x2, y+row_h), 0, 90, fill="#EF4444", width=8)

        cy = y + row_h//2
        flag_w = min(158 if count <= 2 else 132, row_h-48)

        paste_flag(img, a, (x1+42, cy-flag_w//2, x1+42+flag_w, cy+flag_w//2))
        paste_flag(img, b, (x2-42-flag_w, cy-flag_w//2, x2-42, cy+flag_w//2))

        name_font = get_font(34 if count <= 2 else 27)
        draw_text(draw, (x1+42+flag_w//2, y+row_h-38), a, name_font, fill="#FFFFFF", max_width=260)
        draw_text(draw, (x2-42-flag_w//2, y+row_h-38), b, name_font, fill="#FFFFFF", max_width=260)

        tm, period = _ampm_from_time(t)
        time_font = get_font(68 if count == 1 else (60 if count == 2 else 52))
        draw_text(draw, (width//2, cy-18), tm, time_font, fill="#FBBF24")
        if period:
            draw_text(draw, (width//2, cy+46), period, get_font(36), fill="#FBBF24")
        elif count == 1:
            draw_text(draw, (width//2, cy+48), "بتوقيت السعودية", get_font(26), fill="#E5E7EB")

        y += row_h + gap

    draw_text(draw, (width//2, height-44), "المصيف ينقل لكم الحدث", get_font(30), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, f"matches_style2_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_match_results_style2_image(day_name, results):
    ensure_generated_dir()
    count = max(len(results), 1)
    width = 1200

    if count == 1:
        row_h, gap, header_h, footer_h, card_pad = 245, 0, 330, 95, 95
    elif count == 2:
        row_h, gap, header_h, footer_h, card_pad = 215, 28, 320, 95, 115
    elif count <= 4:
        row_h, gap, header_h, footer_h, card_pad = 175, 22, 310, 95, 160
    else:
        row_h, gap, header_h, footer_h, card_pad = 150, 15, 290, 90, 150

    height = max(820, header_h + count * row_h + max(0, count-1) * gap + footer_h)
    img, draw = _style2_canvas(width, height)

    draw_text(draw, (width//2, 62), "MONDIAL AL MASEEF 2026", get_font(30), fill="#FFFFFF")
    draw_text(draw, (width//2, 132), "نتائج اليوم", get_font(80 if count <= 2 else 72), fill="#FFFFFF")
    rounded_rect(draw, (width//2-235, 202, width//2+235, 260), radius=18, fill="#FBBF24", outline="#00000055", width=2)
    draw_text(draw, (width//2, 232), f"اليوم {day_name}", get_font(30), fill="#061633")

    y = 350 if count == 1 else header_h
    for a, sa, sb, b in results:
        x1, x2 = card_pad, width-card_pad
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=32, fill="#0638A5", outline="#14B8F5", width=4)
        draw.line((x1+14, y+row_h-4, x2-14, y+row_h-4), fill="#EF4444", width=4)
        draw.line((x1+170, y+6, x2-170, y+6), fill="#22C55E", width=3)
        draw.arc((x1, y, x1+58, y+58), 180, 270, fill="#14B8F5", width=8)
        draw.arc((x2-58, y+row_h-58, x2, y+row_h), 0, 90, fill="#EF4444", width=8)

        cy = y + row_h//2
        flag_w = min(158 if count <= 2 else 132, row_h-48)
        paste_flag(img, a, (x1+42, cy-flag_w//2, x1+42+flag_w, cy+flag_w//2))
        paste_flag(img, b, (x2-42-flag_w, cy-flag_w//2, x2-42, cy+flag_w//2))

        name_font = get_font(34 if count <= 2 else 27)
        draw_text(draw, (x1+42+flag_w//2, y+row_h-38), a, name_font, fill="#FFFFFF", max_width=260)
        draw_text(draw, (x2-42-flag_w//2, y+row_h-38), b, name_font, fill="#FFFFFF", max_width=260)

        rounded_rect(draw, (width//2-125, cy-50, width//2+125, cy+50), radius=20, fill="#FBBF24", outline="#00000088", width=2)
        draw_text(draw, (width//2, cy), f"{sb} - {sa}", get_font(60 if count <= 2 else 52), fill="#061633")
        y += row_h + gap

    draw_text(draw, (width//2, height-44), "المصيف ينقل لكم الحدث", get_font(30), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, f"results_style2_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_scorers_style2_image(items):
    ensure_generated_dir()
    items = sorted(items, key=lambda x: (-x[1], x[0]))[:12]
    count = max(len(items), 1)
    width = 1200
    row_h = 105 if count <= 8 else 88
    gap = 14 if count <= 8 else 10
    height = max(900, 310 + count*row_h + (count-1)*gap + 90)
    img, draw = _style2_canvas(width, height)
    draw_text(draw, (width//2, 80), "MONDIAL AL MASEEF 2026", get_font(30), fill="#FFFFFF")
    draw_text(draw, (width//2, 150), "هدافين البطولة", get_font(72), fill="#FFFFFF")
    rounded_rect(draw, (width//2-210, 215, width//2+210, 268), radius=18, fill="#FBBF24", outline="#00000055", width=2)
    draw_text(draw, (width//2, 242), "تحديث مستمر", get_font(28), fill="#061633")
    y = 315
    for i, (name, goals, team) in enumerate(items, start=1):
        rounded_rect(draw, (120, y, width-120, y+row_h), radius=24, fill="#0638A5", outline="#14B8F5", width=3)
        cy = y + row_h//2
        draw_text(draw, (1030, cy), str(i), get_font(34), fill="#FBBF24")
        if team:
            paste_flag(img, team, (850, cy-38, 930, cy+38))
        draw_text(draw, (620, cy), name, get_font(34), fill="#FFFFFF", max_width=430)
        draw_text(draw, (240, cy), f"{goals} {'هدف' if goals == 1 else 'أهداف'}", get_font(30), fill="#FBBF24")
        y += row_h + gap
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(28), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, "scorers_style2.png")
    img.save(path, quality=95)
    return path

def create_group_style2_image(group_title, rows):
    ensure_generated_dir()
    count = max(len(rows), 1)
    width = 1200
    row_h = 105 if count <= 6 else 86
    gap = 14 if count <= 6 else 10
    height = max(870, 310 + count*row_h + (count-1)*gap + 90)
    img, draw = _style2_canvas(width, height)
    draw_text(draw, (width//2, 80), "MONDIAL AL MASEEF 2026", get_font(30), fill="#FFFFFF")
    draw_text(draw, (width//2, 150), "ترتيب المجموعة", get_font(72), fill="#FFFFFF")
    rounded_rect(draw, (width//2-220, 215, width//2+220, 268), radius=18, fill="#FBBF24", outline="#00000055", width=2)
    draw_text(draw, (width//2, 242), clean_group_title_for_design(group_title), get_font(30), fill="#061633")
    y = 315
    for i, (team, played, diff, pts) in enumerate(rows, start=1):
        rounded_rect(draw, (120, y, width-120, y+row_h), radius=24, fill="#0638A5", outline="#14B8F5", width=3)
        cy = y + row_h//2
        draw_text(draw, (1040, cy), str(i), get_font(34), fill="#FBBF24")
        paste_flag(img, team, (870, cy-40, 950, cy+40))
        draw_text(draw, (690, cy), team, get_font(34), fill="#FFFFFF", max_width=330)
        draw_text(draw, (420, cy), f"لعب {played}", get_font(24), fill="#E5E7EB")
        draw_text(draw, (300, cy), f"{int(diff):+d}", get_font(28), fill="#E5E7EB")
        draw_text(draw, (190, cy), str(pts), get_font(34), fill="#FBBF24")
        y += row_h + gap
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(28), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, f"group_style2_{_safe_filename(group_title)}.png")
    img.save(path, quality=95)
    return path

def parse_all_groups_text(text):
    groups = []
    current_title = None
    current_rows = []
    for raw in (text or "").splitlines()[1:]:
        line = raw.strip()
        if not line:
            continue
        if "|" not in line:
            if current_title and current_rows:
                groups.append((current_title, current_rows))
                current_rows = []
            current_title = line
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 4 and current_title:
            try:
                current_rows.append((parts[0], int(parts[1]), int(parts[2]), int(parts[3])))
            except Exception:
                pass
    if current_title and current_rows:
        groups.append((current_title, current_rows))
    return groups

def create_all_groups_image(groups):
    ensure_generated_dir()
    width, height = 1800, 2400
    img, draw = _style2_canvas(width, height)
    draw_text(draw, (width//2, 90), "MONDIAL AL MASEEF 2026", get_font(40), fill="#FFFFFF")
    draw_text(draw, (width//2, 165), "ترتيب جميع المجموعات", get_font(72), fill="#FFFFFF")
    cols = 3
    margin_x, gap_x = 80, 38
    card_w = (width - 2*margin_x - (cols-1)*gap_x) // cols
    card_h = 475
    start_y = 260
    gap_y = 42
    for idx, (title, rows) in enumerate(groups[:12]):
        c = (cols - 1) - (idx % cols) if cols == 2 else 0
        r = idx // cols
        x = margin_x + c*(card_w+gap_x)
        y = start_y + r*(card_h+gap_y)
        rounded_rect(draw, (x, y, x+card_w, y+card_h), radius=28, fill="#0638A5EE", outline="#14B8F5", width=3)
        rounded_rect(draw, (x+18, y+18, x+card_w-18, y+68), radius=18, fill="#FBBF24", outline="#00000055", width=1)
        draw_text(draw, (x+card_w//2, y+43), clean_group_title_for_design(title), get_font(26), fill="#061633")
        yy = y + 92
        for pos, (team, played, diff, pts) in enumerate(rows[:4], start=1):
            rounded_rect(draw, (x+20, yy, x+card_w-20, yy+72), radius=16, fill="#061633AA", outline="#FFFFFF22", width=1)
            cy = yy + 36
            draw_text(draw, (x+card_w-45, cy), str(pos), get_font(22), fill="#FBBF24")
            paste_flag(img, team, (x+card_w-125, cy-24, x+card_w-75, cy+24))
            draw_text(draw, (x+card_w-240, cy), team, get_font(22), fill="#FFFFFF", max_width=190)
            draw_text(draw, (x+155, cy), str(pts), get_font(26), fill="#FBBF24")
            draw_text(draw, (x+75, cy), f"{diff:+d}", get_font(21), fill="#E5E7EB")
            yy += 84
    draw_text(draw, (width//2, height-70), "المصيف ينقل لكم الحدث", get_font(36), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, "all_groups_style2.png")
    img.save(path, quality=95)
    return path

def _frame_canvas(width, height):
    img = Image.new("RGB", (width, height), "#07111B")
    d = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height, 1)
        d.line((0, y, width, y), fill=(5, int(18+20*t), int(30+24*t)))
    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)
    od.ellipse((-200, -120, 440, 460), fill=(10,200,120,30))
    od.ellipse((width-360, height-420, width+260, height+160), fill=(70,90,255,40))
    img = Image.alpha_composite(img.convert("RGBA"), overlay.filter(ImageFilter.GaussianBlur(40))).convert("RGB")
    return img, ImageDraw.Draw(img)

def create_match_frame_style_image(day_name, items, is_results=False):
    ensure_generated_dir()
    count = max(len(items), 1)
    width = 1200
    row_h = 126 if count <= 6 else 104
    gap = 18 if count <= 6 else 12
    height = max(620, 220 + count*row_h + (count-1)*gap + 105)
    img, draw = _frame_canvas(width, height)
    draw_text(draw, (width//2, 80), "مونديال المصيف 2026", get_font(50), fill="#FFFFFF")
    draw_text(draw, (width//2, 140), f"{'نتائج' if is_results else 'مباريات'} اليوم {day_name}", get_font(34), fill="#FDE68A")
    y = 215
    for item in items:
        if is_results:
            a, sa, sb, b = item
            center = f"{sb} - {sa}"
        else:
            a, b, t = item
            center = t
        x1, x2 = 120, width-120
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=28, fill="#07110FCC", outline="#22C55E", width=3)
        draw.line((x1+5, y+row_h-5, x1+260, y+row_h-5), fill="#F97316", width=10)
        draw.line((x2-260, y+row_h-5, x2-5, y+row_h-5), fill="#2563EB", width=10)
        draw.line((x1+10, y+5, x2-10, y+5), fill="#22C55E", width=4)
        cy = y + row_h//2
        flag_w = min(84, row_h-36)
        paste_flag(img, a, (x2-160, cy-flag_w//2, x2-160+flag_w, cy+flag_w//2))
        paste_flag(img, b, (x1+80, cy-flag_w//2, x1+80+flag_w, cy+flag_w//2))
        draw_text(draw, (x2-320, cy), a, get_font(30), fill="#FFFFFF", max_width=270)
        draw_text(draw, (x1+280, cy), b, get_font(30), fill="#FFFFFF", max_width=270)
        rounded_rect(draw, (width//2-95, cy-38, width//2+95, cy+38), radius=18, fill="#B8FFF0", outline="#FFFFFFAA", width=2)
        draw_text(draw, (width//2, cy), center, get_font(34), fill="#061633")
        y += row_h + gap
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(28), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"{'results' if is_results else 'matches'}_frame_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path



def build_design_matches_caption(day_name, matches):
    lines = [
        "🏆 مونديال المصيف 2026 🏆",
        f"🔥 مباريات اليوم ( {day_name} ) 🔥",
        ""
    ]
    for a, b, t in matches:
        lines.append(f"{a} × {b} — {t}")
    lines.append("")
    lines.append("المصيف ينقل لكم الحدث")
    return "\n".join(lines)


def build_matches_today_v31_caption(day_name, matches):
    """كابشن خاص لأوامر /مباريات_اليوم و /مباريات_اليوم2:
    يرسل المباريات ثم نموذج المشاركة الرسمي في آخر الرسالة.
    """
    caption = build_design_matches_caption(day_name, matches)
    form = [
        "",
        "📋 نموذج المشاركة الرسمي المعتمد",
        "🏆 تشكيلة الفانتزي - اليوم (    )",
        "🧤 الحارس:",
        " اللاعب 1:",
        " اللاعب 2:",
        " اللاعب 3:",
        "👑 الكابتن :",
    ]
    return caption + "\n" + "\n".join(form)


async def design_matches_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات_ستايل2\nالسابع\nالبرتغال|الكونغو الديمقراطية|8:00 م")
            return
        path = create_matches_style2_image(day_name, matches)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات ستايل2 ❌\n{e}")

async def design_match_results_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, results = parse_match_results_design_text(update.message.text)
        if not results:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_نتائج_مباريات_ستايل2\nالسابع\nالبرتغال|2|1|الكونغو الديمقراطية")
            return
        path = create_match_results_style2_image(day_name, results)
        await send_photo_path(update, path, build_match_results_caption(results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج ستايل2 ❌\n{e}")

async def design_scorers_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        items = parse_scorers_text(update.message.text)
        if not items:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_هدافين_ستايل2\nراؤول خيمينيز|4|المكسيك")
            return
        path = create_scorers_style2_image(items)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين ستايل2 ❌\n{e}")

async def design_group_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        group_title, rows = parse_group_standing_text(update.message.text)
        if not rows:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_ترتيب_مجموعة_ستايل2\nالمجموعة C\nالبرازيل|1|0|1")
            return
        path = create_group_style2_image(group_title, rows)
        await send_photo_path(update, path, build_group_standing_caption(group_title, rows))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم ترتيب المجموعة ستايل2 ❌\n{e}")

async def design_matches_frame_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات_اطار\nالسابع\nالبرتغال|الكونغو الديمقراطية|8:00 م")
            return
        path = create_match_frame_style_image(day_name, matches, False)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات الإطار ❌\n{e}")

async def design_results_frame_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, results = parse_match_results_design_text(update.message.text)
        if not results:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_نتائج_مباريات_اطار\nالسابع\nالبرتغال|2|1|الكونغو الديمقراطية")
            return
        path = create_match_frame_style_image(day_name, results, True)
        await send_photo_path(update, path, build_match_results_caption(results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج الإطار ❌\n{e}")

async def design_all_groups_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        groups = parse_all_groups_text(update.message.text)
        if not groups:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_جميع_المجموعات\nالمجموعة A\nفريق|1|0|3\n...\nالمجموعة B\nفريق|1|0|3")
            return
        path = create_all_groups_image(groups)
        await send_photo_path(update, path, "ترتيب جميع المجموعات ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم جميع المجموعات ❌\n{e}")


# -------------------- V4 FIX: Google/SerpApi live results, safe API checks, no silent hangs --------------------
# هذه الدوال تُعيد تعريف دوال V3 السابقة وتبقي باقي البوت كما هو.


def _http_json_get(url, params=None, headers=None, timeout=10):
    """GET JSON using requests if available, otherwise urllib. Keeps bot alive even if requirements lag."""
    params = params or {}
    headers = headers or {}
    base_headers = {
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15",
        "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
    }
    base_headers.update(headers)
    if requests is not None:
        r = requests.get(url, params=params, headers=base_headers, timeout=timeout)
        try:
            data = r.json()
        except Exception:
            raise RuntimeError(f"رد غير مفهوم من المصدر: HTTP {getattr(r, 'status_code', '')}")
        if int(getattr(r, 'status_code', 200) or 200) >= 400:
            err = data.get("error") if isinstance(data, dict) else None
            raise RuntimeError(str(err or f"HTTP {getattr(r, 'status_code', '')}"))
        return data
    try:
        import urllib.request, urllib.parse
        qs = urllib.parse.urlencode(params, doseq=True)
        full_url = url + (("&" if "?" in url else "?") + qs if qs else "")
        req = urllib.request.Request(full_url, headers=base_headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            enc = resp.headers.get_content_charset() or "utf-8"
            text = raw.decode(enc, errors="replace")
            try:
                return json.loads(text)
            except Exception:
                raise RuntimeError(f"رد غير JSON من المصدر: HTTP {getattr(resp, 'status', '')}")
    except Exception as e:
        raise RuntimeError(str(e)[:220])


def serpapi_search_json(query, hl="ar", gl="sa", timeout=9):
    key = _serpapi_key()
    if not key:
        raise RuntimeError("SERPAPI_KEY غير موجود في Railway")
    data = _http_json_get(
        "https://serpapi.com/search.json",
        params={"engine": "google", "q": query, "hl": hl, "gl": gl, "api_key": key},
        timeout=timeout,
    )
    if isinstance(data, dict) and data.get("error"):
        raise RuntimeError(str(data.get("error"))[:220])
    return data


def _api_football_get(endpoint, params=None):
    key = _api_football_key()
    if not key:
        raise RuntimeError("API_FOOTBALL_KEY غير موجود")
    data = _http_json_get(
        "https://v3.football.api-sports.io" + endpoint,
        params=params or {},
        headers={"x-apisports-key": key},
        timeout=10,
    )
    if isinstance(data, dict) and data.get("errors"):
        errs = data.get("errors")
        if errs:
            raise RuntimeError(str(errs)[:220])
    return data


def _blob_contains_team(blob, team):
    team = canonical_team_name(team) or normalize_name(team)
    variants = [team] + (TEAM_SEARCH_EN.get(team) or [])
    b_low = str(blob).lower()
    b_simple = simple_key(str(blob))
    for v in variants:
        if not v:
            continue
        if str(v).lower() in b_low or simple_key(v) in b_simple:
            return True
    return False


def _match_pair_names(n1, n2, req1, req2):
    return _teams_match(n1, n2, req1, req2) or (
        _blob_contains_team(n1, req1) and _blob_contains_team(n2, req2)
    ) or (
        _blob_contains_team(n1, req2) and _blob_contains_team(n2, req1)
    )


def _score_from_any(item):
    if not isinstance(item, dict):
        return "0"
    for key in ["score", "points", "goals", "result", "value", "display_score", "displayScore"]:
        if key in item and item.get(key) not in (None, ""):
            return _extract_score_value(item.get(key))
    # Some Google nodes use nested score objects
    for key in ["standing", "stats", "score_data"]:
        if isinstance(item.get(key), dict):
            val = _score_from_any(item.get(key))
            if val != "0":
                return val
    return "0"


def _team_name_from_any(item):
    if isinstance(item, str):
        return normalize_name(item)
    if not isinstance(item, dict):
        return ""
    for key in ["name", "title", "team", "short_name", "shortName", "displayName", "display_name", "full_name"]:
        v = item.get(key)
        if isinstance(v, str) and normalize_name(v):
            return canonical_team_name(v) or normalize_name(v)
        if isinstance(v, dict):
            nested = _team_name_from_any(v)
            if nested:
                return nested
    return ""


def _collect_scorers_from_serp_node(node):
    out = []
    def add_text(x):
        s = normalize_name(x)
        if s and ("'" in s or "هدف" in s or "Goal" in s or "⚽" in s or re.search(r"\b\d{1,3}\b", s)):
            out.append(s)
    for n in _walk_json(node):
        if isinstance(n, dict):
            # scorer-ish dicts
            player = n.get("player") or n.get("name") or n.get("title")
            minute = n.get("time") or n.get("minute") or n.get("elapsed")
            event_type = normalize_name(n.get("type") or n.get("event") or n.get("detail") or n.get("description") or "")
            if player and ("goal" in event_type.lower() or "هدف" in event_type or minute):
                txt = normalize_name(player)
                if minute and str(minute) not in txt:
                    txt += f" {minute}'"
                add_text(txt)
            for k in ["text", "description", "subtitle", "detail", "summary"]:
                if isinstance(n.get(k), str):
                    add_text(n.get(k))
        elif isinstance(n, str):
            add_text(n)
    # Clean false positives/long blobs
    clean = []
    for s in out:
        if len(s) <= 80 and s not in clean:
            clean.append(s)
    return clean[:8]


def _status_from_serp_node(node):
    vals = []
    if isinstance(node, dict):
        for k in ["status", "game_status", "match_status", "status_text", "status_line", "state", "period", "time", "date"]:
            if node.get(k) not in (None, ""):
                vals.append(str(node.get(k)))
    raw = " | ".join(vals)
    return _norm_status_ar(raw) if raw else "غير محدد"


def _parse_serp_sports_node(node, req1, req2):
    """Parse most known SerpApi/Google Sports shapes."""
    if not isinstance(node, dict):
        return None
    # Shape 1: teams/players list
    for key in ["teams", "players", "competitors", "participants"]:
        arr = node.get(key)
        if isinstance(arr, list) and len(arr) >= 2:
            parsed = []
            for item in arr[:2]:
                nm = _team_name_from_any(item)
                sc = _score_from_any(item) if isinstance(item, dict) else "0"
                if nm:
                    parsed.append((canonical_team_name(nm) or normalize_name(nm), sc))
            if len(parsed) >= 2 and _match_pair_names(parsed[0][0], parsed[1][0], req1, req2):
                return {
                    "team1": parsed[0][0], "team2": parsed[1][0],
                    "score1": parsed[0][1], "score2": parsed[1][1],
                    "status": _status_from_serp_node(node),
                    "minute": "",
                    "scorers": _collect_scorers_from_serp_node(node),
                    "source": "Google Sports",
                }
    # Shape 2: home/away dicts
    home = node.get("home") or node.get("home_team") or node.get("team_home")
    away = node.get("away") or node.get("away_team") or node.get("team_away")
    if isinstance(home, dict) and isinstance(away, dict):
        n1, n2 = _team_name_from_any(home), _team_name_from_any(away)
        if n1 and n2 and _match_pair_names(n1, n2, req1, req2):
            return {
                "team1": canonical_team_name(n1) or n1, "team2": canonical_team_name(n2) or n2,
                "score1": _score_from_any(home), "score2": _score_from_any(away),
                "status": _status_from_serp_node(node), "minute": "",
                "scorers": _collect_scorers_from_serp_node(node), "source": "Google Sports",
            }
    # Shape 3: explicit names/scores
    n1 = node.get("team1") or node.get("home_name") or node.get("homeTeam")
    n2 = node.get("team2") or node.get("away_name") or node.get("awayTeam")
    if isinstance(n1, dict): n1 = _team_name_from_any(n1)
    if isinstance(n2, dict): n2 = _team_name_from_any(n2)
    if n1 and n2 and _match_pair_names(str(n1), str(n2), req1, req2):
        return {
            "team1": canonical_team_name(n1) or normalize_name(n1),
            "team2": canonical_team_name(n2) or normalize_name(n2),
            "score1": _extract_score_value(node.get("score1") or node.get("home_score") or node.get("homeScore") or 0),
            "score2": _extract_score_value(node.get("score2") or node.get("away_score") or node.get("awayScore") or 0),
            "status": _status_from_serp_node(node), "minute": "",
            "scorers": _collect_scorers_from_serp_node(node), "source": "Google Sports",
        }
    return None


def _serpapi_query_candidates(team1, team2, date_hint=None):
    ar1, ar2 = canonical_team_name(team1) or normalize_name(team1), canonical_team_name(team2) or normalize_name(team2)
    en1s, en2s = team_query_names(ar1), team_query_names(ar2)
    qs = []
    if date_hint:
        qs.append(f"مباراة {ar1} {ar2} {date_hint}")
        qs.append(f"{en1s[1] if len(en1s)>1 else en1s[0]} vs {en2s[1] if len(en2s)>1 else en2s[0]} {date_hint} FIFA World Cup 2026")
    qs.extend([
        f"مباراة {ar1} {ar2}",
        f"{en1s[1] if len(en1s)>1 else en1s[0]} vs {en2s[1] if len(en2s)>1 else en2s[0]}",
        f"{en1s[-1]} {en2s[-1]} score FIFA World Cup 2026",
    ])
    out, seen = [], set()
    for q in qs:
        if q and q not in seen:
            seen.add(q); out.append(q)
    return out


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    last_data = None
    for q in _serpapi_query_candidates(req1, req2, date_hint):
        for hl, gl in [("ar", "sa"), ("en", "us")]:
            try:
                data = serpapi_search_json(q, hl=hl, gl=gl, timeout=8)
                last_data = data
            except Exception:
                continue
            roots = []
            sr = data.get("sports_results") if isinstance(data, dict) else None
            if isinstance(sr, dict):
                roots.append(sr)
                # Prioritize common spotlight/game containers
                for k in ["game_spotlight", "games", "matches", "scoreboard"]:
                    if isinstance(sr.get(k), (dict, list)):
                        roots.append(sr.get(k))
            roots.append(data)
            for root in roots:
                for node in _walk_json(root):
                    obj = _parse_serp_sports_node(node, req1, req2)
                    if obj:
                        return obj
            # Very cautious fallback: sports_results exists and blob contains both teams + score pattern
            blob = json.dumps(sr or data, ensure_ascii=False)
            if _blob_contains_team(blob, req1) and _blob_contains_team(blob, req2):
                # Look for final score in common textual snippets. Keep requested order.
                m = re.search(r"(\d+)\s*[-–]\s*(\d+)", blob)
                if m and len(m.group(1)) <= 2 and len(m.group(2)) <= 2:
                    return {
                        "team1": req1, "team2": req2,
                        "score1": m.group(1), "score2": m.group(2),
                        "status": "حسب نتائج Google", "minute": "",
                        "scorers": _collect_scorers_from_serp_node(sr or data),
                        "source": "Google Search",
                    }
    return None


def _source_mode_sequence(mode):
    m = normalize_name(mode or "official").lower()
    if m in ["سريع", "fast"]:
        return ["google"]
    if m in ["الأحدث", "الاحدث", "latest"]:
        return ["google", "api", "espn", "fifa"]
    return ["api", "fifa", "espn", "google"]


def fetch_live_match_data(team1, team2, mode="official", date_hint=None):
    for src in _source_mode_sequence(mode):
        try:
            if src == "google":
                obj = fetch_match_from_serpapi(team1, team2, date_hint=date_hint)
            elif src == "api":
                obj = fetch_match_from_api_football(team1, team2, date_hint=date_hint)
            elif src == "espn":
                obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
            else:
                obj = fetch_match_from_fifa(team1, team2)
            if obj:
                if not obj.get("status"):
                    obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
                return obj
        except Exception:
            continue
    return None


def fetch_current_groups(mode="official"):
    """لا نعرض أي ترتيب إلا من مصدر واضح. تجنب ترتيب عشوائي."""
    m = normalize_name(mode or "official").lower()
    # ملاحظة: Google Sports standings parsing مختلف حسب البلد/اللغة؛ لا نعتمده إلا لو استطعنا قراءة 8/12 مجموعات بشكل صريح.
    # API-Football official if league ID resolves.
    if m in ["رسمي", "official", "الأحدث", "الاحدث", "latest"]:
        try:
            groups = fetch_standings_from_api_football()
            if groups and len(groups) >= 8:
                return groups, "API-Football"
        except Exception:
            pass
    return [], ""


def _source_help_text(kind, mode):
    if kind == "standings":
        return (
            f"تعذر جلب ترتيب مجموعات مؤكد من مصدر {mode_label_ar(mode)} ❌\n"
            "لن أعرض ترتيبًا غير موثوق.\n\n"
            "للتحديث اليدوي استخدم /قالب_المجموعات ثم /كل_المجموعات."
        )
    return f"تعذر جلب البيانات من مصدر {mode_label_ar(mode)} ❌"



# ==================== V25 FIXTURES DESIGN OVERRIDE ====================
# هذا القسم يتعمد تعريف دوال /مباريات مرة أخيرة قبل التشغيل.
# الهدف:
# - /مباريات 20/06 = نفس تصميم /مباريات_اليوم الرئيسي (تصميم 2)
# - اختيار يوم من الأزرار = يظهر زرين تصميم 1 وتصميم 2
# - تصميم 1 = مضغوط بخلفية التمثال بدون الكلام الكبير فوق
# - تصميم 2 = التصميم الرئيسي المعتمد V31 /مباريات_اليوم
# - /مباريات_مجمعة = تصميم مضغوط لعدة أيام
# - حذف تكرار المباريات داخل نفس اليوم

def _v25_safe_txt(x):
    return str(x or "").strip()


def _v25_dedupe_fixture_matches(matches):
    seen = set()
    out = []
    for m in matches or []:
        t1 = _v25_safe_txt(m.get("team1"))
        t2 = _v25_safe_txt(m.get("team2"))
        tm = _v25_safe_txt(m.get("time"))
        dt = _v25_safe_txt(m.get("date"))
        key = (
            re.sub(r"\s+", " ", t1.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", t2.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", tm).strip(),
            dt,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(dict(m))
    return out


def _v25_fixture_simple_matches(date):
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    simple = []
    for m in rows:
        simple.append((
            _v25_safe_txt(m.get("team1")),
            _v25_safe_txt(m.get("team2")),
            _v25_safe_txt(m.get("time")),
        ))
    return rows, simple


def _v25_fixture_title(date):
    d = _normalize_date_arg(date)
    day = ""
    for x, dy in _fixture_dates():
        if x == d:
            day = dy
            break
    return f"{day} {d[:5]}".strip()


def _fixtures_caption(date_or_title, source="PDF جدول البطولة"):
    return f"{date_or_title}\nالمصدر: {source}\nالمصيف يضعكم بالحدث"


def _v25_compact_bg(w=1080, h=1350):
    """
    خلفية التمثال النظيفة للمضغوط بدون عنوان GAMES الكبير.
    """
    candidates = [
        "games_v31_clean_bg.png",
        os.path.join("assets", "templates", "games_v31_clean_bg.png"),
        "games_v31_full_bg.png",
        os.path.join("assets", "templates", "games_v31_full_bg.png"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                bg = Image.open(p).convert("RGB").resize((w, h))
                ov = Image.new("RGBA", (w, h), (0, 10, 30, 120))
                return Image.alpha_composite(bg.convert("RGBA"), ov).convert("RGB")
            except Exception:
                pass
    return Image.new("RGB", (w, h), "#071329")


def _v25_draw_compact_match(draw, box, m):
    x0, y0, x1, y1 = box
    try:
        draw.rounded_rectangle(box, radius=22, fill=(5,24,58,225), outline=(38,151,255,190), width=2)
    except Exception:
        draw.rectangle(box, fill=(5,24,58), outline=(38,151,255), width=2)

    # الوقت يمين
    draw_text(draw, (x1-35, y0+38), _v25_safe_txt(m.get("time")), get_font(27), fill="#FBBF24", anchor="rm", max_width=190)

    # المباراة بالنص
    text = f"{_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))}"
    draw_text(draw, ((x0+x1)//2, y0+36), text, get_font(30), fill="#FFFFFF", max_width=(x1-x0)-260)

    sub = _v25_safe_txt(m.get("stage"))
    if m.get("group"):
        sub += f" - {_v25_safe_txt(m.get('group'))}"
    draw_text(draw, ((x0+x1)//2, y0+75), sub, get_font(21), fill="#CBD5E1", max_width=(x1-x0)-160)


def _render_fixture_day_compact(date):
    """
    تصميم 1: مضغوط، خلفية التمثال، بدون الكلام الكبير فوق.
    """
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    if not rows:
        return []

    chunks = [rows[i:i+8] for i in range(0, len(rows), 8)]
    paths = []

    for page, chunk in enumerate(chunks, 1):
        h = 1350
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 105
        # شريط التاريخ فقط بدون عنوان كبير
        try:
            draw.rounded_rectangle((120, y, 960, y+62), radius=22, fill=(251,191,36,235))
        except Exception:
            draw.rectangle((120, y, 960, y+62), fill=(251,191,36))
        title = _v25_fixture_title(date)
        if len(chunks) > 1:
            title += f" | {page}/{len(chunks)}"
        draw_text(draw, (540, y+31), title, get_font(34), fill="#061329", max_width=780)

        y += 92
        for m in chunk:
            _v25_draw_compact_match(draw, (70, y, 1010, y+104), m)
            y += 122

        draw.line((250, 1238, 830, 1238), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(30), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_compact_{date.replace('/','_')}_{page}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _render_fixture_day_by_design(date, design=2):
    """
    design=1 => مضغوط
    design=2 => نفس تصميم /مباريات_اليوم الرئيسي
    """
    rows, simple_matches = _v25_fixture_simple_matches(date)
    if not simple_matches:
        return []

    if int(design) == 1:
        return _render_fixture_day_compact(date)

    # تصميم 2: نفس تصميم /مباريات_اليوم
    chunks = [simple_matches[i:i+7] for i in range(0, len(simple_matches), 7)]
    paths = []

    for page_idx, chunk in enumerate(chunks, start=1):
        page_title = _v25_fixture_title(date)
        if len(chunks) > 1:
            page_title = f"{page_title} | {page_idx}/{len(chunks)}"

        # هذا هو التصميم المعتمد نفسه حق /مباريات_اليوم
        path = create_matches_today_v31_full_image(page_title, chunk)
        final_path = os.path.join(
            GENERATED_DIR,
            f"fixtures_day_design2_{date.replace('/','_')}_{page_idx}.png"
        )
        try:
            Image.open(path).save(final_path, quality=96)
            paths.append(final_path)
        except Exception:
            paths.append(path)

    return paths


def render_fixtures_combined_images(dates):
    """
    تصميم مجمع: خلفية التمثال بدون الكلام الكبير فوق.
    إذا كثرت الأيام يقسم كل 3 أيام في صورة.
    """
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    dates = [d for d in dates if _fixtures_for_date(d)]
    # إزالة تكرار التواريخ مع الحفاظ على الترتيب
    clean_dates = []
    for d in dates:
        if d not in clean_dates:
            clean_dates.append(d)
    dates = clean_dates

    if not dates:
        return []

    date_chunks = [dates[i:i+3] for i in range(0, len(dates), 3)]
    paths = []

    for page, dchunk in enumerate(date_chunks, 1):
        rows = []
        for d in dchunk:
            rows.append(("date", d))
            for m in _v25_dedupe_fixture_matches(_fixtures_for_date(d)):
                rows.append(("match", m))

        h = max(1350, 170 + len(rows)*112 + 140)
        h = min(h, 2400)
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 90
        if len(date_chunks) > 1:
            draw_text(draw, (540, 42), f"صفحة {page}/{len(date_chunks)}", get_font(24), fill="#FBBF24")

        for kind, val in rows:
            if y > h - 135:
                break

            if kind == "date":
                try:
                    draw.rounded_rectangle((90, y, 990, y+54), radius=20, fill=(251,191,36,235))
                except Exception:
                    draw.rectangle((90, y, 990, y+54), fill=(251,191,36))
                draw_text(draw, (540, y+27), _v25_fixture_title(val), get_font(31), fill="#061329", max_width=830)
                y += 72
            else:
                _v25_draw_compact_match(draw, (75, y, 1005, y+96), val)
                y += 112

        draw.line((250, h-88, 830, h-88), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, h-48), "المصيف يضعكم بالحدث", get_font(29), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_combined_v25_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _fixtures_day_keyboard(date):
    rows = [
        [
            InlineKeyboardButton("تصميم 1", callback_data=f"fx|render1|{date}"),
            InlineKeyboardButton("تصميم 2", callback_data=f"fx|render2|{date}")
        ]
    ]

    miss = [m for m in _fixtures_for_date(date) if _has_unknown(m)]
    for i, m in enumerate(miss, 1):
        rows.append([InlineKeyboardButton(f"تحديث مباراة {i} — {m.get('time')}", callback_data=f"fx|upd|{m.get('id')}")])

    rows.append([InlineKeyboardButton("رجوع للأيام", callback_data="fx|menu")])
    return InlineKeyboardMarkup(rows)


def _fixtures_dates_keyboard(mode="single", selected=None):
    selected = set(selected or [])
    rows = []
    row = []

    for d, day in _fixture_dates():
        label = f"{'✅ ' if d in selected else ''}{day} {d[:5]}"
        data = f"fx|toggle|{d}" if mode == "multi" else f"fx|day|{d}"
        row.append(InlineKeyboardButton(label, callback_data=data))
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    if mode == "multi":
        rows.append([
            InlineKeyboardButton("تصميم كل يوم", callback_data="fx|render_each"),
            InlineKeyboardButton("تصميم واحد", callback_data="fx|render_combo"),
        ])
        rows.append([
            InlineKeyboardButton("تصفير الاختيار", callback_data="fx|clear"),
            InlineKeyboardButton("رجوع", callback_data="fx|menu"),
        ])
    else:
        rows.append([InlineKeyboardButton("اختيار أكثر من يوم", callback_data="fx|multi")])

    return InlineKeyboardMarkup(rows)


def _fixtures_day_text(date):
    rows, matches = _v25_fixture_simple_matches(date)
    if not matches:
        return "ما فيه مباريات لهذا التاريخ."

    lines = [f"{_v25_fixture_title(date)}", ""]
    for i, m in enumerate(rows, 1):
        lines.append(f"{i}) {_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))} — {_v25_safe_txt(m.get('time'))}")
        extra = []
        if m.get("stage"):
            extra.append(_v25_safe_txt(m.get("stage")))
        if m.get("group"):
            extra.append(_v25_safe_txt(m.get("group")))
        if extra:
            lines.append("   " + " | ".join(extra))
        if _has_unknown(m) and m.get("note"):
            lines.append(f"   {_v25_safe_txt(m.get('note'))}")

    return "\n".join(lines)


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    dates = _extract_fixture_dates_from_text(text)

    if dates:
        # تاريخ واحد: يصمم مباشرة بتصميم 2، نفس /مباريات_اليوم
        if len(dates) == 1:
            d = dates[0]
            wait = await update.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=2)
                if not paths:
                    await wait.edit_text(f"ما فيه مباريات بتاريخ {d}")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(update.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        # أكثر من تاريخ: مجمع
        wait = await update.message.reply_text("⏳ جاري تصميم الأيام المجمعة...")
        try:
            paths = render_fixtures_combined_images(dates)
            if not paths:
                await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
                return
            try:
                await wait.delete()
            except Exception:
                pass
            for p in paths:
                await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
        except Exception as e:
            await wait.edit_text(f"تعذر تصميم الأيام ❌\nالسبب: {str(e)[:400]}")
        return

    await update.message.reply_text("اختر اليوم أو اكتب:\n/مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return

    wait = await update.message.reply_text("⏳ جاري تصميم المباريات المجمعة...")
    try:
        paths = render_fixtures_combined_images(dates)
        if not paths:
            await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
            return
        try:
            await wait.delete()
        except Exception:
            pass
        for p in paths:
            await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
    except Exception as e:
        await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return

    await q.answer()

    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return

    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""

    try:
        if action == "menu":
            await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
            return

        if action == "multi":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "toggle" and len(parts) >= 3:
            d = parts[2]
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if d in sel:
                sel.remove(d)
            else:
                sel.append(d)
            context.user_data["fx_selected_dates"] = sel
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", sel))
            return

        if action == "clear":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "day" and len(parts) >= 3:
            d = parts[2]
            await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
            return

        if action in ["render1", "render2"] and len(parts) >= 3:
            d = parts[2]
            design = 1 if action == "render1" else 2
            wait = await q.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=design)
                if not paths:
                    await wait.edit_text("ما فيه مباريات لهذا اليوم.")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_each":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري تصميم الأيام...")
            try:
                try:
                    await wait.delete()
                except Exception:
                    pass
                for d in sel:
                    for p in _render_fixture_day_by_design(d, design=2):
                        await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await q.message.reply_text(f"تعذر التصميم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_combo":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري التصميم المجمع...")
            try:
                paths = render_fixtures_combined_images(sel)
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption("مباريات مجمعة"))
            except Exception as e:
                await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "upd" and len(parts) >= 3:
            mid = parts[2]
            m = _fixture_by_id(mid)
            if not m:
                await q.message.reply_text("لم أجد المباراة.")
                return

            context.user_data["fixture_update_match_id"] = mid
            await q.message.reply_text(
                f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {m.get('time')}) كذا:\n"
                "الفريق الأول * الفريق الثاني\n\n"
                "مثال: المكسيك * أستراليا\n"
                "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
            )
            return

        await q.message.reply_text("تعذر قراءة الخيار.")
    except Exception as e:
        await q.message.reply_text(f"تعذر تنفيذ خيار المباريات ❌\n{str(e)[:400]}")


async def fixtures_update_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mid = context.user_data.get("fixture_update_match_id")
    if not mid:
        return

    text = (update.message.text or "").strip()

    if "*" in text:
        a, b = [x.strip() for x in text.split("*", 1)]
    elif "×" in text:
        a, b = [x.strip() for x in text.split("×", 1)]
    elif "-" in text:
        a, b = [x.strip() for x in text.split("-", 1)]
    else:
        await update.message.reply_text("اكتبها كذا: الفريق الأول * الفريق الثاني")
        return

    if not a or not b:
        await update.message.reply_text("اكتب اسم الفريقين كاملين.")
        return

    data = _load_fixture_updates()
    data.setdefault(mid, {})
    data[mid]["team1"] = canonical_team_name(a) or normalize_name(a)
    data[mid]["team2"] = canonical_team_name(b) or normalize_name(b)
    _save_fixture_updates(data)

    context.user_data.pop("fixture_update_match_id", None)

    m = _apply_fixture_updates(_fixture_by_id(mid) or {"id": mid})
    await update.message.reply_text(
        f"✅ تم حفظ تحديث المباراة\n"
        f"{m.get('team1')} × {m.get('team2')} — {m.get('time', '')}\n\n"
        f"لن أصمم الآن. وقت ما تبيها اكتب:\n/مباريات {m.get('date', '')}"
    )

# ==================== END V25 FIXTURES DESIGN OVERRIDE ====================


async def api_check_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lines = ["فحص مصادر النتائج:", ""]
    # SerpApi
    if _serpapi_key():
        try:
            data = serpapi_search_json("مباراة المكسيك كوريا الجنوبية", timeout=8)
            lines.append("✅ SERPAPI_KEY موجود والاتصال بقوقل شغال")
            if isinstance(data.get("sports_results"), dict):
                lines.append("✅ Google Sports ظهر في نتيجة الفحص")
            else:
                lines.append("⚠️ الاتصال شغال، لكن لم يظهر كرت Google Sports في هذا الفحص")
        except Exception as e:
            lines.append(f"❌ SerpApi موجود لكن فشل الاتصال: {str(e)[:120]}")
    else:
        lines.append("❌ SERPAPI_KEY غير موجود")
    # API-Football
    if _api_football_key():
        try:
            _api_football_get("/status", {})
            lines.append("✅ API_FOOTBALL_KEY موجود والاتصال شغال")
            lid = get_api_football_league_id()
            if lid:
                lines.append(f"✅ League ID الحالي لكأس العالم: {lid}")
            else:
                lines.append("⚠️ لم يتم تحديد League ID تلقائيًا")
        except Exception as e:
            lines.append(f"❌ API-Football فشل: {str(e)[:120]}")
    else:
        lines.append("❌ API_FOOTBALL_KEY غير موجود")
    await update.message.reply_text("\n".join(lines))


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    q = " ".join(body).strip() or "مباراة المكسيك كوريا الجنوبية"
    msg = await update.message.reply_text(f"⏳ أفحص قوقل: {q}")
    try:
        data = serpapi_search_json(q, timeout=8)
        sr = data.get("sports_results") if isinstance(data, dict) else None
        lines = ["نتيجة فحص قوقل:", f"query: {q}"]
        if isinstance(sr, dict):
            lines.append("✅ sports_results موجود")
            # Try list visible team names
            names = []
            for node in _walk_json(sr):
                if isinstance(node, dict):
                    nm = _team_name_from_any(node)
                    if nm and nm not in names:
                        names.append(nm)
                    if len(names) >= 6:
                        break
            if names:
                lines.append("فرق/أسماء ظهرت: " + "، ".join(names[:6]))
            lines.append("مفاتيح: " + "، ".join(list(sr.keys())[:12]))
        else:
            lines.append("⚠️ لم يظهر sports_results")
        await msg.edit_text("\n".join(lines))
    except Exception as e:
        await msg.edit_text(f"❌ فشل فحص قوقل: {str(e)[:180]}")


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|الاحدث|official|fast|latest)\s*$", text or "", re.I)
        mode = m.group(1) if m else "official"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    msg = await update.message.reply_text(f"⏳ جاري فحص ترتيب المجموعات من مصدر {mode_label_ar(mode)}...")
    groups, source_label = fetch_current_groups(mode)
    if not groups:
        await msg.edit_text(_source_help_text("standings", mode), reply_markup=kb)
        return
    path = create_all_groups_image(groups)
    await msg.delete()
    await send_photo_path_markup(update.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({source_label})", kb)
    await update.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({source_label})"))


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = mode_override
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر السعودية * اسبانيا\nأو\n/مباشر السعودية * اسبانيا * سريع\nأو بتاريخ محدد:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026 * سريع")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري البحث عن {team1} × {team2} من مصدر {mode_label_ar(mode)}...")
    data = fetch_live_match_data(team1, team2, mode, date_hint=date_hint)
    if not data:
        await wait.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\n"
            f"مباراة: {team1} × {team2}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") +
            "\nجرّب زر الأحدث أو اكتب التاريخ.\n"
            "وللتشخيص استخدم: /بحث_قوقل مباراة المكسيك كوريا الجنوبية",
            reply_markup=kb,
        )
        return
    path = render_live_match_card(data, mode_label_ar(mode))
    try:
        await wait.delete()
    except Exception:
        pass
    await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    kb = source_keyboard(context, payload)
    if kind == "standings":
        msg = await query.message.reply_text(f"⏳ جاري فحص ترتيب المجموعات من مصدر {mode_label_ar(mode)}...")
        groups, src = fetch_current_groups(mode)
        if not groups:
            await msg.edit_text(_source_help_text("standings", mode), reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        try:
            await msg.delete()
        except Exception:
            pass
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({src})", kb)
        await query.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({src})"))
        return
    if kind == "live":
        team1, team2 = payload.get("team1"), payload.get("team2")
        msg = await query.message.reply_text(f"⏳ جاري البحث عن {team1} × {team2} من مصدر {mode_label_ar(mode)}...")
        data = fetch_live_match_data(team1, team2, mode, date_hint=payload.get("date_hint"))
        if not data:
            await msg.edit_text(f"تعذر جلب مباراة {team1} × {team2} من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = render_live_match_card(data, mode_label_ar(mode))
        try:
            await msg.delete()
        except Exception:
            pass
        await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")

# -------------------- نهاية V4 FIX --------------------


def main():
    if not TOKEN:
        raise RuntimeError("ضع توكن البوت في متغير البيئة BOT_TOKEN")
    ensure_flags_assets()
    ensure_design_assets()
    load_participants_state()
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"(?i)^/start(?:@\w+)?(?:\s|$)"), start))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:من_انا|معرفي)"), who_am_i))
    app.add_handler(MessageHandler(filters.Document.ALL, remember_last_file))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تفعيل_الصور_التلقائية"), admin_only(enable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إيقاف_الصور_التلقائية"), admin_only(disable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ايقاف_الصور_التلقائية"), admin_only(disable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_اليوم"), daily_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_الترتيب"), overall_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_الاساطير"), legends_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_احصائيات"), dashboard_sheet_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صور_الاحصائيات"), all_dashboard_images_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ملف_الاحصائيات(?:\s|$)"), statistics_pdf_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/بطاقة"), participant_card_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تقرير_الفترة"), period_report_command))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعلان_اليوم"), announcement_day_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ملخص_اليوم"), summary_day_command))
    # V23 أوامر التصميم الإضافي
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_مباريات_ستايل2(?:\s|$)"), admin_only(design_matches_style2_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_نتائج_مباريات_ستايل2(?:\s|$)"), admin_only(design_match_results_style2_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_هدافين_ستايل2(?:\s|$)"), admin_only(design_scorers_style2_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_ترتيب_مجموعة_ستايل2(?:\s|$)"), admin_only(design_group_style2_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_مباريات_اطار(?:\s|$)"), admin_only(design_matches_frame_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_نتائج_مباريات_اطار(?:\s|$)"), admin_only(design_results_frame_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_جميع_المجموعات(?:\s|$)"), admin_only(design_all_groups_command)))

    # أوامر التصاميم بالقالب والتلقائي — الأوامر الأطول قبل الأقصر
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_مباريات_تلقائي(?:\s|$)"), admin_only(design_matches_auto_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_نتائج_مباريات_تلقائي(?:\s|$)"), admin_only(design_match_results_auto_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_ترتيب_مجموعة_تلقائي(?:\s|$)"), admin_only(design_group_standing_auto_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_هدافين_تلقائي(?:\s|$)"), admin_only(design_scorers_auto_command)))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_مباريات(?:\s|$)"), admin_only(design_matches_template_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_نتائج_مباريات(?:\s|$)"), admin_only(design_match_results_template_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_ترتيب_مجموعة(?:\s|$)"), admin_only(design_group_standing_template_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تصميم_هدافين(?:\s|$)"), admin_only(design_scorers_template_command)))

    # اسم قديم للتوافق فقط
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نتائج_مباريات_اليوم(?:\s|$)"), admin_only(design_match_results_template_command)))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:استيراد_ملف|استيراد\s+ملف|استيراد|استيراد_اكسل|استيراد_إكسل|استيراد_excel)(?:\s|$)"), admin_only(import_excel_file)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعتماد_استيراد"), admin_only(approve_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إلغاء_استيراد"), admin_only(cancel_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الغاء_استيراد"), admin_only(cancel_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نسخة_احتياطية"), admin_only(backup_zip)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/استرجاع_نسخة"), admin_only(restore_backup_zip)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الأيام"), admin_only(clean_days)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الايام"), admin_only(clean_days)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الملفات"), admin_only(clean_temp_files)))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/بدء_الكاس(?:\s|$)"), admin_only(start_cup_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/حالة_الكاس(?:\s|$)"), admin_only(cup_status_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نتائج_الكاس(?:\s|$)"), admin_only(cup_results_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مواجهات_الكاس(?:\s|$)"), admin_only(cup_matches_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إعادة_الكاس_من(?:\s|$)"), admin_only(reset_cup_from_day_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعادة_الكاس_من(?:\s|$)"), admin_only(reset_cup_from_day_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الغاء_الكاس(?:\s|$)"), admin_only(cancel_cup_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إلغاء_الكاس(?:\s|$)"), admin_only(cancel_cup_command)))

    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اضافه(?:\s|$)"), admin_only(add_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعتماد_نتائج(?:\s|$)"), admin_only(approve_results_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نتائج"), admin_only(results_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الترتيب_العام"), overall))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ترتيب_نص"), ranking_text))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/?(?:احصائيات|إحصائيات)(?:\s|$)"), dashboard))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(الأيام|الايام)"), list_days))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/فحص(?:\\s|$)"), inspect_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مشاركين"), participants_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اسطورة"), legend_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مقارنة"), compare_days))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_الكل"), admin_only(clear_all)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_يوم"), admin_only(clear_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_نتائج"), admin_only(clear_results)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/استرجاع_آخر"), admin_only(restore_last)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/قفل_يوم"), admin_only(lock_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/فتح_يوم"), admin_only(unlock_day)))
    app.run_polling()



# ===== V22 patched overrides =====

def _clean_display_name(text):
    text = normalize_name(text)
    # إزالة الرموز الشائعة والإيموجي حتى لا تظهر مربعات في الصور
    text = re.sub(r"[\U00010000-\U0010ffff]", "", text)
    text = re.sub(r"[⚽🏆👑🧤🔥✅❌⭐🎯🥇🥈🥉📌📊⚔️😅]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or "-"


def build_daily_summary_text(day):
    data = build_day_summary(day)
    winners = " و ".join(data["winners"]) if data["winners"] else "لا يوجد"
    top_rows = data["sorted_rows"][:5]
    lines = [
        f"📊 ملخص اليوم {ordinal_day(day)}",
        f"👥 عدد المشاركين: {len(data['participants'])}",
        f"🏆 أسطورة اليوم: {winners} — {data['max_score']} نقطة",
        "",
        "🔥 أفضل 5:",
    ]
    for idx, r in enumerate(top_rows, start=1):
        lines.append(f"{idx}. {r['participant']} — {r['total']} نقطة")
    lines.append("")
    lines.extend(matchup_lines(day))
    return "\n".join(lines)


def create_daily_result_image(day, goals_count=None, clean_sheets=None):
    ensure_generated_dir()
    data = build_day_summary(day)
    rows = data["sorted_rows"]
    matchups = generate_matchups_for_day(day)
    match_rows = len(matchups.get("pairs", [])) + (1 if matchups.get("bye") else 0)
    goals_lines = [f"{_clean_display_name(p)} {c}" for p, c in (goals_count or {}).items()] if goals_count else [f"{_clean_display_name(p)} {pts} نقطة" for p, pts in data["player_points"].most_common(5)]
    clean_lines = [ _clean_display_name(x) for x in (clean_sheets or list(data["keeper_points"].keys()) or ["لا يوجد"]) ]
    goals_lines = goals_lines or ["لا يوجد"]
    clean_lines = clean_lines or ["لا يوجد"]

    width = 1400
    base_h = 420 + len(rows) * 56 + max(170, 85 + match_rows * 40) + 250
    height = max(1220, base_h)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, f"فانتزي المصيف 2026 اليوم {ordinal_day(day)}", "أسطورة اليوم وترتيب المشاركين", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=110, accent="#8B5CF6")

    legends = " و ".join(data["winners"]) if data["winners"] else "لا يوجد"
    rounded_rect(draw, (90, fy1 + 18, 1010, fy1 + 150), radius=32, fill="#7C3AEDDD", outline="#FFFFFF40", width=2)
    draw_text(draw, (550, fy1 + 58), "أسطورة اليوم", get_font(28), fill="#FFFFFF")
    draw_text(draw, (550, fy1 + 112), f"{legends} - {data['max_score']} نقطة", get_font(42), fill="#FFF6D6", max_width=850)
    rounded_rect(draw, (1040, fy1 + 18, 1310, fy1 + 150), radius=26, fill="#F59E0BDD", outline="#FFFFFF33", width=2)
    draw_text(draw, (1175, fy1 + 58), "المشاركون", get_font(24), fill="#FFFFFF")
    draw_text(draw, (1175, fy1 + 112), f"{len(data['participants'])}", get_font(46), fill="#FFFFFF")

    y = fy1 + 185
    rounded_rect(draw, (90, y, 1310, y + 56), radius=18, fill="#05070D", outline="#FFFFFF55", width=1)
    headers = [(1180, "المشارك"), (870, "النقاط"), (655, "الحارس"), (430, "الكابتن"), (190, "المركز")]
    for x, t in headers:
        draw_text(draw, (x, y + 28), t, get_font(22), fill="#FDE68A")
    y += 70
    row_h = 50
    for idx, r in enumerate(rows, start=1):
        accent = "#F59E0B" if idx == 1 else "#FFFFFF22"
        fill = "#1A1407" if idx == 1 else ("#0B1020" if idx % 2 else "#10172A")
        rounded_rect(draw, (90, y, 1310, y + row_h), radius=16, fill=fill, outline=accent, width=2 if idx == 1 else 1)
        cy = y + row_h // 2
        draw_text(draw, (1180, cy), _clean_display_name(r["participant"]), get_font(24), fill="#FFFFFF", max_width=300)
        draw_text(draw, (870, cy), str(r["total"]), get_font(28), fill="#FDE68A")
        draw_text(draw, (655, cy), f"+{r['keeper_points']}", get_font(25), fill="#A7F3D0")
        draw_text(draw, (430, cy), f"+{r['captain_points']}", get_font(25), fill="#C4B5FD")
        medal = "🥇" if idx == 1 else "🥈" if idx == 2 else "🥉" if idx == 3 else str(idx)
        draw_text(draw, (190, cy), medal, get_font(24), fill="#FFFFFF")
        y += row_h + 8

    match_box_h = max(170, 85 + match_rows * 38)
    rounded_rect(draw, (90, y + 20, 1310, y + 20 + match_box_h), radius=26, fill="#091122DD", outline="#22C55E", width=2)
    draw_text(draw, (700, y + 56), "مواجهات اليوم", get_font(28), fill="#FDE68A")
    yy = y + 100
    if not matchups.get("pairs") and not matchups.get("bye"):
        draw_text(draw, (700, yy), "لا توجد مواجهات — لا يوجد مشاركون", get_font(23), fill="#FFFFFF")
        yy += 36
    else:
        for p in matchups.get("pairs", []):
            line = f"{_clean_display_name(p['a'])} {p['a_points']} ضد {_clean_display_name(p['b'])} {p['b_points']} الفائز {_clean_display_name(p['winner'])}"
            draw_text(draw, (700, yy), line, get_font(22), fill="#FFFFFF", max_width=1100)
            yy += 36
        if matchups.get("bye"):
            draw_text(draw, (700, yy), f"راحة الجولة: {_clean_display_name(matchups['bye'])}", get_font(22), fill="#FDE68A")
            yy += 36
    y = y + 20 + match_box_h + 28

    box_h = 185
    rounded_rect(draw, (90, y, 675, y + box_h), radius=26, fill="#091122DD", outline="#A855F7", width=2)
    rounded_rect(draw, (725, y, 1310, y + box_h), radius=26, fill="#091122DD", outline="#06B6D4", width=2)
    draw_text(draw, (382, y + 38), "الهدافين", get_font(30), fill="#FFFFFF")
    draw_text(draw, (382, y + 112), "\n".join(goals_lines[:5]), get_font(24), fill="#E5E7EB", max_width=500)
    draw_text(draw, (1018, y + 38), "الكلين شيت", get_font(30), fill="#FFFFFF")
    draw_text(draw, (1018, y + 112), "\n".join(clean_lines[:5]), get_font(24), fill="#E5E7EB", max_width=500)

    footer_event(draw, width, height)
    path = os.path.join(GENERATED_DIR, f"daily_result_day_{day}.png")
    img.save(path, quality=95)
    return path


def create_legends_image(start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    days = stats["days"]
    width = 1200
    count = max(len(days), 1)
    row_h = 72
    gap = 12
    content_h = 86 + count * row_h + max(0, count - 1) * gap
    height = max(820, 250 + content_h + 140)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "سجل أساطير الفانتزي", f"من اليوم {start_day} إلى اليوم {end_day}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=220, bottom_pad=110, accent="#8B5CF6")

    y = fy1 + 24
    rounded_rect(draw, (92, y, width-92, y+56), radius=18, fill="#05070D", outline="#FFFFFF55", width=1)
    draw_text(draw, (1010, y+28), "اليوم", get_font(22), fill="#FDE68A")
    draw_text(draw, (650, y+28), "أسطورة اليوم", get_font(22), fill="#FDE68A")
    draw_text(draw, (220, y+28), "النقاط", get_font(22), fill="#FDE68A")
    y += 72

    if not days:
        rounded_rect(draw, (92, y, width-92, y+70), radius=18, fill="#0B1020", outline="#FFFFFF22", width=1)
        draw_text(draw, (width//2, y+35), "لا توجد أيام في هذا النطاق", get_font(24), fill="#FFFFFF")
    else:
        for idx, day in enumerate(days, start=1):
            info = stats["per_day"].get(day, {})
            winners = " و ".join(_clean_display_name(w) for w in info.get("winners", [])) if info.get("winners") else "لا يوجد"
            max_score = info.get("max_score", 0)
            accent = "#F59E0B" if idx == 1 else v16_accent(idx)
            rounded_rect(draw, (92, y, width-92, y+row_h), radius=20, fill="#0B1020", outline=accent, width=2)
            cy = y + row_h//2
            draw_text(draw, (1010, cy), f"اليوم {ordinal_day(day)}", get_font(24), fill="#FFFFFF")
            draw_text(draw, (650, cy), winners, get_font(24), fill="#FDE68A", max_width=470)
            draw_text(draw, (220, cy), f"{max_score} نقطة", get_font(24), fill="#A7F3D0")
            y += row_h + gap

    footer_event(draw, width, height)
    path = os.path.join(GENERATED_DIR, f"legends_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path


def _rank_map_until(day):
    if day is None or day < 1:
        return {}
    stats = collect_stats(1, day)
    return {name: idx + 1 for idx, name in enumerate(stats.get("ranking", []))}


def _biggest_daily_score(start_day, end_day):
    best_name, best_score, best_day = "-", -1, None
    for day in get_existing_days(start_day, end_day):
        for r in read_day_rows(day):
            if r.get("participated") and r.get("total", 0) > best_score:
                best_name, best_score, best_day = r["participant"], r.get("total", 0), day
    if best_score < 0:
        return "-", 0, "-"
    return best_name, best_score, best_day


def _movement_text(name, old_rank, new_rank):
    if not name or old_rank is None or new_rank is None or old_rank == 999 or new_rank == 999:
        return "-"
    return f"{_clean_display_name(name)}\nمن المركز {old_rank} إلى {new_rank}"


def create_participant_card_image(name, start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    if name not in PARTICIPANTS:
        raise ValueError("اسم المشارك غير موجود")
    rank = stats["ranking"].index(name) + 1 if name in stats["ranking"] else "-"
    total = stats["totals"].get(name, 0)
    day_scores = stats["scores_by_day"].get(name, {})
    best_day = max(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    worst_day = min(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    best_score = day_scores.get(best_day, 0) if best_day != "-" else 0
    worst_score = day_scores.get(worst_day, 0) if worst_day != "-" else 0
    best_cap, best_cap_pts, best_cap_day = participant_best_captain(name, start_day, end_day)
    wins = matchup_wins_map(start_day, end_day).get(name, 0)
    pc = stats["participation_count"].get(name, 0)
    days_count = len(stats["days"])
    pct = f"{round(pc / days_count * 100, 1)}%" if days_count else "0%"

    width, height = 1200, 920
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "بطاقة مشارك فانتزي المصيف", _clean_display_name(name), img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=225, bottom_pad=110, accent="#8B5CF6")
    cards = [
        (90, fy1+20, 360, fy1+145, "المركز", f"#{rank}"),
        (465, fy1+20, 735, fy1+145, "النقاط", str(total)),
        (840, fy1+20, 1110, fy1+145, "أسطورة اليوم", str(stats["daily_wins"].get(name,0))),
        (90, fy1+185, 360, fy1+310, "أفضل يوم", f"{best_day} — {best_score} نقاط"),
        (465, fy1+185, 735, fy1+310, "أسوأ يوم", f"{worst_day} — {worst_score} نقاط"),
        (840, fy1+185, 1110, fy1+310, "نسبة المشاركة", pct),
        (90, fy1+350, 545, fy1+490, "أفضل كابتن", f"{_clean_display_name(best_cap)}\n+{best_cap_pts} نقطة"),
        (655, fy1+350, 1110, fy1+490, "انتصارات المواجهات اليومية", str(wins)),
    ]
    for x1,y1,x2,y2,t,v in cards:
        rounded_rect(draw,(x1,y1,x2,y2), radius=28, fill="#091122DD", outline="#FFFFFF25", width=2)
        draw_text(draw, ((x1+x2)//2, y1+34), t, get_font(26), fill="#E5E7EB")
        draw_text(draw, ((x1+x2)//2, y1+88), v, get_font(34), fill="#FFFFFF", max_width=x2-x1-28)
    footer_event(draw, width, height)
    path = os.path.join(GENERATED_DIR, f"participant_card_{_safe_filename(name)}.png")
    img.save(path, quality=95)
    return path


def create_period_report_image(start_day, end_day):
    ensure_generated_dir()
    days, totals, cap_points, keeper_points, player_impact, player_zero, active_counts = period_stats(start_day, end_day)
    if not days:
        raise ValueError("ما فيه أيام في الفترة")

    stats_range = collect_stats(start_day, end_day)
    champion = totals.most_common(1)[0][0] if totals else "-"
    champ_points = totals.get(champion, 0) if champion != "-" else 0

    start_map = _rank_map_until(start_day - 1)
    end_map = _rank_map_until(end_day)
    rising_name, rising_delta = None, -10**9
    falling_name, falling_delta = None, 10**9
    for name in PARTICIPANTS:
        old_rank = start_map.get(name, 999)
        new_rank = end_map.get(name, 999)
        if old_rank == 999 or new_rank == 999:
            continue
        delta = old_rank - new_rank
        if delta > rising_delta:
            rising_delta, rising_name = delta, name
        if delta < falling_delta:
            falling_delta, falling_name = delta, name

    top_day_name, top_day_score, top_day = _biggest_daily_score(start_day, end_day)
    legends_map = stats_range.get("daily_wins", {})
    top_legend_name = "-"
    top_legend_count = 0
    if legends_map:
        top_legend_name, top_legend_count = max(legends_map.items(), key=lambda x: x[1])

    width = 1400
    top5 = totals.most_common(5)
    height = max(1080, 860 + len(top5) * 62)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, f"تقرير الفترة من اليوم {start_day} إلى {end_day}", "فانتزي المصيف 2026", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=230, bottom_pad=110, accent="#F59E0B")

    cards = [
        (80, fy1+15, 430, fy1+160, "بطل الفترة", f"{_clean_display_name(champion)}\n{champ_points} نقطة"),
        (525, fy1+15, 875, fy1+160, "أكثر اللاعبين صعودًا", _movement_text(rising_name, start_map.get(rising_name), end_map.get(rising_name))),
        (970, fy1+15, 1320, fy1+160, "أكثر اللاعبين تراجعًا", _movement_text(falling_name, start_map.get(falling_name), end_map.get(falling_name))),
        (80, fy1+205, 652, fy1+355, "أعلى نقاط يومية في الفترة", f"{_clean_display_name(top_day_name)}\n{top_day_score} نقطة — الجولة {top_day}"),
        (748, fy1+205, 1320, fy1+355, "أكثر من فاز بأسطورة اليوم", f"{_clean_display_name(top_legend_name)}\n{top_legend_count} مرات"),
    ]
    for x1,y1,x2,y2,t,v in cards:
        rounded_rect(draw,(x1,y1,x2,y2), radius=30, fill="#091122DD", outline="#FFFFFF25", width=2)
        draw_text(draw, ((x1+x2)//2, y1+38), t, get_font(27), fill="#E5E7EB")
        draw_text(draw, ((x1+x2)//2, y1+95), v, get_font(35), fill="#FFFFFF", max_width=x2-x1-30)

    y = fy1 + 405
    draw_text(draw, (700, y), "أفضل 5 في الفترة", get_font(36), fill="#FDE68A")
    y += 55
    for i, (pname, pts) in enumerate(top5, start=1):
        accent = "#F59E0B" if i == 1 else v16_accent(i)
        rounded_rect(draw,(170,y,1230,y+55), radius=18, fill="#0B1020", outline=accent, width=2)
        draw_text(draw,(1110,y+28), f"{i}. {_clean_display_name(pname)}", get_font(29), fill="#FFFFFF", max_width=650)
        draw_text(draw,(300,y+28), f"{pts} نقطة", get_font(29), fill="#FDE68A")
        y += 65

    footer_event(draw, width, height)
    path = os.path.join(GENERATED_DIR, f"period_report_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path



# ============================================================
# V26 FINAL OVERRIDES
# اعتماد التعديلات النهائية:
# - ستايل Games of the Day شبه المرجع بخلفية ثابتة
# - كابتشن نظيف للتصاميم بدون نموذج فانتزي
# - إصلاح صورة اليوم، بطاقة المشارك، سجل الأساطير، تقرير الفترة
# - إيقاف صور الإحصائيات غير المقروءة كجداول
# ============================================================

def paste_event_logo(img, width, y=48):
    # حسب الاعتماد: لا نضع شعار استراحة المصيف على تصاميم الفانتزي
    return

def draw_design_header(draw, width, title, subtitle, img=None):
    draw_text(draw, (width//2, 96), title, get_font(54), fill="#FFFFFF", max_width=width-120)
    draw_text(draw, (width//2, 160), subtitle, get_font(34), fill="#FDE68A", max_width=width-160)
    draw.line((210, 213, width-210, 213), fill="#FFFFFF45", width=1)

def _clean_display_name(value):
    s = str(value or "").strip()
    # إزالة رموز قد تظهر كمربعات داخل الصور
    s = re.sub(r"[\U00010000-\U0010ffff]", "", s)
    s = s.replace("🥇","").replace("🥈","").replace("🥉","")
    s = s.replace("🐐","").replace("👑","").replace("✅","")
    s = s.replace("—", "-").replace("–", "-").replace("|", "-")
    return re.sub(r"\s+", " ", s).strip()

def _safe_rank_label(idx):
    return str(idx)

def build_design_matches_caption(day_name, matches):
    lines = ["🏆 مونديال المصيف 2026 🏆", f"🔥 مباريات اليوم ( {day_name} ) 🔥", ""]
    for a, b, t in matches:
        lines.append(f"{a} × {b} — {t}")
    lines.append("")
    lines.append("المصيف ينقل لكم الحدث")
    return "\n".join(lines)

def build_design_results_caption(day_name, results):
    lines = ["🏆 نتائج مباريات اليوم 🏆", f"اليوم ( {day_name} )", ""]
    for a, sa, sb, b in results:
        lines.append(f"{a} {sa} - {sb} {b}")
    lines.append("")
    lines.append("المصيف ينقل لكم الحدث")
    return "\n".join(lines)

def _games_day_background(width, height):
    img = Image.new("RGB", (width, height), "#06152F")
    draw = ImageDraw.Draw(img)

    # تدرج ثابت أزرق
    for y in range(height):
        t = y / max(1, height)
        r = int(2 + 5*t)
        g = int(14 + 20*t)
        b = int(45 + 35*t)
        draw.line((0, y, width, y), fill=(r, g, b))

    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)

    # دوائر/أشكال الخلفية الثابتة
    od.ellipse((160, -185, width-150, 490), fill=(15, 92, 190, 112))
    od.ellipse((width*0.62, 395, width*1.18, 1010), fill=(0, 102, 230, 50))
    od.ellipse((width*0.68, 720, width*1.12, 1240), fill=(0, 95, 215, 42))
    od.rectangle((width*0.63, 430, width*0.90, 640), fill=(0, 115, 255, 36))
    od.rectangle((width*0.90, 520, width*1.08, 790), fill=(0, 105, 230, 42))
    od.ellipse((width-80, 360, width+250, 720), fill=(0, 115, 255, 55))

    # ظل تمثال/شكل جانبي يسار بروح المرجع، مو شعار رسمي
    stat = Image.new("RGBA", (width, height), (0,0,0,0))
    sd = ImageDraw.Draw(stat)
    sx = int(width*0.08)
    base_y = int(height*0.80)
    blue = (20, 95, 190, 95)
    sd.polygon([(sx-130, base_y+180),(sx+210, base_y+180),(sx+135, base_y-250),(sx-40, base_y-290)], fill=blue)
    sd.ellipse((sx-45, base_y-420, sx+75, base_y-300), fill=(20,95,190,82))
    # تاج بسيط
    for ang in [-70,-45,-20,10,35,60]:
        x2 = sx + 15 + int(115 * (ang/70))
        sd.line((sx+15, base_y-405, x2, base_y-540), fill=(20,95,190,75), width=9)
    # اليد/الشعلة
    sd.line((sx-125, base_y-350, sx-205, base_y-625), fill=(20,95,190,85), width=34)
    sd.ellipse((sx-245, base_y-700, sx-165, base_y-615), fill=(20,95,190,90))
    sd.polygon([(sx-205, base_y-735),(sx-235, base_y-665),(sx-170, base_y-670)], fill=(20,95,190,85))
    stat = stat.filter(ImageFilter.GaussianBlur(2))
    overlay = Image.alpha_composite(overlay, stat)

    overlay = overlay.filter(ImageFilter.GaussianBlur(8))
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    # خطوط خفيفة علوية
    for i in range(0, 70, 7):
        draw.arc((80+i*2, 20+i, width-80-i*2, 720+i), 190, 335, fill="#1D6FEA45", width=2)

    return img, draw

def _draw_games_footer(draw, img, width, height):
    # الراعي الأول: MASEEF SPORTS
    pill_y = height - 142
    pill_w, pill_h = 225, 58
    x1 = width//2 - pill_w - 25
    x2 = width//2 + 25

    for x, label1, label2 in [
        (x1, "MASEEF", "SPORTS"),
        (x2, "MASEEF", "2026"),
    ]:
        rounded_rect(draw, (x, pill_y, x+pill_w, pill_y+pill_h), radius=18, fill="#071633BB", outline="#FFFFFFAA", width=2)
        draw_text(draw, (x+pill_w//2, pill_y+22), label1, get_font(22), fill="#FFFFFF", max_width=pill_w-35)
        draw_text(draw, (x+pill_w//2, pill_y+44), label2, get_font(16), fill="#E5E7EB", max_width=pill_w-35)

    # بادج السعودية يمين
    bx = width - 150
    by = height - 225
    draw_text(draw, (bx+55, by-28), "TIME OF DAY", get_font(18), fill="#FFFFFF")
    rounded_rect(draw, (bx, by, bx+110, by+70), radius=20, fill="#087A36", outline="#A7F3D0", width=1)
    draw_text(draw, (bx+55, by+35), "SA", get_font(30), fill="#FFFFFF")
    rounded_rect(draw, (bx+8, by+76, bx+102, by+128), radius=18, fill="#0B7F36", outline="#A7F3D0", width=1)
    draw_text(draw, (bx+55, by+102), "KSA", get_font(22), fill="#FFFFFF")
    draw_text(draw, (bx+55, by+155), "SAUDI ARABIA", get_font(17), fill="#FFFFFF")

    # النص السفلي
    draw.line((110, height-55, width//2-240, height-55), fill="#D4AF37", width=2)
    draw.line((width//2+240, height-55, width-110, height-55), fill="#D4AF37", width=2)
    draw_text(draw, (width//2, height-55), "المصيف ينقل لكم الحدث", get_font(32), fill="#FBBF24")

def _draw_games_header(draw, width, title, date_text):
    draw_text(draw, (width//2, 70), "MONDIAL AL MASEEF 2026", get_font(34), fill="#FFFFFF")
    draw_text(draw, (width//2, 165), title, get_font(78), fill="#FFFFFF", max_width=width-80)
    rounded_rect(draw, (width//2-230, 255, width//2+230, 312), radius=18, fill="#FBBF24", outline="#00000070", width=2)
    draw_text(draw, (width//2, 285), date_text, get_font(28), fill="#061633", max_width=430)

def create_matches_style2_image(day_name, matches):
    ensure_generated_dir()
    count = max(len(matches), 1)
    width = 1200
    height = 1500 if count >= 4 else 1280
    img, draw = _games_day_background(width, height)

    _draw_games_header(draw, width, "GAMES OF THE DAY", f"اليوم {day_name}")

    if count == 1:
        card_h, gap, y = 245, 0, 490
    elif count == 2:
        card_h, gap, y = 205, 34, 430
    elif count == 3:
        card_h, gap, y = 180, 24, 400
    elif count == 4:
        card_h, gap, y = 155, 22, 400
    else:
        card_h, gap, y = 132, 17, 375

    x1, x2 = 270, 890
    for a, b, t in matches[:6]:
        rounded_rect(draw, (x1, y, x2, y+card_h), radius=26, fill="#0638A5", outline="#13B6EA", width=3)
        draw.line((x1+8, y+card_h-3, x2-8, y+card_h-3), fill="#EF4444", width=4)
        draw.line((x1+80, y+4, x2-80, y+4), fill="#22C55E", width=3)
        draw.line((x1+260, y+4, x1+315, y+4), fill="#FBBF24", width=3)
        draw.arc((x1, y, x1+58, y+58), 180, 270, fill="#14B8F5", width=7)
        draw.arc((x2-58, y+card_h-58, x2, y+card_h), 0, 90, fill="#EF4444", width=7)

        cy = y + card_h//2
        flag_w = min(150, card_h-42)
        paste_flag(img, a, (x1+36, cy-flag_w//2, x1+36+flag_w, cy+flag_w//2))
        paste_flag(img, b, (x2-36-flag_w, cy-flag_w//2, x2-36, cy+flag_w//2))

        name_font = get_font(25 if count >= 4 else 29)
        draw_text(draw, (x1+36+flag_w//2, y+card_h-28), _clean_display_name(a), name_font, fill="#FFFFFF", max_width=210)
        draw_text(draw, (x2-36-flag_w//2, y+card_h-28), _clean_display_name(b), name_font, fill="#FFFFFF", max_width=210)

        tm, period = _ampm_from_time(t)
        draw_text(draw, (width//2, cy-14), tm, get_font(50 if count >= 4 else 62), fill="#FBBF24")
        if period:
            draw_text(draw, (width//2, cy+39), period, get_font(29 if count >= 4 else 36), fill="#FBBF24")
        y += card_h + gap

    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"matches_style2_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_match_results_style2_image(day_name, results):
    ensure_generated_dir()
    count = max(len(results), 1)
    width = 1200
    height = 1500 if count >= 4 else 1280
    img, draw = _games_day_background(width, height)

    _draw_games_header(draw, width, "MATCH RESULTS", f"اليوم {day_name}")

    if count == 1:
        card_h, gap, y = 245, 0, 490
    elif count == 2:
        card_h, gap, y = 205, 34, 430
    elif count == 3:
        card_h, gap, y = 180, 24, 400
    elif count == 4:
        card_h, gap, y = 155, 22, 400
    else:
        card_h, gap, y = 132, 17, 375

    x1, x2 = 270, 890
    for a, sa, sb, b in results[:6]:
        rounded_rect(draw, (x1, y, x2, y+card_h), radius=26, fill="#0638A5", outline="#13B6EA", width=3)
        draw.line((x1+8, y+card_h-3, x2-8, y+card_h-3), fill="#EF4444", width=4)
        draw.line((x1+80, y+4, x2-80, y+4), fill="#22C55E", width=3)
        cy = y + card_h//2
        flag_w = min(145, card_h-46)
        paste_flag(img, a, (x1+36, cy-flag_w//2, x1+36+flag_w, cy+flag_w//2))
        paste_flag(img, b, (x2-36-flag_w, cy-flag_w//2, x2-36, cy+flag_w//2))

        name_font = get_font(25 if count >= 4 else 29)
        draw_text(draw, (x1+36+flag_w//2, y+card_h-28), _clean_display_name(a), name_font, fill="#FFFFFF", max_width=210)
        draw_text(draw, (x2-36-flag_w//2, y+card_h-28), _clean_display_name(b), name_font, fill="#FFFFFF", max_width=210)

        rounded_rect(draw, (width//2-96, cy-42, width//2+96, cy+42), radius=18, fill="#FBBF24", outline="#00000088", width=2)
        draw_text(draw, (width//2, cy), f"{sa} - {sb}", get_font(48 if count >= 4 else 58), fill="#061633")
        y += card_h + gap

    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"results_style2_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path

def create_scorers_style2_image(items):
    ensure_generated_dir()
    items = sorted(items, key=lambda x: (-x[1], x[0]))[:12]
    count = max(len(items), 1)
    width = 1200
    height = 1500
    img, draw = _games_day_background(width, height)
    _draw_games_header(draw, width, "TOP SCORERS", "هدافين البطولة")

    if count <= 3:
        row_h, gap, y = 110, 24, 430
    else:
        row_h, gap, y = 80, 14, 405

    x1, x2 = 230, 950
    for i, (name, goals, team) in enumerate(items, start=1):
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=24, fill="#0638A5", outline=v16_accent(i), width=3)
        cy = y + row_h//2
        draw_text(draw, (x2-55, cy), str(i), get_font(34), fill="#FBBF24")
        if team:
            paste_flag(img, team, (x2-145, cy-30, x2-85, cy+30))
        draw_text(draw, (width//2, cy), _clean_display_name(name), get_font(30 if count <= 3 else 24), fill="#FFFFFF", max_width=420)
        draw_text(draw, (x1+115, cy), f"{goals} أهداف", get_font(27 if count <= 3 else 23), fill="#FBBF24")
        y += row_h + gap

    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, "scorers_style2.png")
    img.save(path, quality=95)
    return path

def create_group_style2_image(group_title, rows):
    ensure_generated_dir()
    width, height = 1200, 1500
    img, draw = _games_day_background(width, height)
    title = clean_group_title_for_design(group_title)
    if not str(title).startswith("المجموعة"):
        title = f"المجموعة {title}"
    _draw_games_header(draw, width, "GROUP STANDINGS", title)

    x1, x2 = 210, 990
    y = 435
    row_h, gap = 105, 18
    for i, (team, played, diff, pts) in enumerate(rows[:6], start=1):
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=24, fill="#0638A5", outline=v16_accent(i), width=3)
        cy = y + row_h//2
        draw_text(draw, (x2-52, cy), str(i), get_font(32), fill="#FBBF24")
        paste_flag(img, team, (x2-145, cy-33, x2-80, cy+33))
        draw_text(draw, (x2-300, cy), _clean_display_name(team), get_font(30), fill="#FFFFFF", max_width=260)
        draw_text(draw, (x1+355, cy), f"لعب {played}", get_font(24), fill="#E5E7EB")
        draw_text(draw, (x1+230, cy), f"{int(diff):+d}", get_font(28), fill="#E5E7EB")
        draw_text(draw, (x1+90, cy), str(pts), get_font(34), fill="#FBBF24")
        y += row_h + gap

    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"group_style2_{_safe_filename(group_title)}.png")
    img.save(path, quality=95)
    return path

def create_all_groups_image(groups):
    ensure_generated_dir()
    width, height = 1800, 2400
    img, draw = _games_day_background(width, height)
    draw_text(draw, (width//2, 90), "MONDIAL AL MASEEF 2026", get_font(40), fill="#FFFFFF")
    draw_text(draw, (width//2, 170), "ALL GROUP STANDINGS", get_font(72), fill="#FFFFFF", max_width=width-160)
    draw_text(draw, (width//2, 235), "ترتيب جميع المجموعات", get_font(36), fill="#FBBF24")

    cols = 3
    margin_x, gap_x = 75, 35
    card_w = (width - 2*margin_x - (cols-1)*gap_x) // cols
    card_h = 470
    start_y = 320
    gap_y = 38

    for idx, (title, rows) in enumerate(groups[:12]):
        c = idx % cols
        r = idx // cols
        x = margin_x + c*(card_w+gap_x)
        y = start_y + r*(card_h+gap_y)
        rounded_rect(draw, (x, y, x+card_w, y+card_h), radius=28, fill="#0638A5EE", outline="#14B8F5", width=3)
        rounded_rect(draw, (x+18, y+18, x+card_w-18, y+68), radius=18, fill="#FBBF24", outline="#00000055", width=1)
        gt = clean_group_title_for_design(title)
        if not str(gt).startswith("المجموعة"):
            gt = f"المجموعة {gt}"
        draw_text(draw, (x+card_w//2, y+43), gt, get_font(26), fill="#061633", max_width=card_w-40)
        yy = y + 92
        for pos, (team, played, diff, pts) in enumerate(rows[:4], start=1):
            rounded_rect(draw, (x+20, yy, x+card_w-20, yy+72), radius=16, fill="#061633AA", outline="#FFFFFF22", width=1)
            cy = yy + 36
            draw_text(draw, (x+card_w-44, cy), str(pos), get_font(22), fill="#FBBF24")
            paste_flag(img, team, (x+card_w-125, cy-24, x+card_w-75, cy+24))
            draw_text(draw, (x+card_w-245, cy), _clean_display_name(team), get_font(21), fill="#FFFFFF", max_width=185)
            draw_text(draw, (x+155, cy), str(pts), get_font(26), fill="#FBBF24")
            draw_text(draw, (x+75, cy), f"{int(diff):+d}", get_font(21), fill="#E5E7EB")
            yy += 84

    draw_text(draw, (width//2, height-70), "المصيف ينقل لكم الحدث", get_font(36), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, "all_groups_style2.png")
    img.save(path, quality=95)
    return path

def create_daily_result_image(day, goals_count=None, clean_sheets=None):
    ensure_generated_dir()
    data = build_day_summary(day)
    rows = data["sorted_rows"]
    matchups = generate_matchups_for_day(day)
    pairs = matchups.get("pairs", [])
    match_rows = len(pairs) + (1 if matchups.get("bye") else 0)

    goals_lines = [f"{_clean_display_name(p)} {c} هدف" for p, c in (goals_count or {}).items()] if goals_count else [f"{_clean_display_name(p)} {pts} نقطة" for p, pts in data["player_points"].most_common(5)]
    clean_lines = [_clean_display_name(x) for x in (clean_sheets or list(data["keeper_points"].keys()) or ["لا يوجد"])]
    goals_lines = goals_lines or ["لا يوجد"]
    clean_lines = clean_lines or ["لا يوجد"]

    width = 1400
    row_h = 44
    match_box_h = max(165, 85 + match_rows * 36)
    height = max(1580, 520 + len(rows)*50 + match_box_h + 300)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, f"فانتزي المصيف 2026 اليوم {ordinal_day(day)}", "أسطورة اليوم وترتيب المشاركين", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=235, bottom_pad=95, accent="#8B5CF6")

    legends = " و ".join(_clean_display_name(w) for w in data["winners"]) if data["winners"] else "لا يوجد"
    rounded_rect(draw, (90, fy1 + 18, 1010, fy1 + 172), radius=32, fill="#7C3AEDDD", outline="#FFFFFF40", width=2)
    draw_text(draw, (550, fy1 + 50), "أسطورة اليوم", get_font(27), fill="#FFFFFF")
    legend_font = 31 if len(legends) > 34 else 36
    draw_text(draw, (550, fy1 + 102), legends, get_font(legend_font), fill="#FFF6D6", max_width=850)
    draw_text(draw, (550, fy1 + 145), f"{data['max_score']} نقطة", get_font(36), fill="#FFF6D6", max_width=850)
    rounded_rect(draw, (1040, fy1 + 18, 1310, fy1 + 172), radius=26, fill="#F59E0BDD", outline="#FFFFFF33", width=2)
    draw_text(draw, (1175, fy1 + 55), "المشاركون", get_font(24), fill="#FFFFFF")
    draw_text(draw, (1175, fy1 + 118), f"{len(data['participants'])}", get_font(44), fill="#FFFFFF")

    y = fy1 + 205
    rounded_rect(draw, (90, y, 1310, y + 50), radius=16, fill="#05070D", outline="#FFFFFF55", width=1)
    headers = [(1180, "المشارك"), (870, "النقاط"), (650, "الحارس"), (430, "الكابتن"), (190, "المركز")]
    for x, t in headers:
        draw_text(draw, (x, y + 25), t, get_font(21), fill="#FDE68A")
    y += 60

    for idx, r in enumerate(rows, start=1):
        accent = "#F59E0B" if idx == 1 else "#FFFFFF22"
        fill = "#1A1407" if idx == 1 else ("#0B1020" if idx % 2 else "#10172A")
        rounded_rect(draw, (90, y, 1310, y + row_h), radius=14, fill=fill, outline=accent, width=2 if idx == 1 else 1)
        cy = y + row_h // 2
        draw_text(draw, (1180, cy), _clean_display_name(r["participant"]), get_font(22), fill="#FFFFFF", max_width=300)
        draw_text(draw, (870, cy), str(r["total"]), get_font(25), fill="#FDE68A")
        draw_text(draw, (650, cy), f"+{r['keeper_points']}", get_font(22), fill="#A7F3D0")
        draw_text(draw, (430, cy), f"+{r['captain_points']}", get_font(22), fill="#C4B5FD")
        draw_text(draw, (190, cy), str(idx), get_font(23), fill="#FFFFFF")
        y += row_h + 6

    rounded_rect(draw, (90, y + 18, 1310, y + 18 + match_box_h), radius=26, fill="#091122DD", outline="#22C55E", width=2)
    draw_text(draw, (700, y + 52), "مواجهات اليوم", get_font(27), fill="#FDE68A")
    yy = y + 88
    if not pairs and not matchups.get("bye"):
        draw_text(draw, (700, yy), "لا توجد مواجهات", get_font(22), fill="#FFFFFF")
    else:
        for p in pairs:
            line = f"{_clean_display_name(p['a'])} {p['a_points']} ضد {_clean_display_name(p['b'])} {p['b_points']} الفائز {_clean_display_name(p['winner'])}"
            draw_text(draw, (700, yy), line, get_font(20), fill="#FFFFFF", max_width=1120)
            yy += 34
        if matchups.get("bye"):
            draw_text(draw, (700, yy), f"راحة الجولة: {_clean_display_name(matchups['bye'])}", get_font(20), fill="#FDE68A")

    y = y + 18 + match_box_h + 26
    box_h = 205
    rounded_rect(draw, (90, y, 675, y + box_h), radius=26, fill="#091122DD", outline="#A855F7", width=2)
    rounded_rect(draw, (725, y, 1310, y + box_h), radius=26, fill="#091122DD", outline="#06B6D4", width=2)
    draw_text(draw, (382, y + 38), "الهدافين", get_font(30), fill="#FFFFFF")
    draw_text(draw, (382, y + 120), "\n".join(goals_lines[:5]), get_font(21), fill="#E5E7EB", max_width=500)
    draw_text(draw, (1018, y + 38), "الكلين شيت", get_font(30), fill="#FFFFFF")
    draw_text(draw, (1018, y + 120), "\n".join(clean_lines[:5]), get_font(21), fill="#E5E7EB", max_width=500)

    draw_text(draw, (width//2, height-48), "المصيف ينقل لكم الحدث", get_font(26), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"daily_result_day_{day}.png")
    img.save(path, quality=95)
    return path

def create_legends_image(start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    days = stats["days"]
    width = 1200
    count = max(len(days), 1)
    row_h = 82
    gap = 12
    height = max(820, 265 + count*(row_h+gap) + 125)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "سجل أساطير الفانتزي", f"من اليوم {start_day} إلى اليوم {end_day}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=220, bottom_pad=100, accent="#8B5CF6")

    y = fy1 + 24
    rounded_rect(draw, (92, y, width-92, y+56), radius=18, fill="#05070D", outline="#FFFFFF55", width=1)
    draw_text(draw, (1010, y+28), "اليوم", get_font(22), fill="#FDE68A")
    draw_text(draw, (650, y+28), "أسطورة اليوم", get_font(22), fill="#FDE68A")
    draw_text(draw, (220, y+28), "النقاط", get_font(22), fill="#FDE68A")
    y += 72

    if not days:
        draw_text(draw, (width//2, y+35), "لا توجد أيام في هذا النطاق", get_font(24), fill="#FFFFFF")
    else:
        for idx, day in enumerate(days, start=1):
            info = stats["per_day"].get(day, {})
            winners = " و ".join(_clean_display_name(w) for w in info.get("winners", [])) if info.get("winners") else "لا يوجد"
            max_score = info.get("max_score", 0)
            accent = "#F59E0B" if idx == 1 else v16_accent(idx)
            rounded_rect(draw, (92, y, width-92, y+row_h), radius=20, fill="#0B1020", outline=accent, width=2)
            cy = y + row_h//2
            draw_text(draw, (1010, cy), f"اليوم {ordinal_day(day)}", get_font(23), fill="#FFFFFF")
            draw_text(draw, (650, cy), winners, get_font(22), fill="#FDE68A", max_width=500)
            draw_text(draw, (220, cy), f"{max_score} نقطة", get_font(23), fill="#A7F3D0")
            y += row_h + gap

    draw_text(draw, (width//2, height-44), "المصيف ينقل لكم الحدث", get_font(24), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"legends_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path

def create_participant_card_image(name, start_day=1, end_day=31):
    ensure_generated_dir()
    stats = collect_stats(start_day, end_day)
    if name not in PARTICIPANTS:
        raise ValueError("اسم المشارك غير موجود")

    rank = stats["ranking"].index(name) + 1 if name in stats["ranking"] else "-"
    total = stats["totals"].get(name, 0)
    day_scores = stats["scores_by_day"].get(name, {})
    best_day = max(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    worst_day = min(day_scores, key=lambda d: day_scores[d]) if day_scores else "-"
    best_score = day_scores.get(best_day, 0) if best_day != "-" else 0
    worst_score = day_scores.get(worst_day, 0) if worst_day != "-" else 0
    best_cap, best_cap_pts, best_cap_day = participant_best_captain(name, start_day, end_day)
    wins = matchup_wins_map(start_day, end_day).get(name, 0)
    pc = stats["participation_count"].get(name, 0)
    days_count = len(stats["days"])
    pct = f"{round(pc / days_count * 100, 1)}%" if days_count else "0%"

    width, height = 1400, 980
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "بطاقة مشارك فانتزي المصيف", _clean_display_name(name), img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=205, bottom_pad=92, accent="#A855F7")

    def stat_card(box, title, value, value_size=42, outline="#FFFFFF35"):
        x1, y1, x2, y2 = box
        rounded_rect(draw, box, radius=28, fill="#091122DD", outline=outline, width=2)
        draw_text(draw, ((x1+x2)//2, y1+34), title, get_font(27), fill="#E5E7EB", max_width=x2-x1-28)
        draw_text(draw, ((x1+x2)//2, (y1+y2)//2+18), value, get_font(value_size), fill="#FFFFFF", max_width=x2-x1-36)

    best_day_text = f"اليوم {best_day}\n{best_score} نقاط" if best_day != "-" else "-"
    worst_day_text = f"اليوم {worst_day}\n{worst_score} نقاط" if worst_day != "-" else "-"
    cap_text = f"{_clean_display_name(best_cap)}\n{best_cap_pts} نقاط" if best_cap else "-"

    stat_card((85, 255, 390, 410), "المركز", f"#{rank}", 52, "#FFFFFF55")
    stat_card((548, 255, 852, 410), "النقاط", str(total), 52, "#F59E0B77")
    stat_card((1015, 255, 1320, 410), "أسطورة اليوم", str(stats["daily_wins"].get(name, 0)), 52, "#A855F777")

    stat_card((85, 455, 390, 615), "أفضل يوم", best_day_text, 35, "#10B98166")
    stat_card((548, 455, 852, 615), "أسوأ يوم", worst_day_text, 35, "#EF444466")
    stat_card((1015, 455, 1320, 615), "نسبة المشاركة", pct, 43, "#06B6D466")

    stat_card((85, 665, 645, 845), "أفضل كابتن", cap_text, 34, "#F59E0B66")
    stat_card((760, 665, 1320, 845), "انتصارات المواجهات اليومية", str(wins), 56, "#06B6D466")

    draw_text(draw, (width//2, 910), f"الفترة من اليوم {start_day} إلى {end_day}", get_font(28), fill="#FDE68A")
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(24), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"participant_card_{_safe_filename(name)}.png")
    img.save(path, quality=95)
    return path

def _period_start_rank_map(start_day):
    before = _rank_map_until(start_day - 1)
    if before:
        return before
    return _rank_map_until(start_day)

def _movement_label(name, old_rank, new_rank):
    if not name or old_rank is None or new_rank is None:
        return "لا يوجد تغير"
    return f"{_clean_display_name(name)}\nمن المركز {old_rank} إلى {new_rank}"

def create_period_report_image(start_day, end_day):
    ensure_generated_dir()
    days, totals, cap_points, keeper_points, player_impact, player_zero, active_counts = period_stats(start_day, end_day)
    if not days:
        raise ValueError("ما فيه أيام في الفترة")

    stats_range = collect_stats(start_day, end_day)
    champion = totals.most_common(1)[0][0] if totals else "-"
    champ_points = totals.get(champion, 0) if champion != "-" else 0

    start_map = _period_start_rank_map(start_day)
    end_map = _rank_map_until(end_day)
    rising_name = falling_name = None
    rising_delta = 0
    falling_delta = 0
    for name in PARTICIPANTS:
        if name not in start_map or name not in end_map:
            continue
        delta = start_map[name] - end_map[name]
        if delta > rising_delta:
            rising_delta, rising_name = delta, name
        if delta < falling_delta:
            falling_delta, falling_name = delta, name

    rising_text = _movement_label(rising_name, start_map.get(rising_name), end_map.get(rising_name)) if rising_name else "لا يوجد تغير"
    falling_text = _movement_label(falling_name, start_map.get(falling_name), end_map.get(falling_name)) if falling_name else "لا يوجد تغير"

    top_day_name, top_day_score, top_day = _biggest_daily_score(start_day, end_day)
    legends_map = stats_range.get("daily_wins", {})
    if legends_map:
        top_legend_name, top_legend_count = max(legends_map.items(), key=lambda x: x[1])
    else:
        top_legend_name, top_legend_count = "-", 0

    width = 1400
    top5 = totals.most_common(5)
    height = max(1080, 865 + len(top5) * 62)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, f"تقرير الفترة من اليوم {start_day} إلى {end_day}", "فانتزي المصيف 2026", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=230, bottom_pad=105, accent="#F59E0B")

    cards = [
        (80, fy1+15, 430, fy1+175, "بطل الفترة", f"{_clean_display_name(champion)}\n{champ_points} نقطة"),
        (525, fy1+15, 875, fy1+175, "أكثر اللاعبين صعودًا", rising_text),
        (970, fy1+15, 1320, fy1+175, "أكثر اللاعبين تراجعًا", falling_text),
        (80, fy1+220, 652, fy1+395, "أعلى نقاط يومية في الفترة", f"{_clean_display_name(top_day_name)}\n{top_day_score} نقطة\nالجولة {top_day}"),
        (748, fy1+220, 1320, fy1+395, "أكثر من فاز بأسطورة اليوم", f"{_clean_display_name(top_legend_name)}\n{top_legend_count} مرات"),
    ]
    for x1,y1,x2,y2,t,v in cards:
        rounded_rect(draw,(x1,y1,x2,y2), radius=30, fill="#091122DD", outline="#FFFFFF25", width=2)
        draw_text(draw, ((x1+x2)//2, y1+40), t, get_font(24), fill="#E5E7EB", max_width=x2-x1-20)
        draw_text(draw, ((x1+x2)//2, y1+108), v, get_font(27), fill="#FFFFFF", max_width=x2-x1-30)

    y = fy1 + 450
    draw_text(draw, (700, y), "أفضل 5 في الفترة", get_font(36), fill="#FDE68A")
    y += 55
    for i, (pname, pts) in enumerate(top5, start=1):
        accent = "#F59E0B" if i == 1 else v16_accent(i)
        rounded_rect(draw,(170,y,1230,y+55), radius=18, fill="#0B1020", outline=accent, width=2)
        draw_text(draw,(1110,y+28), f"{i}. {_clean_display_name(pname)}", get_font(28), fill="#FFFFFF", max_width=650)
        draw_text(draw,(300,y+28), f"{pts} نقطة", get_font(28), fill="#FDE68A")
        y += 65

    draw_text(draw, (width//2, height-45), "المصيف ينقل لكم الحدث", get_font(24), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"period_report_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path

async def dashboard_sheet_image_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        args = update.message.text.split(maxsplit=3)
        start_day = int(args[1]) if len(args) >= 2 and args[1].isdigit() else 1
        end_day = int(args[2]) if len(args) >= 3 and args[2].isdigit() else max(get_existing_days(1, 31) or [1])
        sheet = args[3].strip() if len(args) >= 4 else ""
        if "ترتيب" in sheet:
            path = create_overall_ranking_image(start_day, end_day)
            await send_photo_path(update, path, f"✅ الترتيب العام من اليوم {start_day} إلى {end_day}")
        elif "أساطير" in sheet or "اساطير" in sheet or "سجل" in sheet:
            path = create_legends_image(start_day, end_day)
            await send_photo_path(update, path, f"✅ سجل الأساطير من اليوم {start_day} إلى {end_day}")
        else:
            await update.message.reply_text(
                "صور الإحصائيات كجداول Excel غير معتمدة لأنها غير مقروءة داخل تيليجرام.\n"
                "استخدم /احصائيات للملف، أو /صور_الاحصائيات للصور المعتمدة."
            )
    except Exception as e:
        await update.message.reply_text(f"تعذر تجهيز صورة الإحصائيات ❌\n{e}")

async def all_dashboard_images_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        m = re.search(r"(\d+)\s+(\d+)", update.message.text)
        if m:
            start_day, end_day = int(m.group(1)), int(m.group(2))
        else:
            start_day, end_day = 1, max(get_existing_days(1, 31) or [1])
        await update.message.reply_text("جاري تجهيز الصور المعتمدة فقط ✅")
        await send_photo_path(update, create_overall_ranking_image(start_day, end_day), f"✅ الترتيب العام من اليوم {start_day} إلى {end_day}")
        await send_photo_path(update, create_legends_image(start_day, end_day), f"✅ سجل الأساطير من اليوم {start_day} إلى {end_day}")
        await send_photo_path(update, create_period_report_image(start_day, end_day), f"✅ تقرير الفترة {start_day} - {end_day}")
        await update.message.reply_text("تم إيقاف صور الجداول الصغيرة غير المقروءة. ملف Excel موجود عبر /احصائيات ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر تجهيز صور الإحصائيات ❌\n{e}")

async def design_matches_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات_ستايل2\nالسابع\nالبرتغال|الكونغو الديمقراطية|8:00 م")
            return
        path = create_matches_style2_image(day_name, matches)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات ستايل2 ❌\n{e}")

async def design_match_results_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, results = parse_match_results_design_text(update.message.text)
        if not results:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_نتائج_مباريات_ستايل2\nالسابع\nالبرتغال|2|1|الكونغو الديمقراطية")
            return
        path = create_match_results_style2_image(day_name, results)
        await send_photo_path(update, path, build_design_results_caption(day_name, results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج ستايل2 ❌\n{e}")

async def design_scorers_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        items = parse_scorers_text(update.message.text)
        if not items:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_هدافين_ستايل2\nراؤول خيمينيز|4|المكسيك")
            return
        path = create_scorers_style2_image(items)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين ستايل2 ❌\n{e}")

async def design_group_style2_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        group_title, rows = parse_group_standing_text(update.message.text)
        if not rows:
            await update.message.reply_text("اكتبها كذا:\n/تصميم_ترتيب_مجموعة_ستايل2\nالمجموعة C\nالبرازيل|1|0|1")
            return
        path = create_group_style2_image(group_title, rows)
        await send_photo_path(update, path, build_group_standing_caption(group_title, rows))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم ترتيب المجموعة ستايل2 ❌\n{e}")

async def design_matches_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, matches = parse_matches_text(update.message.text)
    if not matches:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات\nالسادس\nفرنسا|البرازيل|10:00 م\nالأرجنتين|ألمانيا|12:00 ص")
        return
    try:
        path = create_matches_template_image(day_name, matches, use_template=True)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات بالقالب ❌\n{e}")

async def design_matches_auto_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_name, matches = parse_matches_text(update.message.text)
    if not matches:
        await update.message.reply_text("اكتبها كذا:\n/تصميم_مباريات_تلقائي\nالسادس\nفرنسا|البرازيل|10:00 م")
        return
    try:
        path = create_matches_template_image(day_name, matches, use_template=False)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات التلقائي ❌\n{e}")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n/نتائج 5\n/اعتماد_نتائج 5\n/احصائيات\n/احصائيات 1 6\n/ترتيب_نص\n\n"
        "أوامر الصور والتقارير:\n"
        "/صورة_اليوم 6\n/صورة_الترتيب 1 6\n/صورة_الاساطير 1 6\n"
        "/صور_الاحصائيات 1 6\n/بطاقة فارس سالم\n/تقرير_الفترة 1 6\n"
        "/تفعيل_الصور_التلقائية\n/إيقاف_الصور_التلقائية\n\n"
        "أوامر التصاميم:\n"
        "/تصميم_مباريات\n/تصميم_مباريات_تلقائي\n"
        "/تصميم_نتائج_مباريات\n/تصميم_نتائج_مباريات_تلقائي\n"
        "/تصميم_ترتيب_مجموعة\n/تصميم_ترتيب_مجموعة_تلقائي\n"
        "/تصميم_هدافين\n/تصميم_هدافين_تلقائي\n\n"
        "التصميم الإضافي المعتمد:\n"
        "/تصميم_مباريات_ستايل2\n/تصميم_نتائج_مباريات_ستايل2\n"
        "/تصميم_هدافين_ستايل2\n/تصميم_ترتيب_مجموعة_ستايل2\n"
        "/تصميم_مباريات_اطار\n/تصميم_نتائج_مباريات_اطار\n"
        "/تصميم_جميع_المجموعات\n\n"
        "أوامر الفحص والنشر:\n"
        "/الأيام\n/فحص 5\n/مشاركين 5\n/اسطورة 5\n/مقارنة 4 5\n/اعلان_اليوم 5\n/ملخص_اليوم 5\n\n"
        "أوامر الأمان:\n"
        "/مسح_نتائج 5\n/مسح_يوم 5\n/مسح_الكل تأكيد\n/استرجاع_آخر\n/قفل_يوم 5\n/فتح_يوم 5\n/معرفي"
    )


# ============================================================
# V27 FINAL — نظام الستايلات 1-4 + الأوامر المختصرة
# اعتماد المستخدم:
# - كل أوامر التصاميم تدعم ستايل 1-4
# - الأوامر المختصرة: /مباريات /انتهت /مجموعة /هدافين /كل_المجموعات
# - النسخ التلقائية تدعم 1-4
# - /نتائج للفانتزي فقط، و/انتهت لتصميم نتائج المباريات
# - دعم صيغ النجمة * بجانب | للصق السريع
# ============================================================

DESIGN_STYLE_KEY = "design_style"
STYLE_NAMES = {
    1: "ستايل 1 - القديم الموجود",
    2: "ستايل 2 - اللوك الجديد الأزرق",
    3: "ستايل 3 - الإطار الأسود",
    4: "ستايل 4 - اللوك الجديد الفخم",
}


def current_design_style():
    try:
        settings = load_settings()
        style = int(settings.get(DESIGN_STYLE_KEY, 4))
        return style if style in (1, 2, 3, 4) else 4
    except Exception:
        return 4


def save_design_style(style):
    style = int(style)
    if style not in (1, 2, 3, 4):
        raise ValueError("الستايل لازم يكون من 1 إلى 4")
    settings = load_settings()
    settings[DESIGN_STYLE_KEY] = style
    save_settings(settings)
    return style


def command_style(text, forced=None):
    if forced in (1, 2, 3, 4):
        return forced
    first = ((text or "").splitlines() or [""])[0].strip()
    parts = first.split()
    if len(parts) >= 2:
        try:
            s = int(parts[1])
            if s in (1, 2, 3, 4):
                return s
        except Exception:
            pass
    return current_design_style()


def _body_lines_after_command(text):
    lines = [l.strip() for l in (text or "").splitlines() if l.strip()]
    return lines[1:]


def _has_sep(line):
    return any(x in str(line) for x in ["|", "*", "×", " x ", " X "])


def _split_data_parts(line):
    line = normalize_name(line).replace("✕", "×")
    # لا نستبدل حرف x داخل أسماء مثل Mexico؛ فقط نفصل إذا جاء كفاصل مستقل
    if "|" in line:
        return [p.strip() for p in line.split("|") if p.strip()]
    if "*" in line:
        return [p.strip() for p in line.split("*") if p.strip()]
    if "×" in line:
        return [p.strip() for p in line.split("×") if p.strip()]
    if re.search(r"\s+[xX]\s+", line):
        return [p.strip() for p in re.split(r"\s+[xX]\s+", line) if p.strip()]
    return [line]


def _looks_like_match_line(line):
    p = _split_data_parts(line)
    return len(p) >= 2 and not re.search(r"\d+\s*[*×|\-–:]\s*\d+", line or "")


def _looks_like_result_line(line):
    if re.search(r"\d+\s*[*×|\-–:]\s*\d+", line or ""):
        return True
    p = _split_data_parts(line)
    if len(p) >= 4:
        return bool(re.search(r"^-?\d+$", p[1]) and re.search(r"^-?\d+$", p[2]))
    return False


def _maybe_day_and_rows(body_lines, kind="matches"):
    if not body_lines:
        return "اليوم", []
    first = body_lines[0]
    if kind == "results" and _looks_like_result_line(first):
        return "اليوم", body_lines
    if kind == "matches" and _looks_like_match_line(first):
        return "اليوم", body_lines
    return first, body_lines[1:]



EN_WEEKDAYS = {
    'السبت': 'SATURDAY', 'الاحد': 'SUNDAY', 'الأحد': 'SUNDAY', 'الاثنين': 'MONDAY', 'الإثنين': 'MONDAY',
    'الثلاثاء': 'TUESDAY', 'الاربعاء': 'WEDNESDAY', 'الأربعاء': 'WEDNESDAY', 'الخميس': 'THURSDAY',
    'الجمعة': 'FRIDAY', 'الاحد ': 'SUNDAY'
}
EN_MONTHS = {
    1: 'JANUARY', 2: 'FEBRUARY', 3: 'MARCH', 4: 'APRIL', 5: 'MAY', 6: 'JUNE',
    7: 'JULY', 8: 'AUGUST', 9: 'SEPTEMBER', 10: 'OCTOBER', 11: 'NOVEMBER', 12: 'DECEMBER'
}

def _format_design_date_text(raw):
    raw = normalize_name(str(raw or '').strip())
    if not raw:
        return ''
    m = re.match(r'^(\d{1,2})\/(\d{1,2})\/(\d{4})$', raw)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            dt = datetime(y, mo, d)
            return f"{dt.strftime('%A').upper()}, {dt.day} {EN_MONTHS.get(dt.month, dt.strftime('%B').upper())} {dt.year}"
        except Exception:
            return raw
    if raw in EN_WEEKDAYS:
        return EN_WEEKDAYS[raw]
    return raw.upper()


def _display_time_en(t):
    s = normalize_name(str(t or '').strip())
    upper = s.upper()
    if 'فجر' in s or 'ص' in s or 'AM' in upper:
        period = 'AM'
    elif 'م' in s or 'مساء' in s or 'PM' in upper:
        period = 'PM'
    else:
        period = ''
    tm = re.sub(r'(فجرًا|فجراً|فجرا|فجر|صباحًا|صباحا|ص|مساءً|مساء|م|AM|PM|am|pm)', '', s).strip()
    return f"{tm} {period}".strip()



def _display_time_ar(t):
    s = normalize_name(str(t or '').strip())
    upper = s.upper()
    if 'فجر' in s or 'ص' in s or 'AM' in upper:
        period = 'ص'
    elif 'م' in s or 'مساء' in s or 'PM' in upper:
        period = 'م'
    else:
        period = ''
    tm = re.sub(r'(فجرًا|فجراً|فجرا|فجر|صباحًا|صباحا|ص|مساءً|مساء|م|AM|PM|am|pm)', '', s).strip()
    return f"{tm} {period}".strip()
# -------------------- Parsers V27 --------------------

def parse_matches_text(text):
    body = _body_lines_after_command(text)
    day_name, data_lines = _maybe_day_and_rows(body, "matches")
    matches = []
    for line in data_lines:
        parts = _split_data_parts(line)
        if len(parts) >= 3:
            matches.append((normalize_name(parts[0]), normalize_name(parts[1]), normalize_name(parts[2])))
        elif len(parts) == 2:
            # يدعم: فرنسا * البرتغال بدون وقت
            matches.append((normalize_name(parts[0]), normalize_name(parts[1]), ""))
        else:
            # يدعم: فرنسا - البرتغال - 8:00 م
            m = re.match(r"(.+?)\s*[-–]\s*(.+?)(?:\s*[-–]\s*(.+))?$", line)
            if m:
                matches.append((normalize_name(m.group(1)), normalize_name(m.group(2)), normalize_name(m.group(3) or "")))
    return normalize_name(day_name or "اليوم"), matches


def parse_match_results_design_text(text):
    body = _body_lines_after_command(text)
    day_name, data_lines = _maybe_day_and_rows(body, "results")
    results = []
    for line in data_lines:
        line = normalize_name(line)
        parts = _split_data_parts(line)
        if len(parts) >= 4:
            # يدعم: فريق * 2 * 1 * فريق أو فريق|2|1|فريق
            try:
                if re.fullmatch(r"\d+", parts[1]) and re.fullmatch(r"\d+", parts[2]):
                    results.append((normalize_name(parts[0]), int(parts[1]), int(parts[2]), normalize_name(parts[3])))
                    continue
            except Exception:
                pass
        # يدعم الصيغة المطلوبة: فرنسا 2 * 6 البرتغال
        m = re.match(r"(.+?)\s+(\d+)\s*[*×|\-–:]\s*(\d+)\s+(.+)$", line)
        if m:
            results.append((normalize_name(m.group(1)), int(m.group(2)), int(m.group(3)), normalize_name(m.group(4))))
            continue
        # يدعم: فرنسا 2 * البرتغال 6 بشكل احتياطي
        if len(parts) == 2:
            m1 = re.match(r"(.+?)\s+(\d+)\s*$", parts[0])
            m2 = re.match(r"^\s*(\d+)\s+(.+)$", parts[1])
            if m1 and m2:
                results.append((normalize_name(m1.group(1)), int(m1.group(2)), int(m2.group(1)), normalize_name(m2.group(2))))
    return normalize_name(day_name or "اليوم"), results


def parse_scorers_text(text):
    body = _body_lines_after_command(text)
    items = []
    for line in body:
        parts = _split_data_parts(line)
        name = team = ""
        goals = 0
        if len(parts) >= 3:
            # الجديد: ميسي * الأرجنتين * 3
            if re.fullmatch(r"\d+", parts[2]):
                name, team, goals = normalize_name(parts[0]), normalize_name(parts[1]), int(parts[2])
            # القديم: ميسي|3|الأرجنتين
            elif re.fullmatch(r"\d+", parts[1]):
                name, goals, team = normalize_name(parts[0]), int(parts[1]), normalize_name(parts[2])
        elif len(parts) == 2 and re.fullmatch(r"\d+", parts[1]):
            name, goals, team = normalize_name(parts[0]), int(parts[1]), ""
        if name and goals > 0:
            items.append((name, goals, team))
    return sorted(items, key=lambda x: (-x[1], x[0]))



def parse_multi_days_matches_text(text):
    body = _body_lines_after_command(text)
    sections = []
    current_date = None
    current_matches = []
    for line in body:
        line = normalize_name(line)
        if not line:
            continue
        if re.fullmatch(r'\d{1,2}/\d{1,2}/\d{4}', line):
            if current_date and current_matches:
                sections.append((current_date, current_matches))
            current_date = line
            current_matches = []
            continue
        parts = _split_data_parts(line)
        if len(parts) >= 3:
            current_matches.append((normalize_name(parts[0]), normalize_name(parts[1]), normalize_name(parts[2])))
        elif len(parts) == 2:
            current_matches.append((normalize_name(parts[0]), normalize_name(parts[1]), ''))
        else:
            m = re.match(r'(.+?)\s*[*|\-–]\s*(.+?)(?:\s*[*|\-–]\s*(.+))?$', line)
            if m:
                current_matches.append((normalize_name(m.group(1)), normalize_name(m.group(2)), normalize_name(m.group(3) or '')))
    if current_date and current_matches:
        sections.append((current_date, current_matches))
    return sections

def _extract_group_row(line):
    line = normalize_name(line)
    if not line:
        return None
    parts = _split_data_parts(line)
    if len(parts) >= 4:
        team = normalize_name(parts[0])
        nums = re.findall(r"[+-]?\d+", " ".join(parts[1:]))
        if len(nums) >= 3:
            return (team, int(nums[0]), int(nums[1]), int(nums[2]))
    # يدعم: اسكتلندا لعب 1 اهداف 1 نقاط 6
    m = re.match(r"(.+?)\s+لعب\s+([0-9]+)\s+(?:فارق|اهداف|أهداف|\+/-)\s+([+-]?[0-9]+)\s+نقاط\s+([0-9]+)\s*$", line)
    if m:
        return (normalize_name(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4)))
    # يدعم: اسكتلندا 1 +1 6
    m = re.match(r"(.+?)\s+([0-9]+)\s+([+-]?[0-9]+)\s+([0-9]+)\s*$", line)
    if m:
        return (normalize_name(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4)))
    return None


def parse_group_standing_text(text):
    body = _body_lines_after_command(text)
    if not body:
        return "المجموعة", []
    group_title = normalize_name(body[0])
    rows = []
    for line in body[1:]:
        row = _extract_group_row(line)
        if row:
            rows.append(row)
    rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
    return group_title, rows


def parse_all_groups_text(text):
    body = _body_lines_after_command(text)
    groups = []
    current_title = None
    current_rows = []
    for line in body:
        row = _extract_group_row(line)
        if row and current_title:
            current_rows.append(row)
            continue
        # أي سطر ليس صف يعتبر عنوان مجموعة: A / المجموعة A
        if current_title and current_rows:
            current_rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
            groups.append((current_title, current_rows))
            current_rows = []
        current_title = normalize_name(line)
    if current_title and current_rows:
        current_rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
        groups.append((current_title, current_rows))
    return groups


# -------------------- Captions V27 --------------------

def build_design_matches_caption(day_name, matches):
    lines = ["🏆 مونديال المصيف 2026 🏆", f"🔥 مباريات اليوم ( {day_name} ) 🔥", ""]
    for a, b, t in matches:
        lines.append(f"{a} × {b}" + (f" — {t}" if t else ""))
    lines.append("")
    lines.append("المصيف ينقل لكم الحدث")
    return "\n".join(lines)


def build_design_results_caption(day_name, results):
    lines = ["🏆 نتائج مباريات اليوم 🏆", f"اليوم ( {day_name} )", ""]
    for a, sa, sb, b in results:
        lines.append(f"{a} {sa} - {sb} {b}")
    lines.append("")
    lines.append("المصيف ينقل لكم الحدث")
    return "\n".join(lines)


def build_match_results_caption(results):
    # للتوافق مع الأوامر القديمة التي لا تمرر اليوم
    lines = ["🏆 نتائج مباريات اليوم 🏆", ""]
    for a, sa, sb, b in results:
        lines.append(f"{a} {sa} - {sb} {b}")
    lines.append("\nالمصيف ينقل لكم الحدث")
    return "\n".join(lines)


def build_top_scorers_caption(items):
    lines = ["🏆 هدافين البطولة 🏆", ""]
    for i, (name, goals, team) in enumerate(sorted(items, key=lambda x: (-x[1], x[0])), start=1):
        team_part = f" — {team}" if team else ""
        lines.append(f"{i}. {name}{team_part} — {goals} أهداف")
    lines.append("\nالمصيف ينقل لكم الحدث")
    return "\n".join(lines)


# -------------------- New Look Backgrounds/Designs V27 --------------------

def _template_bg_path(style):
    name = "games_style2_bg.png" if int(style) == 2 else "games_style4_bg.png"
    return os.path.join("assets", "templates", name)


def _style4_clean_background(width=1200, height=1500):
    img = Image.new("RGB", (width, height), "#07153A")
    draw = ImageDraw.Draw(img)

    for y in range(height):
        t = y / max(1, height - 1)
        r = int(6 + 8 * t)
        g = int(18 + 18 * t)
        b = int(58 + 34 * t)
        draw.line((0, y, width, y), fill=(r, g, b))

    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)

    # وهج علوي ودائري يمين
    od.ellipse((110, -150, width - 130, 470), fill=(48, 95, 220, 70))
    od.ellipse((width * 0.62, 350, width * 1.12, 840), fill=(45, 96, 220, 55))
    od.ellipse((width * 0.70, 760, width * 1.18, 1280), fill=(32, 82, 205, 52))
    od.rectangle((width * 0.74, 450, width * 0.94, 690), fill=(60, 110, 230, 42))
    od.rectangle((width * 0.88, 550, width * 1.04, 890), fill=(48, 96, 210, 38))
    od.ellipse((width - 40, 420, width + 240, 720), fill=(35, 95, 220, 48))

    # لمسة خطوط/دوائر خفيفة يمين
    for off in [0, 115, 230, 345]:
        od.arc((830, 120 + off, 1160, 410 + off), 255, 105, fill=(89, 140, 255, 42), width=4)

    # ظل تمثال حرية مبسط يسار (بدون أي مربعات/خانات ثابتة)
    stat = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    sd = ImageDraw.Draw(stat)
    sx = int(width * 0.12)
    base_y = int(height * 0.78)
    blue = (20, 72, 180, 108)
    sd.polygon([
        (sx - 145, base_y + 210), (sx + 200, base_y + 210),
        (sx + 145, base_y - 210), (sx - 5, base_y - 255)
    ], fill=blue)
    sd.polygon([
        (sx - 70, base_y - 190), (sx + 60, base_y - 190),
        (sx + 35, base_y - 370), (sx - 45, base_y - 375)
    ], fill=(22, 82, 195, 98))
    sd.ellipse((sx - 45, base_y - 500, sx + 75, base_y - 375), fill=(22, 82, 195, 96))
    for ang in [-72, -45, -20, 10, 38, 62]:
        x2 = sx + 18 + int(120 * (ang / 72))
        sd.line((sx + 18, base_y - 470, x2, base_y - 620), fill=(22, 82, 195, 90), width=10)
    sd.line((sx - 105, base_y - 410, sx - 200, base_y - 705), fill=(22, 82, 195, 92), width=34)
    sd.ellipse((sx - 238, base_y - 782, sx - 155, base_y - 690), fill=(22, 82, 195, 95))
    sd.polygon([(sx - 198, base_y - 814), (sx - 235, base_y - 730), (sx - 160, base_y - 735)], fill=(22, 82, 195, 92))
    stat = stat.filter(ImageFilter.GaussianBlur(1))
    overlay = Image.alpha_composite(overlay, stat)

    overlay = overlay.filter(ImageFilter.GaussianBlur(6))
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    # أقواس علوية خفيفة مثل المرجع
    for i in range(0, 55, 6):
        draw.arc((80 + i * 2, 34 + i, width - 85 - i * 2, 520 + i), 198, 336, fill="#5FA3FF44", width=2)

    return img, draw


def _draw_date_pill_only(draw, width, date_text):
    # شريط التاريخ الديناميكي فقط — بدون إعادة رسم عنوان الخلفية الأصلي.
    rounded_rect(draw, (width//2-240, 276, width//2+240, 334), radius=18, fill="#FBBF24", outline="#00000070", width=2)
    draw_text(draw, (width//2, 306), date_text, get_font(28), fill="#061633", max_width=450)


def _cover_style4_dynamic_areas(img, width, height, date_text):
    # نحافظ على خلفية التمثال والعناوين والرعاة وشعار السعودية،
    # ونغطي فقط تاريخ القالب القديم وخانات المباريات الثابتة.
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)

    # تغطية مستطيل التاريخ الأصفر القديم
    rounded_rect(od, (width//2-255, 266, width//2+255, 344), radius=22, fill="#08245FD8")

    # تغطية خانات المباريات الثابتة فقط مع ترك التمثال والعناصر الجانبية ظاهرة
    rounded_rect(od, (150, 350, 985, 1138), radius=38, fill="#061B58D8")

    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)
    _draw_date_pill_only(draw, width, date_text)
    return img, draw


def _newlook_canvas(style, title, sub_title, width=1200, height=1500):
    style = int(style)

    if style == 4:
        path = _template_bg_path(style)
        if os.path.exists(path):
            try:
                img = Image.open(path).convert("RGB").resize((width, height), Image.LANCZOS)
                return _cover_style4_dynamic_areas(img, width, height, sub_title)
            except Exception:
                pass
        # احتياط فقط إذا لم توجد الخلفية الأصلية
        img, draw = _style4_clean_background(width, height)
        _draw_games_header(draw, width, title, sub_title)
        return img, draw

    path = _template_bg_path(style)
    if os.path.exists(path):
        try:
            img = Image.open(path).convert("RGB").resize((width, height))
            draw = ImageDraw.Draw(img)
            overlay = Image.new("RGBA", (width, height), (0,0,0,0))
            od = ImageDraw.Draw(overlay)
            rounded_rect(od, (55, 45, width-55, 330), radius=30, fill="#06152F88")
            rounded_rect(od, (170, 350, width-170, height-265), radius=36, fill="#03133499")
            img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
            draw = ImageDraw.Draw(img)
        except Exception:
            img, draw = _games_day_background(width, height)
    else:
        img, draw = _games_day_background(width, height)
    _draw_games_header(draw, width, title, sub_title)
    return img, draw


def _newlook_card(draw, box, i=1, style=2):
    base_fill = "#0638A5" if int(style) == 2 else "#061E64"
    accent = v16_accent(i)
    rounded_rect(draw, box, radius=28, fill=base_fill, outline="#13B6EA" if int(style)==2 else "#3B82F6", width=3)
    x1,y1,x2,y2 = box
    draw.line((x1+10, y2-4, x2-10, y2-4), fill="#EF4444", width=4)
    draw.line((x1+90, y1+4, x2-90, y1+4), fill="#22C55E", width=3)
    draw.line((x1+310, y1+4, x1+375, y1+4), fill="#FBBF24", width=3)
    draw.arc((x1, y1, x1+64, y1+64), 180, 270, fill=accent, width=6)
    draw.arc((x2-64, y2-64, x2, y2), 0, 90, fill=accent, width=6)


def _newlook_layout(count):
    if count <= 1:
        return 225, 0, 505
    if count == 2:
        return 200, 34, 430
    if count == 3:
        return 178, 26, 405
    if count == 4:
        return 155, 21, 388
    return 128, 14, 370


def create_matches_newlook_image(day_name, matches, style=2):
    ensure_generated_dir()
    count = max(len(matches), 1)
    width, height = 1200, 1500
    img, draw = _newlook_canvas(style, "GAMES OF THE DAY", _format_design_date_text(day_name), width, height)
    row_h, gap, y = _newlook_layout(count)
    x1, x2 = 240, 960
    for i, (a, b, t) in enumerate(matches[:7], start=1):
        _newlook_card(draw, (x1, y, x2, y+row_h), i, style)
        cy = y + row_h//2
        flag_w = min(148, max(95, row_h - 34))
        # الفريق الأول يمين، الثاني يسار
        paste_flag(img, a, (x2-38-flag_w, cy-flag_w//2, x2-38, cy+flag_w//2))
        paste_flag(img, b, (x1+38, cy-flag_w//2, x1+38+flag_w, cy+flag_w//2))
        draw_text(draw, (x2-38-flag_w//2, y+row_h-25), _clean_display_name(a), get_font(25 if count>=4 else 29), fill="#FFFFFF", max_width=210)
        draw_text(draw, (x1+38+flag_w//2, y+row_h-25), _clean_display_name(b), get_font(25 if count>=4 else 29), fill="#FFFFFF", max_width=210)
        center = t or "VS"
        if t:
            if int(style) == 4:
                draw_text(draw, (width//2, cy), _display_time_en(t), get_font(40 if count>=4 else 48), fill="#FBBF24", max_width=230)
            else:
                tm, period = _ampm_from_time(t)
                draw_text(draw, (width//2, cy-12), tm, get_font(50 if count>=4 else 62), fill="#FBBF24")
                if period:
                    draw_text(draw, (width//2, cy+38), period, get_font(29 if count>=4 else 36), fill="#FBBF24")
        else:
            draw_text(draw, (width//2, cy), "VS", get_font(50), fill="#FBBF24")
        y += row_h + gap
    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"matches_style{style}_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path


def create_match_results_newlook_image(day_name, results, style=2):
    ensure_generated_dir()
    count = max(len(results), 1)
    width, height = 1200, 1500
    img, draw = _newlook_canvas(style, "MATCH RESULTS", _format_design_date_text(day_name), width, height)
    row_h, gap, y = _newlook_layout(count)
    x1, x2 = 240, 960
    for i, (a, sa, sb, b) in enumerate(results[:7], start=1):
        _newlook_card(draw, (x1, y, x2, y+row_h), i, style)
        cy = y + row_h//2
        flag_w = min(145, max(92, row_h - 38))
        paste_flag(img, a, (x2-38-flag_w, cy-flag_w//2, x2-38, cy+flag_w//2))
        paste_flag(img, b, (x1+38, cy-flag_w//2, x1+38+flag_w, cy+flag_w//2))
        draw_text(draw, (x2-38-flag_w//2, y+row_h-25), _clean_display_name(a), get_font(24 if count>=4 else 28), fill="#FFFFFF", max_width=210)
        draw_text(draw, (x1+38+flag_w//2, y+row_h-25), _clean_display_name(b), get_font(24 if count>=4 else 28), fill="#FFFFFF", max_width=210)
        rounded_rect(draw, (width//2-100, cy-44, width//2+100, cy+44), radius=20, fill="#FBBF24", outline="#00000088", width=2)
        draw_text(draw, (width//2, cy), f"{sa} - {sb}", get_font(52 if count>=4 else 60), fill="#061633")
        y += row_h + gap
    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"results_style{style}_{_safe_filename(day_name)}.png")
    img.save(path, quality=95)
    return path


def create_scorers_newlook_image(items, style=2):
    ensure_generated_dir()
    items = sorted(items, key=lambda x: (-x[1], x[0]))[:12]
    count = max(len(items), 1)
    width, height = 1200, 1500
    img, draw = _newlook_canvas(style, "TOP SCORERS", "هدافين البطولة", width, height)
    if count <= 3:
        row_h, gap, y = 122, 26, 430
    elif count <= 6:
        row_h, gap, y = 98, 18, 410
    else:
        row_h, gap, y = 74, 11, 395
    x1, x2 = 210, 990
    for i, (name, goals, team) in enumerate(items, start=1):
        _newlook_card(draw, (x1, y, x2, y+row_h), i, style)
        cy = y + row_h//2
        draw_text(draw, (x2-55, cy), str(i), get_font(32 if count>6 else 38), fill="#FBBF24")
        if team:
            fw = min(84, max(58, row_h-18))
            paste_flag(img, team, (x2-155, cy-fw//2, x2-155+fw, cy+fw//2))
        draw_text(draw, (width//2+55, cy), _clean_display_name(name), get_font(25 if count>6 else 32), fill="#FFFFFF", max_width=430)
        rounded_rect(draw, (x1+35, cy-34, x1+185, cy+34), radius=18, fill="#FBBF24", outline="#00000070", width=1)
        draw_text(draw, (x1+110, cy), str(goals), get_font(40 if count<=6 else 32), fill="#061633")
        draw_text(draw, (x1+250, cy), "أهداف", get_font(22 if count>6 else 26), fill="#FDE68A")
        y += row_h + gap
    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"scorers_style{style}.png")
    img.save(path, quality=95)
    return path


def create_group_newlook_image(group_title, rows, style=2):
    ensure_generated_dir()
    width, height = 1200, 1500
    title = clean_group_title_for_design(group_title)
    if not str(title).startswith("المجموعة"):
        title = f"المجموعة {title}"
    img, draw = _newlook_canvas(style, "GROUP STANDINGS", title, width, height)
    x1, x2 = 190, 1010
    y = 395
    rounded_rect(draw, (x1, y, x2, y+60), radius=20, fill="#061633DD", outline="#FFFFFF44", width=1)
    draw_text(draw, (x2-175, y+30), "المنتخب", get_font(24), fill="#FFFFFF")
    draw_text(draw, (x1+385, y+30), "لعب", get_font(23), fill="#FDE68A")
    draw_text(draw, (x1+245, y+30), "+/-", get_font(23), fill="#FDE68A")
    draw_text(draw, (x1+95, y+30), "نقاط", get_font(23), fill="#FDE68A")
    y += 78
    row_h, gap = 98, 14
    for i, (team, played, diff, pts) in enumerate(rows[:8], start=1):
        _newlook_card(draw, (x1, y, x2, y+row_h), i, style)
        cy = y + row_h//2
        draw_text(draw, (x2-48, cy), str(i), get_font(28), fill="#FBBF24")
        paste_flag(img, team, (x2-135, cy-32, x2-70, cy+32))
        draw_text(draw, (x2-285, cy), _clean_display_name(team), get_font(27), fill="#FFFFFF", max_width=250)
        draw_text(draw, (x1+385, cy), str(played), get_font(27), fill="#FFFFFF")
        draw_text(draw, (x1+245, cy), f"{int(diff):+d}", get_font(27), fill="#E5E7EB")
        draw_text(draw, (x1+95, cy), str(pts), get_font(34), fill="#FBBF24")
        y += row_h + gap
    _draw_games_footer(draw, img, width, height)
    path = os.path.join(GENERATED_DIR, f"group_style{style}_{_safe_filename(group_title)}.png")
    img.save(path, quality=95)
    return path


def _fit_font_to_width(draw, text, start_size, max_width, min_size=16):
    size = int(start_size)
    text = str(text or "")
    while size > int(min_size):
        try:
            if text_width(draw, text, get_font(size)) <= max_width:
                break
        except Exception:
            break
        size -= 1
    return get_font(max(size, int(min_size)))


def create_all_groups_newlook_image(groups, style=2):
    ensure_generated_dir()
    groups = list(groups or [])[:12]
    width = 1800
    cols = 3
    margin_x, gap_x = 72, 34
    card_w = (width - 2*margin_x - (cols-1)*gap_x) // cols

    # ارتفاع كل صف في الشبكة يعتمد على أكبر عدد فرق في مجموعات هذا الصف،
    # عشان نقلل الفراغات الزرقاء الفاضية مع بقاء كل المجموعات في صورة واحدة.
    row_h = 62
    row_gap = 8
    grid_gap_y = 32
    min_card_h = 285
    header_h = 74
    label_h = 34
    bottom_pad = 24

    row_heights = []
    for r in range(4):
        chunk = groups[r*cols:(r+1)*cols]
        max_rows = max([min(len(rows), 4) for _title, rows in chunk] + [1])
        ch = header_h + label_h + 18 + max_rows * row_h + max(0, max_rows-1) * row_gap + bottom_pad
        row_heights.append(max(min_card_h, ch))

    start_y = 300
    footer_space = 110
    height = int(start_y + sum(row_heights) + grid_gap_y*(len(row_heights)-1) + footer_space)
    height = max(1640, min(height, 2320))

    img, draw = _games_day_background(width, height)
    # تغميق خفيف خلف المحتوى لتحسين القراءة بدون تغيير الهوية.
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    rounded_rect(od, (45, 40, width-45, height-45), radius=40, fill="#06152F55", outline="#FFFFFF22", width=2)
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    draw_text(draw, (width//2, 82), "MONDIAL AL MASEEF 2026", _v31_latin_font(42) if '_v31_latin_font' in globals() else get_font(40), fill="#FFFFFF", max_width=900)
    draw_text(draw, (width//2, 160), "ALL GROUP STANDINGS", _v31_latin_font(74) if '_v31_latin_font' in globals() else get_font(72), fill="#FFFFFF", max_width=width-170)
    draw_text(draw, (width//2, 230), "ترتيب جميع المجموعات", get_font(40), fill="#FBBF24", max_width=800)

    y_cursor = start_y
    for idx, (title, rows) in enumerate(groups):
        c = idx % cols
        r = idx // cols
        if c == 0 and idx != 0:
            y_cursor += row_heights[r-1] + grid_gap_y

        x = margin_x + c*(card_w+gap_x)
        y = y_cursor
        card_h = row_heights[r]

        rounded_rect(draw, (x, y, x+card_w, y+card_h), radius=28, fill="#0638A5E8", outline="#14B8F5", width=3)
        rounded_rect(draw, (x+16, y+16, x+card_w-16, y+66), radius=16, fill="#FBBF24", outline="#00000055", width=1)

        gt = clean_group_title_for_design(title)
        if not str(gt).startswith("المجموعة"):
            gt = f"المجموعة {gt}"
        draw_text(draw, (x+card_w//2, y+41), gt, get_font(29), fill="#061633", max_width=card_w-44)

        label_y = y + 88
        draw_text(draw, (x+card_w-64, label_y), "#", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+card_w-205, label_y), "المنتخب", get_font(18), fill="#FDE68A", max_width=190)
        draw_text(draw, (x+195, label_y), "لعب", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+118, label_y), "+/-", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+48, label_y), "نقاط", get_font(18), fill="#FDE68A")

        yy = y + 108
        for pos, (team, played, diff, pts) in enumerate(rows[:4], start=1):
            rounded_rect(draw, (x+16, yy, x+card_w-16, yy+row_h), radius=15, fill="#061633B8", outline="#FFFFFF30", width=1)
            cy = yy + row_h//2
            draw_text(draw, (x+card_w-34, cy), str(pos), get_font(23), fill="#FBBF24")
            fw, fh = 56, 38
            paste_flag(img, team, (x+card_w-100, cy-fh//2, x+card_w-100+fw, cy+fh//2))
            team_name = _clean_display_name(team)
            team_font = _fit_font_to_width(draw, team_name, 24, 190, min_size=17)
            draw_text(draw, (x+card_w-230, cy), team_name, team_font, fill="#FFFFFF", max_width=205)
            draw_text(draw, (x+195, cy), str(played), get_font(23), fill="#FFFFFF")
            draw_text(draw, (x+118, cy), f"{int(diff):+d}", get_font(23), fill="#E5E7EB")
            draw_text(draw, (x+48, cy), str(pts), get_font(29), fill="#FBBF24")
            yy += row_h + row_gap

    draw_text(draw, (width//2, height-50), "المصيف ينقل لكم الحدث", get_font(38), fill="#FBBF24", max_width=700)
    path = os.path.join(GENERATED_DIR, f"all_groups_style{style}.png")
    img.save(path, quality=95)
    return path


# Override old style2 names to use V27 new look.
def create_matches_style2_image(day_name, matches):
    return create_matches_newlook_image(day_name, matches, 2)


def create_match_results_style2_image(day_name, results):
    return create_match_results_newlook_image(day_name, results, 2)


def create_scorers_style2_image(items):
    return create_scorers_newlook_image(items, 2)


def create_group_style2_image(group_title, rows):
    return create_group_newlook_image(group_title, rows, 2)


def create_all_groups_image(groups):
    return create_all_groups_newlook_image(groups, 2)


# -------------------- Style 3 frame designs for all design types --------------------

def create_scorers_frame_style_image(items):
    ensure_generated_dir()
    items = sorted(items, key=lambda x: (-x[1], x[0]))[:12]
    count = max(len(items), 1)
    width = 1200
    row_h = 92 if count <= 8 else 72
    gap = 14 if count <= 8 else 9
    height = max(760, 220 + count*(row_h+gap) + 120)
    img, draw = _frame_canvas(width, height)
    draw_text(draw, (width//2, 80), "مونديال المصيف 2026", get_font(50), fill="#FFFFFF")
    draw_text(draw, (width//2, 140), "هدافين البطولة", get_font(36), fill="#FDE68A")
    y = 215
    for i, (name, goals, team) in enumerate(items, start=1):
        x1, x2 = 110, width-110
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=24, fill="#07110FCC", outline="#22C55E" if i==1 else "#2563EB", width=3)
        cy = y + row_h//2
        draw_text(draw, (x2-45, cy), str(i), get_font(28), fill="#FDE68A")
        if team:
            paste_flag(img, team, (x2-130, cy-30, x2-70, cy+30))
        draw_text(draw, (width//2+95, cy), _clean_display_name(name), get_font(28 if count<=8 else 23), fill="#FFFFFF", max_width=450)
        rounded_rect(draw, (x1+45, cy-30, x1+180, cy+30), radius=16, fill="#B8FFF0", outline="#FFFFFFAA", width=2)
        draw_text(draw, (x1+112, cy), str(goals), get_font(34 if count<=8 else 28), fill="#061633")
        y += row_h + gap
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(28), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, "scorers_frame.png")
    img.save(path, quality=95)
    return path


def create_group_frame_style_image(group_title, rows):
    ensure_generated_dir()
    width = 1200
    count = max(len(rows), 1)
    row_h = 92 if count <= 8 else 74
    gap = 13 if count <= 8 else 8
    height = max(820, 245 + count*(row_h+gap) + 110)
    img, draw = _frame_canvas(width, height)
    title = clean_group_title_for_design(group_title)
    if not str(title).startswith("المجموعة"):
        title = f"المجموعة {title}"
    draw_text(draw, (width//2, 80), "مونديال المصيف 2026", get_font(50), fill="#FFFFFF")
    draw_text(draw, (width//2, 140), title, get_font(36), fill="#FDE68A")
    y = 215
    x1, x2 = 95, width-95
    rounded_rect(draw, (x1, y, x2, y+56), radius=20, fill="#061633", outline="#22C55E", width=2)
    draw_text(draw, (940, y+28), "المنتخب", get_font(24), fill="#FFFFFF")
    draw_text(draw, (530, y+28), "لعب", get_font(22), fill="#FDE68A")
    draw_text(draw, (385, y+28), "+/-", get_font(22), fill="#FDE68A")
    draw_text(draw, (220, y+28), "نقاط", get_font(22), fill="#FDE68A")
    y += 76
    for i, (team, played, diff, pts) in enumerate(rows[:10], start=1):
        rounded_rect(draw, (x1, y, x2, y+row_h), radius=24, fill="#07110FCC", outline="#22C55E" if i==1 else "#2563EB", width=2)
        cy = y + row_h//2
        draw_text(draw, (1045, cy), str(i), get_font(25), fill="#FDE68A")
        paste_flag(img, team, (940, cy-30, 1000, cy+30))
        draw_text(draw, (760, cy), _clean_display_name(team), get_font(27 if count<=8 else 23), fill="#FFFFFF", max_width=310)
        draw_text(draw, (530, cy), str(played), get_font(25), fill="#FFFFFF")
        draw_text(draw, (385, cy), f"{int(diff):+d}", get_font(25), fill="#FFFFFF")
        draw_text(draw, (220, cy), str(pts), get_font(32), fill="#FDE68A")
        y += row_h + gap
    draw_text(draw, (width//2, height-42), "المصيف ينقل لكم الحدث", get_font(28), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"group_frame_{_safe_filename(group_title)}.png")
    img.save(path, quality=95)
    return path


def create_all_groups_frame_style_image(groups):
    # ستايل 3 لجميع المجموعات بنفس روح الإطار لكن أوسع.
    return create_all_groups_newlook_image(groups, 3)



def _multi_days_ar_date_label(raw):
    raw = normalize_name(str(raw or "").strip())
    if not raw:
        return ""
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", raw)
    ar_months = {
        1: "يناير",
        2: "فبراير",
        3: "مارس",
        4: "أبريل",
        5: "مايو",
        6: "يونيو",
        7: "يوليو",
        8: "أغسطس",
        9: "سبتمبر",
        10: "أكتوبر",
        11: "نوفمبر",
        12: "ديسمبر",
    }
    if not m:
        # بدون / أو رموز زخرفية عشان ما تظهر مربعات
        return raw.replace("/", " ")
    d, mo, y = map(int, m.groups())
    ar_days = {
        0: "الاثنين",
        1: "الثلاثاء",
        2: "الأربعاء",
        3: "الخميس",
        4: "الجمعة",
        5: "السبت",
        6: "الأحد",
    }
    try:
        dt = datetime(y, mo, d)
        return f"{ar_days.get(dt.weekday(), '')} {d:02d} {ar_months.get(mo, str(mo).zfill(2))}"
    except Exception:
        return f"{d:02d} {ar_months.get(mo, str(mo).zfill(2))}"

def create_multi_days_matches_image(schedule_blocks, style=4, max_blocks=6, wide_mode=False):
    ensure_generated_dir()

    blocks = list(schedule_blocks or [])[:max_blocks]
    if not blocks:
        raise ValueError("لا توجد مباريات")

    if wide_mode:
        # نسخة 10 أيام: أكبر وأفخم بصريًا، لكنها تبقى صورة واحدة.
        width = 2200
        cols = 2 if len(blocks) > 1 else 1
        margin_x = 105 if cols == 2 else 280
        gap_x = 62
        start_y = 380
        gap_y = 48
        min_height = 1500
        max_height = 3600
        title_size = 92
        sub_size = 48
        date_size = 38
        row_font = 31
        row_h = 74
        card_pad = 26
        max_matches_per_day = 6
        flag_w, flag_h = 64, 42
        team_gap = 190
        time_font = 26
    else:
        width = 1800
        cols = 2 if len(blocks) > 1 else 1
        margin_x = 85 if cols == 2 else 210
        gap_x = 45
        start_y = 330
        gap_y = 36
        min_height = 1180
        max_height = 2400
        title_size = 76
        sub_size = 42
        date_size = 32
        row_font = 26
        row_h = 61
        card_pad = 20
        max_matches_per_day = 6
        flag_w, flag_h = 48, 31
        team_gap = 158
        time_font = 22

    card_w = (width - 2 * margin_x - (cols - 1) * gap_x) // cols

    card_heights = []
    for _date_txt, matches in blocks:
        n = max(1, min(len(matches or []), max_matches_per_day))
        ch = 108 + n * row_h + max(0, n - 1) * 9 + 40
        card_heights.append(max(292, min(760 if wide_mode else 640, ch)))

    rows_count = (len(blocks) + cols - 1) // cols
    row_heights = []
    for r in range(rows_count):
        row_cards = card_heights[r * cols:(r + 1) * cols]
        row_heights.append(max(row_cards) if row_cards else 292)

    height = int(start_y + sum(row_heights) + gap_y * max(0, rows_count - 1) + 165)
    height = max(min_height, min(max_height, height))

    img, draw = _games_day_background(width, height)

    # طبقة فخامة/تعتيم ناعمة فوق الخلفية
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    rounded_rect(od, (46, 46, width - 46, height - 46), radius=50, fill="#03112A77", outline="#FFFFFF28", width=2)
    rounded_rect(od, (60, 60, width - 60, height - 60), radius=44, fill="#061B4938", outline="#14B8F555", width=2)
    try:
        od.ellipse((width-420, 90, width+150, 620), fill="#0EA5E950")
        od.ellipse((-260, 220, 300, 830), fill="#1D4ED850")
        overlay = overlay.filter(ImageFilter.GaussianBlur(12))
    except Exception:
        pass
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    # عنوان أفخم
    rounded_rect(draw, (width//2 - 560, 64, width//2 + 560, 222), radius=42, fill="#061633CC", outline="#FFFFFF22", width=2)
    draw_text(draw, (width // 2, 112), "مباريات الأيام", get_font(title_size), fill="#FFFFFF", max_width=width - 180)
    draw_text(draw, (width // 2, 192), "جدول مباريات عدة أيام", get_font(sub_size), fill="#FBBF24", max_width=980)
    draw.line((width // 2 - 470, 260, width // 2 + 470, 260), fill="#FBBF2477", width=3)

    # يبدأ من اليمين لليسار
    x_positions = [margin_x] if cols == 1 else [width - margin_x - card_w, margin_x]
    y_cursor = start_y
    for idx, (date_txt, matches) in enumerate(blocks):
        c = idx % cols
        r = idx // cols
        if c == 0 and idx != 0:
            y_cursor += row_heights[r - 1] + gap_y

        x = x_positions[c]
        y = y_cursor
        card_h = row_heights[r]

        # كرت اليوم
        rounded_rect(draw, (x, y, x + card_w, y + card_h), radius=34, fill="#052C83EE", outline="#38BDF8", width=4)
        rounded_rect(draw, (x + 6, y + 6, x + card_w - 6, y + card_h - 6), radius=30, fill="#06163355", outline="#1D4ED880", width=2)

        # شريط التاريخ بدون رموز تسبب مربعات
        rounded_rect(draw, (x + card_pad, y + 18, x + card_w - card_pad, y + 88), radius=20, fill="#FBBF24", outline="#FFFFFF55", width=2)
        draw.line((x + card_pad + 28, y + 82, x + card_w - card_pad - 28, y + 82), fill="#7C2D1288", width=2)
        draw_text(draw, (x + card_w // 2, y + 53), _multi_days_ar_date_label(date_txt), get_font(date_size), fill="#061633", max_width=card_w - 70)

        yy = y + 112
        match_list = list(matches or [])[:max_matches_per_day]
        if not match_list:
            rounded_rect(draw, (x + 30, yy, x + card_w - 30, yy + row_h), radius=16, fill="#07132FEE", outline="#FFFFFF28", width=1)
            draw_text(draw, (x + card_w // 2, yy + row_h // 2), "لا توجد مباريات", get_font(row_font), fill="#FFFFFF")
        else:
            for row_index, (team1, team2, t) in enumerate(match_list):
                row_fill = "#07132FEE" if row_index % 2 == 0 else "#0A1D44EE"
                rounded_rect(draw, (x + 30, yy, x + card_w - 30, yy + row_h), radius=16, fill=row_fill, outline="#38BDF840", width=2)

                tx = x + card_w // 2
                tbox_w = 140 if wide_mode else 116
                rounded_rect(draw, (tx - tbox_w // 2, yy + 11, tx + tbox_w // 2, yy + row_h - 11), radius=14, fill="#081123", outline="#FBBF24", width=2)
                draw_text(draw, (tx, yy + row_h // 2), _display_time_ar(t), get_font(time_font), fill="#FBBF24", max_width=tbox_w-12)

                # أعلام أكبر وأوضح مع fallback إذا العلم ناقص
                try:
                    _v31_paste_flag(img, team1, (x + card_w - 104, yy + (row_h-flag_h)//2, x + card_w - 104 + flag_w, yy + (row_h-flag_h)//2 + flag_h))
                except Exception:
                    try:
                        paste_flag(img, team1, (x + card_w - 104, yy + (row_h-flag_h)//2, x + card_w - 104 + flag_w, yy + (row_h-flag_h)//2 + flag_h))
                    except Exception:
                        pass
                try:
                    _v31_paste_flag(img, team2, (x + 40, yy + (row_h-flag_h)//2, x + 40 + flag_w, yy + (row_h-flag_h)//2 + flag_h))
                except Exception:
                    try:
                        paste_flag(img, team2, (x + 40, yy + (row_h-flag_h)//2, x + 40 + flag_w, yy + (row_h-flag_h)//2 + flag_h))
                    except Exception:
                        pass

                draw_text(draw, (x + card_w - team_gap, yy + row_h // 2), normalize_name(team1), get_font(row_font), fill="#FFFFFF", max_width=240 if wide_mode else 190)
                draw_text(draw, (x + team_gap, yy + row_h // 2), normalize_name(team2), get_font(row_font), fill="#FFFFFF", max_width=240 if wide_mode else 190)
                yy += row_h + 9

    draw.line((width//2 - 240, height - 86, width//2 - 90, height - 86), fill="#FBBF2455", width=3)
    draw.line((width//2 + 90, height - 86, width//2 + 240, height - 86), fill="#FBBF2455", width=3)
    draw_text(draw, (width // 2, height - 60), "المصيف ينقل لكم الحدث", get_font(34 if wide_mode else 30), fill="#FBBF24")
    suffix = "_10" if wide_mode else ""
    out = os.path.join(GENERATED_DIR, f"multi_days_schedule_ar{suffix}.png")
    img.save(out, quality=96)
    return out


async def multi_days_matches_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        style = command_style(update.message.text, 4)
        blocks = parse_multi_days_matches_text(update.message.text)
        if not blocks:
            await update.message.reply_text(
                "اكتبها كذا:\n"
                "/مباريات_الأيام 4\n"
                "24/06/2026\n"
                "الهلال * النصر * 8:00 م\n"
                "الهلال * الشباب * 10:00 م\n\n"
                "25/06/2026\n"
                "الأهلي * الاتحاد * 9:00 م"
            )
            return
        path = create_multi_days_matches_image(blocks, style, max_blocks=6, wide_mode=False)
        await send_photo_path(update, path, "✅ جدول مباريات عدة أيام")
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات الأيام ❌\n{e}")


# V31 background template filenames
V31_FULL_BG = "games_v31_full_bg.png"
V31_CLEAN_BG = "games_v31_clean_bg.png"

def _v31_latin_font(size):
    # خط إنجليزي احتياطي للعناوين والتاريخ حتى ما تطلع مربعات لو خط العربي ما يدعم Latin.
    candidates = [
        "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "Arial.ttf",
    ]
    try:
        for fp in candidates:
            if os.path.exists(fp):
                return ImageFont.truetype(fp, size)
    except Exception:
        pass
    return get_font(size)


def _v31_template_path(name):
    try:
        p = template_path(name)
        if p and os.path.exists(p):
            return p
    except Exception:
        pass
    return os.path.join("assets", "templates", name)


def _v31_load_bg(name, width=1200, height=1500):
    path = _v31_template_path(name)
    if Image and os.path.exists(path):
        try:
            img = Image.open(path).convert("RGB").resize((width, height), Image.LANCZOS)
            return img, ImageDraw.Draw(img)
        except Exception:
            pass
    # احتياط لو الخلفية ناقصة: خلفية زرقاء نظيفة بدون كراش
    img = Image.new("RGB", (width, height), "#061633")
    draw = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height - 1, 1)
        r = int(2 + 4 * t)
        g = int(13 + 16 * t)
        b = int(36 + 42 * t)
        draw.line((0, y, width, y), fill=(r, g, b))
    return img, draw


def _v31_safe_date(day_name):
    try:
        return _format_design_date_text(day_name)
    except Exception:
        return str(day_name or "").strip()


def _v31_time_label(t):
    t = str(t or "").strip()
    if not t:
        return "VS"
    try:
        return _display_time_en(t)
    except Exception:
        pass
    try:
        tm, period = _ampm_from_time(t)
        return f"{tm} {period}".strip()
    except Exception:
        return t


def _v31_layout(count, clean=False):
    """
    توزيع V31 بعد الاعتماد: نستغل المساحة الوسطى أكثر ونقلل الفراغ تحت الكروت،
    مع الحفاظ على العنوان والتذييل الثابتين داخل الخلفية.
    """
    count = max(1, min(int(count or 1), 7))
    if count == 1:
        return 240, 0, 555 if not clean else 600
    if count == 2:
        return 205, 38, 490 if not clean else 520
    if count == 3:
        return 172, 30, 455 if not clean else 465
    if count == 4:
        return 152, 22, 425 if not clean else 430
    if count == 5:
        # مساحة أكبر للكروت حتى تظهر الأعلام أوضح بدون تغيير الهوية
        return 140, 10, 410 if not clean else 410
    if count == 6:
        # ننزل كروت المباريات قليلًا لتقليل الفراغ السفلي وتوسيط التاريخ بصريًا
        return 126, 8, 408 if not clean else 410
    return 100, 10, 390 if not clean else 390


def _v31_fit_ar_font(draw, text, start_size, max_width, min_size=20):
    """يكبر الخط العربي قدر الإمكان بدون خروجه من مساحة الفريق داخل كرت V31."""
    text = _clean_display_name(text)
    size = int(start_size)
    while size > int(min_size):
        try:
            if text_width(draw, text, get_font(size)) <= max_width:
                break
        except Exception:
            break
        size -= 1
    return get_font(max(size, int(min_size)))


def _v31_draw_team_name(draw, pos, name, font, max_width, fill="#FFFFFF"):
    """ظل خفيف + اسم واضح داخل المربع، بدون تغيير شكل التصميم."""
    x, y = pos
    clean_name = _clean_display_name(name)
    try:
        draw_text(draw, (x + 2, y + 2), clean_name, font, fill="#000814", max_width=max_width)
    except Exception:
        pass
    draw_text(draw, (x, y), clean_name, font, fill=fill, max_width=max_width)


def _v31_paste_flag(img, team_name, box):
    """
    لصق أعلام V31 بحجم أوضح وآمن.
    - يكبر العلم صراحة بدل thumbnail.
    - يقص الفراغ الشفاف حول العلم إذا موجود، عشان العلم يبان كبير داخل الكرت.
    - إذا صار أي خطأ يرجع للدالة الأصلية بدون ما يطيح البوت.
    """
    x1, y1, x2, y2 = [int(v) for v in box]
    w = max(1, x2 - x1)
    h = max(1, y2 - y1)
    try:
        path = flag_path_for(team_name)
        if path and os.path.exists(path):
            flag = Image.open(path).convert("RGBA")

            # بعض ملفات الأعلام فيها فراغ شفاف، نقصّه عشان العلم يكبر فعليًا.
            try:
                alpha = flag.getchannel("A")
                bbox = alpha.getbbox()
                if bbox:
                    flag = flag.crop(bbox)
            except Exception:
                pass

            fw, fh = flag.size
            if fw > 0 and fh > 0:
                scale = min(w / fw, h / fh)
                nw = max(1, int(fw * scale))
                nh = max(1, int(fh * scale))
                flag = flag.resize((nw, nh), Image.LANCZOS)
                px = x1 + (w - nw) // 2
                py = y1 + (h - nh) // 2
                img.paste(flag, (px, py), flag)
                return
    except Exception:
        pass
    try:
        paste_flag(img, team_name, (x1, y1, x2, y2))
    except Exception:
        pass

def _v31_card(img, draw, box, idx, team_a, team_b, time_text, count):
    x1, y1, x2, y2 = [int(v) for v in box]
    cy = (y1 + y2) // 2
    row_h = y2 - y1
    card_w = x2 - x1

    # الكرت الأساسي — نفس الهوية، فقط تنسيق داخلي أوضح.
    rounded_rect(draw, (x1, y1, x2, y2), radius=30, fill="#061633", outline="#2F80FF", width=3)
    draw.line((x1+32, y1+5, x2-32, y1+5), fill="#FBBF24", width=3)
    draw.line((x1+32, y2-5, x2-32, y2-5), fill="#1257D6", width=3)

    # رقم المباراة
    badge_size = 30 if row_h >= 112 else 27
    rounded_rect(draw, (x2-badge_size-10, y1+14, x2-10, y1+14+badge_size), radius=10, fill="#FBBF24", outline="#FFFFFF33", width=1)
    draw_text(draw, (x2-10-badge_size//2, y1+14+badge_size//2), str(idx), _v31_latin_font(18), fill="#061633", max_width=badge_size)

    # ضبط نهائي: تصغير الأعلام 15٪ من الوضع الحالي.
    # تبقى واضحة وكبيرة، لكن بدون ما تزاحم أسماء المنتخبات أو رقم المباراة.
    if count >= 5:
        flag_h = min(70, row_h - 16)
        flag_w = int(flag_h * 1.55)
        side_pad = 22
    elif count == 4:
        flag_h = min(78, row_h - 18)
        flag_w = int(flag_h * 1.55)
        side_pad = 23
    else:
        flag_h = min(85, row_h - 20)
        flag_w = int(flag_h * 1.55)
        side_pad = 24

    # الفريق الأول يمين، الفريق الثاني يسار
    right_flag_box = (x2-side_pad-flag_w, cy-flag_h//2, x2-side_pad, cy+flag_h//2)
    left_flag_box = (x1+side_pad, cy-flag_h//2, x1+side_pad+flag_w, cy+flag_h//2)
    _v31_paste_flag(img, team_a, right_flag_box)
    _v31_paste_flag(img, team_b, left_flag_box)

    # مربع الوقت بالنص: ثابت وواضح، ولا يأكل مساحة أسماء المنتخبات.
    time_w = 176 if count <= 3 else (158 if count <= 5 else 146)
    time_h = 70 if row_h >= 115 else 62
    cx = (x1 + x2) // 2
    rounded_rect(draw, (cx-time_w//2, cy-time_h//2, cx+time_w//2, cy+time_h//2), radius=18, fill="#020A1B", outline="#FBBF24", width=2)
    draw_text(draw, (cx, cy), _v31_time_label(time_text), _v31_latin_font(31 if count <= 4 else 27), fill="#FBBF24", max_width=time_w-16)

    # أسماء المنتخبات: خط عربي أوضح وأكبر، مع تصغير تلقائي إذا طال الاسم.
    base_team_size = 38 if count <= 2 else (36 if count == 3 else (34 if count == 4 else (31 if count <= 6 else 27)))

    right_text_left = cx + time_w//2 + 18
    right_text_right = right_flag_box[0] - 10
    left_text_left = left_flag_box[2] + 10
    left_text_right = cx - time_w//2 - 18
    right_text_width = max(95, int(right_text_right - right_text_left))
    left_text_width = max(95, int(left_text_right - left_text_left))
    right_text_x = (right_text_left + right_text_right) // 2
    left_text_x = (left_text_left + left_text_right) // 2

    font_a = _v31_fit_ar_font(draw, team_a, base_team_size, right_text_width, min_size=22 if count <= 6 else 20)
    font_b = _v31_fit_ar_font(draw, team_b, base_team_size, left_text_width, min_size=22 if count <= 6 else 20)
    _v31_draw_team_name(draw, (right_text_x, cy), team_a, font_a, right_text_width)
    _v31_draw_team_name(draw, (left_text_x, cy), team_b, font_b, left_text_width)


def _v31_draw_date_pill(draw, width, y, day_name):
    date_txt = _v31_safe_date(day_name)
    if not date_txt:
        return
    x1, x2 = 345, width - 345
    rounded_rect(draw, (x1, y-31, x2, y+31), radius=22, fill="#061633", outline="#FBBF24", width=2)
    font_obj = get_font(30) if re.search(r"[\u0600-\u06FF]", date_txt) else _v31_latin_font(30)
    draw_text(draw, (width//2, y), date_txt, font_obj, fill="#FFFFFF", max_width=x2-x1-30)


def create_matches_today_v31_full_image(day_name, matches):
    ensure_generated_dir()
    matches = list(matches or [])[:7]
    count = max(len(matches), 1)
    width, height = 1200, 1500
    img, draw = _v31_load_bg(V31_FULL_BG, width, height)

    # الخلفية فيها العنوان والتذييل ثابتين؛ نضيف التاريخ والكروت فقط.
    _v31_draw_date_pill(draw, width, 365, day_name)

    row_h, gap, y = _v31_layout(count, clean=False)
    x1, x2 = 212, 1058
    for idx, (a, b, t) in enumerate(matches, start=1):
        _v31_card(img, draw, (x1, y, x2, y+row_h), idx, a, b, t, count)
        y += row_h + gap

    path = os.path.join(GENERATED_DIR, "matches_today_v31_full.png")
    img.save(path, quality=96)
    return path


def create_matches_today_v31_clean_image(day_name, matches):
    ensure_generated_dir()
    matches = list(matches or [])[:7]
    count = max(len(matches), 1)
    width, height = 1200, 1500
    img, draw = _v31_load_bg(V31_CLEAN_BG, width, height)

    # النسخة النظيفة: العنوان كامل ديناميكي من البوت.
    draw_text(draw, (width//2, 70), "MONDIAL AL MASEEF 2026", _v31_latin_font(36), fill="#FFFFFF", max_width=760)
    draw_text(draw, (width//2, 165), "GAMES OF THE DAY", _v31_latin_font(86), fill="#FFFFFF", max_width=940)
    _v31_draw_date_pill(draw, width, 312, day_name)

    row_h, gap, y = _v31_layout(count, clean=True)
    x1, x2 = 212, 1058
    for idx, (a, b, t) in enumerate(matches, start=1):
        _v31_card(img, draw, (x1, y, x2, y+row_h), idx, a, b, t, count)
        y += row_h + gap

    draw_text(draw, (width//2, height-70), "المصيف ينقل لكم الحدث", get_font(36), fill="#FBBF24", max_width=620)
    path = os.path.join(GENERATED_DIR, "matches_today_v31_clean.png")
    img.save(path, quality=96)
    return path


async def matches_today_v31_full_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text(
                "اكتبها كذا:\n"
                "/مباريات_اليوم\n"
                "24/06/2026\n"
                "السعودية * فرنسا * 8:00 م\n"
                "إسبانيا * الأرجنتين * 11:00 م"
            )
            return
        path = create_matches_today_v31_full_image(day_name, matches)
        await send_photo_path(update, path, build_matches_today_v31_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات اليوم V31 ❌\n{e}")


async def matches_today_v31_clean_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text(
                "اكتبها كذا:\n"
                "/مباريات_اليوم2\n"
                "24/06/2026\n"
                "السعودية * فرنسا * 8:00 م\n"
                "إسبانيا * الأرجنتين * 11:00 م"
            )
            return
        path = create_matches_today_v31_clean_image(day_name, matches)
        await send_photo_path(update, path, build_matches_today_v31_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات اليوم V31-2 ❌\n{e}")


# -------------------- Render by style --------------------

def render_matches_by_style(day_name, matches, style, auto=False):
    style = int(style)
    if style == 1:
        return create_matches_template_image(day_name, matches, use_template=not auto)
    if style == 2:
        return create_matches_newlook_image(day_name, matches, 2)
    if style == 3:
        return create_match_frame_style_image(day_name, matches, False)
    return create_matches_newlook_image(day_name, matches, 4)


def render_results_by_style(day_name, results, style, auto=False):
    style = int(style)
    if style == 1:
        return create_match_results_template_image(day_name, results, use_template=not auto)
    if style == 2:
        return create_match_results_newlook_image(day_name, results, 2)
    if style == 3:
        return create_match_frame_style_image(day_name, results, True)
    return create_match_results_newlook_image(day_name, results, 4)


def render_scorers_by_style(items, style, auto=False):
    style = int(style)
    if style == 1:
        return create_top_scorers_template_image(items, use_template=not auto)
    if style == 2:
        return create_scorers_newlook_image(items, 2)
    if style == 3:
        return create_scorers_frame_style_image(items)
    return create_scorers_newlook_image(items, 4)


def render_group_by_style(group_title, rows, style, auto=False):
    style = int(style)
    if style == 1:
        return create_group_standing_image(group_title, rows, use_template=not auto)
    if style == 2:
        return create_group_newlook_image(group_title, rows, 2)
    if style == 3:
        return create_group_frame_style_image(group_title, rows)
    return create_group_newlook_image(group_title, rows, 4)


def render_all_groups_by_style(groups, style, auto=False):
    style = int(style)
    if style == 1:
        return create_all_groups_newlook_image(groups, 1)
    if style == 3:
        return create_all_groups_frame_style_image(groups)
    if style == 4:
        return create_all_groups_newlook_image(groups, 4)
    return create_all_groups_newlook_image(groups, 2)


# -------------------- Commands V27 --------------------

async def set_style_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nums = get_numbers(update.message.text)
    if not nums or nums[0] not in (1, 2, 3, 4):
        await update.message.reply_text("اكتبها كذا:\n/اعتماد_ستايل 4\n\nالخيارات: 1، 2، 3، 4")
        return
    style = save_design_style(nums[0])
    await update.message.reply_text(f"تم اعتماد الستايل {style} ✅\n{STYLE_NAMES[style]}")


async def get_style_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    style = current_design_style()
    await update.message.reply_text(f"الستايل الحالي: {style} ✅\n{STYLE_NAMES.get(style, '')}")


async def matches_command_v27(update: Update, context: ContextTypes.DEFAULT_TYPE, forced_style=None, auto=False):
    try:
        style = command_style(update.message.text, forced_style)
        day_name, matches = parse_matches_text(update.message.text)
        if not matches:
            await update.message.reply_text("اكتبها كذا:\n/مباريات 4\nالسابع\nفرنسا * البرتغال * 8:00 م\nإنجلترا * كرواتيا * 11:00 م")
            return
        path = render_matches_by_style(day_name, matches, style, auto=auto)
        await send_photo_path(update, path, build_design_matches_caption(day_name, matches))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم المباريات ❌\n{e}")


async def results_command_v27(update: Update, context: ContextTypes.DEFAULT_TYPE, forced_style=None, auto=False):
    try:
        style = command_style(update.message.text, forced_style)
        day_name, results = parse_match_results_design_text(update.message.text)
        if not results:
            await update.message.reply_text("اكتبها كذا:\n/انتهت 4\nالسابع\nفرنسا 2 * 6 البرتغال\nإنجلترا 3 * 0 كرواتيا")
            return
        path = render_results_by_style(day_name, results, style, auto=auto)
        await send_photo_path(update, path, build_design_results_caption(day_name, results))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم نتائج المباريات ❌\n{e}")


async def scorers_command_v27(update: Update, context: ContextTypes.DEFAULT_TYPE, forced_style=None, auto=False):
    try:
        style = command_style(update.message.text, forced_style)
        items = parse_scorers_text(update.message.text)
        if not items:
            await update.message.reply_text("اكتبها كذا:\n/هدافين 4\nميسي * الأرجنتين * 3\nرونالدو * البرتغال * 6")
            return
        path = render_scorers_by_style(items, style, auto=auto)
        await send_photo_path(update, path, build_top_scorers_caption(items))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم الهدافين ❌\n{e}")


async def group_command_v27(update: Update, context: ContextTypes.DEFAULT_TYPE, forced_style=None, auto=False):
    try:
        style = command_style(update.message.text, forced_style)
        group_title, rows = parse_group_standing_text(update.message.text)
        if not rows:
            await update.message.reply_text("اكتبها كذا:\n/مجموعة 4\nC\nاسكتلندا * لعب 1 * فارق +1 * نقاط 6\nالمغرب * لعب 1 * فارق 0 * نقاط 1")
            return
        path = render_group_by_style(group_title, rows, style, auto=auto)
        await send_photo_path(update, path, build_group_standing_caption(group_title, rows))
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم ترتيب المجموعة ❌\n{e}")


async def all_groups_command_v27(update: Update, context: ContextTypes.DEFAULT_TYPE, forced_style=None, auto=False):
    try:
        style = command_style(update.message.text, forced_style)
        groups = parse_all_groups_text(update.message.text)
        if not groups:
            await update.message.reply_text("اكتبها كذا:\n/كل_المجموعات 4\nA\nالمكسيك * 1 * +2 * 3\nكوريا الجنوبية * 1 * +1 * 3\n\nB\nكندا * 1 * 0 * 1")
            return
        path = render_all_groups_by_style(groups, style, auto=auto)
        await send_photo_path(update, path, "ترتيب جميع المجموعات ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم جميع المجموعات ❌\n{e}")


# Wrappers عشان MessageHandler يستدعي دالة بدون بارامترات إضافية
async def short_matches_command(update, context):
    return await matches_command_v27(update, context)
async def short_matches_auto_command(update, context):
    return await matches_command_v27(update, context, auto=True)
async def short_results_command(update, context):
    return await results_command_v27(update, context)
async def short_results_auto_command(update, context):
    return await results_command_v27(update, context, auto=True)
async def short_scorers_command(update, context):
    return await scorers_command_v27(update, context)
async def short_scorers_auto_command(update, context):
    return await scorers_command_v27(update, context, auto=True)
async def short_group_command(update, context):
    return await group_command_v27(update, context)
async def short_group_auto_command(update, context):
    return await group_command_v27(update, context, auto=True)
async def short_all_groups_command(update, context):
    return await all_groups_command_v27(update, context)
async def short_all_groups_auto_command(update, context):
    return await all_groups_command_v27(update, context, auto=True)

# Old commands routed to the new 1-4 style system.
async def design_matches_template_command(update, context):
    return await matches_command_v27(update, context, auto=False)
async def design_matches_auto_command(update, context):
    return await matches_command_v27(update, context, auto=True)
async def design_match_results_template_command(update, context):
    return await results_command_v27(update, context, auto=False)
async def design_match_results_auto_command(update, context):
    return await results_command_v27(update, context, auto=True)
async def design_group_standing_template_command(update, context):
    return await group_command_v27(update, context, auto=False)
async def design_group_standing_auto_command(update, context):
    return await group_command_v27(update, context, auto=True)
async def design_scorers_template_command(update, context):
    return await scorers_command_v27(update, context, auto=False)
async def design_scorers_auto_command(update, context):
    return await scorers_command_v27(update, context, auto=True)

async def design_matches_style2_command(update, context):
    return await matches_command_v27(update, context, forced_style=2)
async def design_match_results_style2_command(update, context):
    return await results_command_v27(update, context, forced_style=2)
async def design_scorers_style2_command(update, context):
    return await scorers_command_v27(update, context, forced_style=2)
async def design_group_style2_command(update, context):
    return await group_command_v27(update, context, forced_style=2)
async def design_matches_frame_command(update, context):
    return await matches_command_v27(update, context, forced_style=3)
async def design_results_frame_command(update, context):
    return await results_command_v27(update, context, forced_style=3)
async def design_all_groups_command(update, context):
    return await all_groups_command_v27(update, context)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n/نتائج 5\n/اعتماد_نتائج 5\n/احصائيات\n/احصائيات 1 6\n/الترتيب_العام\n/الترتيب_العام 1 6\n/ترتيب_نص\n\n"
        "أوامر الصور والتقارير:\n"
        "/صورة_اليوم 6\n/صورة_الترتيب 1 6\n/صورة_الاساطير 1 6\n/صورة_احصائيات 1 6 لوحة عامة\n/صور_الاحصائيات 1 6\n"
        "/بطاقة فارس سالم\n/تقرير_الفترة 1 4\n/تفعيل_الصور_التلقائية\n/إيقاف_الصور_التلقائية\n\n"
        "أوامر الستايل:\n"
        "/اعتماد_ستايل 1\n/اعتماد_ستايل 2\n/اعتماد_ستايل 3\n/اعتماد_ستايل 4\n/الستايل\n\n"
        "أوامر التصاميم المختصرة:\n"
        "/مباريات 4\n/انتهت 4\n/مجموعة 4\n/هدافين 4\n/كل_المجموعات 4\n\n"
        "الأوامر التلقائية المختصرة:\n"
        "/مباريات_تلقائي 4\n/انتهت_تلقائي 4\n/مجموعة_تلقائي 4\n/هدافين_تلقائي 4\n/كل_المجموعات_تلقائي 4\n\n"
        "أوامر التصاميم القديمة باقية وتدعم 1-4:\n"
        "/تصميم_مباريات 4\n/تصميم_مباريات_تلقائي 4\n/تصميم_نتائج_مباريات 4\n/تصميم_نتائج_مباريات_تلقائي 4\n"
        "/تصميم_ترتيب_مجموعة 4\n/تصميم_ترتيب_مجموعة_تلقائي 4\n/تصميم_هدافين 4\n/تصميم_هدافين_تلقائي 4\n"
        "/تصميم_مباريات_ستايل2\n/تصميم_نتائج_مباريات_ستايل2\n/تصميم_هدافين_ستايل2\n/تصميم_ترتيب_مجموعة_ستايل2\n"
        "/تصميم_مباريات_اطار\n/تصميم_نتائج_مباريات_اطار\n/تصميم_جميع_المجموعات 4\n\n"
        "أوامر الكأس:\n"
        "/بدء_الكاس 7\n/حالة_الكاس\n/مواجهات_الكاس\n/نتائج_الكاس\n/إعادة_الكاس_من 7\n/الغاء_الكاس تأكيد\n\n"
        "أوامر الفحص والنشر:\n"
        "/الأيام\n/فحص 5\n/مشاركين 5\n/اسطورة 5\n/مقارنة 4 5\n/اعلان_اليوم 5\n/ملخص_اليوم 5\n\n"
        "أوامر الاستيراد والنسخ:\n"
        "/استيراد_ملف\n/اعتماد_استيراد\n/إلغاء_استيراد\n/نسخة_احتياطية\n/استرجاع_نسخة\n/تنظيف_الأيام\n/تنظيف_الملفات\n\n"
        "أوامر الأمان:\n"
        "/مسح_نتائج 5\n/مسح_يوم 5\n/مسح_الكل تأكيد\n/استرجاع_آخر\n/قفل_يوم 5\n/فتح_يوم 5\n/معرفي\n\n"
        "مهم: /نتائج للفانتزي فقط، و/انتهت لتصميم نتائج المباريات."
    )






async def multi_days_matches10_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        style = command_style(update.message.text, 4)
        blocks = parse_multi_days_matches_text(update.message.text)
        if not blocks:
            await update.message.reply_text(
                "اكتبها كذا:\n"
                "/مباريات_الأيام10 4\n"
                "24/06/2026\n"
                "الهلال * النصر * 8:00 م\n"
                "الهلال * الشباب * 10:00 م\n\n"
                "25/06/2026\n"
                "الأهلي * الاتحاد * 9:00 م"
            )
            return
        path = create_multi_days_matches_image(blocks, style, max_blocks=10, wide_mode=True)
        await send_photo_path(update, path, "✅ جدول مباريات عدة أيام — نسخة 10 أيام")
    except Exception as e:
        await update.message.reply_text(f"تعذر تصميم مباريات الأيام 10 ❌\n{e}")

def _parse_range_from_text(text):
    nums = get_numbers(text or "")
    existing = get_existing_days(1, 999)
    if not existing:
        return 1, 1
    if len(nums) >= 2:
        start_day, end_day = int(nums[0]), int(nums[1])
    elif len(nums) == 1:
        start_day, end_day = 1, int(nums[0])
    else:
        start_day, end_day = min(existing), max(existing)

    if start_day > end_day:
        start_day, end_day = end_day, start_day
    start_day = max(min(existing), start_day)
    end_day = min(max(existing), end_day)
    return start_day, end_day


def _stats_active_participants(start_day, end_day):
    stats = collect_stats(start_day, end_day)
    ranking = stats.get("ranking", []) or PARTICIPANTS[:]
    totals = stats.get("totals", {})
    active_counts = stats.get("participation_count", {})
    active = [p for p in ranking if totals.get(p, 0) > 0 or active_counts.get(p, 0) > 0]
    return active or ranking or PARTICIPANTS[:]


def create_statistics_cover_image(start_day, end_day):
    ensure_generated_dir()
    width, height = 1400, 1800
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "فانتزي المصيف 2026", "الملف الإحصائي الرسمي", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=260, bottom_pad=120, accent="#F59E0B")

    rounded_rect(draw, (110, fy1+40, width-110, fy1+300), radius=38, fill="#091122DD", outline="#FFFFFF33", width=2)
    draw_text(draw, (width//2, fy1+105), "فانتزي المصيف 2026", get_font(64), fill="#FFFFFF", max_width=1000)
    draw_text(draw, (width//2, fy1+190), "التقرير الإحصائي الشامل", get_font(42), fill="#FDE68A", max_width=900)

    rounded_rect(draw, (170, fy1+360, width-170, fy1+505), radius=28, fill="#7C3AEDDD", outline="#FFFFFF33", width=2)
    draw_text(draw, (width//2, fy1+412), f"الفترة من اليوم {start_day} إلى اليوم {end_day}", get_font(38), fill="#FFFFFF", max_width=880)
    draw_text(draw, (width//2, fy1+468), "تم توليد الملف مباشرة من بيانات الإكسل", get_font(28), fill="#E5E7EB", max_width=760)

    features = [
        "الترتيب العام",
        "سجل الأساطير",
        "تقرير الفترة",
        "ملخصات الأيام",
        "تفاصيل اختيارات اللاعبين",
        "بطاقات المشاركين",
    ]
    y = fy1 + 600
    for item in features:
        rounded_rect(draw, (200, y, width-200, y+86), radius=24, fill="#0B1020", outline="#FFFFFF22", width=1)
        draw_text(draw, (width//2, y+43), item, get_font(30), fill="#FFFFFF", max_width=800)
        y += 102

    draw_text(draw, (width//2, height-60), "المصيف ينقل لكم الحدث", get_font(28), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"statistics_cover_{start_day}_{end_day}.png")
    img.save(path, quality=95)
    return path


def create_day_choices_image(day):
    ensure_generated_dir()
    rows = read_day_rows(day, data_only=True)
    if not rows:
        raise ValueError(f"لا توجد بيانات لليوم {day}")

    rows = sorted(rows, key=lambda r: (-int(r.get("total", 0)), _clean_display_name(r.get("participant"))))

    width = 1600
    row_h = 84
    top = 345
    bottom = 96
    height = max(1360, top + len(rows) * (row_h + 10) + bottom)
    img, draw = design_canvas(None, width, height, "purple")
    draw_design_header(draw, width, "تفاصيل اختيارات اللاعبين", f"فانتزي المصيف 2026\nاليوم {ordinal_day(day)}", img)
    fx1, fy1, fx2, fy2 = draw_broadcast_inner_frame(draw, width, height, top=255, bottom_pad=96, accent="#06B6D4")

    cols = [
        (1450, "المشارك", 240),
        (1215, "الحارس", 145),
        (1010, "اللاعب 1", 145),
        (805, "اللاعب 2", 145),
        (600, "اللاعب 3", 145),
        (390, "الكابتن", 145),
        (155, "المجموع", 95),
    ]
    y = fy1 + 26
    rounded_rect(draw, (65, y, width - 65, y + 60), radius=18, fill="#05070D", outline="#FFFFFF55", width=1)
    for x, title, _mw in cols:
        draw_text(draw, (x, y + 30), title, get_font(22), fill="#FDE68A")
    y += 76

    for idx, r in enumerate(rows, start=1):
        accent = "#F59E0B" if idx == 1 else ("#FFFFFF22" if idx % 2 else "#FFFFFF15")
        fill = "#1A1407" if idx == 1 else ("#0B1020" if idx % 2 else "#10172A")
        rounded_rect(draw, (65, y, width - 65, y + row_h), radius=18, fill=fill, outline=accent, width=2 if idx == 1 else 1)
        cy = y + row_h // 2
        items = [
            (1450, _clean_display_name(r.get("participant")), 240, "#FFFFFF"),
            (1215, _clean_display_name(r.get("keeper")), 145, "#FFFFFF"),
            (1010, _clean_display_name(r.get("p1")), 145, "#FFFFFF"),
            (805, _clean_display_name(r.get("p2")), 145, "#FFFFFF"),
            (600, _clean_display_name(r.get("p3")), 145, "#FFFFFF"),
            (390, _clean_display_name(r.get("captain")), 145, "#FFFFFF"),
            (155, str(int(r.get("total", 0))), 95, "#FDE68A"),
        ]
        for x, txt, mw, color in items:
            draw_text(draw, (x, cy), txt or "-", get_font(23 if x != 155 else 26), fill=color, max_width=mw)
        y += row_h + 10

    draw_text(draw, (width // 2, height - 44), "المصيف ينقل لكم الحدث", get_font(24), fill="#FFFFFF")
    path = os.path.join(GENERATED_DIR, f"day_choices_{day}.png")
    img.save(path, quality=95)
    return path

def build_statistics_image_paths(start_day, end_day, include_cover=False):
    days = get_existing_days(start_day, end_day)
    if not days:
        raise ValueError("لا توجد أيام في هذا النطاق")

    paths = []
    if include_cover:
        paths.append(create_statistics_cover_image(start_day, end_day))

    paths.append(create_overall_ranking_image(start_day, end_day))
    paths.append(create_legends_image(start_day, end_day))
    paths.append(create_period_report_image(start_day, end_day))

    for day in days:
        paths.append(create_daily_result_image(day))

    for day in days:
        paths.append(create_day_choices_image(day))

    for name in _stats_active_participants(start_day, end_day):
        paths.append(create_participant_card_image(name, start_day, end_day))

    return paths


def _pdf_ready_image(path, max_w=900):
    """تحويل الصورة لصفحة PDF خفيفة عشان تيليقرام/Railway ما يعلقون."""
    img = Image.open(path)
    if img.mode not in ("RGB", "L"):
        if "A" in img.getbands():
            bg = Image.new("RGB", img.size, "white")
            bg.paste(img, mask=img.split()[-1])
            img = bg
        else:
            img = img.convert("RGB")
    else:
        img = img.convert("RGB")

    if img.width > max_w:
        nh = int(img.height * (max_w / img.width))
        img = img.resize((max_w, nh), Image.LANCZOS)
    return img


def _save_images_as_pdf(paths, out_path, max_w=900, quality=65):
    images = [_pdf_ready_image(p, max_w=max_w) for p in paths]
    try:
        images[0].save(
            out_path,
            "PDF",
            save_all=True,
            append_images=images[1:],
            resolution=80.0,
            quality=quality,
            optimize=True,
        )
    finally:
        for im in images:
            try:
                im.close()
            except Exception:
                pass


def create_statistics_pdf(start_day, end_day):
    ensure_generated_dir()
    paths = build_statistics_image_paths(start_day, end_day, include_cover=True)
    if not paths:
        raise ValueError("لا توجد صفحات PDF للتجهيز")

    # نحاول أكثر من ضغط؛ الهدف أن الملف ينرسل من تيليقرام بدون تعليق.
    attempts = [
        (900, 65, "normal"),
        (760, 60, "small"),
        (620, 55, "mini"),
    ]

    last_path = None
    for max_w, quality, label in attempts:
        out_path = os.path.join(GENERATED_DIR, f"fantasy_stats_{start_day}_{end_day}_{label}.pdf")
        _save_images_as_pdf(paths, out_path, max_w=max_w, quality=quality)
        last_path = out_path
        size_mb = os.path.getsize(out_path) / (1024 * 1024)
        if size_mb <= 45 or label == "mini":
            return out_path, len(paths), size_mb

    size_mb = os.path.getsize(last_path) / (1024 * 1024)
    return last_path, len(paths), size_mb


async def all_dashboard_images_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        start_day, end_day = _parse_range_from_text(update.message.text)
        await update.message.reply_text(f"جاري تجهيز صور الإحصائيات من اليوم {start_day} إلى {end_day} ✅")
        paths = build_statistics_image_paths(start_day, end_day, include_cover=False)
        for p in paths:
            await send_photo_path(update, p)
        await update.message.reply_text(f"تم إرسال {len(paths)} صورة إحصائية ✅")
    except Exception as e:
        await update.message.reply_text(f"تعذر تجهيز صور الإحصائيات ❌\n{e}")


async def statistics_pdf_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        start_day, end_day = _parse_range_from_text(update.message.text)
        await update.message.reply_text(f"جاري تجهيز ملف الإحصائيات PDF من اليوم {start_day} إلى {end_day} ⏳")

        pdf_path, page_count, size_mb = await asyncio.to_thread(create_statistics_pdf, start_day, end_day)

        if not os.path.exists(pdf_path):
            await update.message.reply_text("تعذر تجهيز ملف الإحصائيات ❌\nملف PDF لم يتم إنشاؤه.")
            return

        with open(pdf_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"فانتزي_المصيف_2026_{start_day}_{end_day}.pdf",
                caption=f"✅ ملف الإحصائيات جاهز\nعدد الصفحات: {page_count}\nالحجم: {size_mb:.1f} MB",
                read_timeout=240,
                write_timeout=240,
                connect_timeout=30,
                pool_timeout=30,
            )
    except Exception as e:
        await update.message.reply_text(f"تعذر تجهيز ملف الإحصائيات ❌\n{e}")


async def test_pdf_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """اختبار سريع لإرسال PDF بدون الاعتماد على بيانات الإكسل."""
    try:
        ensure_generated_dir()
        img = Image.new("RGB", (900, 1200), "#071426")
        draw = ImageDraw.Draw(img)
        draw_text(draw, (450, 390), "اختبار ملف PDF", get_font(58), fill="#FFFFFF", max_width=760)
        draw_text(draw, (450, 500), "إذا وصلك هذا الملف فإرسال PDF شغال", get_font(34), fill="#FBBF24", max_width=760)
        path = os.path.join(GENERATED_DIR, "test_pdf_send.pdf")
        img.save(path, "PDF", resolution=80.0, quality=65, optimize=True)
        with open(path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="اختبار_PDF.pdf",
                caption="✅ اختبار إرسال PDF شغال",
                read_timeout=120,
                write_timeout=120,
            )
    except Exception as e:
        await update.message.reply_text(f"تعذر اختبار PDF ❌\n{e}")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "البوت جاهز ✅\n\n"
        "الأوامر الأساسية:\n"
        "/اضافه 5\n/نتائج 5\n/اعتماد_نتائج 5\n/احصائيات\n/احصائيات 1 6\n/الترتيب_العام\n/الترتيب_العام 1 6\n/ترتيب_نص\n\n"
        "أوامر الصور والتقارير:\n"
        "/صورة_اليوم 6\n/صورة_الترتيب 1 6\n/صورة_الاساطير 1 6\n/صورة_احصائيات 1 6 لوحة عامة\n"
        "/صور_الاحصائيات\n/صور_الاحصائيات 1 6\n/ملف_الاحصائيات\n/ملف_الاحصائيات 1 6\n"
        "/بطاقة فارس سالم\n/تقرير_الفترة 1 4\n/تفعيل_الصور_التلقائية\n/إيقاف_الصور_التلقائية\n\n"
        "أوامر المباريات:\n"
        "/مباريات_اليوم\n/مباريات_اليوم2\n/مباريات_الأيام\n/مباريات_الايام\n/كل_المجموعات\n\n"
        "مهم: /نتائج للفانتزي فقط، و/انتهت لتصميم نتائج المباريات."
    )



# -------------------- ميزات الأخبار والمتأهلين والمباشر V2 --------------------

QUALIFIED32_FILE = "qualified32_state.json"
SPORTS_CACHE_FILE = "sports_cache_state.json"

WORLD_CUP_GROUPS = [
    ("A", ["المكسيك", "جنوب أفريقيا", "كوريا الجنوبية", "التشيك"]),
    ("B", ["كندا", "البوسنة والهرسك", "قطر", "سويسرا"]),
    ("C", ["البرازيل", "المغرب", "هايتي", "اسكتلندا"]),
    ("D", ["الولايات المتحدة", "باراغواي", "أستراليا", "تركيا"]),
    ("E", ["ألمانيا", "كوراساو", "ساحل العاج", "الإكوادور"]),
    ("F", ["هولندا", "اليابان", "السويد", "تونس"]),
    ("G", ["بلجيكا", "مصر", "إيران", "نيوزيلندا"]),
    ("H", ["إسبانيا", "الرأس الأخضر", "السعودية", "أوروجواي"]),
    ("I", ["فرنسا", "السنغال", "العراق", "النرويج"]),
    ("J", ["الأرجنتين", "الجزائر", "النمسا", "الأردن"]),
    ("K", ["البرتغال", "الكونغو الديمقراطية", "أوزبكستان", "كولومبيا"]),
    ("L", ["إنجلترا", "كرواتيا", "غانا", "بنما"]),
]

WORLD_CUP_TEAMS = [t for _, teams in WORLD_CUP_GROUPS for t in teams]

TEAM_ALIASES = {
    "أمريكا": "الولايات المتحدة",
    "امريكا": "الولايات المتحدة",
    "اميركا": "الولايات المتحدة",
    "أميركا": "الولايات المتحدة",
    "usa": "الولايات المتحدة",
    "unitedstates": "الولايات المتحدة",
    "الولاياتالمتحده": "الولايات المتحدة",
    "الولاياتالمتحدة": "الولايات المتحدة",
    "اسكتلندا": "اسكتلندا",
    "إسكتلندا": "اسكتلندا",
    "سكوتلندا": "اسكتلندا",
    "الباراغواي": "باراغواي",
    "بارجواي": "باراغواي",
    "اوروغواي": "أوروجواي",
    "اوروجواي": "أوروجواي",
    "الاوروغواي": "أوروجواي",
    "كوتديفوار": "ساحل العاج",
    "كوتديفوارى": "ساحل العاج",
    "ساحلالعاج": "ساحل العاج",
    "هولند": "هولندا",
    "انجلترا": "إنجلترا",
    "كوريا": "كوريا الجنوبية",
    "كورياالجنوبية": "كوريا الجنوبية",
    "كورياالجنوبية": "كوريا الجنوبية",
    "راسالخضراء": "الرأس الأخضر",
    "الرأسالخضر": "الرأس الأخضر",
    "جنوبافريقيا": "جنوب أفريقيا",
    "البوسنة": "البوسنة والهرسك",
    "الكونغو": "الكونغو الديمقراطية",
}


# أسماء إنجليزية شائعة للمصادر الخارجية مثل ESPN / FIFA
TEAM_ALIASES.update({
    "Mexico": "المكسيك", "South Africa": "جنوب أفريقيا", "Korea Republic": "كوريا الجنوبية", "South Korea": "كوريا الجنوبية", "Czechia": "التشيك",
    "Canada": "كندا", "Bosnia and Herzegovina": "البوسنة والهرسك", "Bosnia-Herzegovina": "البوسنة والهرسك", "Qatar": "قطر", "Switzerland": "سويسرا",
    "Brazil": "البرازيل", "Morocco": "المغرب", "Haiti": "هايتي", "Scotland": "اسكتلندا",
    "United States": "الولايات المتحدة", "USA": "الولايات المتحدة", "Paraguay": "باراغواي", "Australia": "أستراليا", "Turkey": "تركيا", "Türkiye": "تركيا",
    "Germany": "ألمانيا", "Curacao": "كوراساو", "Curaçao": "كوراساو", "Ivory Coast": "ساحل العاج", "Cote d'Ivoire": "ساحل العاج", "Ecuador": "الإكوادور",
    "Netherlands": "هولندا", "Japan": "اليابان", "Sweden": "السويد", "Tunisia": "تونس",
    "Belgium": "بلجيكا", "Egypt": "مصر", "Iran": "إيران", "IR Iran": "إيران", "New Zealand": "نيوزيلندا",
    "Spain": "إسبانيا", "Cape Verde": "الرأس الأخضر", "Cabo Verde": "الرأس الأخضر", "Saudi Arabia": "السعودية", "Uruguay": "أوروجواي",
    "France": "فرنسا", "Senegal": "السنغال", "Iraq": "العراق", "Norway": "النرويج",
    "Argentina": "الأرجنتين", "Algeria": "الجزائر", "Austria": "النمسا", "Jordan": "الأردن",
    "Portugal": "البرتغال", "DR Congo": "الكونغو الديمقراطية", "Congo DR": "الكونغو الديمقراطية", "Uzbekistan": "أوزبكستان", "Colombia": "كولومبيا",
    "England": "إنجلترا", "Croatia": "كرواتيا", "Ghana": "غانا", "Panama": "بنما",
})

TEAM_THEME_MAP = {
    "السعودية": ("#0F5132", "#22C55E", "#FFFFFF"),
    "المكسيك": ("#0B5130", "#CE1126", "#FFFFFF"),
    "الأرجنتين": ("#5ABCEB", "#FFFFFF", "#F6C945"),
    "فرنسا": ("#0F172A", "#2563EB", "#FFFFFF"),
    "إسبانيا": ("#991B1B", "#F59E0B", "#FDE68A"),
    "البرازيل": ("#0F5132", "#FDE047", "#FFFFFF"),
    "المغرب": ("#7F1D1D", "#15803D", "#FFFFFF"),
    "البرتغال": ("#14532D", "#B91C1C", "#FDE68A"),
    "إنجلترا": ("#111827", "#FFFFFF", "#DC2626"),
    "ألمانيا": ("#111827", "#B91C1C", "#FBBF24"),
    "كرواتيا": ("#991B1B", "#FFFFFF", "#1D4ED8"),
    "مصر": ("#991B1B", "#111827", "#FFFFFF"),
    "الولايات المتحدة": ("#1E3A8A", "#DC2626", "#FFFFFF"),
}

SPORTS_SOURCE_LABELS = {
    "official": "رسمي",
    "fast": "سريع",
    "latest": "الأحدث",
}


def simple_key(name):
    s = normalize_name(name).lower()
    rep = {
        "أ": "ا", "إ": "ا", "آ": "ا", "ة": "ه", "ى": "ي",
        "ؤ": "و", "ئ": "ي",
    }
    for a, b in rep.items():
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9\u0600-\u06FF]+", "", s)
    return s


def all_teams_canonical_map():
    data = {}
    for t in WORLD_CUP_TEAMS:
        data[simple_key(t)] = t
    for k, v in TEAM_ALIASES.items():
        data[simple_key(k)] = v
    # دعم أسماء الأعلام المعروفة في flags_map.json
    try:
        flags = load_flags_map()
        for k in flags.keys():
            can = flags.get(k)
            if k:
                kk = simple_key(k)
                if kk not in data and can:
                    data[kk] = can if can in WORLD_CUP_TEAMS else data.get(simple_key(can), can)
    except Exception:
        pass
    return data


def canonical_team_name(name):
    if not name:
        return None
    key = simple_key(name)
    data = all_teams_canonical_map()
    exact = data.get(key)
    if exact:
        return exact
    # تقريب ذكي
    options = list(data.keys())
    close = difflib.get_close_matches(key, options, n=1, cutoff=0.80)
    if close:
        return data.get(close[0])
    return None


def grouped_teams_text(with_points=False):
    lines = []
    for group, teams in WORLD_CUP_GROUPS:
        lines.append(f"المجموعة {group}")
        for t in teams:
            if with_points:
                lines.append(f"{t} — 0 نقطة")
            else:
                lines.append(t)
        lines.append("")
    return "\n".join(lines).strip()


def groups_template_text():
    lines = ["/كل_المجموعات"]
    for group, teams in WORLD_CUP_GROUPS:
        lines.append(group)
        for t in teams:
            lines.append(f"{t} * 0 * 0 * 0")
        lines.append("")
    return "\n".join(lines).strip()


def unsupported_team_message(name):
    header = f"ما قدرت أتعرف على المنتخب: {normalize_name(name)} ❌\n\n"
    keys = [t for t in WORLD_CUP_TEAMS]
    suggestions = difflib.get_close_matches(simple_key(name), [simple_key(x) for x in keys], n=5, cutoff=0.55)
    if suggestions:
        rev = {simple_key(x): x for x in keys}
        sug_names = [rev[s] for s in suggestions if s in rev]
        if sug_names:
            header += "هل تقصد:\n- " + "\n- ".join(sug_names) + "\n\n"
    header += "اكتب /المنتخبات لعرض كل الـ 48 منتخب المعتمدة."
    return header


def load_qualified32_state():
    if os.path.exists(QUALIFIED32_FILE):
        try:
            with open(QUALIFIED32_FILE, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if isinstance(obj, dict):
                teams = obj.get("teams", [])
                if isinstance(teams, list):
                    return [canonical_team_name(x) or normalize_name(x) for x in teams if normalize_name(x)]
        except Exception:
            pass
    return []


def save_qualified32_state(teams):
    data = {"teams": [normalize_name(x) for x in teams[:32]]}
    with open(QUALIFIED32_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _qualified32_template_path():
    candidates = [
        os.path.join("assets", "templates", "qualified32_template.png"),
        os.path.join("assets", "templates", "qualified32_template.jpg"),
        "qualified32_template.png",
        "/mnt/data/qualified32_template.png",
        "/mnt/data/كأس_العالم_2026_الفرق_المتأهلة.png",
    ]
    for p in candidates:
        try:
            if p and os.path.exists(p):
                return p
        except Exception:
            pass
    return None


def _paste_flag_stretched(base, team_name, box, radius=8):
    """خاص بلوحة المتأهلين: العلم يملأ المربع الأبيض بشكل عامودي مثل المرجع."""
    path = flag_path_for(team_name)
    x1, y1, x2, y2 = [int(v) for v in box]
    w, h = max(1, x2-x1), max(1, y2-y1)
    if path and os.path.exists(path):
        try:
            flag = Image.open(path).convert("RGBA").resize((w, h), Image.LANCZOS)
            mask = Image.new("L", (w, h), 0)
            md = ImageDraw.Draw(mask)
            try:
                md.rounded_rectangle((0, 0, w, h), radius=radius, fill=255)
            except Exception:
                md.rectangle((0, 0, w, h), fill=255)
            base.paste(flag, (x1, y1), mask)
            return
        except Exception:
            pass
    d = ImageDraw.Draw(base)
    rounded_rect(d, (x1, y1, x2, y2), radius=radius, fill="#F8FAFC", outline="#CBD5E1", width=1)
    draw_text(d, ((x1+x2)//2, (y1+y2)//2), (normalize_name(team_name) or "?")[:2], get_font(20), fill="#061633")


def _fit_font_for_box(draw, text, max_w, start_size=17, min_size=10):
    for size in range(start_size, min_size-1, -1):
        f = get_font(size)
        try:
            if text_width(draw, text, f) <= max_w:
                return f
        except Exception:
            pass
    return get_font(min_size)


def render_qualified32_board(teams):
    """لوحة المتأهلين الرسمية: تستخدم PNG ثابت مثل المرجع، وتبدأ من أعلى اليمين."""
    ensure_generated_dir()
    tpl = _qualified32_template_path()
    if tpl:
        img = Image.open(tpl).convert("RGB")
        if img.size != (1086, 1448):
            img = img.resize((1086, 1448), Image.LANCZOS)
    else:
        width, height = 1086, 1448
        img = Image.new("RGB", (width, height), "#F8FAFC")
        d = ImageDraw.Draw(img)
        draw_text(d, (width//2, 180), "كأس العالم 2026", get_font(62), fill="#061633")
        draw_text(d, (width//2, 238), "المنتخبات المتأهلة إلى دور الـ 32", get_font(34), fill="#061633")
        for r in range(4):
            for c in range(8):
                x = 31 + c*130
                y = 531 + r*198
                rounded_rect(d, (x, y, x+112, y+182), radius=13, fill="#FFFFFF", outline="#94A3B8", width=2)
                rounded_rect(d, (x, y+145, x+112, y+182), radius=10, fill="#061633")
    draw = ImageDraw.Draw(img)
    x_ranges = [(31,142), (159,270), (287,400), (417,531), (547,661), (679,792), (809,922), (939,1053)]
    footer_starts = [677, 875, 1072, 1269]
    white_tops = [531, 729, 926, 1123]
    footer_h = 37
    order = []
    for r in range(4):
        for x1, x2 in reversed(x_ranges):
            order.append((x1, x2, white_tops[r], footer_starts[r]))
    for idx, team in enumerate((teams or [])[:32]):
        x1, x2, yt, yf = order[idx]
        flag_box = (x1+5, yt+7, x2-5, yf-4)
        _paste_flag_stretched(img, team, flag_box, radius=8)
        name = normalize_name(team)
        font = _fit_font_for_box(draw, name, (x2-x1)-8, start_size=17, min_size=10)
        draw_text(draw, ((x1+x2)//2, yf + footer_h//2), name, font, fill="#FDE68A", max_width=(x2-x1)-8)
    out = os.path.join(GENERATED_DIR, "qualified32_board.png")
    img.save(out, quality=96)
    return out


def parse_command_body_lines(text):
    lines = (text or "").splitlines()
    if not lines:
        return []
    first = lines[0]
    rem = first.split(maxsplit=1)
    body = []
    if len(rem) > 1 and rem[1].strip():
        body.append(rem[1].strip())
    for line in lines[1:]:
        if normalize_name(line):
            body.append(normalize_name(line))
    return body


def parse_news_input(text):
    body = parse_command_body_lines(text)
    if not body:
        return None, ""
    if len(body) >= 2:
        maybe_team = canonical_team_name(body[0])
        if maybe_team:
            return maybe_team, "\n".join(body[1:]).strip()
    # إذا سطر واحد وكان فريق فقط، نرفض لأن الخبر ناقص
    if len(body) == 1 and canonical_team_name(body[0]):
        return canonical_team_name(body[0]), ""
    return None, "\n".join(body).strip()


def team_theme(team, mode="خبر"):
    if team and team in TEAM_THEME_MAP:
        return TEAM_THEME_MAP[team]
    if mode == "عاجل":
        return ("#7F1D1D", "#F59E0B", "#FFFFFF")
    if mode == "تأهل":
        return ("#14532D", "#FBBF24", "#FFFFFF")
    if mode == "إقصاء":
        return ("#1F2937", "#B91C1C", "#E5E7EB")
    return ("#0F172A", "#38BDF8", "#FFFFFF")


def news_header_title(kind):
    return {
        "خبر": "خبر",
        "عاجل": "عاجل",
        "تأهل": "رسميًا",
        "إقصاء": "إقصاء",
    }.get(kind, "خبر")


def _draw_glow_card(draw, box, radius, fill, outline, width=3):
    x1, y1, x2, y2 = box
    try:
        for i in range(3):
            pad = 8 + i*8
            rounded_rect(draw, (x1-pad, y1-pad, x2+pad, y2+pad), radius=radius+pad, fill=None, outline=outline, width=2)
    except Exception:
        pass
    rounded_rect(draw, box, radius=radius, fill=fill, outline=outline, width=width)


def _dynamic_news_font(draw, text, max_w, max_lines=4, start=58, minimum=32):
    text = normalize_name(text)
    for size in range(start, minimum-1, -2):
        f = get_font(size)
        lines = wrap_text(draw, text, f, max_w)
        if len(lines) <= max_lines:
            return f, lines
    f = get_font(minimum)
    return f, wrap_text(draw, text, f, max_w)


def render_news_card(kind, team, body):
    """تصميم خبر/عاجل فخم جدًا، مربع الخبر هو بطل التصميم."""
    ensure_generated_dir()
    width, height = 1200, 1350
    base, accent, txt2 = team_theme(team, kind)
    label = news_header_title(kind)
    if kind == "عاجل":
        badge = "#DC2626"; accent2 = "#F59E0B"
    elif kind == "تأهل":
        badge = "#16A34A"; accent2 = "#FBBF24"
    elif kind == "إقصاء":
        badge = "#991B1B"; accent2 = "#FB923C"
    else:
        badge = "#0EA5E9"; accent2 = "#FDE68A"
    try:
        rgb1 = tuple(int(base.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        rgb1 = (15, 23, 42)
    img = Image.new("RGB", (width, height), base)
    draw = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height - 1, 1)
        rgb2 = (2, 6, 18)
        rr = int(rgb1[0]*(1-t)+rgb2[0]*t)
        gg = int(rgb1[1]*(1-t)+rgb2[1]*t)
        bb = int(rgb1[2]*(1-t)+rgb2[2]*t)
        draw.line((0, y, width, y), fill=(rr, gg, bb))
    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)
    try:
        od.ellipse((width-500, 90, width+120, 710), fill=(255,255,255,30))
        od.ellipse((-260, 690, 380, 1320), fill=(255,255,255,18))
        od.rectangle((0, 0, width, 210), fill=(0,0,0,80))
        for i in range(7):
            od.arc((70+i*30, 55+i*22, width-70+i*30, 520+i*22), 185, 352, fill=(255,255,255,16), width=3)
        overlay = overlay.filter(ImageFilter.GaussianBlur(8))
    except Exception:
        pass
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)
    rounded_rect(draw, (44, 52, width-44, height-76), radius=48, fill="#020617B8", outline="#FFFFFF66", width=3)
    rounded_rect(draw, (62, 70, width-62, height-94), radius=42, fill="#02061766", outline=accent, width=4)
    rounded_rect(draw, (width//2-190, 105, width//2+190, 180), radius=24, fill=badge, outline=accent2, width=2)
    draw_text(draw, (width//2, 142), label, get_font(43), fill="#FFFFFF")
    if team:
        rounded_rect(draw, (width//2-185, 220, width//2+185, 388), radius=34, fill="#FFFFFFF2", outline="#FFFFFF88", width=2)
        paste_flag(img, team, (width//2-160, 238, width//2+160, 366))
        draw_text(draw, (width//2, 452), team, get_font(55), fill="#FFFFFF", max_width=980)
        card_top = 548
    else:
        draw_text(draw, (width//2, 275), "كأس العالم 2026", get_font(68), fill="#FFFFFF")
        draw_text(draw, (width//2, 340), label, get_font(38), fill=accent2)
        card_top = 500
    card = (90, card_top, width-90, height-205)
    _draw_glow_card(draw, card, 36, "#071633F2", accent, width=4)
    draw.line((card[0]+60, card[1]+58, card[2]-60, card[1]+58), fill="#FFFFFF70", width=2)
    rounded_rect(draw, (width//2-85, card[1]+42, width//2+85, card[1]+76), radius=15, fill=badge, outline="#FFFFFF33", width=1)
    max_w = card[2] - card[0] - 130
    f, lines = _dynamic_news_font(draw, body, max_w, max_lines=4, start=56, minimum=34)
    line_h = int(font_size(f, 40) * 1.42)
    total_h = line_h * len(lines)
    y0 = (card[1] + card[3])//2 - total_h//2 + 18
    for i, line in enumerate(lines[:5]):
        draw_text(draw, (width//2, y0 + i*line_h), line, f, fill="#FFFFFF", max_width=max_w)
    draw.line((240, height-146, width-240, height-146), fill="#FFFFFF66", width=2)
    draw_text(draw, (width//2, height-108), "المصيف يضعكم بالحدث", get_font(32), fill="#FDE68A")
    out = os.path.join(GENERATED_DIR, f"news_{_safe_filename(kind)}_{_safe_filename(team or 'worldcup')}.png")
    img.save(out, quality=96)
    return out


async def send_photo_path_markup(message, path, caption=None, reply_markup=None):
    with open(path, "rb") as f:
        await message.reply_photo(photo=f, caption=caption or "", reply_markup=reply_markup)


def load_sports_cache():
    if os.path.exists(SPORTS_CACHE_FILE):
        try:
            with open(SPORTS_CACHE_FILE, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if isinstance(obj, dict):
                return obj
        except Exception:
            pass
    return {}


def save_sports_cache(data):
    with open(SPORTS_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def store_source_request(context, payload):
    store = context.bot_data.setdefault("sports_source_requests", {})
    key = datetime.now().strftime("%Y%m%d%H%M%S%f")[-12:]
    store[key] = payload
    # تنظيف خفيف
    if len(store) > 80:
        for k in list(store.keys())[:-60]:
            store.pop(k, None)
    return key


def source_keyboard(context, payload):
    token = store_source_request(context, payload)
    rows = [[
        InlineKeyboardButton("رسمي", callback_data=f"sportsrc|{token}|official"),
        InlineKeyboardButton("سريع", callback_data=f"sportsrc|{token}|fast"),
        InlineKeyboardButton("الأحدث", callback_data=f"sportsrc|{token}|latest"),
    ]]
    return InlineKeyboardMarkup(rows)


class _SimpleHTTPResponse:
    def __init__(self, text, status_code=200, headers=None):
        self.text = text
        self.status_code = status_code
        self.headers = headers or {}
    def json(self):
        return json.loads(self.text)
    def raise_for_status(self):
        if int(self.status_code or 0) >= 400:
            raise RuntimeError(f"HTTP Error {self.status_code}")


def _requests_get(url, timeout=16):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
    }
    if requests is not None:
        return requests.get(url, timeout=timeout, headers=headers)
    try:
        import urllib.request
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            enc = resp.headers.get_content_charset() or "utf-8"
            txt = raw.decode(enc, errors="replace")
            return _SimpleHTTPResponse(txt, getattr(resp, "status", 200), dict(resp.headers))
    except Exception:
        raise RuntimeError("تعذر الاتصال بالمصدر")

def _extract_json_from_next_data(html):
    m = re.search(r'<script[^>]*id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None


def _iter_dict_nodes(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from _iter_dict_nodes(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from _iter_dict_nodes(item)


def _normalize_status_text(text):
    t = normalize_name(text)
    if not t:
        return ""
    if any(x in t.lower() for x in ["in progress", "live"]):
        return "مباشر"
    if any(x in t.lower() for x in ["halftime", "half-time"]):
        return "بين الشوطين"
    if any(x in t.lower() for x in ["final", "full time", "ft"]):
        return "انتهت المباراة"
    if any(x in t.lower() for x in ["scheduled", "not started"]):
        return "لم تبدأ"
    return t


def _parse_espn_match_from_event(event):
    try:
        comp = (event.get("competitions") or [event])[0]
        competitors = comp.get("competitors") or []
        if len(competitors) < 2:
            return None
        a = competitors[0]
        b = competitors[1]
        team1 = canonical_team_name(a.get("team", {}).get("displayName")) or normalize_name(a.get("team", {}).get("displayName"))
        team2 = canonical_team_name(b.get("team", {}).get("displayName")) or normalize_name(b.get("team", {}).get("displayName"))
        score1 = str(a.get("score", 0))
        score2 = str(b.get("score", 0))
        status = comp.get("status") or event.get("status") or {}
        detail = normalize_name(((status.get("type") or {}).get("detail")) or status.get("displayClock") or "")
        short_detail = normalize_name(((status.get("type") or {}).get("shortDetail")) or "")
        state = normalize_name(((status.get("type") or {}).get("name")) or ((status.get("type") or {}).get("state")) or "")
        status_ar = _normalize_status_text(detail or short_detail or state)
        scorers = []
        for d in comp.get("details", []) or []:
            if isinstance(d, dict):
                txt = normalize_name(d.get("text") or d.get("detail") or "")
                if txt:
                    scorers.append(txt)
        minute = detail or short_detail
        return {
            "team1": team1,
            "team2": team2,
            "score1": score1,
            "score2": score2,
            "status": status_ar,
            "minute": minute,
            "scorers": scorers[:8],
            "source": "ESPN",
        }
    except Exception:
        return None


def _teams_match(event_team1, event_team2, req1, req2):
    c1 = canonical_team_name(event_team1) or normalize_name(event_team1)
    c2 = canonical_team_name(event_team2) or normalize_name(event_team2)
    r1 = canonical_team_name(req1) or normalize_name(req1)
    r2 = canonical_team_name(req2) or normalize_name(req2)
    return (c1 == r1 and c2 == r2) or (c1 == r2 and c2 == r1)


def _espn_date_candidates(extra_date=None, days_back=3, days_forward=2):
    dates = []
    if extra_date:
        dates.append(extra_date)
    today = datetime.utcnow().date()
    for off in range(-days_back, days_forward + 1):
        dates.append((today + timedelta(days=off)).strftime("%Y%m%d"))
    for d in ["20260611", "20260612", "20260613", "20260614", "20260615", "20260616", "20260617", "20260618", "20260619", "20260620"]:
        dates.append(d)
    seen, out = set(), []
    for d in dates:
        if d and d not in seen:
            seen.add(d); out.append(d)
    return out


def _espn_scoreboard_urls(date_value):
    return [
        f"https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard?dates={date_value}&limit=300",
        f"https://site.api.espn.com/apis/site/v2/sports/soccer/all/scoreboard?dates={date_value}&limit=300",
    ]


def _fetch_espn_events_by_date(date_value):
    for url in _espn_scoreboard_urls(date_value):
        try:
            r = _requests_get(url, timeout=14)
            if int(getattr(r, "status_code", 200) or 200) >= 400:
                continue
            data = r.json()
            events = data.get("events", []) if isinstance(data, dict) else []
            if events is not None:
                return events
        except Exception:
            continue
    return []


def _parse_live_date_from_text(text):
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text or "")
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    return f"{y:04d}{mo:02d}{d:02d}"


def fetch_match_from_espn(team1, team2, date_hint=None):
    """مصدر سريع: يبحث مباشر + اليوم + آخر 24/48 ساعة + أيام قريبة، ويجلب المنتهية أيضًا."""
    best = None
    for d in _espn_date_candidates(date_hint, days_back=3, days_forward=2):
        for event in _fetch_espn_events_by_date(d):
            obj = _parse_espn_match_from_event(event)
            if obj and _teams_match(obj["team1"], obj["team2"], team1, team2):
                status = (obj.get("status") or "") + " " + (obj.get("minute") or "")
                if "مباشر" in status or "LIVE" in status.upper():
                    return obj
                best = obj
    return best


def _parse_fifa_match_candidates(obj, req1, req2):
    for node in _iter_dict_nodes(obj):
        keys = {k.lower() for k in node.keys()}
        if not keys:
            continue
        # استخراج أسماء الفرق من أي شكل معروف
        team_names = []
        for key in ["homeTeam", "awayTeam", "home_team", "away_team", "team1", "team2"]:
            if key in node:
                val = node.get(key)
                if isinstance(val, dict):
                    team_names.append(val.get("name") or val.get("shortName") or val.get("displayName"))
                else:
                    team_names.append(val)
        if len(team_names) < 2 and "participants" in node and isinstance(node.get("participants"), list):
            parts = node.get("participants")
            for p in parts[:2]:
                if isinstance(p, dict):
                    team_names.append(p.get("name") or p.get("shortName") or ((p.get("team") or {}).get("name") if isinstance(p.get("team"), dict) else None))
        if len(team_names) < 2 and "teams" in node and isinstance(node.get("teams"), list):
            for p in node.get("teams")[:2]:
                if isinstance(p, dict):
                    team_names.append(p.get("name") or p.get("shortName") or ((p.get("team") or {}).get("name") if isinstance(p.get("team"), dict) else None))
        if len(team_names) >= 2 and _teams_match(team_names[0], team_names[1], req1, req2):
            s1 = node.get("homeScore") or node.get("score1") or node.get("home_score")
            s2 = node.get("awayScore") or node.get("score2") or node.get("away_score")
            if isinstance(s1, dict):
                s1 = s1.get("score") or s1.get("value") or s1.get("goals")
            if isinstance(s2, dict):
                s2 = s2.get("score") or s2.get("value") or s2.get("goals")
            status = node.get("status") or node.get("matchStatus") or node.get("state") or node.get("stage") or ""
            minute = node.get("minute") or node.get("clock") or node.get("time") or ""
            scorers = []
            for k in ["details", "events", "goals"]:
                if k in node and isinstance(node.get(k), list):
                    for item in node.get(k):
                        if isinstance(item, dict):
                            txt = item.get("text") or item.get("description") or item.get("detail") or item.get("name")
                            if txt:
                                scorers.append(normalize_name(txt))
            return {
                "team1": canonical_team_name(team_names[0]) or normalize_name(team_names[0]),
                "team2": canonical_team_name(team_names[1]) or normalize_name(team_names[1]),
                "score1": str(s1 if s1 is not None else 0),
                "score2": str(s2 if s2 is not None else 0),
                "status": _normalize_status_text(status),
                "minute": normalize_name(minute),
                "scorers": scorers[:8],
                "source": "FIFA",
            }
    return None


def fetch_match_from_fifa(team1, team2):
    url = "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/scores-fixtures"
    r = _requests_get(url)
    txt = r.text
    data = _extract_json_from_next_data(txt)
    if data:
        obj = _parse_fifa_match_candidates(data, team1, team2)
        if obj:
            return obj
    # fallback: raw regex is intentionally light; if no reliable data نجد nothing.
    return None


def fetch_live_match_data(team1, team2, mode="official", date_hint=None):
    mode = normalize_name(mode or "official").lower()
    if mode in ["رسمي", "official"]:
        sources = [lambda a,b: fetch_match_from_fifa(a, b), lambda a,b: fetch_match_from_espn(a, b, date_hint)]
    else:
        sources = [lambda a,b: fetch_match_from_espn(a, b, date_hint), lambda a,b: fetch_match_from_fifa(a, b)]
    for fn in sources:
        try:
            obj = fn(team1, team2)
            if obj:
                if not obj.get("status"):
                    obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
                return obj
        except Exception:
            continue
    return None


def _parse_espn_standings_json(data):
    groups = []
    entries = data.get("children") or data.get("groups") or data.get("standings") or []
    for grp in entries:
        title = normalize_name(grp.get("name") or grp.get("abbreviation") or grp.get("title") or "")
        title = title if title.startswith("المجموعة") else (f"المجموعة {title}" if title else "")
        rows = []
        for item in grp.get("standings", []) or grp.get("entries", []) or []:
            team = None
            if isinstance(item.get("team"), dict):
                team = item.get("team", {}).get("displayName") or item.get("team", {}).get("name")
            elif isinstance(item.get("team"), str):
                team = item.get("team")
            elif isinstance(item.get("stats"), dict):
                team = item.get("name")
            stats = item.get("stats", []) or []
            stat_map = {}
            if isinstance(stats, list):
                for s in stats:
                    if isinstance(s, dict):
                        key = normalize_name(s.get("name") or s.get("abbreviation") or "").lower()
                        stat_map[key] = s.get("value") if s.get("value") is not None else s.get("displayValue")
            elif isinstance(stats, dict):
                stat_map = {normalize_name(k).lower(): v for k, v in stats.items()}
            played = stat_map.get("played") or stat_map.get("p") or stat_map.get("gamesplayed") or 0
            gd = stat_map.get("pointdifferential") or stat_map.get("gd") or stat_map.get("goaldifference") or 0
            pts = stat_map.get("points") or stat_map.get("pts") or 0
            team = canonical_team_name(team) or normalize_name(team)
            if team:
                try:
                    played = int(played)
                except Exception:
                    pass
                try:
                    gd = int(gd)
                except Exception:
                    pass
                try:
                    pts = int(pts)
                except Exception:
                    pass
                rows.append((team, played, gd, pts))
        if title and rows:
            rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
            groups.append((title, rows))
    return groups


def fetch_standings_from_espn():
    """نحسب ترتيب المجموعات من نتائج ESPN، وهذا أكثر ثباتًا من endpoint standings وقت البطولة."""
    table = {t: {"played": 0, "gd": 0, "pts": 0, "gf": 0, "ga": 0} for t in WORLD_CUP_TEAMS}
    today = datetime.utcnow().date()
    start = datetime(2026, 6, 11).date()
    end = max(today + timedelta(days=2), start)
    dates = []
    d = start
    while d <= end:
        dates.append(d.strftime("%Y%m%d"))
        d += timedelta(days=1)
    found_any = False
    seen_events = set()
    for dv in dates:
        for event in _fetch_espn_events_by_date(dv):
            eid = str(event.get("id") or event.get("uid") or f"{dv}-{len(seen_events)}")
            if eid in seen_events:
                continue
            seen_events.add(eid)
            obj = _parse_espn_match_from_event(event)
            if not obj:
                continue
            t1 = canonical_team_name(obj.get("team1")) or obj.get("team1")
            t2 = canonical_team_name(obj.get("team2")) or obj.get("team2")
            if t1 not in table or t2 not in table:
                continue
            status = normalize_name(obj.get("status") or obj.get("minute") or "").lower()
            try:
                s1 = int(float(obj.get("score1", 0)))
                s2 = int(float(obj.get("score2", 0)))
            except Exception:
                continue
            if any(x in status for x in ["scheduled", "لم تبدأ", "not started"]):
                continue
            found_any = True
            table[t1]["played"] += 1; table[t2]["played"] += 1
            table[t1]["gf"] += s1; table[t1]["ga"] += s2; table[t1]["gd"] += s1 - s2
            table[t2]["gf"] += s2; table[t2]["ga"] += s1; table[t2]["gd"] += s2 - s1
            if s1 > s2:
                table[t1]["pts"] += 3
            elif s2 > s1:
                table[t2]["pts"] += 3
            else:
                table[t1]["pts"] += 1; table[t2]["pts"] += 1
    if not found_any:
        return []
    groups = []
    for g, teams in WORLD_CUP_GROUPS:
        rows = []
        for t in teams:
            r = table[t]
            rows.append((t, r["played"], r["gd"], r["pts"]))
        rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
        groups.append((f"المجموعة {g}", rows))
    return groups


def _parse_fifa_standings_candidates(obj):
    group_map = defaultdict(list)
    for node in _iter_dict_nodes(obj):
        team = None
        group = None
        played = gd = pts = None
        # team
        if isinstance(node.get("team"), dict):
            team = node.get("team", {}).get("name") or node.get("team", {}).get("shortName")
        team = team or node.get("teamName") or node.get("name")
        # stats
        group = node.get("group") or node.get("groupName") or node.get("pool") or node.get("letter")
        played = node.get("played") or node.get("gamesPlayed") or node.get("matches")
        gd = node.get("goalDifference") or node.get("gd") or node.get("goal_diff")
        pts = node.get("points") or node.get("pts")
        if team and group is not None and pts is not None:
            team = canonical_team_name(team) or normalize_name(team)
            g = normalize_name(group)
            if not g.startswith("المجموعة"):
                g = f"المجموعة {g}"
            try:
                played = int(played)
            except Exception:
                played = 0
            try:
                gd = int(gd)
            except Exception:
                gd = 0
            try:
                pts = int(pts)
            except Exception:
                pts = 0
            row = (team, played, gd, pts)
            if row not in group_map[g]:
                group_map[g].append(row)
    groups = []
    for g, rows in group_map.items():
        rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
        groups.append((g, rows))
    groups.sort(key=lambda x: x[0])
    return groups


def fetch_standings_from_fifa():
    url = "https://www.fifa.com/en/tournaments/mens/worldcup/canadamexicousa2026/standings"
    r = _requests_get(url)
    txt = r.text
    data = _extract_json_from_next_data(txt)
    if data:
        groups = _parse_fifa_standings_candidates(data)
        if groups:
            return groups
    return []


def fetch_current_groups(mode="official"):
    mode = normalize_name(mode or "official").lower()
    if mode in ["رسمي", "official"]:
        order = [(fetch_standings_from_fifa, "FIFA"), (fetch_standings_from_espn, "ESPN")]
    elif mode in ["سريع", "fast"]:
        order = [(fetch_standings_from_espn, "ESPN"), (fetch_standings_from_fifa, "FIFA")]
    else:
        order = [(fetch_standings_from_espn, "ESPN"), (fetch_standings_from_fifa, "FIFA")]
    errors = []
    for fn, label in order:
        try:
            groups = fn()
            if groups:
                return groups, label
        except Exception as e:
            errors.append(str(e))
    if errors:
        raise RuntimeError(" | ".join(errors[:2]))
    return [], ""


def build_groups_text(groups, source_label=None):
    lines = []
    for title, rows in groups:
        lines.append(title)
        for i, (team, played, gd, pts) in enumerate(rows, start=1):
            lines.append(f"{i}- {team} — لعب {played} | فارق {gd} | {pts} نقطة")
        lines.append("")
    if source_label:
        lines.append(f"المصدر: {source_label}")
    return "\n".join(lines).strip()


def render_live_match_card(match, mode_label="رسمي"):
    ensure_generated_dir()
    width, height = 1200, 1350
    img = _style4_clean_background(width, height)
    draw = ImageDraw.Draw(img)
    draw_text(draw, (width//2, 100), "مباشر الآن", get_font(62), fill="#FFFFFF")
    draw_text(draw, (width//2, 165), f"المصدر: {mode_label}", get_font(28), fill="#FDE68A")

    rounded_rect(draw, (80, 235, width-80, 820), radius=40, fill="#07132FDD", outline="#38BDF855", width=2)
    paste_flag(img, match["team1"], (160, 320, 310, 420))
    paste_flag(img, match["team2"], (width-310, 320, width-160, 420))
    draw_text(draw, (245, 470), match["team1"], get_font(40), fill="#FFFFFF", max_width=280)
    draw_text(draw, (width-245, 470), match["team2"], get_font(40), fill="#FFFFFF", max_width=280)
    draw_text(draw, (width//2, 430), f"{match['score1']} - {match['score2']}", get_font(96), fill="#FFFFFF")
    rounded_rect(draw, (width//2-170, 520, width//2+170, 590), radius=22, fill="#FBBF24", outline="#FFFFFF33", width=1)
    status_line = match.get("minute") or match.get("status") or ""
    if match.get("status") and match.get("minute") and match.get("status") not in status_line:
        status_line = f"{match['status']} — {match['minute']}"
    draw_text(draw, (width//2, 555), status_line, get_font(30), fill="#061633")

    y = 640
    scorers = match.get("scorers") or []
    if scorers:
        draw_text(draw, (width//2, y), "الأحداث / الهدافون", get_font(34), fill="#FDE68A")
        y += 60
        for item in scorers[:6]:
            rounded_rect(draw, (120, y-18, width-120, y+36), radius=18, fill="#0B1E46", outline="#FFFFFF22", width=1)
            draw_text(draw, (width//2, y+10), normalize_name(item), get_font(28), fill="#FFFFFF", max_width=880)
            y += 64
    else:
        draw_text(draw, (width//2, 690), match.get("status") or "متابعة مباشرة", get_font(34), fill="#FFFFFF")

    footer_event(draw, width, height)
    out = os.path.join(GENERATED_DIR, f"live_{_safe_filename(match['team1'])}_{_safe_filename(match['team2'])}.png")
    img.save(out, quality=95)
    return out


def parse_live_command_text(text):
    body = parse_command_body_lines(text)
    raw = " ".join(body)
    parts = [normalize_name(x) for x in raw.split("*") if normalize_name(x)]
    mode = "official"
    date_hint = _parse_live_date_from_text(raw)
    if len(parts) >= 3:
        mode_part = normalize_name(parts[-1])
        if mode_part in ["رسمي", "سريع", "الأحدث", "official", "fast", "latest"]:
            mode = mode_part
            parts = parts[:-1]
        elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", mode_part):
            date_hint = _parse_live_date_from_text(mode_part)
            parts = parts[:-1]
    if len(parts) < 2:
        return None, None, mode, date_hint
    return canonical_team_name(parts[0]) or normalize_name(parts[0]), canonical_team_name(parts[1]) or normalize_name(parts[1]), mode, date_hint


def mode_label_ar(mode):
    mode = normalize_name(mode or "official").lower()
    if mode in ["fast", "سريع"]:
        return "سريع"
    if mode in ["latest", "الأحدث"]:
        return "الأحدث"
    return "رسمي"


def build_live_caption(match, mode_label="رسمي"):
    lines = [f"{match['team1']} {match['score1']} - {match['score2']} {match['team2']}"]
    status_line = match.get("minute") or match.get("status") or ""
    if match.get("status") and match.get("minute") and match.get("status") not in status_line:
        status_line = f"{match['status']} — {match['minute']}"
    if status_line:
        lines.append(status_line)
    if match.get("scorers"):
        lines.append("")
        lines.append("الهدافون / الأحداث:")
        for s in match.get("scorers", [])[:6]:
            lines.append(f"- {normalize_name(s)}")
    lines.append("")
    lines.append(f"المصدر: {mode_label}")
    return "\n".join(lines)


async def teams_supported_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(grouped_teams_text(with_points=False))


async def groups_points_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(grouped_teams_text(with_points=True))


async def groups_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(groups_template_text())


async def qualified_show_board_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    teams = load_qualified32_state()
    path = render_qualified32_board(teams)
    await send_photo_path(update, path, f"لوحة المتأهلين الحالية ✅\nعدد المنتخبات المضافة: {len(teams)}/32")


async def qualified_reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    save_qualified32_state([])
    path = render_qualified32_board([])
    await send_photo_path(update, path, "تمت إعادة لوحة المتأهلين إلى وضعها الفارغ ✅")


async def qualified_add_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    if not body:
        await update.message.reply_text("اكتبها كذا:\n/متأهل المكسيك")
        return
    team_raw = body[0]
    team = canonical_team_name(team_raw)
    if not team:
        await update.message.reply_text(unsupported_team_message(team_raw))
        return
    teams = load_qualified32_state()
    if team in teams:
        await update.message.reply_text(f"{team} موجود مسبقًا في اللوحة ✅")
        return
    if len(teams) >= 32:
        await update.message.reply_text("تم اكتمال قائمة المنتخبات المتأهلة إلى دور الـ32 ✅")
        return
    teams.append(team)
    save_qualified32_state(teams)
    path = render_qualified32_board(teams)
    await send_photo_path(update, path, f"تمت إضافة {team} ✅\nالمجموع الحالي: {len(teams)}/32")


async def qualified_add_many_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    if not body:
        await update.message.reply_text("اكتبها كذا:\n/متأهلين\nالمكسيك\nاليابان\nفرنسا")
        return
    if len(body) == 1 and ("," in body[0] or "،" in body[0]):
        names = [normalize_name(x) for x in re.split(r"[,،]", body[0]) if normalize_name(x)]
    else:
        names = body
    teams = load_qualified32_state()
    added, invalid, skipped = [], [], []
    for raw in names:
        team = canonical_team_name(raw)
        if not team:
            invalid.append(raw)
            continue
        if team in teams or team in added:
            skipped.append(team)
            continue
        if len(teams) + len(added) >= 32:
            break
        added.append(team)
    teams.extend(added)
    save_qualified32_state(teams)
    path = render_qualified32_board(teams)
    cap = [f"تم تحديث لوحة المتأهلين ✅\nالمجموع الحالي: {len(teams)}/32"]
    if added:
        cap.append("\nتمت الإضافة:\n- " + "\n- ".join(added))
    if skipped:
        cap.append("\nموجود مسبقًا:\n- " + "\n- ".join(skipped[:8]))
    if invalid:
        cap.append("\nتعذر التعرف على:\n- " + "\n- ".join(invalid[:8]))
    await send_photo_path(update, path, "\n".join(cap))


async def qualified_remove_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    if not body:
        await update.message.reply_text("اكتبها كذا:\n/حذف_متأهل المكسيك")
        return
    team_raw = body[0]
    team = canonical_team_name(team_raw)
    if not team:
        await update.message.reply_text(unsupported_team_message(team_raw))
        return
    teams = load_qualified32_state()
    if team not in teams:
        await update.message.reply_text(f"{team} غير موجود في لوحة المتأهلين.")
        return
    teams = [x for x in teams if x != team]
    save_qualified32_state(teams)
    path = render_qualified32_board(teams)
    await send_photo_path(update, path, f"تم حذف {team} من اللوحة ✅\nالمجموع الحالي: {len(teams)}/32")


async def news_generic_command(update: Update, context: ContextTypes.DEFAULT_TYPE, kind="خبر"):
    team, body = parse_news_input(update.message.text)
    if not body:
        sample = f"/{kind}\nالسعودية\nاستعادت صدارة المجموعة"
        if kind in ["خبر", "عاجل"]:
            sample += f"\n\nأو بدون منتخب:\n/{kind}\nتأجيل مباراة اليوم 30 دقيقة"
        await update.message.reply_text(f"اكتبها كذا:\n{sample}")
        return
    try:
        path = render_news_card(kind, team, body)
    except Exception as e:
        await update.message.reply_text(f"تعذر إنشاء تصميم {kind} ❌\nالسبب: {str(e)[:160]}")
        return
    caption = f"{kind} ✅"
    if team:
        caption += f"\n{team}"
    await send_photo_path(update, path, caption)


async def news_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await news_generic_command(update, context, "خبر")


async def urgent_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await news_generic_command(update, context, "عاجل")


async def qualified_news_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await news_generic_command(update, context, "تأهل")


async def eliminated_news_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await news_generic_command(update, context, "إقصاء")


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|official|fast|latest)\s*$", text or "", re.I)
        mode = m.group(1) if m else "official"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    try:
        groups, source_label = fetch_current_groups(mode)
        if not groups:
            await update.message.reply_text("تعذر جلب ترتيب المجموعات من المصدر الحالي.\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        caption = f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)}"
        await send_photo_path_markup(update.message, path, caption, kb)
        await update.message.reply_text(build_groups_text(groups, mode_label_ar(mode)))
    except Exception:
        await update.message.reply_text(
            f"تعذر جلب ترتيب المجموعات من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:",
            reply_markup=kb,
        )

async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = mode_override
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر السعودية * اسبانيا\nأو\n/مباشر السعودية * اسبانيا * سريع\nأو بتاريخ محدد:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    try:
        data = fetch_live_match_data(team1, team2, mode, date_hint=date_hint)
        if not data:
            await update.message.reply_text(
                f"ما لقيت مباراة {team1} ضد {team2} ضمن المباشر أو مباريات آخر 48 ساعة.\n"
                f"جرب تختار مصدر آخر أو اكتب التاريخ.\n\nاختر مصدر آخر:",
                reply_markup=kb,
            )
            return
        path = render_live_match_card(data, mode_label_ar(mode))
        await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
    except Exception:
        await update.message.reply_text(
            f"تعذر جلب المباراة من المصدر الحالي ❌\nمباراة: {team1} × {team2}\nالمصدر الحالي: {mode_label_ar(mode)}\n\nاختر مصدر آخر:",
            reply_markup=kb,
        )


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    try:
        if kind == "standings":
            groups, _src = fetch_current_groups(mode)
            kb = source_keyboard(context, payload)
            if not groups:
                await query.message.reply_text(f"تعذر جلب ترتيب المجموعات من مصدر {mode_label_ar(mode)}.\n\nاختر مصدر آخر:", reply_markup=kb)
                return
            path = create_all_groups_image(groups)
            await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)}", kb)
            await query.message.reply_text(build_groups_text(groups, mode_label_ar(mode)))
        elif kind == "live":
            team1 = payload.get("team1")
            team2 = payload.get("team2")
            kb = source_keyboard(context, payload)
            data = fetch_live_match_data(team1, team2, mode, date_hint=payload.get("date_hint"))
            if not data:
                await query.message.reply_text(f"ما لقيت مباراة {team1} ضد {team2} من مصدر {mode_label_ar(mode)}.\n\nاختر مصدر آخر:", reply_markup=kb)
                return
            path = render_live_match_card(data, mode_label_ar(mode))
            await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        else:
            await query.message.reply_text("تعذر تحديد نوع الطلب.")
    except Exception:
        if kind == "live":
            await query.message.reply_text(f"تعذر جلب البيانات من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=source_keyboard(context, payload))
        elif kind == "standings":
            await query.message.reply_text(f"تعذر جلب ترتيب المجموعات من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=source_keyboard(context, payload))
        else:
            await query.message.reply_text("تعذر جلب البيانات الآن.")


# -------------------- V3 إصلاح نهائي: SerpApi + API-Football + أخبار أفخم + لوحة 32 --------------------

def _env(name, default=""):
    try:
        return (os.getenv(name) or default or "").strip()
    except Exception:
        return default or ""

TEAM_SEARCH_EN = {
    "المكسيك": ["Mexico"],
    "جنوب أفريقيا": ["South Africa"],
    "كوريا الجنوبية": ["South Korea", "Korea Republic"],
    "التشيك": ["Czechia", "Czech Republic"],
    "كندا": ["Canada"],
    "البوسنة والهرسك": ["Bosnia and Herzegovina", "Bosnia-Herzegovina"],
    "قطر": ["Qatar"],
    "سويسرا": ["Switzerland"],
    "البرازيل": ["Brazil"],
    "المغرب": ["Morocco"],
    "هايتي": ["Haiti"],
    "اسكتلندا": ["Scotland"],
    "الولايات المتحدة": ["United States", "USA"],
    "باراغواي": ["Paraguay"],
    "أستراليا": ["Australia"],
    "تركيا": ["Turkey", "Türkiye"],
    "ألمانيا": ["Germany"],
    "كوراساو": ["Curacao", "Curaçao"],
    "ساحل العاج": ["Ivory Coast", "Cote d'Ivoire"],
    "الإكوادور": ["Ecuador"],
    "هولندا": ["Netherlands", "Holland"],
    "اليابان": ["Japan"],
    "السويد": ["Sweden"],
    "تونس": ["Tunisia"],
    "بلجيكا": ["Belgium"],
    "مصر": ["Egypt"],
    "إيران": ["Iran", "IR Iran"],
    "نيوزيلندا": ["New Zealand"],
    "إسبانيا": ["Spain"],
    "الرأس الأخضر": ["Cape Verde", "Cabo Verde"],
    "السعودية": ["Saudi Arabia", "KSA"],
    "أوروجواي": ["Uruguay"],
    "فرنسا": ["France"],
    "السنغال": ["Senegal"],
    "العراق": ["Iraq"],
    "النرويج": ["Norway"],
    "الأرجنتين": ["Argentina"],
    "الجزائر": ["Algeria"],
    "النمسا": ["Austria"],
    "الأردن": ["Jordan"],
    "البرتغال": ["Portugal"],
    "الكونغو الديمقراطية": ["DR Congo", "Congo DR", "Democratic Republic of the Congo"],
    "أوزبكستان": ["Uzbekistan"],
    "كولومبيا": ["Colombia"],
    "إنجلترا": ["England"],
    "كرواتيا": ["Croatia"],
    "غانا": ["Ghana"],
    "بنما": ["Panama"],
}

# aliases extra for API names
try:
    for _ar, _ens in TEAM_SEARCH_EN.items():
        for _en in _ens:
            TEAM_ALIASES[_en] = _ar
except Exception:
    pass


def team_query_name(team):
    team = canonical_team_name(team) or normalize_name(team)
    arr = TEAM_SEARCH_EN.get(team) or [team]
    return arr[0]


def team_query_names(team):
    team = canonical_team_name(team) or normalize_name(team)
    return [team] + (TEAM_SEARCH_EN.get(team) or [])


def _as_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        if isinstance(v, str):
            v = re.sub(r"[^0-9\-]", "", v)
            if v in ("", "-"):
                return default
        return int(float(v))
    except Exception:
        return default


def _norm_status_ar(status, short=None, elapsed=None):
    raw = normalize_name(status or short or "")
    low = raw.lower()
    if short in ("FT", "AET", "PEN") or any(x in low for x in ["final", "full", "انته", "نهاية"]):
        return "انتهت المباراة"
    if short in ("NS", "TBD") or any(x in low for x in ["not started", "scheduled", "لم تبدأ"]):
        return "لم تبدأ"
    if short in ("1H", "2H", "LIVE", "HT", "ET", "BT") or any(x in low for x in ["live", "half", "مباشر"]):
        if elapsed:
            return f"مباشر — {elapsed}'"
        return "مباشر الآن"
    return raw or (f"{elapsed}'" if elapsed else "غير محدد")


def _safe_get(d, *keys, default=None):
    cur = d
    for k in keys:
        if isinstance(cur, dict):
            cur = cur.get(k)
        else:
            return default
    return default if cur is None else cur


def _crop_alpha_content(img):
    try:
        if img.mode != "RGBA":
            img = img.convert("RGBA")
        alpha = img.getchannel("A")
        bbox = alpha.getbbox()
        if bbox:
            return img.crop(bbox)
    except Exception:
        pass
    return img.convert("RGBA")


def _rounded_mask(w, h, radius=8):
    mask = Image.new("L", (max(1, w), max(1, h)), 0)
    md = ImageDraw.Draw(mask)
    try:
        md.rounded_rectangle((0, 0, w-1, h-1), radius=radius, fill=255)
    except Exception:
        md.rectangle((0, 0, w, h), fill=255)
    return mask


def _paste_flag_stretched(base, team_name, box, radius=8):
    """لوحة المتأهلين فقط: علم عامودي يملأ الجزء الأبيض، بدون خلفية سوداء أو شفافية."""
    path = flag_path_for(team_name)
    x1, y1, x2, y2 = [int(v) for v in box]
    w, h = max(1, x2-x1), max(1, y2-y1)
    d = ImageDraw.Draw(base)
    # خلفية بيضاء دائماً حتى لا تظهر مربعات سوداء من PNG الشفاف
    try:
        rounded_rect(d, (x1, y1, x2, y2), radius=radius, fill="#FFFFFF", outline="#E5E7EB", width=1)
    except Exception:
        d.rectangle((x1, y1, x2, y2), fill="#FFFFFF")
    if path and os.path.exists(path):
        try:
            flag = Image.open(path).convert("RGBA")
            flag = _crop_alpha_content(flag)
            # المطلوب هنا عامودي: نمدد العلم على كامل المربع الأبيض، مثل مثال المكسيك.
            flag = flag.resize((w, h), Image.LANCZOS)
            canvas = Image.new("RGBA", (w, h), (255, 255, 255, 255))
            canvas.paste(flag, (0, 0), flag if flag.mode == "RGBA" else None)
            mask = _rounded_mask(w, h, radius)
            base.paste(canvas.convert("RGB"), (x1, y1), mask)
            return
        except Exception:
            pass
    draw_text(d, ((x1+x2)//2, (y1+y2)//2), (normalize_name(team_name) or "?")[:2], get_font(22), fill="#061633")


def render_qualified32_board(teams):
    """لوحة المتأهلين الرسمية: PNG ثابت + بداية من أعلى اليمين + علم عامودي."""
    ensure_generated_dir()
    tpl = _qualified32_template_path()
    if tpl:
        img = Image.open(tpl).convert("RGB")
        # القالب المعتمد 1086×1448، ولو اختلف نحافظ على نفس النسبة.
        if img.size != (1086, 1448):
            img = img.resize((1086, 1448), Image.LANCZOS)
    else:
        width, height = 1086, 1448
        img = Image.new("RGB", (width, height), "#F8FAFC")
        d = ImageDraw.Draw(img)
        draw_text(d, (width//2, 180), "كأس العالم 2026", get_font(62), fill="#061633")
        draw_text(d, (width//2, 238), "المنتخبات المتأهلة إلى دور الـ 32", get_font(34), fill="#061633")
        for r in range(4):
            for c in range(8):
                x = 31 + c*130
                y = 531 + r*198
                rounded_rect(d, (x, y, x+112, y+182), radius=13, fill="#FFFFFF", outline="#94A3B8", width=2)
                rounded_rect(d, (x, y+145, x+112, y+182), radius=10, fill="#061633")
    draw = ImageDraw.Draw(img)
    # إحداثيات القالب المعتمد، مرتبة بصريًا من اليسار، وسنقلبها للبدء من اليمين.
    x_ranges = [(31,142), (159,270), (287,400), (417,531), (547,661), (679,792), (809,922), (939,1053)]
    white_tops = [531, 729, 926, 1123]
    footer_starts = [677, 875, 1072, 1269]
    footer_h = 37
    order = []
    for r in range(4):
        for x1, x2 in reversed(x_ranges):
            order.append((x1, x2, white_tops[r], footer_starts[r]))
    for idx, team in enumerate((teams or [])[:32]):
        x1, x2, yt, yf = order[idx]
        # الجزء الأبيض فقط، ويترك الشريط الأزرق للاسم.
        flag_box = (x1+3, yt+5, x2-3, yf-3)
        _paste_flag_stretched(img, team, flag_box, radius=8)
        name = normalize_name(team)
        font = _fit_font_for_box(draw, name, (x2-x1)-6, start_size=16, min_size=9)
        draw_text(draw, ((x1+x2)//2, yf + footer_h//2), name, font, fill="#FDE68A", max_width=(x2-x1)-7)
    out = os.path.join(GENERATED_DIR, "qualified32_board.png")
    img.save(out, quality=96)
    return out


def _fit_font_by_lines(draw, text, max_width, max_height, start=54, min_size=24):
    for size in range(start, min_size-1, -2):
        f = get_font(size)
        lines = wrap_text(draw, text, f, max_width)
        line_h = int(size * 1.32)
        if len(lines) * line_h <= max_height:
            return f, lines, line_h
    f = get_font(min_size)
    lines = wrap_text(draw, text, f, max_width)
    return f, lines, int(min_size * 1.30)


def render_news_card(kind, team, body):
    """تصميم خبر/عاجل V3: أوضح، أقل زحمة، مربع الخبر هو بطل التصميم."""
    ensure_generated_dir()
    width, height = 1200, 1350
    base, accent, _ = team_theme(team, kind)
    label = news_header_title(kind)
    if kind == "عاجل":
        badge, accent2 = "#DC2626", "#F59E0B"
    elif kind == "تأهل":
        badge, accent2 = "#16A34A", "#FBBF24"
    elif kind == "إقصاء":
        badge, accent2 = "#991B1B", "#FB923C"
    else:
        badge, accent2 = "#0EA5E9", "#FDE68A"
    try:
        rgb1 = tuple(int(base.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
    except Exception:
        rgb1 = (15, 23, 42)
    img = Image.new("RGB", (width, height), base)
    draw = ImageDraw.Draw(img)
    for y in range(height):
        t = y / max(height - 1, 1)
        rgb2 = (2, 6, 18)
        rr = int(rgb1[0]*(1-t)+rgb2[0]*t)
        gg = int(rgb1[1]*(1-t)+rgb2[1]*t)
        bb = int(rgb1[2]*(1-t)+rgb2[2]*t)
        draw.line((0, y, width, y), fill=(rr, gg, bb))
    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)
    try:
        od.ellipse((width-520, 80, width+140, 760), fill=(255,255,255,30))
        od.ellipse((-260, 730, 380, 1360), fill=(255,255,255,18))
        od.rectangle((0, 0, width, 230), fill=(0,0,0,72))
        for i in range(5):
            od.arc((95+i*35, 72+i*22, width-95+i*35, 540+i*22), 188, 350, fill=(255,255,255,13), width=3)
        overlay = overlay.filter(ImageFilter.GaussianBlur(7))
    except Exception:
        pass
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)

    # شارة نوع الخبر
    rounded_rect(draw, (width//2-220, 88, width//2+220, 168), radius=28, fill=badge, outline=accent2, width=2)
    draw_text(draw, (width//2, 128), label, get_font(46), fill="#FFFFFF")

    if team:
        # علم كبير بدون مربع أبيض ثقيل
        rounded_rect(draw, (width//2-170, 222, width//2+170, 362), radius=30, fill="#FFFFFFF5", outline="#FFFFFF99", width=2)
        paste_flag(img, team, (width//2-145, 238, width//2+145, 346))
        draw_text(draw, (width//2, 425), team, get_font(58), fill="#FFFFFF", max_width=990)
        card_top = 520
    else:
        draw_text(draw, (width//2, 285), "كأس العالم 2026", get_font(68), fill="#FFFFFF")
        draw_text(draw, (width//2, 352), "MONDIAL AL MASEEF", get_font(26), fill="#FDE68A")
        card_top = 470

    card_left, card_right = 82, width-82
    card_bottom = height - 185
    # ظل وهالة
    shadow = Image.new("RGBA", (width, height), (0,0,0,0))
    sd = ImageDraw.Draw(shadow)
    sd.rounded_rectangle((card_left+10, card_top+12, card_right+10, card_bottom+12), radius=48, fill=(0,0,0,120))
    shadow = shadow.filter(ImageFilter.GaussianBlur(12))
    img = Image.alpha_composite(img.convert("RGBA"), shadow).convert("RGB")
    draw = ImageDraw.Draw(img)
    rounded_rect(draw, (card_left, card_top, card_right, card_bottom), radius=46, fill="#061633EE", outline=accent, width=4)
    rounded_rect(draw, (card_left+12, card_top+12, card_right-12, card_bottom-12), radius=38, fill="#0B1737CC", outline="#FFFFFF33", width=2)
    # زخرفة بسيطة لا تغطي النص
    draw.line((card_left+72, card_top+72, card_right-72, card_top+72), fill="#FFFFFF55", width=2)
    rounded_rect(draw, (width//2-82, card_top+51, width//2+82, card_top+90), radius=16, fill=badge, outline=accent2, width=1)

    max_w = card_right - card_left - 170
    max_h = card_bottom - card_top - 190
    font, lines, line_h = _fit_font_by_lines(draw, body, max_w, max_h, start=58 if len(body) < 45 else 52, min_size=28)
    total_h = len(lines) * line_h
    y0 = card_top + (card_bottom-card_top)//2 - total_h//2 + 10
    for i, line in enumerate(lines):
        draw_text(draw, (width//2, y0 + i*line_h), line, font, fill="#FFFFFF")

    draw.line((250, height-122, width-250, height-122), fill="#FFFFFF55", width=2)
    draw_text(draw, (width//2, height-88), "المصيف يضعكم بالحدث", get_font(30), fill="#FDE68A")
    out = os.path.join(GENERATED_DIR, f"news_{_safe_filename(kind)}_{_safe_filename(team or 'worldcup')}.png")
    img.save(out, quality=96)
    return out


def _serpapi_key():
    return _env("SERPAPI_KEY")


def serpapi_search_json(query, hl="ar", gl="sa"):
    key = _serpapi_key()
    if not key:
        raise RuntimeError("SERPAPI_KEY غير موجود في Railway")
    if requests is None:
        raise RuntimeError("مكتبة requests غير متوفرة")
    r = requests.get(
        "https://serpapi.com/search.json",
        params={"engine": "google", "q": query, "hl": hl, "gl": gl, "api_key": key},
        timeout=24,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    try:
        data = r.json()
    except Exception:
        raise RuntimeError(f"SerpApi رجع رد غير مفهوم: {r.status_code}")
    if r.status_code >= 400 or data.get("error"):
        raise RuntimeError(str(data.get("error") or f"SerpApi HTTP {r.status_code}"))
    return data


def _looks_like_team_name(name):
    can = canonical_team_name(name)
    return can if can in WORLD_CUP_TEAMS else None


def _extract_score_value(v):
    if isinstance(v, (int, float)):
        return str(int(v))
    if isinstance(v, str):
        m = re.search(r"\d+", v)
        return m.group(0) if m else "0"
    return "0"


def _parse_serp_sports_node(node, req1, req2):
    if not isinstance(node, dict):
        return None
    teams = None
    for key in ["teams", "players"]:
        if isinstance(node.get(key), list) and len(node.get(key)) >= 2:
            teams = node.get(key)
            break
    # بعض نتائج Google ترجع score بشكل مصفوفة
    if teams:
        parsed = []
        for item in teams[:2]:
            if not isinstance(item, dict):
                continue
            nm = item.get("name") or item.get("title") or item.get("team") or item.get("short_name")
            sc = item.get("score") or item.get("points") or item.get("result") or item.get("goals") or 0
            parsed.append((canonical_team_name(nm) or normalize_name(nm), _extract_score_value(sc)))
        if len(parsed) >= 2 and _teams_match(parsed[0][0], parsed[1][0], req1, req2):
            status = node.get("status") or node.get("game_status") or node.get("match_status") or node.get("state") or node.get("time") or node.get("date") or ""
            status_ar = _norm_status_ar(status)
            # لو Google كتب نهائية/Final ضمن status أو date
            scorers = []
            for container in [node.get("timeline"), node.get("events"), node.get("game_spotlight"), node.get("scorers")]:
                if isinstance(container, list):
                    for e in container:
                        if isinstance(e, dict):
                            txt = e.get("text") or e.get("title") or e.get("description") or e.get("name")
                            tm = e.get("time") or e.get("minute")
                            if txt:
                                scorers.append(f"{normalize_name(txt)}" + (f" {tm}'" if tm and str(tm) not in str(txt) else ""))
                        elif isinstance(e, str):
                            scorers.append(e)
                elif isinstance(container, dict):
                    for e in container.values():
                        if isinstance(e, str) and ("'" in e or "هدف" in e or "Goal" in e):
                            scorers.append(e)
            return {
                "team1": parsed[0][0], "team2": parsed[1][0],
                "score1": parsed[0][1], "score2": parsed[1][1],
                "status": status_ar or "غير محدد",
                "minute": "",
                "scorers": list(dict.fromkeys([normalize_name(x) for x in scorers if normalize_name(x)]))[:8],
                "source": "Google Sports",
            }
    # fallback: title + score
    title = normalize_name(node.get("title") or node.get("match") or "")
    if title and _teams_match(title, title, req1, req2):
        return None
    return None


def _walk_json(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from _walk_json(v)
    elif isinstance(obj, list):
        for x in obj:
            yield from _walk_json(x)


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    if not _serpapi_key():
        return None
    ar1, ar2 = canonical_team_name(team1) or team1, canonical_team_name(team2) or team2
    en1, en2 = team_query_name(ar1), team_query_name(ar2)
    queries = []
    if date_hint:
        queries.append(f"مباراة {ar1} {ar2} {date_hint} كأس العالم 2026")
        queries.append(f"{en1} vs {en2} {date_hint} FIFA World Cup 2026")
    queries.extend([
        f"مباراة {ar1} {ar2}",
        f"{en1} vs {en2} FIFA World Cup 2026",
        f"{en1} {en2} score",
    ])
    seen = set()
    for q in queries:
        if q in seen:
            continue
        seen.add(q)
        try:
            data = serpapi_search_json(q)
        except Exception:
            continue
        # sports_results أولًا
        roots = []
        if isinstance(data.get("sports_results"), dict):
            roots.append(data.get("sports_results"))
        roots.append(data)
        for root in roots:
            for node in _walk_json(root):
                obj = _parse_serp_sports_node(node, ar1, ar2)
                if obj:
                    return obj
        # fallback من snippets إذا ظهرت نتيجة مثل 1-0
        text_blob = json.dumps(data, ensure_ascii=False)
        if all(x in text_blob for x in [en1.split()[0], en2.split()[0]]) or (ar1 in text_blob and ar2 in text_blob):
            m = re.search(r"(\d+)\s*[-–]\s*(\d+)", text_blob)
            if m:
                return {"team1": ar1, "team2": ar2, "score1": m.group(1), "score2": m.group(2), "status": "حسب نتائج Google", "minute": "", "scorers": [], "source": "Google Search"}
    return None


def _api_football_key():
    return _env("API_FOOTBALL_KEY")


def _api_football_season():
    return _env("API_FOOTBALL_SEASON", "2026")


def _api_football_get(endpoint, params=None):
    key = _api_football_key()
    if not key:
        raise RuntimeError("API_FOOTBALL_KEY غير موجود")
    if requests is None:
        raise RuntimeError("مكتبة requests غير متوفرة")
    url = "https://v3.football.api-sports.io" + endpoint
    r = requests.get(url, params=params or {}, headers={"x-apisports-key": key}, timeout=22)
    try:
        data = r.json()
    except Exception:
        raise RuntimeError(f"API-Football رجع رد غير مفهوم: {r.status_code}")
    if r.status_code >= 400:
        raise RuntimeError(f"API-Football HTTP {r.status_code}")
    if data.get("errors"):
        # بعض الأحيان errors تكون dict أو list
        if isinstance(data.get("errors"), (list, dict)) and data.get("errors"):
            raise RuntimeError(str(data.get("errors"))[:240])
    return data


def get_api_football_league_id():
    val = _env("API_FOOTBALL_LEAGUE_ID")
    if val:
        return val
    cache = load_sports_cache()
    cached = str(cache.get("api_football_worldcup_league_id") or "").strip()
    if cached:
        return cached
    season = _api_football_season()
    # ابحث عن World Cup في الموسم المطلوب
    for params in [{"name": "World Cup", "season": season}, {"search": "World Cup", "season": season}, {"search": "FIFA World Cup"}]:
        try:
            data = _api_football_get("/leagues", params)
            for item in data.get("response", []) or []:
                lg = item.get("league") or {}
                name = normalize_name(lg.get("name"))
                if "World Cup" in name and "Women" not in name and "U20" not in name and "U17" not in name:
                    lid = lg.get("id")
                    if lid:
                        cache["api_football_worldcup_league_id"] = str(lid)
                        save_sports_cache(cache)
                        return str(lid)
        except Exception:
            continue
    return ""


def _api_fixture_to_match(fx):
    fixture = fx.get("fixture") or {}
    teams = fx.get("teams") or {}
    goals = fx.get("goals") or {}
    home = teams.get("home") or {}
    away = teams.get("away") or {}
    t1 = canonical_team_name(home.get("name")) or normalize_name(home.get("name"))
    t2 = canonical_team_name(away.get("name")) or normalize_name(away.get("name"))
    status_obj = fixture.get("status") or {}
    elapsed = status_obj.get("elapsed")
    status_ar = _norm_status_ar(status_obj.get("long"), status_obj.get("short"), elapsed)
    return {
        "team1": t1,
        "team2": t2,
        "score1": str(goals.get("home") if goals.get("home") is not None else 0),
        "score2": str(goals.get("away") if goals.get("away") is not None else 0),
        "status": status_ar,
        "minute": f"{elapsed}'" if elapsed else "",
        "scorers": [],
        "source": "API-Football",
        "fixture_id": fixture.get("id"),
        "date": fixture.get("date"),
    }


def _fetch_api_football_events(fixture_id):
    if not fixture_id:
        return []
    try:
        data = _api_football_get("/fixtures/events", {"fixture": fixture_id})
        out = []
        for e in data.get("response", []) or []:
            if normalize_name(e.get("type")).lower() == "goal" or "Goal" in normalize_name(e.get("detail")):
                player = _safe_get(e, "player", "name", default="")
                elapsed = _safe_get(e, "time", "elapsed", default="")
                team = _safe_get(e, "team", "name", default="")
                out.append(f"{player} {elapsed}'" + (f" — {canonical_team_name(team) or normalize_name(team)}" if team else ""))
        return [normalize_name(x) for x in out if normalize_name(x)][:8]
    except Exception:
        return []


def fetch_match_from_api_football(team1, team2, date_hint=None):
    if not _api_football_key():
        return None
    lid = get_api_football_league_id()
    season = _api_football_season()
    dates = []
    if date_hint:
        dates.append(date_hint)
    else:
        today = datetime.utcnow().date()
        for delta in [0, -1, -2, 1]:
            dates.append((today + timedelta(days=delta)).isoformat())
    seen = set()
    for d in dates:
        if d in seen:
            continue
        seen.add(d)
        params = {"date": d, "season": season}
        if lid:
            params["league"] = lid
        try:
            data = _api_football_get("/fixtures", params)
            for fx in data.get("response", []) or []:
                obj = _api_fixture_to_match(fx)
                if _teams_match(obj.get("team1"), obj.get("team2"), team1, team2):
                    obj["scorers"] = _fetch_api_football_events(obj.get("fixture_id"))
                    return obj
        except Exception:
            continue
    # live fallback
    try:
        data = _api_football_get("/fixtures", {"live": "all"})
        for fx in data.get("response", []) or []:
            obj = _api_fixture_to_match(fx)
            if _teams_match(obj.get("team1"), obj.get("team2"), team1, team2):
                obj["scorers"] = _fetch_api_football_events(obj.get("fixture_id"))
                return obj
    except Exception:
        pass
    return None


def _validate_worldcup_groups(groups):
    if not groups or len(groups) < 8:
        return []
    allowed = set(WORLD_CUP_TEAMS)
    valid_groups = []
    seen = set()
    for title, rows in groups:
        if not rows:
            continue
        clean_rows = []
        for row in rows:
            team = canonical_team_name(row[0]) or normalize_name(row[0])
            if team not in allowed:
                continue
            played, gd, pts = _as_int(row[1]), _as_int(row[2]), _as_int(row[3])
            clean_rows.append((team, played, gd, pts))
            seen.add(team)
        if clean_rows:
            clean_rows.sort(key=lambda x: (x[3], x[2], x[0]), reverse=True)
            valid_groups.append((title, clean_rows))
    # لا نعرض ترتيب مشكوك؛ لازم عدد كبير من فرق كأس العالم يظهر
    if len(seen) < 24:
        return []
    return valid_groups


def fetch_standings_from_api_football():
    if not _api_football_key():
        return []
    lid = get_api_football_league_id()
    if not lid:
        return []
    data = _api_football_get("/standings", {"league": lid, "season": _api_football_season()})
    resp = data.get("response", []) or []
    groups = []
    for item in resp:
        league = item.get("league") or {}
        for group_rows in league.get("standings", []) or []:
            rows = []
            title = ""
            for r in group_rows or []:
                group_name = normalize_name(r.get("group") or "")
                if group_name and not title:
                    # Group A -> المجموعة A
                    m = re.search(r"([A-L])\b", group_name, re.I)
                    title = f"المجموعة {m.group(1).upper()}" if m else group_name
                team = canonical_team_name(_safe_get(r, "team", "name", default=""))
                if not team:
                    continue
                rows.append((team, _as_int(r.get("all", {}).get("played") or r.get("played")), _as_int(r.get("goalsDiff")), _as_int(r.get("points"))))
            if rows:
                groups.append((title or "المجموعة", rows))
    return _validate_worldcup_groups(groups)


def fetch_standings_from_serpapi():
    # Google Sports standings parsing is not always returned in one stable shape.
    # We only accept it when we can validate it against the 48 approved teams.
    if not _serpapi_key():
        return []
    queries = ["ترتيب مجموعات كأس العالم 2026", "FIFA World Cup 2026 group standings"]
    for q in queries:
        try:
            data = serpapi_search_json(q)
        except Exception:
            continue
        groups = []
        # Common SerpApi shapes: sports_results > standings/tables/groups
        candidates = []
        sr = data.get("sports_results") if isinstance(data, dict) else None
        if isinstance(sr, dict):
            candidates.append(sr)
        candidates.append(data)
        for root in candidates:
            for node in _walk_json(root):
                # group standings list
                for key in ["standings", "table", "tables", "groups"]:
                    val = node.get(key) if isinstance(node, dict) else None
                    if isinstance(val, list):
                        # val may be list of groups or rows
                        for sub in val:
                            if isinstance(sub, dict) and isinstance(sub.get("teams"), list):
                                title = normalize_name(sub.get("title") or sub.get("name") or sub.get("group") or "المجموعة")
                                rows = []
                                for t in sub.get("teams") or []:
                                    if not isinstance(t, dict):
                                        continue
                                    nm = t.get("name") or t.get("team") or t.get("title")
                                    team = canonical_team_name(nm)
                                    if not team:
                                        continue
                                    pts = t.get("points") or t.get("pts") or t.get("score") or 0
                                    played = t.get("played") or t.get("matches") or t.get("mp") or 0
                                    gd = t.get("goal_difference") or t.get("gd") or 0
                                    rows.append((team, _as_int(played), _as_int(gd), _as_int(pts)))
                                if rows:
                                    groups.append((title, rows))
        groups = _validate_worldcup_groups(groups)
        if groups:
            return groups
    return []


def fetch_live_match_data(team1, team2, mode="official", date_hint=None):
    mode = normalize_name(mode or "official").lower()
    if mode in ["سريع", "fast"]:
        sources = [lambda a,b: fetch_match_from_serpapi(a, b, date_hint), lambda a,b: fetch_match_from_api_football(a, b, date_hint), lambda a,b: fetch_match_from_fifa(a, b)]
    elif mode in ["الأحدث", "latest"]:
        sources = [lambda a,b: fetch_match_from_serpapi(a, b, date_hint), lambda a,b: fetch_match_from_api_football(a, b, date_hint), lambda a,b: fetch_match_from_fifa(a, b)]
    else:
        sources = [lambda a,b: fetch_match_from_api_football(a, b, date_hint), lambda a,b: fetch_match_from_fifa(a, b), lambda a,b: fetch_match_from_serpapi(a, b, date_hint)]
    for fn in sources:
        try:
            obj = fn(team1, team2)
            if obj:
                if not obj.get("status"):
                    obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
                return obj
        except Exception:
            continue
    return None


def fetch_current_groups(mode="official"):
    mode = normalize_name(mode or "official").lower()
    if mode in ["سريع", "fast"]:
        order = [(fetch_standings_from_serpapi, "Google Sports"), (fetch_standings_from_api_football, "API-Football")]
    elif mode in ["الأحدث", "latest"]:
        order = [(fetch_standings_from_serpapi, "Google Sports"), (fetch_standings_from_api_football, "API-Football")]
    else:
        order = [(fetch_standings_from_api_football, "API-Football")]
    for fn, label in order:
        try:
            groups = fn()
            groups = _validate_worldcup_groups(groups)
            if groups:
                return groups, label
        except Exception:
            continue
    # لا رجوع لبيانات ESPN العامة حتى لا يظهر ترتيب خاطئ.
    return [], ""


def parse_live_command_text(text):
    body = parse_command_body_lines(text)
    raw = " ".join(body)
    parts = [normalize_name(x) for x in raw.split("*") if normalize_name(x)]
    mode = "official"
    date_hint = _parse_live_date_from_text(raw)
    # آخر جزء قد يكون مصدر أو تاريخ
    while len(parts) >= 3:
        last = normalize_name(parts[-1])
        if last in ["رسمي", "سريع", "الأحدث", "official", "fast", "latest"]:
            mode = last
            parts = parts[:-1]
            continue
        d = _parse_live_date_from_text(last)
        if d:
            date_hint = d
            parts = parts[:-1]
            continue
        break
    if len(parts) < 2:
        return None, None, mode, date_hint
    return canonical_team_name(parts[0]) or normalize_name(parts[0]), canonical_team_name(parts[1]) or normalize_name(parts[1]), mode, date_hint


def build_live_caption(match, mode_label="رسمي"):
    lines = [f"{match['team1']} {match['score1']} - {match['score2']} {match['team2']}"]
    status_line = match.get("minute") or match.get("status") or ""
    if match.get("status") and match.get("minute") and match.get("status") not in status_line:
        status_line = f"{match['status']} — {match['minute']}"
    if status_line:
        lines.append(status_line)
    if match.get("scorers"):
        lines.append("")
        lines.append("الهدافون / الأحداث:")
        for s in match.get("scorers", [])[:6]:
            lines.append(f"- {normalize_name(s)}")
    lines.append("")
    src = match.get("source") or mode_label
    lines.append(f"المصدر: {mode_label} ({src})")
    return "\n".join(lines)



# ==================== V25 FIXTURES DESIGN OVERRIDE ====================
# هذا القسم يتعمد تعريف دوال /مباريات مرة أخيرة قبل التشغيل.
# الهدف:
# - /مباريات 20/06 = نفس تصميم /مباريات_اليوم الرئيسي (تصميم 2)
# - اختيار يوم من الأزرار = يظهر زرين تصميم 1 وتصميم 2
# - تصميم 1 = مضغوط بخلفية التمثال بدون الكلام الكبير فوق
# - تصميم 2 = التصميم الرئيسي المعتمد V31 /مباريات_اليوم
# - /مباريات_مجمعة = تصميم مضغوط لعدة أيام
# - حذف تكرار المباريات داخل نفس اليوم

def _v25_safe_txt(x):
    return str(x or "").strip()


def _v25_dedupe_fixture_matches(matches):
    seen = set()
    out = []
    for m in matches or []:
        t1 = _v25_safe_txt(m.get("team1"))
        t2 = _v25_safe_txt(m.get("team2"))
        tm = _v25_safe_txt(m.get("time"))
        dt = _v25_safe_txt(m.get("date"))
        key = (
            re.sub(r"\s+", " ", t1.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", t2.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", tm).strip(),
            dt,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(dict(m))
    return out


def _v25_fixture_simple_matches(date):
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    simple = []
    for m in rows:
        simple.append((
            _v25_safe_txt(m.get("team1")),
            _v25_safe_txt(m.get("team2")),
            _v25_safe_txt(m.get("time")),
        ))
    return rows, simple


def _v25_fixture_title(date):
    d = _normalize_date_arg(date)
    day = ""
    for x, dy in _fixture_dates():
        if x == d:
            day = dy
            break
    return f"{day} {d[:5]}".strip()


def _fixtures_caption(date_or_title, source="PDF جدول البطولة"):
    return f"{date_or_title}\nالمصدر: {source}\nالمصيف يضعكم بالحدث"


def _v25_compact_bg(w=1080, h=1350):
    """
    خلفية التمثال النظيفة للمضغوط بدون عنوان GAMES الكبير.
    """
    candidates = [
        "games_v31_clean_bg.png",
        os.path.join("assets", "templates", "games_v31_clean_bg.png"),
        "games_v31_full_bg.png",
        os.path.join("assets", "templates", "games_v31_full_bg.png"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                bg = Image.open(p).convert("RGB").resize((w, h))
                ov = Image.new("RGBA", (w, h), (0, 10, 30, 120))
                return Image.alpha_composite(bg.convert("RGBA"), ov).convert("RGB")
            except Exception:
                pass
    return Image.new("RGB", (w, h), "#071329")


def _v25_draw_compact_match(draw, box, m):
    x0, y0, x1, y1 = box
    try:
        draw.rounded_rectangle(box, radius=22, fill=(5,24,58,225), outline=(38,151,255,190), width=2)
    except Exception:
        draw.rectangle(box, fill=(5,24,58), outline=(38,151,255), width=2)

    # الوقت يمين
    draw_text(draw, (x1-35, y0+38), _v25_safe_txt(m.get("time")), get_font(27), fill="#FBBF24", anchor="rm", max_width=190)

    # المباراة بالنص
    text = f"{_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))}"
    draw_text(draw, ((x0+x1)//2, y0+36), text, get_font(30), fill="#FFFFFF", max_width=(x1-x0)-260)

    sub = _v25_safe_txt(m.get("stage"))
    if m.get("group"):
        sub += f" - {_v25_safe_txt(m.get('group'))}"
    draw_text(draw, ((x0+x1)//2, y0+75), sub, get_font(21), fill="#CBD5E1", max_width=(x1-x0)-160)


def _render_fixture_day_compact(date):
    """
    تصميم 1: مضغوط، خلفية التمثال، بدون الكلام الكبير فوق.
    """
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    if not rows:
        return []

    chunks = [rows[i:i+8] for i in range(0, len(rows), 8)]
    paths = []

    for page, chunk in enumerate(chunks, 1):
        h = 1350
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 105
        # شريط التاريخ فقط بدون عنوان كبير
        try:
            draw.rounded_rectangle((120, y, 960, y+62), radius=22, fill=(251,191,36,235))
        except Exception:
            draw.rectangle((120, y, 960, y+62), fill=(251,191,36))
        title = _v25_fixture_title(date)
        if len(chunks) > 1:
            title += f" | {page}/{len(chunks)}"
        draw_text(draw, (540, y+31), title, get_font(34), fill="#061329", max_width=780)

        y += 92
        for m in chunk:
            _v25_draw_compact_match(draw, (70, y, 1010, y+104), m)
            y += 122

        draw.line((250, 1238, 830, 1238), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(30), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_compact_{date.replace('/','_')}_{page}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _render_fixture_day_by_design(date, design=2):
    """
    design=1 => مضغوط
    design=2 => نفس تصميم /مباريات_اليوم الرئيسي
    """
    rows, simple_matches = _v25_fixture_simple_matches(date)
    if not simple_matches:
        return []

    if int(design) == 1:
        return _render_fixture_day_compact(date)

    # تصميم 2: نفس تصميم /مباريات_اليوم
    chunks = [simple_matches[i:i+7] for i in range(0, len(simple_matches), 7)]
    paths = []

    for page_idx, chunk in enumerate(chunks, start=1):
        page_title = _v25_fixture_title(date)
        if len(chunks) > 1:
            page_title = f"{page_title} | {page_idx}/{len(chunks)}"

        # هذا هو التصميم المعتمد نفسه حق /مباريات_اليوم
        path = create_matches_today_v31_full_image(page_title, chunk)
        final_path = os.path.join(
            GENERATED_DIR,
            f"fixtures_day_design2_{date.replace('/','_')}_{page_idx}.png"
        )
        try:
            Image.open(path).save(final_path, quality=96)
            paths.append(final_path)
        except Exception:
            paths.append(path)

    return paths


def render_fixtures_combined_images(dates):
    """
    تصميم مجمع: خلفية التمثال بدون الكلام الكبير فوق.
    إذا كثرت الأيام يقسم كل 3 أيام في صورة.
    """
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    dates = [d for d in dates if _fixtures_for_date(d)]
    # إزالة تكرار التواريخ مع الحفاظ على الترتيب
    clean_dates = []
    for d in dates:
        if d not in clean_dates:
            clean_dates.append(d)
    dates = clean_dates

    if not dates:
        return []

    date_chunks = [dates[i:i+3] for i in range(0, len(dates), 3)]
    paths = []

    for page, dchunk in enumerate(date_chunks, 1):
        rows = []
        for d in dchunk:
            rows.append(("date", d))
            for m in _v25_dedupe_fixture_matches(_fixtures_for_date(d)):
                rows.append(("match", m))

        h = max(1350, 170 + len(rows)*112 + 140)
        h = min(h, 2400)
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 90
        if len(date_chunks) > 1:
            draw_text(draw, (540, 42), f"صفحة {page}/{len(date_chunks)}", get_font(24), fill="#FBBF24")

        for kind, val in rows:
            if y > h - 135:
                break

            if kind == "date":
                try:
                    draw.rounded_rectangle((90, y, 990, y+54), radius=20, fill=(251,191,36,235))
                except Exception:
                    draw.rectangle((90, y, 990, y+54), fill=(251,191,36))
                draw_text(draw, (540, y+27), _v25_fixture_title(val), get_font(31), fill="#061329", max_width=830)
                y += 72
            else:
                _v25_draw_compact_match(draw, (75, y, 1005, y+96), val)
                y += 112

        draw.line((250, h-88, 830, h-88), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, h-48), "المصيف يضعكم بالحدث", get_font(29), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_combined_v25_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _fixtures_day_keyboard(date):
    rows = [
        [
            InlineKeyboardButton("تصميم 1", callback_data=f"fx|render1|{date}"),
            InlineKeyboardButton("تصميم 2", callback_data=f"fx|render2|{date}")
        ]
    ]

    miss = [m for m in _fixtures_for_date(date) if _has_unknown(m)]
    for i, m in enumerate(miss, 1):
        rows.append([InlineKeyboardButton(f"تحديث مباراة {i} — {m.get('time')}", callback_data=f"fx|upd|{m.get('id')}")])

    rows.append([InlineKeyboardButton("رجوع للأيام", callback_data="fx|menu")])
    return InlineKeyboardMarkup(rows)


def _fixtures_dates_keyboard(mode="single", selected=None):
    selected = set(selected or [])
    rows = []
    row = []

    for d, day in _fixture_dates():
        label = f"{'✅ ' if d in selected else ''}{day} {d[:5]}"
        data = f"fx|toggle|{d}" if mode == "multi" else f"fx|day|{d}"
        row.append(InlineKeyboardButton(label, callback_data=data))
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    if mode == "multi":
        rows.append([
            InlineKeyboardButton("تصميم كل يوم", callback_data="fx|render_each"),
            InlineKeyboardButton("تصميم واحد", callback_data="fx|render_combo"),
        ])
        rows.append([
            InlineKeyboardButton("تصفير الاختيار", callback_data="fx|clear"),
            InlineKeyboardButton("رجوع", callback_data="fx|menu"),
        ])
    else:
        rows.append([InlineKeyboardButton("اختيار أكثر من يوم", callback_data="fx|multi")])

    return InlineKeyboardMarkup(rows)


def _fixtures_day_text(date):
    rows, matches = _v25_fixture_simple_matches(date)
    if not matches:
        return "ما فيه مباريات لهذا التاريخ."

    lines = [f"{_v25_fixture_title(date)}", ""]
    for i, m in enumerate(rows, 1):
        lines.append(f"{i}) {_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))} — {_v25_safe_txt(m.get('time'))}")
        extra = []
        if m.get("stage"):
            extra.append(_v25_safe_txt(m.get("stage")))
        if m.get("group"):
            extra.append(_v25_safe_txt(m.get("group")))
        if extra:
            lines.append("   " + " | ".join(extra))
        if _has_unknown(m) and m.get("note"):
            lines.append(f"   {_v25_safe_txt(m.get('note'))}")

    return "\n".join(lines)


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    dates = _extract_fixture_dates_from_text(text)

    if dates:
        # تاريخ واحد: يصمم مباشرة بتصميم 2، نفس /مباريات_اليوم
        if len(dates) == 1:
            d = dates[0]
            wait = await update.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=2)
                if not paths:
                    await wait.edit_text(f"ما فيه مباريات بتاريخ {d}")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(update.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        # أكثر من تاريخ: مجمع
        wait = await update.message.reply_text("⏳ جاري تصميم الأيام المجمعة...")
        try:
            paths = render_fixtures_combined_images(dates)
            if not paths:
                await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
                return
            try:
                await wait.delete()
            except Exception:
                pass
            for p in paths:
                await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
        except Exception as e:
            await wait.edit_text(f"تعذر تصميم الأيام ❌\nالسبب: {str(e)[:400]}")
        return

    await update.message.reply_text("اختر اليوم أو اكتب:\n/مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return

    wait = await update.message.reply_text("⏳ جاري تصميم المباريات المجمعة...")
    try:
        paths = render_fixtures_combined_images(dates)
        if not paths:
            await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
            return
        try:
            await wait.delete()
        except Exception:
            pass
        for p in paths:
            await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
    except Exception as e:
        await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return

    await q.answer()

    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return

    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""

    try:
        if action == "menu":
            await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
            return

        if action == "multi":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "toggle" and len(parts) >= 3:
            d = parts[2]
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if d in sel:
                sel.remove(d)
            else:
                sel.append(d)
            context.user_data["fx_selected_dates"] = sel
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", sel))
            return

        if action == "clear":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "day" and len(parts) >= 3:
            d = parts[2]
            await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
            return

        if action in ["render1", "render2"] and len(parts) >= 3:
            d = parts[2]
            design = 1 if action == "render1" else 2
            wait = await q.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=design)
                if not paths:
                    await wait.edit_text("ما فيه مباريات لهذا اليوم.")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_each":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري تصميم الأيام...")
            try:
                try:
                    await wait.delete()
                except Exception:
                    pass
                for d in sel:
                    for p in _render_fixture_day_by_design(d, design=2):
                        await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await q.message.reply_text(f"تعذر التصميم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_combo":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري التصميم المجمع...")
            try:
                paths = render_fixtures_combined_images(sel)
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption("مباريات مجمعة"))
            except Exception as e:
                await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "upd" and len(parts) >= 3:
            mid = parts[2]
            m = _fixture_by_id(mid)
            if not m:
                await q.message.reply_text("لم أجد المباراة.")
                return

            context.user_data["fixture_update_match_id"] = mid
            await q.message.reply_text(
                f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {m.get('time')}) كذا:\n"
                "الفريق الأول * الفريق الثاني\n\n"
                "مثال: المكسيك * أستراليا\n"
                "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
            )
            return

        await q.message.reply_text("تعذر قراءة الخيار.")
    except Exception as e:
        await q.message.reply_text(f"تعذر تنفيذ خيار المباريات ❌\n{str(e)[:400]}")


async def fixtures_update_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mid = context.user_data.get("fixture_update_match_id")
    if not mid:
        return

    text = (update.message.text or "").strip()

    if "*" in text:
        a, b = [x.strip() for x in text.split("*", 1)]
    elif "×" in text:
        a, b = [x.strip() for x in text.split("×", 1)]
    elif "-" in text:
        a, b = [x.strip() for x in text.split("-", 1)]
    else:
        await update.message.reply_text("اكتبها كذا: الفريق الأول * الفريق الثاني")
        return

    if not a or not b:
        await update.message.reply_text("اكتب اسم الفريقين كاملين.")
        return

    data = _load_fixture_updates()
    data.setdefault(mid, {})
    data[mid]["team1"] = canonical_team_name(a) or normalize_name(a)
    data[mid]["team2"] = canonical_team_name(b) or normalize_name(b)
    _save_fixture_updates(data)

    context.user_data.pop("fixture_update_match_id", None)

    m = _apply_fixture_updates(_fixture_by_id(mid) or {"id": mid})
    await update.message.reply_text(
        f"✅ تم حفظ تحديث المباراة\n"
        f"{m.get('team1')} × {m.get('team2')} — {m.get('time', '')}\n\n"
        f"لن أصمم الآن. وقت ما تبيها اكتب:\n/مباريات {m.get('date', '')}"
    )

# ==================== END V25 FIXTURES DESIGN OVERRIDE ====================


async def api_check_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lines = ["فحص مصادر النتائج:", ""]
    # SerpApi
    if _serpapi_key():
        try:
            data = serpapi_search_json("مباراة المكسيك كوريا الجنوبية")
            lines.append("✅ SERPAPI_KEY موجود والاتصال بقوقل شغال")
            if isinstance(data.get("sports_results"), dict):
                lines.append("✅ Google Sports ظهر في نتيجة الفحص")
            else:
                lines.append("⚠️ الاتصال شغال، لكن لم يظهر كرت Google Sports في هذا الفحص")
        except Exception as e:
            lines.append(f"❌ SerpApi موجود لكن فشل الاتصال: {str(e)[:120]}")
    else:
        lines.append("❌ SERPAPI_KEY غير موجود")
    # API-Football
    if _api_football_key():
        try:
            status = _api_football_get("/status", {})
            lines.append("✅ API_FOOTBALL_KEY موجود والاتصال شغال")
            lid = get_api_football_league_id()
            if lid:
                lines.append(f"✅ League ID الحالي لكأس العالم: {lid}")
            else:
                lines.append("⚠️ لم يتم تحديد League ID تلقائيًا")
        except Exception as e:
            lines.append(f"❌ API-Football فشل: {str(e)[:120]}")
    else:
        lines.append("❌ API_FOOTBALL_KEY غير موجود")
    await update.message.reply_text("\n".join(lines))


def _source_help_text(kind, mode):
    if kind == "standings":
        return f"تعذر جلب ترتيب مجموعات مؤكد من مصدر {mode_label_ar(mode)} ❌\nلن أعرض ترتيبًا غير موثوق.\n\nاختر مصدر آخر أو استخدم /قالب_المجموعات للتحديث اليدوي."
    return f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\nجرب المصدر السريع أو اكتب التاريخ."


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|official|fast|latest)\s*$", text or "", re.I)
        mode = m.group(1) if m else "official"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    groups, source_label = fetch_current_groups(mode)
    if not groups:
        await update.message.reply_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    path = create_all_groups_image(groups)
    caption = f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({source_label})"
    await send_photo_path_markup(update.message, path, caption, kb)
    await update.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({source_label})"))


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = mode_override
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر السعودية * اسبانيا\nأو\n/مباشر السعودية * اسبانيا * سريع\nأو بتاريخ محدد:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    data = fetch_live_match_data(team1, team2, mode, date_hint=date_hint)
    if not data:
        await update.message.reply_text(
            f"تعذر جلب المباراة من المصدر الحالي ❌\nمباراة: {team1} × {team2}\nالمصدر الحالي: {mode_label_ar(mode)}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") + "\nاختر مصدر آخر:",
            reply_markup=kb,
        )
        return
    path = render_live_match_card(data, mode_label_ar(mode))
    await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    kb = source_keyboard(context, payload)
    if kind == "standings":
        groups, src = fetch_current_groups(mode)
        if not groups:
            await query.message.reply_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({src})", kb)
        await query.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({src})"))
        return
    if kind == "live":
        team1, team2 = payload.get("team1"), payload.get("team2")
        data = fetch_live_match_data(team1, team2, mode, date_hint=payload.get("date_hint"))
        if not data:
            await query.message.reply_text(f"تعذر جلب مباراة {team1} × {team2} من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = render_live_match_card(data, mode_label_ar(mode))
        await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")



# ==================== V5 FINAL PATCH: live sources + participants + news card polish ====================

def _env_value(name, default=""):
    try:
        return (os.getenv(name) or default or "").strip()
    except Exception:
        return default or ""


def serpapi_search_json(query, hl="ar", gl="sa", timeout=8, **kwargs):
    """SerpApi Google JSON. يقبل timeout حتى لا يتعطل /بحث_قوقل و/مباشر."""
    key = _serpapi_key() if '_serpapi_key' in globals() else _env_value('SERPAPI_KEY')
    if not key:
        raise RuntimeError("SERPAPI_KEY غير موجود في Railway")
    params = {"engine": "google", "q": query, "hl": hl, "gl": gl, "api_key": key}
    params.update(kwargs or {})
    return _http_json_get("https://serpapi.com/search.json", params=params, timeout=timeout)


def _norm_source_mode(mode):
    m = normalize_name(mode or "").lower().strip()
    table = {
        "سريع": "google", "fast": "google", "google": "google", "قوقل": "google", "جوجل": "google",
        "365": "365", "٣٦٥": "365", "365scores": "365", "365score": "365",
        "كورة": "kooora", "كوره": "kooora", "kooora": "kooora", "koora": "kooora",
        "رسمي": "official", "official": "official", "api": "official", "api-football": "official",
        "الأحدث": "latest", "الاحدث": "latest", "latest": "latest",
    }
    return table.get(m, m or "official")


def mode_label_ar(mode):
    m = _norm_source_mode(mode)
    return {
        "google": "قوقل",
        "365": "365",
        "kooora": "كورة",
        "official": "رسمي",
        "latest": "الأحدث",
    }.get(m, normalize_name(mode or "رسمي"))


def source_keyboard(context, payload):
    token = store_source_request(context, payload)
    rows = [
        [
            InlineKeyboardButton("قوقل", callback_data=f"sportsrc|{token}|google"),
            InlineKeyboardButton("365", callback_data=f"sportsrc|{token}|365"),
            InlineKeyboardButton("كورة", callback_data=f"sportsrc|{token}|kooora"),
        ],
        [
            InlineKeyboardButton("رسمي", callback_data=f"sportsrc|{token}|official"),
            InlineKeyboardButton("الأحدث", callback_data=f"sportsrc|{token}|latest"),
        ],
    ]
    return InlineKeyboardMarkup(rows)


def parse_live_command_text(text):
    body = parse_command_body_lines(text)
    raw = " ".join(body)
    parts = [normalize_name(x) for x in raw.split("*") if normalize_name(x)]
    mode = "google"
    date_hint = _parse_live_date_from_text(raw) if '_parse_live_date_from_text' in globals() else None
    source_words = {"رسمي", "سريع", "الأحدث", "الاحدث", "official", "fast", "latest", "google", "قوقل", "جوجل", "365", "٣٦٥", "365scores", "كورة", "كوره", "kooora", "koora", "api"}
    while len(parts) >= 3:
        last = normalize_name(parts[-1])
        if last.lower() in [x.lower() for x in source_words] or last in source_words:
            mode = last
            parts = parts[:-1]
            continue
        d = _parse_live_date_from_text(last) if '_parse_live_date_from_text' in globals() else None
        if d:
            date_hint = d
            parts = parts[:-1]
            continue
        break
    if len(parts) < 2:
        return None, None, _norm_source_mode(mode), date_hint
    return canonical_team_name(parts[0]) or normalize_name(parts[0]), canonical_team_name(parts[1]) or normalize_name(parts[1]), _norm_source_mode(mode), date_hint


def _team_variants_for_text(team):
    can = canonical_team_name(team) or normalize_name(team)
    vals = [can]
    try:
        vals += TEAM_SEARCH_EN.get(can, []) or []
    except Exception:
        pass
    vals += [simple_key(v) for v in list(vals)]
    out = []
    for v in vals:
        v = normalize_name(v)
        if v and v not in out:
            out.append(v)
    return out


def _blob_has_both_teams(blob, team1, team2):
    low = str(blob).lower()
    simp = simple_key(str(blob))
    def has(team):
        for v in _team_variants_for_text(team):
            if v.lower() in low or simple_key(v) in simp:
                return True
        return False
    return has(team1) and has(team2)


def _score_from_text_near_teams(text_blob, team1, team2):
    """Fallback: يلقط نتيجة مثل Mexico 1-0 Korea أو المكسيك 1 - 0 كوريا."""
    blob = str(text_blob)
    if not _blob_has_both_teams(blob, team1, team2):
        return None
    patterns = [
        r"(\d{1,2})\s*[-–:]\s*(\d{1,2})",
        r"(\d{1,2})\s+vs\s+(\d{1,2})",
    ]
    for pat in patterns:
        for m in re.finditer(pat, blob, flags=re.I):
            a, b = m.group(1), m.group(2)
            if a is not None and b is not None:
                status = "انتهت المباراة" if re.search(r"final|full.?time|نهاية|انته", blob, re.I) else "حسب المصدر"
                return {"score1": a, "score2": b, "status": status}
    return None


def _parse_serp_sports_node_v5(node, req1, req2, source_name="Google Sports"):
    try:
        obj = _parse_serp_sports_node(node, req1, req2)
        if obj:
            obj["source"] = source_name if obj.get("source") in (None, "Google Sports", "Google Search") else obj.get("source")
            return obj
    except Exception:
        pass
    if not isinstance(node, dict):
        return None
    # More Google variants: game_spotlight sometimes has team_comparison / game_status and team scores nested.
    names, scores = [], []
    for n in _walk_json(node):
        if isinstance(n, dict):
            nm = _team_name_from_any(n) if '_team_name_from_any' in globals() else ""
            if nm and any(_blob_has_both_teams(nm, nm, req) for req in [req1, req2]):
                if nm not in names:
                    names.append(nm)
                    scores.append(_score_from_any(n) if '_score_from_any' in globals() else "0")
    if len(names) >= 2 and _match_pair_names(names[0], names[1], req1, req2):
        return {
            "team1": canonical_team_name(names[0]) or names[0], "team2": canonical_team_name(names[1]) or names[1],
            "score1": scores[0] if scores else "0", "score2": scores[1] if len(scores) > 1 else "0",
            "status": _status_from_serp_node(node) if '_status_from_serp_node' in globals() else "حسب المصدر",
            "minute": "", "scorers": _collect_scorers_from_serp_node(node) if '_collect_scorers_from_serp_node' in globals() else [],
            "source": source_name,
        }
    return None


def _serpapi_query_candidates(team1, team2, date_hint=None, source="google"):
    ar1, ar2 = canonical_team_name(team1) or normalize_name(team1), canonical_team_name(team2) or normalize_name(team2)
    en1 = team_query_name(ar1) if 'team_query_name' in globals() else ar1
    en2 = team_query_name(ar2) if 'team_query_name' in globals() else ar2
    qs = []
    if source == "365":
        base = [
            f"site:365scores.com {ar1} {ar2} كأس العالم 2026",
            f"site:365scores.com {en1} {en2} FIFA World Cup 2026 score",
        ]
    elif source == "kooora":
        base = [
            f"site:kooora.com {ar1} {ar2} كأس العالم 2026",
            f"site:kooora.com {en1} {en2} FIFA World Cup 2026 score",
        ]
    else:
        base = [
            f"مباراة {ar1} {ar2}",
            f"{en1} vs {en2} FIFA World Cup 2026",
            f"{en1} {en2} score",
        ]
    if date_hint:
        qs.extend([f"{q} {date_hint}" for q in base])
    qs.extend(base)
    out, seen = [], set()
    for q in qs:
        q = normalize_name(q)
        if q and q not in seen:
            seen.add(q); out.append(q)
    return out


def _fetch_match_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    source_name = {"google": "Google Sports", "365": "365Scores عبر Google", "kooora": "Kooora عبر Google"}.get(source, "Google")
    for q in _serpapi_query_candidates(req1, req2, date_hint, source=source):
        for hl, gl in [("ar", "sa"), ("en", "us")]:
            data = serpapi_search_json(q, hl=hl, gl=gl, timeout=8)
            roots = []
            sr = data.get("sports_results") if isinstance(data, dict) else None
            if isinstance(sr, dict):
                roots.append(sr)
                for k in ["game_spotlight", "games", "matches", "scoreboard", "team_standings"]:
                    if isinstance(sr.get(k), (dict, list)):
                        roots.append(sr.get(k))
            roots.append(data)
            for root in roots:
                for node in _walk_json(root):
                    obj = _parse_serp_sports_node_v5(node, req1, req2, source_name=source_name)
                    if obj:
                        return obj
            # Snippets fallback from organic results / answer box
            blob = json.dumps(data, ensure_ascii=False)
            s = _score_from_text_near_teams(blob, req1, req2)
            if s:
                return {
                    "team1": req1, "team2": req2, "score1": s["score1"], "score2": s["score2"],
                    "status": s.get("status") or "حسب المصدر", "minute": "", "scorers": [], "source": source_name,
                }
    return None


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")


def fetch_match_from_365(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="365")


def fetch_match_from_kooora(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="kooora")


def _source_mode_sequence(mode):
    m = _norm_source_mode(mode)
    if m == "google":
        return ["google"]
    if m == "365":
        return ["365"]
    if m == "kooora":
        return ["kooora"]
    if m == "official":
        return ["api", "fifa", "espn", "google"]
    if m == "latest":
        return ["google", "365", "kooora", "api", "fifa", "espn"]
    return ["google", "365", "kooora", "api"]


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    errors = []
    for src in _source_mode_sequence(mode):
        try:
            if src == "google":
                obj = fetch_match_from_serpapi(team1, team2, date_hint=date_hint)
            elif src == "365":
                obj = fetch_match_from_365(team1, team2, date_hint=date_hint)
            elif src == "kooora":
                obj = fetch_match_from_kooora(team1, team2, date_hint=date_hint)
            elif src == "api":
                obj = fetch_match_from_api_football(team1, team2, date_hint=date_hint)
            elif src == "espn":
                obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
            else:
                obj = fetch_match_from_fifa(team1, team2)
            if obj:
                obj.setdefault("status", "حسب المصدر")
                obj.setdefault("source", mode_label_ar(src))
                return obj
        except Exception as e:
            errors.append(f"{src}: {str(e)[:90]}")
            continue
    return None


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = parse_command_body_lines(update.message.text)
    query = " ".join(q).strip() or "مباراة المكسيك كوريا الجنوبية"
    try:
        data = serpapi_search_json(query, timeout=8)
        sr = data.get("sports_results") if isinstance(data, dict) else None
        lines = [f"✅ وصلنا إلى قوقل عبر SerpApi", f"البحث: {query}"]
        if isinstance(sr, dict):
            lines.append("✅ sports_results موجود")
            title = sr.get("title") or sr.get("league") or sr.get("game") or ""
            if title:
                lines.append(f"العنوان: {title}")
            blob = json.dumps(sr, ensure_ascii=False)
            m = re.search(r"(.{0,60}\d{1,2}\s*[-–:]\s*\d{1,2}.{0,60})", blob)
            if m:
                lines.append(f"لقطة نتيجة: {m.group(1)}")
        else:
            lines.append("⚠️ لم يظهر sports_results في هذا البحث")
        await update.message.reply_text("\n".join(lines[:12]))
    except Exception as e:
        await update.message.reply_text(f"❌ فشل فحص قوقل: {str(e)[:200]}")


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = _norm_source_mode(mode_override)
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر السعودية * إسبانيا\nأو\n/مباشر المكسيك * كوريا الجنوبية * سريع\nأو بتاريخ محدد:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026 * سريع")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait_msg = await update.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}" + (f"\nالتاريخ: {date_hint}" if date_hint else ""))
    data = fetch_live_match_data(team1, team2, mode, date_hint=date_hint)
    if not data:
        await wait_msg.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\nمباراة: {team1} × {team2}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") + "\nاختر مصدر آخر:",
            reply_markup=kb,
        )
        return
    try:
        await wait_msg.delete()
    except Exception:
        pass
    path = render_live_match_card(data, mode_label_ar(mode))
    await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    mode = _norm_source_mode(mode)
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    kb = source_keyboard(context, payload)
    if kind == "standings":
        groups, src = fetch_current_groups(mode)
        if not groups:
            await query.message.reply_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({src})", kb)
        await query.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({src})"))
        return
    if kind == "live":
        team1, team2 = payload.get("team1"), payload.get("team2")
        msg = await query.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
        data = fetch_live_match_data(team1, team2, mode, date_hint=payload.get("date_hint"))
        if not data:
            await msg.edit_text(f"تعذر جلب مباراة {team1} × {team2} من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        try:
            await msg.delete()
        except Exception:
            pass
        path = render_live_match_card(data, mode_label_ar(mode))
        await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|الاحدث|official|fast|latest|google|قوقل|365|٣٦٥|كورة|كوره|kooora)\s*$", text or "", re.I)
        mode = m.group(1) if m else "latest"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    groups, source_label = fetch_current_groups(mode)
    if not groups:
        await update.message.reply_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    path = create_all_groups_image(groups)
    caption = f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({source_label})"
    await send_photo_path_markup(update.message, path, caption, kb)
    await update.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({source_label})"))


def _paste_flag_cover(base, team_name, box, radius=24):
    path = flag_path_for(team_name)
    x1, y1, x2, y2 = [int(v) for v in box]
    w, h = x2-x1, y2-y1
    if path and os.path.exists(path):
        try:
            flag = Image.open(path).convert("RGBA")
            flag = _crop_alpha_content(flag) if '_crop_alpha_content' in globals() else flag
            scale = max(w / max(flag.width,1), h / max(flag.height,1))
            nw, nh = int(flag.width * scale + .5), int(flag.height * scale + .5)
            flag = flag.resize((nw, nh), Image.LANCZOS)
            left, top = max(0, (nw-w)//2), max(0, (nh-h)//2)
            flag = flag.crop((left, top, left+w, top+h))
            mask = _rounded_mask(w, h, radius) if '_rounded_mask' in globals() else None
            base.paste(flag, (x1, y1), mask or flag)
            return
        except Exception:
            pass
    d = ImageDraw.Draw(base)
    rounded_rect(d, box, radius=radius, fill="#FFFFFF", outline="#FFFFFFAA", width=2)
    draw_text(d, ((x1+x2)//2, (y1+y2)//2), normalize_name(team_name)[:2], get_font(44), fill="#0B1635")


def render_news_card(kind, team, body):
    """V5: خبر/عاجل — العلم يملأ الخانة والصندوق يتناسب مع طول الخبر."""
    ensure_generated_dir()
    width, height = 1200, 1350
    base, accent, _ = team_theme(team, kind)
    label = news_header_title(kind)
    if kind == "عاجل":
        badge, accent2 = "#DC2626", "#F59E0B"
    elif kind in ["تأهل", "رسميًا"]:
        badge, accent2 = "#16A34A", "#F59E0B"
    elif kind in ["إقصاء", "اقصاء"]:
        badge, accent2 = "#B91C1C", "#FB923C"
    else:
        badge, accent2 = "#2563EB", "#38BDF8"

    _bg = _style4_clean_background(width, height)
    img = _bg[0] if isinstance(_bg, tuple) else _bg
    overlay = Image.new("RGBA", (width, height), (0,0,0,0))
    od = ImageDraw.Draw(overlay)
    # tint by team/accent
    try:
        rgb = tuple(int((base or '#123456').lstrip('#')[i:i+2],16) for i in (0,2,4))
    except Exception:
        rgb = (8, 31, 22)
    od.rectangle((0,0,width,height), fill=rgb+(116,))
    # subtle spotlight
    for r, alpha in [(540,42),(390,34),(250,26)]:
        od.ellipse((width-r-80, 80, width+180, 80+r+260), fill=(255,255,255,alpha))
    img = Image.alpha_composite(img.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(img)

    # top badge
    rounded_rect(draw, (width//2-170, 95, width//2+170, 170), radius=26, fill=badge, outline=accent2, width=2)
    draw_text(draw, (width//2, 133), label, get_font(42), fill="#FFFFFF")

    title = canonical_team_name(team) if team else None
    if title:
        flag_box = (width//2-175, 230, width//2+175, 375)
        rounded_rect(draw, flag_box, radius=28, fill="#FFFFFF", outline="#FFFFFF", width=2)
        _paste_flag_cover(img, title, flag_box, radius=28)
        draw_text(draw, (width//2, 445), title, get_font(52), fill="#FFFFFF", max_width=700)
    else:
        draw_text(draw, (width//2, 330), "كأس العالم 2026", get_font(62), fill="#FFFFFF", max_width=850)

    # smart story card size
    body = clean_draw_text(body or "")
    maxw = 840
    size = 58 if len(body) <= 25 else 52 if len(body) <= 55 else 44 if len(body) <= 95 else 38
    font = get_font(size)
    temp_draw = ImageDraw.Draw(img)
    lines = wrap_text(temp_draw, body, font, maxw) if 'wrap_text' in globals() else [body]
    line_h = int(size * 1.42)
    content_h = max(1, len(lines)) * line_h
    card_h = min(430, max(210, content_h + 120))
    card_y1 = 565 if title else 520
    card_y2 = card_y1 + card_h
    rounded_rect(draw, (90, card_y1, width-90, card_y2), radius=34, fill="#07132FEE", outline=accent2, width=4)
    rounded_rect(draw, (110, card_y1+18, width-110, card_y2-18), radius=26, fill=None, outline="#FFFFFF77", width=2)
    # small top ornament
    rounded_rect(draw, (width//2-75, card_y1-16, width//2+75, card_y1+18), radius=14, fill=badge, outline=accent2, width=1)
    draw.line((150, card_y1+48, width//2-90, card_y1+48), fill="#FFFFFF77", width=2)
    draw.line((width//2+90, card_y1+48, width-150, card_y1+48), fill="#FFFFFF77", width=2)
    start_y = (card_y1 + card_y2)//2 - content_h//2 + line_h//2 - 5
    for i, line in enumerate(lines):
        draw_text(draw, (width//2, start_y + i*line_h), line, font, fill="#FFFFFF", max_width=maxw)

    fy = min(height-118, card_y2 + 95)
    draw.line((220, fy, width-220, fy), fill="#FFFFFFBB", width=2)
    draw_text(draw, (width//2, fy+42), "المصيف يضعكم بالحدث", get_font(26), fill="#FDE68A")
    path = os.path.join(GENERATED_DIR, f"news_{_safe_filename(kind)}_{_safe_filename(title or 'worldcup')}.png")
    img.convert("RGB").save(path, quality=96)
    return path


PARTICIPANTS_STATE_FILE = "participants_state.json"


def _save_participants_state():
    try:
        with open(PARTICIPANTS_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"participants": PARTICIPANTS}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_participants_state():
    try:
        if os.path.exists(PARTICIPANTS_STATE_FILE):
            with open(PARTICIPANTS_STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            arr = [normalize_name(x) for x in data.get("participants", []) if normalize_name(x)]
            if arr:
                PARTICIPANTS[:] = arr
                return
    except Exception:
        pass
    _save_participants_state()


async def add_participant_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = normalize_name(parse_command_body(update.message.text)) if 'parse_command_body' in globals() else normalize_name(re.sub(r"^/\S+", "", update.message.text or ""))
    if not name:
        await update.message.reply_text("اكتبها كذا:\n/إضافة_متسابق عبدالله محمد")
        return
    if name in PARTICIPANTS:
        await update.message.reply_text(f"✅ {name} موجود مسبقًا في قائمة المتسابقين")
        return
    PARTICIPANTS.append(name)
    _save_participants_state()
    await update.message.reply_text(f"✅ تمت إضافة المتسابق: {name}\nالمجموع الحالي: {len(PARTICIPANTS)}")


async def remove_participant_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = normalize_name(parse_command_body(update.message.text)) if 'parse_command_body' in globals() else normalize_name(re.sub(r"^/\S+", "", update.message.text or ""))
    if not name:
        await update.message.reply_text("اكتبها كذا:\n/حذف_متسابق عبدالله محمد")
        return
    # exact or close
    target = name if name in PARTICIPANTS else None
    if not target:
        close = difflib.get_close_matches(name, PARTICIPANTS, n=1, cutoff=0.75)
        target = close[0] if close else None
    if not target:
        await update.message.reply_text(f"ما لقيت المتسابق: {name} ❌")
        return
    PARTICIPANTS.remove(target)
    _save_participants_state()
    await update.message.reply_text(f"✅ تم حذف المتسابق: {target}\nالمجموع الحالي: {len(PARTICIPANTS)}")


async def participants_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lines = [f"👥 المتسابقون ({len(PARTICIPANTS)}):"]
    for i, name in enumerate(PARTICIPANTS, start=1):
        lines.append(f"{i}- {name}")
    await update.message.reply_text("\n".join(lines))

# ==================== END V5 FINAL PATCH ====================

# ==================== V6 HOTFIX: live sources stability + Google/365/Kooora parsing ====================
import asyncio as _asyncio_v6


def _norm_source_mode(mode):
    m = normalize_name(mode or "").lower().strip()
    table = {
        "سريع": "google", "fast": "google", "google": "google", "قوقل": "google", "جوجل": "google",
        "365": "365", "٣٦٥": "365", "365scores": "365", "365score": "365",
        "كورة": "kooora", "كوره": "kooora", "كووورة": "kooora", "كووره": "kooora", "kooora": "kooora", "koora": "kooora",
        "رسمي": "official", "official": "official", "api": "official", "api-football": "official", "api football": "official",
        "الأحدث": "latest", "الاحدث": "latest", "latest": "latest",
    }
    return table.get(m, m or "google")


def mode_label_ar(mode):
    m = _norm_source_mode(mode)
    return {"google": "قوقل", "365": "365", "kooora": "كورة", "official": "رسمي", "latest": "الأحدث"}.get(m, normalize_name(mode or "قوقل"))


def source_keyboard(context, payload):
    token = store_source_request(context, payload)
    rows = [
        [
            InlineKeyboardButton("قوقل", callback_data=f"sportsrc|{token}|google"),
            InlineKeyboardButton("365", callback_data=f"sportsrc|{token}|365"),
            InlineKeyboardButton("كورة", callback_data=f"sportsrc|{token}|kooora"),
        ],
        [
            InlineKeyboardButton("رسمي", callback_data=f"sportsrc|{token}|official"),
            InlineKeyboardButton("الأحدث", callback_data=f"sportsrc|{token}|latest"),
        ],
    ]
    return InlineKeyboardMarkup(rows)


def _parse_live_date_token_v6(token):
    token = normalize_name(token)
    if not token:
        return None
    # dd/mm/yyyy -> yyyy-mm-dd
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", token)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    # yyyy-mm-dd stays
    if re.match(r"^\d{4}-\d{2}-\d{2}$", token):
        return token
    try:
        return _parse_live_date_from_text(token)
    except Exception:
        return None


def parse_live_command_text(text):
    body = parse_command_body_lines(text)
    raw = " ".join(body)
    parts = [normalize_name(x) for x in raw.split("*") if normalize_name(x)]
    mode = "google"
    date_hint = _parse_live_date_token_v6(raw)
    source_words = {"رسمي", "سريع", "الأحدث", "الاحدث", "official", "fast", "latest", "google", "قوقل", "جوجل", "365", "٣٦٥", "365scores", "كورة", "كوره", "كووورة", "كووره", "kooora", "koora", "api"}
    # Peel any source/date tokens from the end, in any order.
    while len(parts) >= 3:
        last = normalize_name(parts[-1])
        if last.lower() in {x.lower() for x in source_words} or last in source_words:
            mode = last
            parts.pop()
            continue
        d = _parse_live_date_token_v6(last)
        if d:
            date_hint = d
            parts.pop()
            continue
        break
    if len(parts) < 2:
        return None, None, _norm_source_mode(mode), date_hint
    t1 = canonical_team_name(parts[0]) or normalize_name(parts[0])
    t2 = canonical_team_name(parts[1]) or normalize_name(parts[1])
    return t1, t2, _norm_source_mode(mode), date_hint


def _source_mode_sequence(mode):
    m = _norm_source_mode(mode)
    if m == "google":
        return ["google"]
    if m == "365":
        return ["365", "google"]
    if m == "kooora":
        return ["kooora", "google"]
    if m == "official":
        return ["api", "fifa", "espn", "google"]
    if m == "latest":
        return ["google", "365", "kooora", "api", "fifa", "espn"]
    return ["google", "365", "kooora", "api"]


def _team_name_from_any_v6(item):
    if isinstance(item, str):
        return canonical_team_name(item) or normalize_name(item)
    if not isinstance(item, dict):
        return ""
    keys = [
        "name", "title", "team", "team_name", "teamName", "short_name", "shortName",
        "displayName", "display_name", "full_name", "fullName", "participant", "competitor",
        "homeTeam", "awayTeam", "home_team", "away_team", "home", "away",
    ]
    for key in keys:
        v = item.get(key)
        if isinstance(v, str) and normalize_name(v):
            return canonical_team_name(v) or normalize_name(v)
        if isinstance(v, dict):
            nested = _team_name_from_any_v6(v)
            if nested:
                return nested
    # Some Google result cards use aria/alt labels.
    for key in ["alt", "aria_label", "aria-label", "subtitle", "label"]:
        v = item.get(key)
        if isinstance(v, str):
            for tm in WORLD_CUP_TEAMS:
                if _blob_has_both_teams(v, tm, tm):
                    return tm
    return ""


def _score_from_any_v6(item):
    if not isinstance(item, dict):
        return None
    score_keys = [
        "score", "points", "goals", "result", "value", "display_score", "displayScore",
        "home_score", "away_score", "homeScore", "awayScore", "goalsFor", "goals_for",
    ]
    for key in score_keys:
        if key in item and item.get(key) not in (None, ""):
            try:
                return _extract_score_value(item.get(key))
            except Exception:
                return str(item.get(key)).strip()
    for key in ["standing", "stats", "score_data", "scoreData", "statistics"]:
        if isinstance(item.get(key), dict):
            val = _score_from_any_v6(item.get(key))
            if val is not None:
                return val
    return None


def _team_matches_req_v6(name, req):
    try:
        return _blob_contains_team(name, req)
    except Exception:
        name = simple_key(name)
        req = simple_key(req)
        return bool(name and req and (name in req or req in name))


def _collect_scores_by_team_v6(root, req1, req2):
    scores = {}
    names = {}
    for node in _walk_json(root):
        if not isinstance(node, dict):
            continue
        nm = _team_name_from_any_v6(node)
        if not nm:
            continue
        sc = _score_from_any_v6(node)
        if sc is None:
            continue
        if _team_matches_req_v6(nm, req1):
            scores["req1"] = sc
            names["req1"] = canonical_team_name(nm) or normalize_name(nm)
        elif _team_matches_req_v6(nm, req2):
            scores["req2"] = sc
            names["req2"] = canonical_team_name(nm) or normalize_name(nm)
    if "req1" in scores and "req2" in scores:
        return {
            "team1": canonical_team_name(req1) or names.get("req1") or req1,
            "team2": canonical_team_name(req2) or names.get("req2") or req2,
            "score1": scores["req1"],
            "score2": scores["req2"],
        }
    return None


def _status_from_serp_v6(root):
    vals = []
    for node in _walk_json(root):
        if not isinstance(node, dict):
            continue
        for k in ["status", "game_status", "match_status", "status_text", "statusText", "status_line", "state", "period", "time", "date"]:
            v = node.get(k)
            if isinstance(v, str) and normalize_name(v) and len(v) < 80:
                vals.append(v)
    raw = " | ".join(vals[:4])
    if raw:
        try:
            return _norm_status_ar(raw)
        except Exception:
            return normalize_name(raw)
    return "غير محدد"


def _parse_serp_sports_node_v6(node, req1, req2, source_name="Google Sports"):
    # Existing parser first.
    try:
        obj = _parse_serp_sports_node(node, req1, req2)
        if obj:
            obj["source"] = source_name
            return obj
    except Exception:
        pass
    if not isinstance(node, (dict, list)):
        return None
    # New robust team-score collector for Google Sports game_spotlight.
    pair = _collect_scores_by_team_v6(node, req1, req2)
    if pair:
        pair.update({
            "status": _status_from_serp_v6(node),
            "minute": "",
            "scorers": _collect_scorers_from_serp_node(node) if '_collect_scorers_from_serp_node' in globals() else [],
            "source": source_name,
        })
        return pair
    # If Google returned teams without scores, do not invent a score.
    return None


def _serpapi_query_candidates(team1, team2, date_hint=None, source="google"):
    ar1, ar2 = canonical_team_name(team1) or normalize_name(team1), canonical_team_name(team2) or normalize_name(team2)
    en1 = team_query_name(ar1) if 'team_query_name' in globals() else ar1
    en2 = team_query_name(ar2) if 'team_query_name' in globals() else ar2
    if source == "365":
        base = [
            f"site:365scores.com {ar1} {ar2} كأس العالم 2026",
            f"site:365scores.com {en1} {en2} FIFA World Cup 2026 score",
        ]
    elif source == "kooora":
        base = [
            f"site:kooora.com {ar1} {ar2} كأس العالم 2026",
            f"site:kooora.com {en1} {en2} FIFA World Cup 2026 score",
        ]
    else:
        base = [
            f"مباراة {ar1} {ar2}",
            f"{en1} vs {en2} FIFA World Cup 2026",
        ]
    queries = []
    if date_hint:
        queries.extend([f"{q} {date_hint}" for q in base])
    queries.extend(base)
    out = []
    for q in queries:
        q = normalize_name(q)
        if q and q not in out:
            out.append(q)
    return out[:3]


def _fetch_match_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    source_name = {"google": "Google Sports", "365": "365Scores عبر Google", "kooora": "Kooora عبر Google"}.get(source, "Google")
    # Limit attempts so Telegram never feels frozen.
    attempts = []
    for q in _serpapi_query_candidates(req1, req2, date_hint, source=source):
        if source == "google":
            attempts.append((q, "ar", "sa"))
            if len(attempts) < 2:
                attempts.append((q, "en", "us"))
        else:
            attempts.append((q, "ar", "sa"))
    for q, hl, gl in attempts[:3]:
        data = serpapi_search_json(q, hl=hl, gl=gl, timeout=7)
        roots = []
        sr = data.get("sports_results") if isinstance(data, dict) else None
        if isinstance(sr, dict):
            roots.append(sr)
            for k in ["game_spotlight", "games", "matches", "scoreboard", "team_standings", "players", "teams"]:
                if isinstance(sr.get(k), (dict, list)):
                    roots.append(sr.get(k))
        roots.append(data)
        for root in roots:
            obj = _parse_serp_sports_node_v6(root, req1, req2, source_name=source_name)
            if obj:
                return obj
        blob = json.dumps(data, ensure_ascii=False)
        # Organic snippets fallback, only if both teams are clearly present.
        if _blob_has_both_teams(blob, req1, req2):
            s = _score_from_text_near_teams(blob, req1, req2) if '_score_from_text_near_teams' in globals() else None
            if s:
                return {
                    "team1": req1, "team2": req2,
                    "score1": s.get("score1"), "score2": s.get("score2"),
                    "status": s.get("status") or "حسب المصدر", "minute": "",
                    "scorers": _collect_scorers_from_serp_node(data) if '_collect_scorers_from_serp_node' in globals() else [],
                    "source": source_name,
                }
    return None


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")


def fetch_match_from_365(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="365")


def fetch_match_from_kooora(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="kooora")


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    for src in _source_mode_sequence(mode):
        try:
            if src == "google":
                obj = fetch_match_from_serpapi(team1, team2, date_hint=date_hint)
            elif src == "365":
                obj = fetch_match_from_365(team1, team2, date_hint=date_hint)
            elif src == "kooora":
                obj = fetch_match_from_kooora(team1, team2, date_hint=date_hint)
            elif src == "api":
                obj = fetch_match_from_api_football(team1, team2, date_hint=date_hint)
            elif src == "fifa":
                obj = fetch_match_from_fifa(team1, team2)
            else:
                obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
            if obj:
                obj["source"] = obj.get("source") or mode_label_ar(src)
                if not obj.get("status"):
                    obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
                return obj
        except Exception as e:
            # Do not crash or freeze the bot. Try next source.
            continue
    return None


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = _norm_source_mode(mode_override)
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر المكسيك * كوريا الجنوبية * قوقل\nأو:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026 * قوقل")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, date_hint), timeout=18)
    except Exception:
        data = None
    if not data:
        await wait.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\n"
            f"مباراة: {team1} × {team2}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") +
            "\nاختر مصدر آخر أو جرّب الأمر:\n/بحث_قوقل مباراة المكسيك كوريا الجنوبية",
            reply_markup=kb,
        )
        return
    try:
        path = render_live_match_card(data, mode_label_ar(mode))
        await wait.delete()
        await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
    except Exception:
        await wait.edit_text(build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kb = source_keyboard(context, payload)
    if payload.get("kind") == "standings":
        groups, src = fetch_current_groups(mode)
        if not groups:
            await query.message.reply_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({src})", kb)
        await query.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({src})"))
        return
    if payload.get("kind") == "live":
        team1, team2 = payload.get("team1"), payload.get("team2")
        msg = await query.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
        try:
            data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, payload.get("date_hint")), timeout=18)
        except Exception:
            data = None
        if not data:
            await msg.edit_text(f"تعذر جلب مباراة {team1} × {team2} من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        try:
            path = render_live_match_card(data, mode_label_ar(mode))
            await msg.delete()
            await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        except Exception:
            await msg.edit_text(build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    query = " ".join(body).strip() or "مباراة المكسيك كوريا الجنوبية"
    msg = await update.message.reply_text(f"⏳ أفحص قوقل: {query}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(serpapi_search_json, query, "ar", "sa", 8), timeout=12)
        sr = data.get("sports_results") if isinstance(data, dict) else None
        lines = ["نتيجة فحص قوقل:", f"query: {query}"]
        if isinstance(sr, dict):
            lines.append("✅ sports_results موجود")
            lines.append("مفاتيح: " + "، ".join(list(sr.keys())[:12]))
            # Try parse the requested/known Mexico-Korea match if present.
            obj = None
            for root in [sr, data]:
                obj = _parse_serp_sports_node_v6(root, "المكسيك", "كوريا الجنوبية", source_name="Google Sports")
                if obj:
                    break
            if obj:
                lines.append(f"✅ قرأت المباراة: {obj['team1']} {obj['score1']} - {obj['score2']} {obj['team2']}")
                lines.append(f"الحالة: {obj.get('status','')}")
            else:
                names = []
                for node in _walk_json(sr):
                    if isinstance(node, dict):
                        nm = _team_name_from_any_v6(node)
                        if nm and nm not in names:
                            names.append(nm)
                    if len(names) >= 8:
                        break
                if names:
                    lines.append("أسماء ظهرت: " + "، ".join(names[:8]))
                lines.append("⚠️ وصلنا لقوقل لكن لم أقرأ النتيجة من البنية الحالية")
        else:
            lines.append("⚠️ لم يظهر sports_results")
        await msg.edit_text("\n".join(lines[:12]))
    except Exception as e:
        await msg.edit_text(f"❌ فشل فحص قوقل: {str(e)[:220]}")

# ==================== END V6 HOTFIX ====================


def main():
    if not TOKEN:
        raise RuntimeError("ضع توكن البوت في متغير البيئة BOT_TOKEN")
    ensure_flags_assets()
    ensure_design_assets()
    app = ApplicationBuilder().token(TOKEN).build()

    # أساسيات
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"(?i)^/start(?:@\w+)?(?:\s|$)"), start))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:من_انا|معرفي)"), who_am_i))
    app.add_handler(MessageHandler(filters.Document.ALL, remember_last_file))

    # صور وتقارير الفانتزي المعتمدة
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تفعيل_الصور_التلقائية"), admin_only(enable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إيقاف_الصور_التلقائية"), admin_only(disable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ايقاف_الصور_التلقائية"), admin_only(disable_auto_images)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_اليوم"), daily_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_الترتيب"), overall_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_الاساطير"), legends_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صورة_احصائيات"), dashboard_sheet_image_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/صور_الاحصائيات"), all_dashboard_images_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:ملف_الاحصائيات|ملف_الإحصائيات|ملف_احصائيات|ملف_إحصائيات|pdf_الاحصائيات|PDF_الاحصائيات)(?:\s|$)"), statistics_pdf_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/بطاقة"), participant_card_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تقرير_الفترة"), period_report_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعلان_اليوم"), announcement_day_command))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ملخص_اليوم"), summary_day_command))

    # V31 + جدول PDF الجديد — الأوامر الجديدة فقط
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_اليوم2(?:\s|$)"), admin_only(matches_today_v31_clean_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_اليوم(?:\s|$)"), admin_only(matches_today_v31_full_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_الأيام10(?:\s|$)"), admin_only(multi_days_matches10_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_الايام10(?:\s|$)"), admin_only(multi_days_matches10_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_الأيام(?:\s|$)"), admin_only(multi_days_matches_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_الايام(?:\s|$)"), admin_only(multi_days_matches_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_ناقصة(?:\s|$)"), admin_only(fixtures_missing_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مراجعة_مباراة(?:\s|$)"), admin_only(fixtures_review_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات_مجمعة(?:\s|$)"), admin_only(fixtures_combined_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباريات(?:\s|$)"), admin_only(fixtures_command)))
    app.add_handler(CallbackQueryHandler(fixtures_callback, pattern=r"^fx\|"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_only(fixtures_update_text_handler)))

    # استيراد ونسخ
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:استيراد_ملف|استيراد\s+ملف|استيراد|استيراد_اكسل|استيراد_إكسل|استيراد_excel)(?:\s|$)"), admin_only(import_excel_file)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعتماد_استيراد"), admin_only(approve_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إلغاء_استيراد"), admin_only(cancel_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الغاء_استيراد"), admin_only(cancel_import)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نسخة_احتياطية"), admin_only(backup_zip)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/استرجاع_نسخة"), admin_only(restore_backup_zip)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الأيام"), admin_only(clean_days)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الايام"), admin_only(clean_days)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تنظيف_الملفات"), admin_only(clean_temp_files)))

    # الكأس الداخلية محفوظة حسب اعتمادك
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/بدء_الكاس(?:\s|$)"), admin_only(start_cup_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/حالة_الكاس(?:\s|$)"), admin_only(cup_status_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نتائج_الكاس(?:\s|$)"), admin_only(cup_results_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مواجهات_الكاس(?:\s|$)"), admin_only(cup_matches_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إعادة_الكاس_من(?:\s|$)"), admin_only(reset_cup_from_day_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعادة_الكاس_من(?:\s|$)"), admin_only(reset_cup_from_day_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الغاء_الكاس(?:\s|$)"), admin_only(cancel_cup_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إلغاء_الكاس(?:\s|$)"), admin_only(cancel_cup_command)))

    # فانتزي أساسي + إدارة المتسابقين
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:إضافة_متسابق|اضافة_متسابق|إضافه_متسابق|اضافه_متسابق|إضافة\s+متسابق|اضافة\s+متسابق|إضافه\s+متسابق|اضافه\s+متسابق)(?:\s|$)"), admin_only(add_participant_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(?:حذف_متسابق|ازالة_متسابق|إزالة_متسابق|حذف\s+متسابق|ازالة\s+متسابق|إزالة\s+متسابق)(?:\s|$)"), admin_only(remove_participant_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/المتسابقين(?:\s|$)"), admin_only(participants_list_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اضافه(?:\s|$)"), admin_only(add_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعتماد_نتائج(?:\s|$)"), admin_only(approve_results_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/نتائج"), admin_only(results_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/الترتيب_العام"), overall))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ترتيب_نص"), ranking_text))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/?(?:احصائيات|إحصائيات)(?:\s|$)"), dashboard))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/(الأيام|الايام)"), list_days))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/فحص(?:\\s|$)"), inspect_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مشاركين"), participants_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اسطورة"), legend_day))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مقارنة"), compare_days))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_الكل"), admin_only(clear_all)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_يوم"), admin_only(clear_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مسح_نتائج"), admin_only(clear_results)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/استرجاع_آخر"), admin_only(restore_last)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/قفل_يوم"), admin_only(lock_day)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/فتح_يوم"), admin_only(unlock_day)))

    # الأخبار والمتأهلين والمباشر والترتيب
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/فحص_api(?:\s|$)"), admin_only(api_check_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/بحث_قوقل(?:\s|$)"), admin_only(google_search_debug_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/المنتخبات(?:\s|$)"), admin_only(teams_supported_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/المجموعات(?:\s|$)"), admin_only(groups_points_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/قالب_المجموعات(?:\s|$)"), admin_only(groups_template_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/لوحة_المتأهلين(?:\s|$)"), admin_only(qualified_show_board_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إعادة_المتأهلين(?:\s|$)"), admin_only(qualified_reset_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اعادة_المتأهلين(?:\s|$)"), admin_only(qualified_reset_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/حذف_متأهل(?:\s|$)"), admin_only(qualified_remove_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/متأهلين(?:\s|$)"), admin_only(qualified_add_many_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/متأهل(?:\s|$)"), admin_only(qualified_add_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/تأهل(?:\s|$)"), admin_only(qualified_news_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/إقصاء(?:\s|$)"), admin_only(eliminated_news_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/اقصاء(?:\s|$)"), admin_only(eliminated_news_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/عاجل(?:\s|$)"), admin_only(urgent_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/خبر(?:\s|$)"), admin_only(news_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/ترتيب_المجموعات_الان(?:\s|$)"), admin_only(current_groups_now_command)))
    app.add_handler(MessageHandler(filters.TEXT & filters.Regex(r"^/مباشر(?:\s|$)"), admin_only(live_match_command)))
    app.add_handler(CallbackQueryHandler(sports_source_callback, pattern=r"^sportsrc\|"))

    app.run_polling()



# ==================== V8 FINAL PATCH: validated live score + scorers details ====================
# الهدف: منع قراءة التاريخ كأنه نتيجة، تثبيت نتيجة Google/رسمي فقط، ثم جلب الهدافين من التفاصيل.


def _v8_score_int(x):
    try:
        s = str(x).strip()
        # لا نقبل نصوص طويلة أو تواريخ داخل النتيجة
        if not re.fullmatch(r"\d{1,2}", s):
            return None
        return int(s)
    except Exception:
        return None


def _v8_valid_score_obj(obj):
    if not isinstance(obj, dict):
        return False
    a = _v8_score_int(obj.get("score1"))
    b = _v8_score_int(obj.get("score2"))
    if a is None or b is None:
        return False
    # مباريات كرة القدم الطبيعية؛ نرفض أرقام التاريخ مثل 26-06
    if a < 0 or b < 0 or a > 15 or b > 15:
        return False
    t1 = normalize_name(obj.get("team1") or "")
    t2 = normalize_name(obj.get("team2") or "")
    if not t1 or not t2:
        return False
    return True


def _v8_clean_scorer_line(s):
    s = normalize_name(s or "")
    if not s:
        return ""
    # تقليل النصوص الطويلة أو snippets العامة
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 85:
        return ""
    low = s.lower()
    # تجاهل نصوص تاريخ/بطولة/عامّة
    bad_words = ["fifa", "world cup", "كأس العالم", "المجموعة", "group", "video", "highlight", "ملخص", "reuters"]
    if any(w in low for w in bad_words):
        return ""
    # المقبول: فيه دقيقة/هدف/Goal أو اسم قصير مع رقم دقيقة
    if "'" in s or "هدف" in s or "goal" in low or re.search(r"\b\d{1,3}\b", s):
        # نمنع السطور اللي هي أرقام فقط
        if re.fullmatch(r"[\d\s\-/:]+", s):
            return ""
        return s
    return ""


def _v8_sanitize_scorers(items):
    out = []
    for item in items or []:
        s = _v8_clean_scorer_line(item)
        if s and s not in out:
            out.append(s)
    return out[:8]


def _v8_enrich_scorers_from_serp(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return []
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    queries = _serpapi_query_candidates(req1, req2, date_hint=None, source=source)
    # نضيف بحث تفاصيل واضح مع التاريخ إن وجد
    en1 = team_query_name(req1) if 'team_query_name' in globals() else req1
    en2 = team_query_name(req2) if 'team_query_name' in globals() else req2
    extra = [
        f"تفاصيل مباراة {req1} {req2} من سجل",
        f"{en1} {en2} goals scorers FIFA World Cup 2026",
    ]
    if date_hint:
        extra.insert(0, f"تفاصيل مباراة {req1} {req2} {date_hint} من سجل")
    queries = extra + queries
    seen = set()
    for q in queries[:4]:
        q = normalize_name(q)
        if not q or q in seen:
            continue
        seen.add(q)
        try:
            data = serpapi_search_json(q, hl="ar", gl="sa", timeout=7)
        except Exception:
            continue
        sr = data.get("sports_results") if isinstance(data, dict) else None
        roots = []
        if isinstance(sr, dict):
            roots.append(sr)
            for key in ["game_spotlight", "timeline", "events", "scorers", "news_results", "organic_results"]:
                if isinstance(sr.get(key), (dict, list)):
                    roots.append(sr.get(key))
        roots.append(data)
        collected = []
        for root in roots:
            try:
                collected.extend(_collect_scorers_from_serp_node(root))
            except Exception:
                pass
        clean = _v8_sanitize_scorers(collected)
        if clean:
            return clean
    return []


def _v8_enrich_match_details(obj, team1, team2, date_hint=None, preferred_source=None):
    """أضف الهدافين من Google/365/Kooora/ESPN بدون تعطيل النتيجة."""
    if not isinstance(obj, dict):
        return obj
    obj["score1"] = str(_v8_score_int(obj.get("score1")) if _v8_score_int(obj.get("score1")) is not None else obj.get("score1"))
    obj["score2"] = str(_v8_score_int(obj.get("score2")) if _v8_score_int(obj.get("score2")) is not None else obj.get("score2"))
    existing = _v8_sanitize_scorers(obj.get("scorers") or [])
    if existing:
        obj["scorers"] = existing
        obj["scorers_source"] = obj.get("source") or preferred_source or "المصدر"
        return obj
    order = []
    if preferred_source in ["365", "kooora", "google"]:
        order.append(preferred_source)
    order += ["google", "365", "kooora"]
    seen = set()
    for source in order:
        if source in seen:
            continue
        seen.add(source)
        scorers = _v8_enrich_scorers_from_serp(team1, team2, date_hint=date_hint, source=source)
        if scorers:
            obj["scorers"] = scorers
            obj["scorers_source"] = {"google":"Google Sports", "365":"365Scores", "kooora":"Kooora"}.get(source, source)
            return obj
    # ESPN غالبًا عنده details إذا المباراة موجودة
    try:
        espn_obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
        scorers = _v8_sanitize_scorers((espn_obj or {}).get("scorers") or [])
        if scorers:
            obj["scorers"] = scorers
            obj["scorers_source"] = "ESPN"
            return obj
    except Exception:
        pass
    obj["scorers"] = []
    return obj


def _v8_parse_serp_sports_root(root, req1, req2, source_name="Google Sports"):
    """قراءة صارمة: لا تقبل أرقام تواريخ كأنها نتيجة."""
    if not isinstance(root, (dict, list)):
        return None
    # 1) اجمع scores من team dicts
    try:
        pair = _collect_scores_by_team_v6(root, req1, req2)
        if pair:
            pair.update({
                "status": _status_from_serp_v6(root),
                "minute": "",
                "scorers": _v8_sanitize_scorers(_collect_scorers_from_serp_node(root) if '_collect_scorers_from_serp_node' in globals() else []),
                "source": source_name,
            })
            if _v8_valid_score_obj(pair):
                return pair
    except Exception:
        pass
    # 2) جرّب parser القديم لكن مع فلترة صارمة
    try:
        obj = _parse_serp_sports_node(root, req1, req2)
        if obj:
            obj["source"] = source_name
            obj["scorers"] = _v8_sanitize_scorers(obj.get("scorers") or [])
            if _v8_valid_score_obj(obj):
                return obj
    except Exception:
        pass
    # 3) snippets فقط لو النتيجة منطقية جدًا وبالقرب من الفريقين
    try:
        blob = json.dumps(root, ensure_ascii=False)
        if _blob_has_both_teams(blob, req1, req2):
            s = _score_from_text_near_teams(blob, req1, req2) if '_score_from_text_near_teams' in globals() else None
            if s:
                obj = {
                    "team1": canonical_team_name(req1) or req1,
                    "team2": canonical_team_name(req2) or req2,
                    "score1": s.get("score1"),
                    "score2": s.get("score2"),
                    "status": s.get("status") or _status_from_serp_v6(root),
                    "minute": "",
                    "scorers": _v8_sanitize_scorers(_collect_scorers_from_serp_node(root) if '_collect_scorers_from_serp_node' in globals() else []),
                    "source": source_name,
                }
                if _v8_valid_score_obj(obj):
                    return obj
    except Exception:
        pass
    return None


def _serpapi_query_candidates(team1, team2, date_hint=None, source="google"):
    """V8: نخلي البحث العام قبل التاريخ حتى لا يلتقط 26-06 كأهداف."""
    ar1, ar2 = canonical_team_name(team1) or normalize_name(team1), canonical_team_name(team2) or normalize_name(team2)
    en1 = team_query_name(ar1) if 'team_query_name' in globals() else ar1
    en2 = team_query_name(ar2) if 'team_query_name' in globals() else ar2
    if source == "365":
        base = [
            f"site:365scores.com {ar1} {ar2} كأس العالم 2026",
            f"site:365scores.com {en1} {en2} FIFA World Cup 2026 score",
            f"365Scores {ar1} {ar2} النتيجة من سجل",
        ]
    elif source == "kooora":
        base = [
            f"site:kooora.com {ar1} {ar2} كأس العالم 2026",
            f"site:kooora.com {en1} {en2} FIFA World Cup 2026 score",
            f"كورة {ar1} {ar2} النتيجة من سجل",
        ]
    else:
        base = [
            f"مباراة {ar1} {ar2}",
            f"{en1} vs {en2} FIFA World Cup 2026",
            f"{en1} {en2} score",
        ]
    queries = list(base)
    if date_hint:
        queries += [f"{q} {date_hint}" for q in base]
    out = []
    for q in queries:
        q = normalize_name(q)
        if q and q not in out:
            out.append(q)
    return out[:5]


def _fetch_match_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    source_name = {"google": "Google Sports", "365": "365Scores عبر Google", "kooora": "Kooora عبر Google"}.get(source, "Google")
    attempts = []
    for q in _serpapi_query_candidates(req1, req2, date_hint, source=source):
        if source == "google":
            attempts.append((q, "ar", "sa"))
            attempts.append((q, "en", "us"))
        else:
            attempts.append((q, "ar", "sa"))
            attempts.append((q, "en", "us"))
    seen = set()
    for q, hl, gl in attempts[:6]:
        key = (q, hl, gl)
        if key in seen:
            continue
        seen.add(key)
        try:
            data = serpapi_search_json(q, hl=hl, gl=gl, timeout=7)
        except Exception:
            continue
        roots = []
        sr = data.get("sports_results") if isinstance(data, dict) else None
        if isinstance(sr, dict):
            roots.append(sr)
            for k in ["game_spotlight", "games", "matches", "scoreboard"]:
                if isinstance(sr.get(k), (dict, list)):
                    roots.append(sr.get(k))
        roots.append(data)
        for root in roots:
            obj = _v8_parse_serp_sports_root(root, req1, req2, source_name=source_name)
            if obj and _v8_valid_score_obj(obj):
                return _v8_enrich_match_details(obj, req1, req2, date_hint=date_hint, preferred_source=source)
    return None


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")


def fetch_match_from_365(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="365")


def fetch_match_from_kooora(team1, team2, date_hint=None):
    return _fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="kooora")


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    """V8: النتيجة المؤكدة أولًا، ثم التفاصيل والهدافين."""
    norm = _norm_source_mode(mode)
    # المصدر المختار 365/كورة: نجربه، لكن إذا أعطى نتيجة غير مؤكدة نعتمد قوقل/رسمي ونستخدمه للتفاصيل فقط.
    sequence = _source_mode_sequence(norm)
    fallback_primary = []
    if norm in ["365", "kooora"]:
        fallback_primary = ["google", "espn", "api"]
    for src in sequence + fallback_primary:
        try:
            if src == "google":
                obj = fetch_match_from_serpapi(team1, team2, date_hint=date_hint)
            elif src == "365":
                obj = fetch_match_from_365(team1, team2, date_hint=date_hint)
            elif src == "kooora":
                obj = fetch_match_from_kooora(team1, team2, date_hint=date_hint)
            elif src == "api":
                obj = fetch_match_from_api_football(team1, team2, date_hint=date_hint)
            elif src == "fifa":
                obj = fetch_match_from_fifa(team1, team2)
            else:
                obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
            if obj and _v8_valid_score_obj(obj):
                obj["source"] = obj.get("source") or mode_label_ar(src)
                if not obj.get("status"):
                    obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
                # لو المصدر المختار 365/كورة لكن اعتمدنا مصدر آخر، اجلب التفاصيل منه
                pref = norm if norm in ["365", "kooora"] else src
                return _v8_enrich_match_details(obj, team1, team2, date_hint=date_hint, preferred_source=pref)
        except Exception:
            continue
    return None


def build_live_caption(match, mode_label="رسمي"):
    lines = [f"{match['team1']} {match['score1']} - {match['score2']} {match['team2']}"]
    status_line = match.get("minute") or match.get("status") or ""
    if match.get("status") and match.get("minute") and match.get("status") not in status_line:
        status_line = f"{match['status']} — {match['minute']}"
    if status_line:
        lines.append(status_line)
    lines.append("")
    lines.append("⚽ الهدافون:")
    scorers = _v8_sanitize_scorers(match.get("scorers") or [])
    if scorers:
        for s in scorers[:8]:
            lines.append(f"- {s}")
        if match.get("scorers_source"):
            lines.append(f"تفاصيل الأهداف: {match.get('scorers_source')}")
    else:
        lines.append("غير متوفرة حاليًا من المصادر")
    lines.append("")
    src = match.get("source") or mode_label
    lines.append(f"المصدر: {mode_label} ({src})")
    return "\n".join(lines)


def render_live_match_card(match, mode_label="رسمي"):
    ensure_generated_dir()
    width, height = 1200, 1350
    img = _style4_clean_background(width, height)
    draw = ImageDraw.Draw(img)
    draw_text(draw, (width//2, 100), "مباشر الآن", get_font(62), fill="#FFFFFF")
    draw_text(draw, (width//2, 165), f"المصدر: {mode_label}", get_font(28), fill="#FDE68A")

    rounded_rect(draw, (80, 235, width-80, 840), radius=40, fill="#07132FDD", outline="#38BDF855", width=2)
    paste_flag(img, match["team1"], (155, 315, 315, 425))
    paste_flag(img, match["team2"], (width-315, 315, width-155, 425))
    draw_text(draw, (250, 485), match["team1"], get_font(38), fill="#FFFFFF", max_width=300)
    draw_text(draw, (width-250, 485), match["team2"], get_font(38), fill="#FFFFFF", max_width=300)
    draw_text(draw, (width//2, 430), f"{match['score1']} - {match['score2']}", get_font(104), fill="#FFFFFF")
    rounded_rect(draw, (width//2-190, 530, width//2+190, 602), radius=22, fill="#FBBF24", outline="#FFFFFF33", width=1)
    status_line = match.get("minute") or match.get("status") or ""
    if match.get("status") and match.get("minute") and match.get("status") not in status_line:
        status_line = f"{match['status']} — {match['minute']}"
    draw_text(draw, (width//2, 566), status_line or "متابعة مباشرة", get_font(30), fill="#061633", max_width=330)

    scorers = _v8_sanitize_scorers(match.get("scorers") or [])
    draw_text(draw, (width//2, 655), "الهدافون", get_font(36), fill="#FDE68A")
    y = 710
    if scorers:
        for item in scorers[:6]:
            rounded_rect(draw, (120, y-24, width-120, y+34), radius=18, fill="#0B1E46", outline="#FFFFFF22", width=1)
            draw_text(draw, (width//2, y+5), item, get_font(27), fill="#FFFFFF", max_width=880)
            y += 68
    else:
        rounded_rect(draw, (160, y-25, width-160, y+40), radius=20, fill="#0B1E46", outline="#FFFFFF22", width=1)
        draw_text(draw, (width//2, y+7), "تفاصيل الهدافين غير متوفرة حاليًا", get_font(28), fill="#FFFFFF", max_width=820)

    footer_event(draw, width, height)
    out = os.path.join(GENERATED_DIR, f"live_{_safe_filename(match['team1'])}_{_safe_filename(match['team2'])}.png")
    img.save(out, quality=95)
    return out


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    query = " ".join(body).strip() or "مباراة المكسيك كوريا الجنوبية"
    msg = await update.message.reply_text(f"⏳ أفحص قوقل: {query}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(serpapi_search_json, query, "ar", "sa", 8), timeout=12)
        sr = data.get("sports_results") if isinstance(data, dict) else None
        lines = ["نتيجة فحص قوقل:", f"query: {query}"]
        if isinstance(sr, dict):
            lines.append("✅ sports_results موجود")
            lines.append("مفاتيح: " + "، ".join(list(sr.keys())[:12]))
            # نحاول قراءة المباراة بين أول فريقين معروفين في النص، وإلا المكسيك/كوريا للتجربة.
            req1, req2 = "المكسيك", "كوريا الجنوبية"
            txt = normalize_name(query)
            found = []
            for tm in WORLD_CUP_TEAMS:
                if _blob_contains_team(txt, tm):
                    found.append(tm)
                if len(found) >= 2:
                    break
            if len(found) >= 2:
                req1, req2 = found[0], found[1]
            obj = _v8_parse_serp_sports_root(sr, req1, req2, source_name="Google Sports")
            if obj:
                lines.append(f"✅ قرأت المباراة: {obj['team1']} {obj['score1']} - {obj['score2']} {obj['team2']}")
                lines.append(f"الحالة: {obj.get('status','')}")
                sc = _v8_sanitize_scorers(obj.get("scorers") or [])
                if sc:
                    lines.append("الهدافون: " + "، ".join(sc[:4]))
            else:
                lines.append("⚠️ وصلنا لقوقل لكن لم أقرأ نتيجة مؤكدة من البنية الحالية")
        else:
            lines.append("⚠️ لم يظهر sports_results")
        await msg.edit_text("\n".join(lines[:14]))
    except Exception as e:
        await msg.edit_text(f"❌ فشل فحص قوقل: {str(e)[:220]}")


# ==================== V9 FINAL LIVE IMAGE + GOALS + SOURCES PATCH ====================
# اعتماد V9:
# - /مباشر يرسل صورة بخلفية التمثال وليس نص فقط.
# - النتيجة لا تُقبل إذا كانت تاريخًا أو رقمًا غير منطقيًا مثل 26-06.
# - الهدافون لا يُقبلون من نصوص عشوائية/تعليقات؛ فقط صيغ موثوقة فيها دقيقة هدف أو dict goal.
# - 365 وكورة للتفاصيل/الهدافين، ولا يغيرون نتيجة قوقل/رسمي إذا كانت مشكوك فيها.
# - يوضح المصدر المطلوب والمصدر الفعلي.

BAD_SCORER_HINTS_V9 = [
    "google sports", "google", "sports_results", "video", "highlight", "carousel", "ملخص", "ملخصات",
    "مباراة", "ولا اروع", "أروع", "اعجبتني", "أعجبتني", "كثير", "2h", "1h", "ago", "قبل",
    "reuters", "النتيجة", "انتهت", "مباشر", "اليوم", "غداً", "غدا", "أمس", "امس", "حسب", "المصدر",
    "http", "www", "site:", "kooora", "365scores", "espn", "serpapi", "تشكيلات", "إحصاءات", "الاحصاءات",
]


def _v9_int_score(v):
    try:
        s = str(v).strip()
        # لا تقبل رقم فيه سنة أو تاريخ أو وقت
        if re.search(r"20\d{2}|/|:", s):
            return None
        if not re.fullmatch(r"\d{1,2}", s):
            return None
        n = int(s)
        # نتائج كرة القدم الواقعية هنا، ونمنع 26/06 من التاريخ
        if 0 <= n <= 15:
            return n
    except Exception:
        pass
    return None


def _v9_valid_score_obj(obj):
    if not isinstance(obj, dict):
        return False
    a = _v9_int_score(obj.get("score1"))
    b = _v9_int_score(obj.get("score2"))
    if a is None or b is None:
        return False
    # لا تقبل 26-06 أو 18-06 أو أرقام شبيهة بالتاريخ
    if a > 15 or b > 15:
        return False
    obj["score1"] = str(a)
    obj["score2"] = str(b)
    return True


def _v9_clean_source_name(src):
    src = normalize_name(src or "")
    if not src:
        return "غير محدد"
    return src.replace(" عبر Google", "")


def _v9_status_line(obj):
    status = normalize_name((obj or {}).get("status") or "")
    minute = normalize_name((obj or {}).get("minute") or "")
    if minute and minute not in status:
        # لو الدقيقة شكلها رقم فقط خله مع علامة دقيقة
        if re.fullmatch(r"\d{1,3}", minute):
            minute = minute + "'"
        return f"{status} — {minute}" if status else minute
    return status or minute or "متابعة مباشرة"


def _v9_is_strict_goal_text(s):
    s = normalize_name(s)
    if not s or len(s) > 70:
        return False
    low = s.lower()
    if any(bad in low for bad in BAD_SCORER_HINTS_V9):
        return False
    # لازم تظهر دقيقة هدف: 50' أو 50’ أو 50 دقيقة، أو كلمة هدف مع اسم قصير
    has_minute = bool(re.search(r"(?:^|\s)\d{1,3}\s*(?:['’′]|د|دقيقة|min|m)(?:\s|$)", s, re.I))
    has_goal_word = bool(re.search(r"\b(goal|هدف|penalty|ركلة جزاء|own goal|هدف عكسي)\b", s, re.I))
    # لا نأخذ نص فيه أرقام كثيرة/سنوات/تواريخ
    if re.search(r"20\d{2}|\d{1,2}/\d{1,2}|\d+h", s):
        return False
    return has_minute or (has_goal_word and len(s.split()) <= 7)


def _v9_normalize_goal_text(s):
    s = clean_draw_text(normalize_name(s))
    s = s.replace("’", "'").replace("′", "'")
    s = re.sub(r"\s+", " ", s).strip(" -—•،,.")
    # اختصر "Goal - Luis Romo 50'" إلى شكل مفهوم إن أمكن
    s = re.sub(r"^(goal|هدف)\s*[-:–]?\s*", "", s, flags=re.I)
    # 50 min -> 50'
    s = re.sub(r"\b(\d{1,3})\s*(?:min|m|دقيقة|د)\b", r"\1'", s, flags=re.I)
    return s


def _v9_sanitize_scorers(items):
    clean = []
    for item in items or []:
        if isinstance(item, dict):
            name = normalize_name(item.get("player") or item.get("name") or item.get("title") or item.get("athlete") or "")
            minute = normalize_name(item.get("minute") or item.get("time") or item.get("elapsed") or item.get("displayTime") or "")
            detail = normalize_name(item.get("type") or item.get("detail") or item.get("event") or item.get("description") or "")
            if name and (minute or "goal" in detail.lower() or "هدف" in detail):
                if minute and not str(minute).endswith("'"):
                    minute = re.sub(r"\D+", "", minute) or minute
                    if re.fullmatch(r"\d{1,3}", minute):
                        minute += "'"
                s = f"{name} {minute}".strip()
            else:
                s = ""
        else:
            s = str(item or "")
        s = _v9_normalize_goal_text(s)
        if _v9_is_strict_goal_text(s) and s not in clean:
            clean.append(s)
        if len(clean) >= 8:
            break
    return clean


def _v9_collect_goal_details_from_json(root):
    out = []
    try:
        for node in _walk_json(root):
            if isinstance(node, dict):
                # أوضح حالة: كائن هدف/حدث فيه لاعب ودقيقة
                detail = normalize_name(node.get("type") or node.get("event") or node.get("detail") or node.get("description") or node.get("play") or "")
                name = node.get("player") or node.get("name") or node.get("title") or node.get("athlete")
                minute = node.get("minute") or node.get("time") or node.get("elapsed") or node.get("displayTime")
                if name and (minute or "goal" in detail.lower() or "هدف" in detail):
                    out.append({"name": name, "minute": minute, "detail": detail})
                # بعض المصادر تضع أحداث الأهداف كنصوص قصيرة
                for k in ["text", "summary", "subtitle", "description", "detail"]:
                    v = node.get(k)
                    if isinstance(v, str) and _v9_is_strict_goal_text(v):
                        out.append(v)
            elif isinstance(node, str) and _v9_is_strict_goal_text(node):
                out.append(node)
    except Exception:
        pass
    return _v9_sanitize_scorers(out)


def _v9_serp_goal_queries(team1, team2, date_hint=None, source="google"):
    c1 = canonical_team_name(team1) or normalize_name(team1)
    c2 = canonical_team_name(team2) or normalize_name(team2)
    en1 = (TEAM_SEARCH_EN.get(c1) or [c1])[0]
    en2 = (TEAM_SEARCH_EN.get(c2) or [c2])[0]
    base_ar = f"مباراة {c1} {c2} الهدافين"
    base_en = f"{en1} {en2} scorers goals"
    if date_hint:
        base_ar += f" {date_hint}"
        base_en += f" {date_hint}"
    if source == "365":
        return [f"site:365scores.com {base_ar}", f"site:365scores.com {base_en}"]
    if source == "kooora":
        return [f"site:kooora.com {base_ar}", f"site:kooora.com {base_en}"]
    return [base_ar, base_en]


def _v9_fetch_scorers_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return []
    for q in _v9_serp_goal_queries(team1, team2, date_hint, source=source)[:2]:
        for hl, gl in [("ar", "sa"), ("en", "us")]:
            try:
                data = serpapi_search_json(q, hl=hl, gl=gl, timeout=7)
            except Exception:
                continue
            sr = data.get("sports_results") if isinstance(data, dict) else None
            # فقط JSON منظم أو نصوص فيها دقيقة هدف بوضوح
            scorers = _v9_collect_goal_details_from_json(sr or data)
            if scorers:
                return scorers
    return []


def _v9_fetch_scorers_from_espn(team1, team2, date_hint=None):
    try:
        obj = fetch_match_from_espn(team1, team2, date_hint=date_hint)
        return _v9_sanitize_scorers((obj or {}).get("scorers") or [])
    except Exception:
        return []


def _v9_fetch_goal_details(team1, team2, date_hint=None, preferred_source="google"):
    # 365 وكورة للتفاصيل أولاً لو المستخدم اختارهم، ثم ESPN/Google.
    sequence = []
    if preferred_source == "365":
        sequence = ["365", "kooora", "espn", "google"]
    elif preferred_source == "kooora":
        sequence = ["kooora", "365", "espn", "google"]
    else:
        sequence = ["espn", "google", "365", "kooora"]
    for src in sequence:
        if src == "espn":
            scorers = _v9_fetch_scorers_from_espn(team1, team2, date_hint)
        else:
            scorers = _v9_fetch_scorers_from_serp_source(team1, team2, date_hint, source=src)
        scorers = _v9_sanitize_scorers(scorers)
        if scorers:
            return scorers, mode_label_ar(src) if src in ["365", "kooora"] else ("ESPN" if src == "espn" else "Google Sports")
    return [], ""


def _v9_fetch_primary_from_source(team1, team2, src, date_hint=None):
    if src == "google":
        return fetch_match_from_serpapi(team1, team2, date_hint=date_hint)
    if src == "365":
        return fetch_match_from_365(team1, team2, date_hint=date_hint)
    if src == "kooora":
        return fetch_match_from_kooora(team1, team2, date_hint=date_hint)
    if src == "api":
        return fetch_match_from_api_football(team1, team2, date_hint=date_hint)
    if src == "fifa":
        return fetch_match_from_fifa(team1, team2)
    return fetch_match_from_espn(team1, team2, date_hint=date_hint)


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    """V9: نتيجة مؤكدة + تفاصيل هدافين موثوقة + منع التاريخ والكاش القديم."""
    norm = _norm_source_mode(mode)
    requested_label = mode_label_ar(norm)
    # 365/كورة نحاولهم، لكن لو النتيجة مشكوك فيها نثبت من Google/ESPN.
    seq = _source_mode_sequence(norm)
    if norm in ["365", "kooora"]:
        seq = [norm, "google", "espn", "api"]
    elif norm == "latest":
        seq = ["google", "espn", "365", "kooora", "api"]
    elif norm == "official":
        seq = ["espn", "api", "google"]
    elif norm == "google":
        seq = ["google", "espn"]

    seen = set()
    for src in seq:
        if src in seen:
            continue
        seen.add(src)
        try:
            obj = _v9_fetch_primary_from_source(team1, team2, src, date_hint=date_hint)
        except Exception:
            continue
        if not obj or not _v9_valid_score_obj(obj):
            continue
        obj["source"] = _v9_clean_source_name(obj.get("source") or ("ESPN" if src == "espn" else mode_label_ar(src)))
        obj["requested_source"] = requested_label
        obj["actual_source"] = obj["source"]
        if not obj.get("status"):
            obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
        # الهدافين من المصدر نفسه أولاً، لكن بفلتر صارم.
        scorers = _v9_sanitize_scorers(obj.get("scorers") or [])
        details_source = obj.get("source") if scorers else ""
        if not scorers:
            pref = norm if norm in ["365", "kooora"] else src
            scorers, details_source = _v9_fetch_goal_details(team1, team2, date_hint=date_hint, preferred_source=pref)
        obj["scorers"] = scorers
        obj["goals_source"] = details_source or "غير متوفر"
        return obj
    return None


def _v9_live_bg(width, height):
    # خلفية التمثال المعتمدة إن وجدت، وإلا fallback نظيف.
    try:
        img, draw = _v31_load_bg(V31_CLEAN_BG, width, height)
        # تغميق خفيف حتى النص يطلع واضح
        overlay = Image.new("RGBA", (width, height), (0, 6, 20, 80))
        img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
        return img, ImageDraw.Draw(img)
    except Exception:
        try:
            img = _style4_clean_background(width, height)
            return img, ImageDraw.Draw(img)
        except Exception:
            img = Image.new("RGB", (width, height), "#061633")
            return img, ImageDraw.Draw(img)


def _v9_draw_flag_fit(img, team, box):
    try:
        # علم كبير ومناسب داخل خانته
        _v31_paste_flag(img, team, box)
    except Exception:
        try:
            paste_flag(img, team, box)
        except Exception:
            pass


def render_live_match_card(match, mode_label="رسمي"):
    ensure_generated_dir()
    width, height = 1200, 1500
    img, draw = _v9_live_bg(width, height)

    # عنوان
    draw_text(draw, (width//2, 88), "MONDIAL AL MASEEF 2026", _v31_latin_font(34), fill="#FFFFFF", max_width=760)
    draw_text(draw, (width//2, 160), "LIVE MATCH", _v31_latin_font(74), fill="#FFFFFF", max_width=900)
    draw_text(draw, (width//2, 222), "متابعة المباراة", get_font(34), fill="#FDE68A", max_width=720)

    # الكرت الرئيسي
    main = (70, 290, width-70, 1020)
    try:
        _draw_glow_card(draw, main, 42, "#061633E8", "#38BDF8", width=4)
    except Exception:
        rounded_rect(draw, main, radius=42, fill="#061633", outline="#38BDF8", width=4)

    # المنتخبات والأعلام
    flag_y1, flag_y2 = 365, 485
    left_flag = (130, flag_y1, 310, flag_y2)
    right_flag = (width-310, flag_y1, width-130, flag_y2)
    rounded_rect(draw, (left_flag[0]-12, left_flag[1]-10, left_flag[2]+12, left_flag[3]+10), radius=28, fill="#FFFFFFE8", outline="#FFFFFF55", width=2)
    rounded_rect(draw, (right_flag[0]-12, right_flag[1]-10, right_flag[2]+12, right_flag[3]+10), radius=28, fill="#FFFFFFE8", outline="#FFFFFF55", width=2)
    _v9_draw_flag_fit(img, match.get("team1", ""), left_flag)
    _v9_draw_flag_fit(img, match.get("team2", ""), right_flag)
    draw_text(draw, (220, 540), match.get("team1", ""), get_font(40), fill="#FFFFFF", max_width=330)
    draw_text(draw, (width-220, 540), match.get("team2", ""), get_font(40), fill="#FFFFFF", max_width=330)

    # النتيجة الكبيرة
    score = f"{match.get('score1','0')} - {match.get('score2','0')}"
    draw_text(draw, (width//2, 463), score, _v31_latin_font(116), fill="#FFFFFF", max_width=360)

    # الحالة
    status = _v9_status_line(match)
    rounded_rect(draw, (width//2-230, 590, width//2+230, 662), radius=24, fill="#FBBF24", outline="#FFFFFF55", width=2)
    draw_text(draw, (width//2, 626), status, get_font(30), fill="#061633", max_width=410)

    # الهدافين
    scorers = _v9_sanitize_scorers(match.get("scorers") or [])
    details_top = 720
    draw_text(draw, (width//2, details_top), "الهدافون", get_font(40), fill="#FDE68A", max_width=480)
    y = details_top + 70
    if scorers:
        for s in scorers[:5]:
            rounded_rect(draw, (130, y-28, width-130, y+38), radius=22, fill="#07132FEA", outline="#FFFFFF33", width=2)
            draw_text(draw, (width//2, y+5), f"⚽ {s}", get_font(30), fill="#FFFFFF", max_width=880)
            y += 78
    else:
        rounded_rect(draw, (150, y-30, width-150, y+42), radius=24, fill="#07132FEA", outline="#FFFFFF33", width=2)
        draw_text(draw, (width//2, y+6), "تفاصيل الهدافين غير متوفرة حاليًا", get_font(30), fill="#FFFFFF", max_width=820)
        y += 86

    # مصادر
    actual = match.get("actual_source") or match.get("source") or mode_label
    requested = match.get("requested_source") or mode_label
    goals_src = match.get("goals_source") or "غير متوفر"
    src_text = f"المصدر الفعلي: {actual}"
    if requested and requested != actual:
        src_text += f"  |  المطلوب: {requested}"
    draw_text(draw, (width//2, 1115), src_text, get_font(28), fill="#D7E7FF", max_width=950)
    draw_text(draw, (width//2, 1160), f"تفاصيل الأهداف: {goals_src}", get_font(26), fill="#FDE68A", max_width=860)

    # تذييل
    draw.line((230, height-160, width-230, height-160), fill="#FFFFFF66", width=2)
    draw_text(draw, (width//2, height-120), "المصيف يضعكم بالحدث", get_font(34), fill="#FDE68A", max_width=650)

    out = os.path.join(GENERATED_DIR, f"live_v9_{_safe_filename(match.get('team1','team1'))}_{_safe_filename(match.get('team2','team2'))}.png")
    try:
        img.save(out, quality=96)
    except TypeError:
        img.save(out)
    return out


def build_live_caption(match, mode_label="رسمي"):
    lines = [f"{match.get('team1','')} {match.get('score1','0')} - {match.get('score2','0')} {match.get('team2','')}"]
    st = _v9_status_line(match)
    if st:
        lines.append(st)
    scorers = _v9_sanitize_scorers(match.get("scorers") or [])
    lines.append("")
    lines.append("⚽ الهدافون:")
    if scorers:
        lines.extend([f"- {s}" for s in scorers[:5]])
    else:
        lines.append("غير متوفرين حاليًا من المصادر")
    actual = match.get("actual_source") or match.get("source") or mode_label
    requested = match.get("requested_source") or mode_label
    lines.append("")
    if requested and requested != actual:
        lines.append(f"المصدر المطلوب: {requested}")
    lines.append(f"المصدر الفعلي: {actual}")
    if match.get("goals_source"):
        lines.append(f"تفاصيل الأهداف: {match.get('goals_source')}")
    return "\n".join(lines)


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text if getattr(update, 'message', None) else "")
    if mode_override:
        mode = _norm_source_mode(mode_override)
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر المكسيك * كوريا الجنوبية * قوقل\nأو:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026 * قوقل")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, date_hint), timeout=22)
    except Exception:
        data = None
    if not data:
        await wait.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\n"
            f"مباراة: {team1} × {team2}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") +
            "\nاختر مصدر آخر أو جرّب الأمر:\n/بحث_قوقل مباراة المكسيك كوريا الجنوبية",
            reply_markup=kb,
        )
        return
    path = None
    try:
        path = render_live_match_card(data, mode_label_ar(mode))
    except Exception:
        path = None
    if path and os.path.exists(path):
        try:
            await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
            try:
                await wait.delete()
            except Exception:
                pass
            return
        except Exception as e:
            try:
                await wait.edit_text(f"تم جلب المباراة لكن تعذر إرسال الصورة ❌\n{str(e)[:120]}\n\n" + build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)
                return
            except Exception:
                pass
    await wait.edit_text(build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    kb = source_keyboard(context, payload)
    if kind == "standings":
        msg = await query.message.reply_text(f"⏳ جاري فحص ترتيب المجموعات من مصدر {mode_label_ar(mode)}...")
        groups, src = fetch_current_groups(mode)
        if not groups:
            await msg.edit_text(_source_help_text("standings", mode), reply_markup=kb)
            return
        path = create_all_groups_image(groups)
        try:
            await msg.delete()
        except Exception:
            pass
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر الحالي: {mode_label_ar(mode)} ({src})", kb)
        await query.message.reply_text(build_groups_text(groups, f"{mode_label_ar(mode)} ({src})"))
        return
    if kind == "live":
        team1, team2 = payload.get("team1"), payload.get("team2")
        msg = await query.message.reply_text(f"⏳ جاري البحث عن {team1} × {team2} من مصدر {mode_label_ar(mode)}...")
        try:
            data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, payload.get("date_hint")), timeout=22)
        except Exception:
            data = None
        if not data:
            await msg.edit_text(f"تعذر جلب مباراة {team1} × {team2} من مصدر {mode_label_ar(mode)} ❌\n\nاختر مصدر آخر:", reply_markup=kb)
            return
        path = None
        try:
            path = render_live_match_card(data, mode_label_ar(mode))
        except Exception:
            path = None
        if path and os.path.exists(path):
            try:
                await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
                try:
                    await msg.delete()
                except Exception:
                    pass
                return
            except Exception as e:
                await msg.edit_text(f"تم جلب المباراة لكن تعذر إرسال الصورة ❌\n{str(e)[:120]}\n\n" + build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)
                return
        await msg.edit_text(build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")

# ==================== END V9 FINAL PATCH ====================

# ==================== END V8 FINAL PATCH ====================


# ==================== V10 FINAL PATCH: Google standings + live clean image/caption ====================
# اعتماد V10:
# - ترتيب المجموعات يبدأ باستعلام "ترتيبات كأس العالم" من Google Sports/SerpApi.
# - المصدر لا يظهر داخل صورة /مباشر، ويظهر فقط بالكابشن.
# - محاولة استخراج الهدافين ورابط الملخص من كرت قوقل، ثم 365/كورة كاحتياط.
# - أسفل تصاميم المجموعات: المصيف يضعكم بالحدث.

V10_GOOGLE_STANDINGS_QUERIES = [
    "ترتيبات كأس العالم",
    "ترتيب مجموعات كأس العالم 2026",
    "FIFA World Cup 2026 standings",
    "World Cup 2026 group standings",
]


def _v10_int_any(v, default=0):
    try:
        if isinstance(v, bool):
            return default
        if isinstance(v, (int, float)):
            return int(v)
        s = normalize_name(str(v or ""))
        s = s.replace("+", "")
        m = re.search(r"-?\d+", s)
        return int(m.group(0)) if m else default
    except Exception:
        return default


def _v10_group_letter(title):
    t = normalize_name(title or "")
    ar_map = {"أ":"A", "ا":"A", "ب":"B", "ج":"C", "د":"D", "هـ":"E", "ه":"E", "و":"F", "ز":"G", "ح":"H", "ط":"I", "ي":"J", "ك":"K", "ل":"L"}
    m = re.search(r"(?:Group|المجموعة|مجموعة)\s*([A-Lأابجدهـوزحطيكل])", t, re.I)
    if m:
        ch = m.group(1).upper()
        return ar_map.get(ch, ch)
    # أحيانًا يكون العنوان فقط: المجموعة أ أو Group A داخل حقل name/title
    for ar, en in ar_map.items():
        if re.search(rf"(?:المجموعة|مجموعة)\s*{re.escape(ar)}\b", t):
            return en
    return ""


def _v10_group_title(title, fallback=""):
    letter = _v10_group_letter(title) or _v10_group_letter(fallback)
    if letter:
        # نخلي الحروف العربية في التصميم حسب الموجود في بقية البوت
        arabic = {"A":"أ", "B":"ب", "C":"ج", "D":"د", "E":"هـ", "F":"و", "G":"ز", "H":"ح", "I":"ط", "J":"ي", "K":"ك", "L":"ل"}.get(letter, letter)
        return f"المجموعة {arabic}"
    return normalize_name(title or fallback or "المجموعة")


def _v10_get_nested_team_name(node):
    if not isinstance(node, dict):
        return ""
    direct_keys = ["name", "team", "title", "team_name", "participant", "competitor", "displayName", "shortName"]
    for k in direct_keys:
        v = node.get(k)
        if isinstance(v, str) and normalize_name(v):
            return normalize_name(v)
        if isinstance(v, dict):
            for kk in ["name", "displayName", "shortName", "title"]:
                if isinstance(v.get(kk), str) and normalize_name(v.get(kk)):
                    return normalize_name(v.get(kk))
    for k in ["team", "competitor", "participant", "club", "country"]:
        v = node.get(k)
        if isinstance(v, dict):
            nm = _v10_get_nested_team_name(v)
            if nm:
                return nm
    return ""


def _v10_stats_map_from_any(node):
    """يحاول تحويل أي stats/list/labels من SerpApi/Google إلى قاموس بسيط."""
    stats = {}
    if not isinstance(node, dict):
        return stats
    # القيم المباشرة
    for k, v in node.items():
        lk = normalize_name(str(k)).lower()
        if isinstance(v, (str, int, float)):
            stats[lk] = v
    # قوائم stats الشائعة
    for key in ["stats", "statistics", "columns", "values", "details"]:
        val = node.get(key)
        if isinstance(val, list):
            for item in val:
                if isinstance(item, dict):
                    label = normalize_name(item.get("name") or item.get("label") or item.get("title") or item.get("displayName") or item.get("type") or "").lower()
                    value = item.get("value") if "value" in item else item.get("displayValue") if "displayValue" in item else item.get("rank")
                    if label:
                        stats[label] = value
                elif isinstance(item, str):
                    pass
        elif isinstance(val, dict):
            for kk, vv in val.items():
                stats[normalize_name(str(kk)).lower()] = vv
    return stats


def _v10_pick_stat(stats, keys, default=0):
    # exact then fuzzy Arabic/English
    for k in keys:
        if k in stats:
            return _v10_int_any(stats.get(k), default)
    for sk, sv in stats.items():
        for k in keys:
            if k and (k in sk or sk in k):
                return _v10_int_any(sv, default)
    return default


def _v10_parse_standing_row(node):
    if not isinstance(node, dict):
        return None
    team_raw = _v10_get_nested_team_name(node)
    team = canonical_team_name(team_raw) or normalize_name(team_raw)
    if not team or team not in set(WORLD_CUP_TEAMS):
        return None
    stats = _v10_stats_map_from_any(node)
    played = _v10_pick_stat(stats, [
        "played", "matches played", "mp", "p", "games played", "لعب", "ل", "لعبت"
    ], 0)
    pts = _v10_pick_stat(stats, [
        "points", "pts", "pt", "pnts", "نقاط", "النقاط", "نقطة"
    ], 0)
    gd = _v10_pick_stat(stats, [
        "goal difference", "goal_difference", "gd", "+/-", "difference", "diff", "فارق", "فرق", "فارق الاهداف", "فارق الأهداف"
    ], 0)
    gf = _v10_pick_stat(stats, ["goals for", "gf", "له", "له اهداف", "أهداف له"], None)
    ga = _v10_pick_stat(stats, ["goals against", "ga", "عليه", "أهداف عليه"], None)
    try:
        if (gd == 0 or gd is None) and gf is not None and ga is not None:
            gd = int(gf) - int(ga)
    except Exception:
        pass
    # بعض بنى Google تعطي القيم تحت مفاتيح قصيرة جدًا بالعربي، نحاول التقاطها من display text
    if pts == 0:
        blob = json.dumps(node, ensure_ascii=False)
        m = re.search(r"(?:نقاط|points|pts)\D{0,12}(\d{1,2})", blob, re.I)
        if m:
            pts = _v10_int_any(m.group(1), 0)
    return (team, int(played or 0), int(gd or 0), int(pts or 0))


def _v10_rows_from_container(container):
    rows = []
    if isinstance(container, list):
        iterable = container
    elif isinstance(container, dict):
        iterable = None
        for key in ["teams", "rows", "entries", "standings", "table", "rankings", "ranking", "competitors", "items"]:
            val = container.get(key)
            if isinstance(val, list) and val:
                # خذ أول قائمة تبدو كصفوف فرق
                if any(_v10_parse_standing_row(x) for x in val if isinstance(x, dict)):
                    iterable = val
                    break
        if iterable is None:
            iterable = []
    else:
        iterable = []
    for item in iterable:
        row = _v10_parse_standing_row(item)
        if row and row[0] not in [r[0] for r in rows]:
            rows.append(row)
    # ترتيب قوقل غالبًا موجود جاهز؛ إذا لا يوجد ترتيب واضح نرتب بالنقاط والفارق
    if rows:
        rows = sorted(rows, key=lambda r: (r[3], r[2], r[1]), reverse=True)
    return rows


def _v10_extract_groups_from_google_json(data):
    allowed = set(WORLD_CUP_TEAMS)
    found = []
    seen_titles = set()

    def add_group(title, rows):
        rows = [r for r in rows or [] if r and r[0] in allowed]
        if len(rows) < 2:
            return
        title = _v10_group_title(title)
        # منع تكرار نفس المجموعة إذا رجعت من أكثر من مكان
        key = title + "|" + ",".join([r[0] for r in rows[:4]])
        if key in seen_titles:
            return
        seen_titles.add(key)
        found.append((title, rows[:4]))

    roots = []
    if isinstance(data, dict):
        sr = data.get("sports_results")
        if isinstance(sr, dict):
            roots.append(sr)
            for k in ["standings", "tables", "table", "groups", "rankings", "ranking", "league", "tournament"]:
                if isinstance(sr.get(k), (dict, list)):
                    roots.append(sr.get(k))
        roots.append(data)
    else:
        roots.append(data)

    # 1) حاويات واضحة لمجموعة كاملة
    for root in roots:
        for node in _walk_json(root):
            if not isinstance(node, dict):
                continue
            title = node.get("title") or node.get("name") or node.get("group") or node.get("label") or node.get("stage") or ""
            if _v10_group_letter(title) or any(k in node for k in ["teams", "rows", "entries", "standings", "table", "rankings"]):
                rows = _v10_rows_from_container(node)
                if rows:
                    add_group(title or node.get("group") or "المجموعة", rows)
            # 2) مفاتيح groups/tables فيها قائمة مجموعات
            for key in ["groups", "tables", "standings", "rankings", "table"]:
                val = node.get(key)
                if isinstance(val, list):
                    for sub in val:
                        if isinstance(sub, dict):
                            title2 = sub.get("title") or sub.get("name") or sub.get("group") or title
                            rows2 = _v10_rows_from_container(sub)
                            if rows2:
                                add_group(title2, rows2)

    # 3) صفوف مسطحة فيها حقل group
    flat = {}
    for root in roots:
        for node in _walk_json(root):
            if not isinstance(node, dict):
                continue
            row = _v10_parse_standing_row(node)
            if not row:
                continue
            g = node.get("group") or node.get("pool") or node.get("stage") or node.get("division") or ""
            gtitle = _v10_group_title(g)
            if gtitle:
                flat.setdefault(gtitle, [])
                if row[0] not in [r[0] for r in flat[gtitle]]:
                    flat[gtitle].append(row)
    for title, rows in flat.items():
        if len(rows) >= 2:
            add_group(title, rows)

    # ترتيب المجموعات من أ إلى ل
    order = {"أ":0,"ب":1,"ج":2,"د":3,"هـ":4,"ه":4,"و":5,"ز":6,"ح":7,"ط":8,"ي":9,"ك":10,"ل":11,
             "A":0,"B":1,"C":2,"D":3,"E":4,"F":5,"G":6,"H":7,"I":8,"J":9,"K":10,"L":11}
    def gkey(gr):
        title = gr[0]
        m = re.search(r"المجموعة\s*([أابجدهـوزحطيكلA-L])", title, re.I)
        return order.get(m.group(1).upper() if m else "", 99)
    found.sort(key=gkey)
    return found


def _v10_validate_google_groups(groups, allow_partial=True):
    clean = []
    seen_teams = set()
    for title, rows in groups or []:
        out_rows = []
        for row in rows or []:
            if not row or len(row) < 4:
                continue
            team = canonical_team_name(row[0]) or normalize_name(row[0])
            if team not in set(WORLD_CUP_TEAMS):
                continue
            played, gd, pts = _v10_int_any(row[1]), _v10_int_any(row[2]), _v10_int_any(row[3])
            out_rows.append((team, played, gd, pts))
            seen_teams.add(team)
        if len(out_rows) >= 2:
            clean.append((_v10_group_title(title), out_rows[:4]))
    if not clean:
        return []
    if allow_partial:
        # إذا قوقل رجع مجموعتين أو 8 فرق على الأقل نعرضها كجزئي بدل الرفض.
        if len(clean) >= 2 or len(seen_teams) >= 8:
            return clean[:12]
        return []
    if len(clean) >= 8 and len(seen_teams) >= 24:
        return clean[:12]
    return []


def fetch_standings_from_serpapi():
    if not _serpapi_key():
        return []
    errors = []
    for q in V10_GOOGLE_STANDINGS_QUERIES:
        for hl, gl in [("ar", "sa"), ("en", "us")]:
            try:
                data = serpapi_search_json(q, hl=hl, gl=gl, timeout=9)
                groups = _v10_extract_groups_from_google_json(data)
                groups = _v10_validate_google_groups(groups, allow_partial=True)
                if groups:
                    return groups
            except Exception as e:
                errors.append(str(e)[:80])
                continue
    return []


def fetch_current_groups(mode="latest"):
    norm = _norm_source_mode(mode)
    # بعد اعتماد الرابط: قوقل هو المصدر الأساسي للترتيب في كل الأوضاع، الرسمي/API احتياط فقط.
    if norm in ["official", "api"]:
        order = [(fetch_standings_from_serpapi, "Google Sports"), (fetch_standings_from_api_football, "API-Football")]
    else:
        order = [(fetch_standings_from_serpapi, "Google Sports"), (fetch_standings_from_api_football, "API-Football")]
    for fn, label in order:
        try:
            groups = fn()
            if label == "Google Sports":
                groups = _v10_validate_google_groups(groups, allow_partial=True)
            else:
                groups = _validate_worldcup_groups(groups)
            if groups:
                return groups, label
        except Exception:
            continue
    return [], ""


def _v10_extract_highlight_link(root):
    best = ""
    best_score = -1
    try:
        for node in _walk_json(root):
            if not isinstance(node, dict):
                continue
            blob = json.dumps(node, ensure_ascii=False).lower()
            score = 0
            if any(w in blob for w in ["highlight", "highlights", "ملخص", "فيديو", "video"]):
                score += 5
            if any(w in blob for w in ["youtube", "youtu.be", "fifa.com", "google"]):
                score += 2
            for k in ["link", "url", "source", "watch_link", "video_link", "href"]:
                v = node.get(k)
                if isinstance(v, str) and v.startswith("http"):
                    if score > best_score:
                        best, best_score = v, score
            # أحيانًا الرابط داخل nested dict
            for k in ["video", "highlight", "thumbnail", "content"]:
                v = node.get(k)
                if isinstance(v, dict):
                    for kk in ["link", "url", "href"]:
                        vv = v.get(kk)
                        if isinstance(vv, str) and vv.startswith("http") and score >= best_score:
                            best, best_score = vv, score
    except Exception:
        pass
    return best if best_score >= 3 else ""


def _v10_goal_candidate_strings(root):
    vals = []
    try:
        for node in _walk_json(root):
            if isinstance(node, dict):
                # كائن هدف واضح
                name = node.get("player") or node.get("scorer") or node.get("athlete") or node.get("name") or node.get("title")
                minute = node.get("minute") or node.get("time") or node.get("elapsed") or node.get("displayTime")
                detail = node.get("detail") or node.get("description") or node.get("type") or node.get("event") or ""
                if name and minute:
                    vals.append(f"{name} {minute} {detail}")
                for k in ["text", "summary", "subtitle", "description", "detail", "title", "name"]:
                    v = node.get(k)
                    if isinstance(v, str):
                        vals.append(v)
            elif isinstance(node, str):
                vals.append(node)
    except Exception:
        pass
    return vals


def _v10_extract_goal_lines_from_text(s):
    s = normalize_name(s or "")
    if not s:
        return []
    s = s.replace("’", "'").replace("′", "'")
    out = []
    # شكل: كاميرون بورغس 11' (هدف في مرماه)
    for m in re.finditer(r"([A-Za-zÀ-ÿء-ي][A-Za-zÀ-ÿء-ي\s\-'.]{2,42}?)\s+(\d{1,3})\s*['’′](?:\s*\(([^)]{1,28})\))?", s):
        name = normalize_name(m.group(1)).strip(" -—•،,.")
        minute = m.group(2)
        extra = normalize_name(m.group(3) or "")
        if len(name.split()) > 5:
            continue
        line = f"{name} {minute}'" + (f" {extra}" if extra else "")
        out.append(line)
    # شكل: 11' كاميرون بورغس
    for m in re.finditer(r"(\d{1,3})\s*['’′]\s+([A-Za-zÀ-ÿء-ي][A-Za-zÀ-ÿء-ي\s\-'.]{2,42})", s):
        minute = m.group(1)
        name = normalize_name(m.group(2)).strip(" -—•،,.")
        if len(name.split()) <= 5:
            out.append(f"{name} {minute}'")
    return out


def _v10_collect_goal_details_from_google_card(root):
    collected = []
    for s in _v10_goal_candidate_strings(root):
        collected.extend(_v10_extract_goal_lines_from_text(s))
    clean = _v9_sanitize_scorers(collected)
    # تنقية أخيرة من السطور العامة
    bad = ["نهاية", "المباراة", "ترتيب", "المجموعة", "كأس", "fifa", "world cup", "ملخص", "video"]
    out = []
    for line in clean:
        low = line.lower()
        if any(b in low for b in bad):
            continue
        if line not in out:
            out.append(line)
    return out[:8]


def _v10_fetch_match_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    source_name = {"google": "Google Sports", "365": "365Scores عبر Google", "kooora": "Kooora عبر Google"}.get(source, "Google")
    attempts = []
    for q in _serpapi_query_candidates(req1, req2, date_hint, source=source):
        if source == "google":
            attempts.append((q, "ar", "sa"))
            attempts.append((q, "en", "us"))
        else:
            attempts.append((q, "ar", "sa"))
    seen_q = set()
    final_attempts = []
    for item in attempts:
        key = item[0] + "|" + item[1] + "|" + item[2]
        if key not in seen_q:
            seen_q.add(key)
            final_attempts.append(item)
    for q, hl, gl in final_attempts[:5]:
        try:
            data = serpapi_search_json(q, hl=hl, gl=gl, timeout=8)
        except Exception:
            continue
        roots = []
        sr = data.get("sports_results") if isinstance(data, dict) else None
        if isinstance(sr, dict):
            roots.append(sr)
            for k in ["game_spotlight", "games", "matches", "scoreboard", "players", "teams", "events", "timeline"]:
                if isinstance(sr.get(k), (dict, list)):
                    roots.append(sr.get(k))
        roots.append(data)
        for root in roots:
            obj = None
            try:
                obj = _parse_serp_sports_node_v6(root, req1, req2, source_name=source_name)
            except Exception:
                obj = None
            if obj and _v9_valid_score_obj(obj):
                scorers = _v10_collect_goal_details_from_google_card(data)
                if scorers:
                    obj["scorers"] = scorers
                    obj["goals_source"] = "Google Sports" if source == "google" else mode_label_ar(source)
                else:
                    obj["scorers"] = _v9_sanitize_scorers(obj.get("scorers") or [])
                link = _v10_extract_highlight_link(data)
                if link:
                    obj["highlight_url"] = link
                obj["source"] = source_name
                return obj
        blob = json.dumps(data, ensure_ascii=False)
        if _blob_has_both_teams(blob, req1, req2):
            s = _score_from_text_near_teams(blob, req1, req2) if '_score_from_text_near_teams' in globals() else None
            if s:
                obj = {
                    "team1": req1, "team2": req2,
                    "score1": s.get("score1"), "score2": s.get("score2"),
                    "status": s.get("status") or _status_from_serp_v6(data), "minute": "",
                    "scorers": _v10_collect_goal_details_from_google_card(data),
                    "source": source_name,
                }
                if obj.get("scorers"):
                    obj["goals_source"] = "Google Sports" if source == "google" else mode_label_ar(source)
                link = _v10_extract_highlight_link(data)
                if link:
                    obj["highlight_url"] = link
                if _v9_valid_score_obj(obj):
                    return obj
    return None


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")


def fetch_match_from_365(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="365")


def fetch_match_from_kooora(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="kooora")


def _v9_fetch_goal_details(team1, team2, date_hint=None, preferred_source="google"):
    # V10: قوقل أولاً للهدافين، ثم 365/كورة، ثم ESPN احتياط أخير.
    if preferred_source == "365":
        sequence = ["365", "google", "kooora", "espn"]
    elif preferred_source == "kooora":
        sequence = ["kooora", "google", "365", "espn"]
    else:
        sequence = ["google", "365", "kooora", "espn"]
    for src in sequence:
        if src == "espn":
            scorers = _v9_fetch_scorers_from_espn(team1, team2, date_hint)
        else:
            # نستفيد من كرت المباراة نفسه لو كان Google
            try:
                obj = _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source=("google" if src == "google" else src))
                scorers = _v9_sanitize_scorers((obj or {}).get("scorers") or [])
            except Exception:
                scorers = []
            if not scorers:
                scorers = _v9_fetch_scorers_from_serp_source(team1, team2, date_hint, source=src)
        scorers = _v9_sanitize_scorers(scorers)
        if scorers:
            return scorers, mode_label_ar(src) if src in ["365", "kooora"] else ("ESPN" if src == "espn" else "Google Sports")
    return [], ""


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    norm = _norm_source_mode(mode)
    requested_label = mode_label_ar(norm)
    if norm in ["365", "kooora"]:
        seq = [norm, "google", "espn", "api"]
    elif norm == "latest":
        seq = ["google", "espn", "365", "kooora", "api"]
    elif norm == "official":
        # API-Football قد يكون موقوف؛ ESPN ثم Google
        seq = ["espn", "google", "api"]
    elif norm == "google":
        seq = ["google", "espn"]
    else:
        seq = ["google", "espn", "api"]
    seen = set()
    for src in seq:
        if src in seen:
            continue
        seen.add(src)
        try:
            obj = _v9_fetch_primary_from_source(team1, team2, src, date_hint=date_hint)
        except Exception:
            obj = None
        if not obj or not _v9_valid_score_obj(obj):
            continue
        obj["source"] = _v9_clean_source_name(obj.get("source") or ("ESPN" if src == "espn" else mode_label_ar(src)))
        obj["requested_source"] = requested_label
        obj["actual_source"] = obj["source"]
        if not obj.get("status"):
            obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
        scorers = _v9_sanitize_scorers(obj.get("scorers") or [])
        details_source = obj.get("goals_source") or (obj.get("source") if scorers else "")
        if not scorers:
            pref = norm if norm in ["365", "kooora"] else src
            scorers, details_source = _v9_fetch_goal_details(team1, team2, date_hint=date_hint, preferred_source=pref)
        obj["scorers"] = scorers
        obj["goals_source"] = details_source or "غير متوفر"
        # إذا جاء رابط ملخص من Google احتفظ به حتى لو المصدر الفعلي ESPN.
        if not obj.get("highlight_url") and src != "google" and _serpapi_key():
            try:
                gobj = _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")
                if gobj and gobj.get("highlight_url"):
                    obj["highlight_url"] = gobj.get("highlight_url")
            except Exception:
                pass
        return obj
    return None


def render_live_match_card(match, mode_label="رسمي"):
    ensure_generated_dir()
    width, height = 1200, 1500
    img, draw = _v9_live_bg(width, height)

    draw_text(draw, (width//2, 88), "MONDIAL AL MASEEF 2026", _v31_latin_font(34), fill="#FFFFFF", max_width=760)
    draw_text(draw, (width//2, 160), "LIVE MATCH", _v31_latin_font(74), fill="#FFFFFF", max_width=900)
    draw_text(draw, (width//2, 222), "متابعة المباراة", get_font(34), fill="#FDE68A", max_width=720)

    main = (70, 290, width-70, 1070)
    try:
        _draw_glow_card(draw, main, 42, "#061633E8", "#38BDF8", width=4)
    except Exception:
        rounded_rect(draw, main, radius=42, fill="#061633", outline="#38BDF8", width=4)

    flag_y1, flag_y2 = 365, 485
    left_flag = (130, flag_y1, 310, flag_y2)
    right_flag = (width-310, flag_y1, width-130, flag_y2)
    rounded_rect(draw, (left_flag[0]-12, left_flag[1]-10, left_flag[2]+12, left_flag[3]+10), radius=28, fill="#FFFFFFE8", outline="#FFFFFF55", width=2)
    rounded_rect(draw, (right_flag[0]-12, right_flag[1]-10, right_flag[2]+12, right_flag[3]+10), radius=28, fill="#FFFFFFE8", outline="#FFFFFF55", width=2)
    _v9_draw_flag_fit(img, match.get("team1", ""), left_flag)
    _v9_draw_flag_fit(img, match.get("team2", ""), right_flag)
    draw_text(draw, (220, 540), match.get("team1", ""), get_font(40), fill="#FFFFFF", max_width=330)
    draw_text(draw, (width-220, 540), match.get("team2", ""), get_font(40), fill="#FFFFFF", max_width=330)

    score = f"{match.get('score1','0')} - {match.get('score2','0')}"
    draw_text(draw, (width//2, 463), score, _v31_latin_font(116), fill="#FFFFFF", max_width=360)

    status = _v9_status_line(match)
    rounded_rect(draw, (width//2-245, 590, width//2+245, 662), radius=24, fill="#FBBF24", outline="#FFFFFF55", width=2)
    draw_text(draw, (width//2, 626), status, get_font(30), fill="#061633", max_width=450)

    scorers = _v9_sanitize_scorers(match.get("scorers") or [])
    details_top = 720
    draw_text(draw, (width//2, details_top), "الهدافون", get_font(42), fill="#FDE68A", max_width=480)
    y = details_top + 74
    if scorers:
        for s in scorers[:5]:
            rounded_rect(draw, (120, y-30, width-120, y+42), radius=22, fill="#07132FEA", outline="#FFFFFF33", width=2)
            draw_text(draw, (width//2, y+6), f"⚽ {s}", get_font(31), fill="#FFFFFF", max_width=900)
            y += 82
    else:
        rounded_rect(draw, (150, y-30, width-150, y+42), radius=24, fill="#07132FEA", outline="#FFFFFF33", width=2)
        draw_text(draw, (width//2, y+6), "تفاصيل الهدافين غير متوفرة حاليًا", get_font(30), fill="#FFFFFF", max_width=820)
        y += 86

    # لا نعرض المصدر داخل الصورة حسب اعتماد V10.
    draw.line((230, height-160, width-230, height-160), fill="#FFFFFF66", width=2)
    draw_text(draw, (width//2, height-120), "المصيف يضعكم بالحدث", get_font(34), fill="#FDE68A", max_width=650)

    out = os.path.join(GENERATED_DIR, f"live_v10_{_safe_filename(match.get('team1','team1'))}_{_safe_filename(match.get('team2','team2'))}.png")
    try:
        img.save(out, quality=96)
    except TypeError:
        img.save(out)
    return out


def build_live_caption(match, mode_label="رسمي"):
    lines = [f"{match.get('team1','')} {match.get('score1','0')} - {match.get('score2','0')} {match.get('team2','')}"]
    st = _v9_status_line(match)
    if st:
        lines.append(st)
    scorers = _v9_sanitize_scorers(match.get("scorers") or [])
    lines.append("")
    lines.append("⚽ الهدافون:")
    if scorers:
        lines.extend([f"- {s}" for s in scorers[:6]])
    else:
        lines.append("غير متوفرين حاليًا من المصادر")
    actual = match.get("actual_source") or match.get("source") or mode_label
    requested = match.get("requested_source") or mode_label
    lines.append("")
    lines.append(f"المصدر المطلوب: {requested}")
    lines.append(f"المصدر الفعلي: {actual}")
    if match.get("goals_source"):
        lines.append(f"تفاصيل الأهداف: {match.get('goals_source')}")
    link = match.get("highlight_url") or match.get("summary_url") or ""
    if link:
        lines.append(f"🎥 ملخص المباراة: {link}")
    else:
        lines.append("🎥 ملخص المباراة: غير متوفر حاليًا من قوقل")
    return "\n".join(lines)


def create_all_groups_newlook_image(groups, style=2):
    ensure_generated_dir()
    groups = list(groups or [])[:12]
    width = 1800
    cols = 3
    margin_x, gap_x = 72, 34
    card_w = (width - 2*margin_x - (cols-1)*gap_x) // cols
    row_h = 62
    row_gap = 8
    grid_gap_y = 32
    min_card_h = 285
    header_h = 74
    label_h = 34
    bottom_pad = 24
    row_heights = []
    for r in range(4):
        chunk = groups[r*cols:(r+1)*cols]
        max_rows = max([min(len(rows), 4) for _title, rows in chunk] + [1])
        ch = header_h + label_h + 18 + max_rows * row_h + max(0, max_rows-1) * row_gap + bottom_pad
        row_heights.append(max(min_card_h, ch))
    start_y = 300
    footer_space = 110
    height = int(start_y + sum(row_heights) + grid_gap_y*(len(row_heights)-1) + footer_space)
    height = max(1640, min(height, 2320))
    img, draw = _games_day_background(width, height)
    overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    rounded_rect(od, (45, 40, width-45, height-45), radius=40, fill="#06152F55", outline="#FFFFFF22", width=2)
    img = Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")
    draw = ImageDraw.Draw(img)
    draw_text(draw, (width//2, 82), "MONDIAL AL MASEEF 2026", _v31_latin_font(42) if '_v31_latin_font' in globals() else get_font(40), fill="#FFFFFF", max_width=900)
    draw_text(draw, (width//2, 160), "ALL GROUP STANDINGS", _v31_latin_font(74) if '_v31_latin_font' in globals() else get_font(72), fill="#FFFFFF", max_width=width-170)
    draw_text(draw, (width//2, 230), "ترتيب جميع المجموعات", get_font(40), fill="#FBBF24", max_width=800)
    y_cursor = start_y
    for idx, (title, rows) in enumerate(groups):
        c = idx % cols
        r = idx // cols
        if c == 0 and idx != 0:
            y_cursor += row_heights[r-1] + grid_gap_y
        x = margin_x + c*(card_w+gap_x)
        y = y_cursor
        card_h = row_heights[r]
        rounded_rect(draw, (x, y, x+card_w, y+card_h), radius=28, fill="#0638A5E8", outline="#14B8F5", width=3)
        rounded_rect(draw, (x+16, y+16, x+card_w-16, y+66), radius=16, fill="#FBBF24", outline="#00000055", width=1)
        gt = clean_group_title_for_design(title)
        if not str(gt).startswith("المجموعة"):
            gt = f"المجموعة {gt}"
        draw_text(draw, (x+card_w//2, y+41), gt, get_font(29), fill="#061633", max_width=card_w-44)
        label_y = y + 88
        draw_text(draw, (x+card_w-64, label_y), "#", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+card_w-205, label_y), "المنتخب", get_font(18), fill="#FDE68A", max_width=190)
        draw_text(draw, (x+195, label_y), "لعب", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+118, label_y), "+/-", get_font(18), fill="#FDE68A")
        draw_text(draw, (x+48, label_y), "نقاط", get_font(18), fill="#FDE68A")
        yy = y + 108
        for pos, row in enumerate(rows[:4], start=1):
            team, played, diff, pts = row[0], _v10_int_any(row[1]), _v10_int_any(row[2]), _v10_int_any(row[3])
            rounded_rect(draw, (x+16, yy, x+card_w-16, yy+row_h), radius=15, fill="#061633B8", outline="#FFFFFF30", width=1)
            cy = yy + row_h//2
            draw_text(draw, (x+card_w-34, cy), str(pos), get_font(23), fill="#FBBF24")
            fw, fh = 56, 38
            try:
                paste_flag(img, team, (x+card_w-100, cy-fh//2, x+card_w-100+fw, cy+fh//2))
            except Exception:
                pass
            team_name = _clean_display_name(team)
            team_font = _fit_font_to_width(draw, team_name, 24, 190, min_size=17)
            draw_text(draw, (x+card_w-230, cy), team_name, team_font, fill="#FFFFFF", max_width=205)
            draw_text(draw, (x+195, cy), str(played), get_font(23), fill="#FFFFFF")
            draw_text(draw, (x+118, cy), f"{int(diff):+d}", get_font(23), fill="#E5E7EB")
            draw_text(draw, (x+48, cy), str(pts), get_font(29), fill="#FBBF24")
            yy += row_h + row_gap
    draw_text(draw, (width//2, height-50), "المصيف يضعكم بالحدث", get_font(38), fill="#FBBF24", max_width=700)
    path = os.path.join(GENERATED_DIR, f"all_groups_style{style}.png")
    img.save(path, quality=95)
    return path


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|الاحدث|official|fast|latest|google|قوقل|365|٣٦٥|كورة|كوره|kooora)\s*$", text or "", re.I)
        mode = m.group(1) if m else "latest"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري جلب ترتيب المجموعات من قوقل...")
    try:
        groups, source_label = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_current_groups, mode), timeout=24)
    except Exception:
        groups, source_label = [], ""
    if not groups:
        await wait.edit_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    path = create_all_groups_image(groups)
    caption = f"ترتيب المجموعات الآن ✅\nالمصدر: {source_label}\nالاستعلام المعتمد: ترتيبات كأس العالم"
    try:
        await wait.delete()
    except Exception:
        pass
    await send_photo_path_markup(update.message, path, caption, kb)
    await update.message.reply_text(build_groups_text(groups, source_label))

# ==================== END V10 FINAL PATCH ====================



# ==================== V11 FINAL PATCH: robust Google rankings + live uses debug parser ====================
# اعتماد V11:
# - /ترتيب_المجموعات_الان يقرأ sports_results.rankings من Google/SerpApi فعليًا.
# - /مباشر * قوقل يستخدم نفس parser الناجح في /بحث_قوقل مع game_spotlight.
# - رسالة الخطأ لا تسحب مباراة قديمة، بل تقترح نفس مباراة المستخدم.
# - المصدر يبقى في الكابشن فقط، وأسفل التصاميم: المصيف يضعكم بالحدث.

V11_STANDINGS_QUERIES = [
    "ترتيبات كأس العالم",
    "ترتيب مجموعات كأس العالم 2026",
    "FIFA World Cup 2026 standings",
    "World Cup 2026 group standings",
]

V11_GROUP_BY_TEAM = {}
for _g, _teams in WORLD_CUP_GROUPS:
    for _t in _teams:
        V11_GROUP_BY_TEAM[_t] = _g


def _v11_to_int(v, default=None):
    try:
        if v is None or isinstance(v, bool):
            return default
        if isinstance(v, (int, float)):
            return int(v)
        s = normalize_name(str(v))
        s = s.replace("+", "")
        m = re.search(r"-?\d+", s)
        return int(m.group(0)) if m else default
    except Exception:
        return default


def _v11_group_title_from_letter(letter):
    letter = (letter or "").upper()
    arabic = {"A":"أ", "B":"ب", "C":"ج", "D":"د", "E":"هـ", "F":"و", "G":"ز", "H":"ح", "I":"ط", "J":"ي", "K":"ك", "L":"ل"}.get(letter, letter)
    return f"المجموعة {arabic}" if arabic else "المجموعة"


def _v11_group_letter_from_title(title):
    return _v10_group_letter(title) if '_v10_group_letter' in globals() else ""


def _v11_team_from_value(v):
    if isinstance(v, str):
        c = canonical_team_name(v)
        if c in WORLD_CUP_TEAMS:
            return c
    if isinstance(v, dict):
        for k in ["name", "displayName", "display_name", "shortName", "short_name", "title", "teamName", "team_name", "country"]:
            c = canonical_team_name(v.get(k)) if isinstance(v.get(k), str) else None
            if c in WORLD_CUP_TEAMS:
                return c
        for k in ["team", "competitor", "participant", "club", "country", "entity"]:
            c = _v11_team_from_value(v.get(k))
            if c:
                return c
    return ""


def _v11_team_from_row(row, allow_blob=False):
    # لا نقرأ من blob للحاويات الكبيرة إلا عند الحاجة، حتى لا يلتقط أول منتخب في مجموعة كاملة.
    if isinstance(row, dict):
        for k in ["team", "competitor", "participant", "club", "country", "entity"]:
            c = _v11_team_from_value(row.get(k))
            if c:
                return c
        for k in ["name", "displayName", "display_name", "shortName", "short_name", "title", "teamName", "team_name"]:
            v = row.get(k)
            if isinstance(v, str):
                c = canonical_team_name(v)
                if c in WORLD_CUP_TEAMS:
                    return c
    elif isinstance(row, (list, tuple)):
        for item in row:
            c = _v11_team_from_value(item)
            if c:
                return c
    elif isinstance(row, str):
        allow_blob = True
    if allow_blob:
        try:
            blob = normalize_name(json.dumps(row, ensure_ascii=False) if not isinstance(row, str) else row)
            hits = []
            for tm in WORLD_CUP_TEAMS:
                if simple_key(tm) and simple_key(tm) in simple_key(blob):
                    hits.append(tm)
            if len(hits) == 1:
                return hits[0]
        except Exception:
            pass
    return ""


def _v11_stats_dict(row):
    stats = {}
    if not isinstance(row, dict):
        return stats

    def add(k, v):
        kk = simple_key(str(k))
        if kk:
            stats[kk] = v

    for k, v in row.items():
        if isinstance(v, (str, int, float)):
            add(k, v)

    for key in ["stats", "statistics", "columns", "details", "record", "records"]:
        val = row.get(key)
        if isinstance(val, dict):
            for k, v in val.items():
                add(k, v)
        elif isinstance(val, list):
            for it in val:
                if isinstance(it, dict):
                    label = it.get("name") or it.get("label") or it.get("title") or it.get("displayName") or it.get("abbreviation") or it.get("type")
                    value = it.get("value") if "value" in it else it.get("displayValue") if "displayValue" in it else it.get("rank") if "rank" in it else it.get("summary")
                    if label is not None and value is not None:
                        add(label, value)

    # Google/SerpApi أحيانًا يعطي columns + values متوازية
    cols = row.get("columns") or row.get("headers") or row.get("labels")
    vals = row.get("values") or row.get("row") or row.get("cells")
    if isinstance(cols, list) and isinstance(vals, list):
        for c, v in zip(cols, vals):
            label = c.get("name") if isinstance(c, dict) else c
            value = v.get("value") if isinstance(v, dict) else v
            if label is not None:
                add(label, value)
    return stats


def _v11_pick(stats, aliases, default=0):
    alias_keys = [simple_key(a) for a in aliases]
    for a in alias_keys:
        if a in stats:
            val = _v11_to_int(stats.get(a), None)
            if val is not None:
                return val
    for sk, sv in stats.items():
        for a in alias_keys:
            if a and (a == sk or a in sk or sk in a):
                val = _v11_to_int(sv, None)
                if val is not None:
                    return val
    return default


def _v11_parse_standing_row(row):
    team = _v11_team_from_row(row, allow_blob=False)
    if not team:
        # صفوف list أو string غالبًا قصيرة وآمنة للبحث النصي
        team = _v11_team_from_row(row, allow_blob=isinstance(row, (list, tuple, str)))
    if team not in WORLD_CUP_TEAMS:
        return None

    played = gd = pts = None
    if isinstance(row, dict):
        stats = _v11_stats_dict(row)
        played = _v11_pick(stats, ["played", "matches played", "games played", "mp", "gp", "p", "played games", "لعب", "ل", "لعبت"], None)
        pts = _v11_pick(stats, ["points", "pts", "pt", "pnts", "نقاط", "النقاط", "نقطة"], None)
        gd = _v11_pick(stats, ["goal difference", "goaldifference", "gd", "+/-", "diff", "difference", "فارق", "فرق", "فارق الأهداف", "فارق الاهداف"], None)
        gf = _v11_pick(stats, ["goals for", "goalsfor", "gf", "for", "له", "اهداف له", "أهداف له"], None)
        ga = _v11_pick(stats, ["goals against", "goalsagainst", "ga", "against", "عليه", "اهداف عليه", "أهداف عليه"], None)
        if gd is None and gf is not None and ga is not None:
            try:
                gd = int(gf) - int(ga)
            except Exception:
                gd = None
        # إذا كانت القيم داخل list بدون تسميات
        if played is None or pts is None:
            nums = []
            for k in ["values", "row", "cells"]:
                v = row.get(k)
                if isinstance(v, list):
                    for item in v:
                        itemv = item.get("value") if isinstance(item, dict) else item
                        n = _v11_to_int(itemv, None)
                        if n is not None:
                            nums.append(n)
            if nums:
                # غالبًا: rank, played, win, draw, loss, gd/gf, pts أو played,...,pts
                if len(nums) >= 7:
                    played = nums[1] if played is None else played
                    gd = nums[-2] if gd is None else gd
                    pts = nums[-1] if pts is None else pts
                elif len(nums) >= 4:
                    played = nums[0] if played is None else played
                    gd = nums[-2] if gd is None else gd
                    pts = nums[-1] if pts is None else pts
    else:
        vals = row if isinstance(row, (list, tuple)) else [row]
        nums = []
        for item in vals:
            if canonical_team_name(str(item)):
                continue
            n = _v11_to_int(item, None)
            if n is not None:
                nums.append(n)
        if len(nums) >= 7:
            played, gd, pts = nums[1], nums[-2], nums[-1]
        elif len(nums) >= 4:
            played, gd, pts = nums[0], nums[-2], nums[-1]
        elif len(nums) >= 2:
            played, pts = nums[0], nums[-1]
            gd = 0
    if played is None:
        played = 0
    if gd is None:
        gd = 0
    if pts is None:
        # إذا ما فيه نقاط واضحة غالبًا هذا ليس صف ترتيب موثوق
        return None
    return (team, int(played), int(gd), int(pts))


def _v11_rows_from_any(container):
    rows = []
    def add(row):
        pr = _v11_parse_standing_row(row)
        if pr and pr[0] not in [x[0] for x in rows]:
            rows.append(pr)
    if isinstance(container, list):
        for item in container:
            if isinstance(item, (dict, list, tuple, str)):
                add(item)
            if isinstance(item, dict):
                for k in ["teams", "rows", "entries", "standings", "table", "rankings", "ranking", "competitors", "items", "values"]:
                    if isinstance(item.get(k), list):
                        for sub in item.get(k):
                            add(sub)
    elif isinstance(container, dict):
        # القيم الواضحة أولًا
        for k in ["teams", "rows", "entries", "standings", "table", "rankings", "ranking", "competitors", "items", "values"]:
            val = container.get(k)
            if isinstance(val, list):
                for item in val:
                    add(item)
        # ربما dict نفسه صف
        add(container)
    if rows:
        rows.sort(key=lambda r: (r[3], r[2], r[1]), reverse=True)
    return rows


def _v11_group_rows_by_official_groups(rows):
    buckets = {g: [] for g, _ in WORLD_CUP_GROUPS}
    for r in rows or []:
        team = r[0]
        g = V11_GROUP_BY_TEAM.get(team)
        if not g:
            continue
        if team not in [x[0] for x in buckets[g]]:
            buckets[g].append(r)
    out = []
    for g, _teams in WORLD_CUP_GROUPS:
        rr = buckets.get(g) or []
        if len(rr) >= 2:
            rr.sort(key=lambda r: (r[3], r[2], r[1]), reverse=True)
            out.append((_v11_group_title_from_letter(g), rr[:4]))
    return out


def _v11_extract_groups_from_rankings(rankings):
    groups = []
    flat_rows = []

    def try_add_group(title, obj):
        rows = _v11_rows_from_any(obj)
        if rows:
            # لو العنوان مو مجموعة، نوزع على المجموعات الرسمية بدل تجميع كل الترتيب في كرت واحد
            gl = _v11_group_letter_from_title(title or "")
            if gl:
                groups.append((_v11_group_title_from_letter(gl), rows[:4]))
            else:
                flat_rows.extend(rows)

    if isinstance(rankings, dict):
        # Dict بمفاتيح أسماء مجموعات
        for k, v in rankings.items():
            if isinstance(v, (list, dict)):
                try_add_group(k, v)
        try_add_group(rankings.get("title") or rankings.get("name") or "", rankings)
    elif isinstance(rankings, list):
        for item in rankings:
            if isinstance(item, dict):
                title = item.get("title") or item.get("name") or item.get("group") or item.get("label") or item.get("stage") or ""
                try_add_group(title, item)
                # أحيانًا المجموعة داخل children/items
                for k in ["groups", "tables", "standings", "rankings", "table", "items", "children"]:
                    val = item.get(k)
                    if isinstance(val, list):
                        for sub in val:
                            if isinstance(sub, dict):
                                t2 = sub.get("title") or sub.get("name") or sub.get("group") or title
                                try_add_group(t2, sub)
            elif isinstance(item, (list, tuple, str)):
                pr = _v11_parse_standing_row(item)
                if pr:
                    flat_rows.append(pr)
    # أضف المجموعات الموزعة من flat
    groups += _v11_group_rows_by_official_groups(flat_rows)
    # تنظيف التكرار
    seen = set()
    out = []
    for title, rows in groups:
        clean_rows = []
        for r in rows:
            if r and r[0] in WORLD_CUP_TEAMS and r[0] not in [x[0] for x in clean_rows]:
                clean_rows.append(r)
        if len(clean_rows) < 2:
            continue
        key = title + "|" + ",".join(r[0] for r in clean_rows)
        if key not in seen:
            seen.add(key)
            out.append((title, clean_rows[:4]))
    return out


def _v11_extract_groups_from_google_json(data):
    sr = data.get("sports_results") if isinstance(data, dict) else None
    candidates = []
    if isinstance(sr, dict):
        # أهم حالة ظهرت عند المستخدم: sports_results.rankings
        if "rankings" in sr:
            candidates.extend(_v11_extract_groups_from_rankings(sr.get("rankings")))
        for k in ["standings", "tables", "table", "groups", "ranking", "league", "tournament"]:
            if isinstance(sr.get(k), (dict, list)):
                candidates.extend(_v11_extract_groups_from_rankings(sr.get(k)))
    # بحث عام داخل JSON
    flat_rows = []
    for node in _walk_json(sr or data):
        if isinstance(node, dict):
            title = node.get("title") or node.get("name") or node.get("group") or node.get("label") or node.get("stage") or ""
            gl = _v11_group_letter_from_title(title)
            rows = _v11_rows_from_any(node)
            if rows:
                if gl:
                    candidates.append((_v11_group_title_from_letter(gl), rows[:4]))
                else:
                    flat_rows.extend(rows)
        elif isinstance(node, list):
            rows = _v11_rows_from_any(node)
            if rows:
                flat_rows.extend(rows)
    candidates.extend(_v11_group_rows_by_official_groups(flat_rows))
    # تنظيف وترتيب
    order = {g: i for i, (g, _) in enumerate(WORLD_CUP_GROUPS)}
    seen = set()
    out = []
    for title, rows in candidates:
        clean = []
        for r in rows:
            if r and r[0] in WORLD_CUP_TEAMS and r[0] not in [x[0] for x in clean]:
                clean.append((r[0], int(r[1]), int(r[2]), int(r[3])))
        if len(clean) < 2:
            continue
        gl = _v11_group_letter_from_title(title) or (V11_GROUP_BY_TEAM.get(clean[0][0]) or "")
        title = _v11_group_title_from_letter(gl) if gl else _v10_group_title(title)
        key = title + "|" + ",".join(x[0] for x in clean[:4])
        if key in seen:
            continue
        seen.add(key)
        clean.sort(key=lambda r: (r[3], r[2], r[1]), reverse=True)
        out.append((title, clean[:4]))
    def g_sort(item):
        title = item[0]
        gl = _v11_group_letter_from_title(title)
        return order.get(gl, 99)
    out.sort(key=g_sort)
    return out[:12]


def fetch_standings_from_serpapi():
    if not _serpapi_key():
        return []
    for q in V11_STANDINGS_QUERIES:
        for hl, gl in [("ar", "sa"), ("en", "us")]:
            try:
                data = serpapi_search_json(q, hl=hl, gl=gl, timeout=10)
                groups = _v11_extract_groups_from_google_json(data)
                if groups and (len(groups) >= 2 or sum(len(r) for _t, r in groups) >= 8):
                    return groups
            except Exception:
                continue
    return []


def fetch_current_groups(mode="latest"):
    # قوقل هو المصدر الأساسي الثابت بعد ظهور rankings في SerpApi
    for fn, label in [(fetch_standings_from_serpapi, "Google Sports"), (fetch_standings_from_api_football, "API-Football")]:
        try:
            groups = fn()
            if groups:
                return groups[:12], label
        except Exception:
            continue
    return [], ""


def _v11_google_match_queries(team1, team2, date_hint=None, source="google"):
    ar1 = canonical_team_name(team1) or normalize_name(team1)
    ar2 = canonical_team_name(team2) or normalize_name(team2)
    en1 = team_query_name(ar1) if 'team_query_name' in globals() else ar1
    en2 = team_query_name(ar2) if 'team_query_name' in globals() else ar2
    if source == "365":
        base = [
            f"site:365scores.com {ar1} {ar2} كأس العالم 2026",
            f"365Scores {ar1} {ar2} النتيجة من سجل",
            f"365Scores {en1} vs {en2} FIFA World Cup 2026",
        ]
    elif source == "kooora":
        base = [
            f"site:kooora.com {ar1} {ar2} كأس العالم 2026",
            f"كورة {ar1} {ar2} النتيجة من سجل",
            f"Kooora {en1} vs {en2} FIFA World Cup 2026",
        ]
    else:
        base = [
            f"مباراة {ar1} {ar2}",
            f"{ar1} ضد {ar2}",
            f"{ar1} × {ar2}",
            f"{en1} vs {en2} FIFA World Cup 2026",
            f"{en1} {en2} score",
        ]
    if date_hint:
        base = base + [f"{q} {date_hint}" for q in list(base)]
    out = []
    for q in base:
        q = normalize_name(q)
        if q and q not in out:
            out.append(q)
    return out[:8]


def _v11_parse_google_match_data(data, req1, req2, source_name="Google Sports"):
    sr = data.get("sports_results") if isinstance(data, dict) else None
    parsers = []
    if '_v8_parse_serp_sports_root' in globals():
        parsers.append(_v8_parse_serp_sports_root)
    if '_parse_serp_sports_node_v6' in globals():
        parsers.append(_parse_serp_sports_node_v6)
    roots = []
    if isinstance(sr, dict):
        roots.append(sr)
        for k in ["game_spotlight", "games", "matches", "scoreboard", "events", "timeline"]:
            if isinstance(sr.get(k), (dict, list)):
                roots.append(sr.get(k))
    roots.append(data)
    for root in roots:
        for parser in parsers:
            try:
                obj = parser(root, req1, req2, source_name=source_name)
            except TypeError:
                try:
                    obj = parser(root, req1, req2)
                    if obj:
                        obj["source"] = source_name
                except Exception:
                    obj = None
            except Exception:
                obj = None
            if obj and _v9_valid_score_obj(obj):
                obj["source"] = source_name
                scorers = _v10_collect_goal_details_from_google_card(data) if '_v10_collect_goal_details_from_google_card' in globals() else []
                if scorers:
                    obj["scorers"] = scorers
                    obj["goals_source"] = "Google Sports"
                link = _v10_extract_highlight_link(data) if '_v10_extract_highlight_link' in globals() else ""
                if link:
                    obj["highlight_url"] = link
                return obj
    # fallback من النص الكامل، مثل /بحث_قوقل الناجح
    try:
        blob = json.dumps(data, ensure_ascii=False)
        if _blob_has_both_teams(blob, req1, req2):
            s = _score_from_text_near_teams(blob, req1, req2) if '_score_from_text_near_teams' in globals() else None
            if s:
                obj = {
                    "team1": req1,
                    "team2": req2,
                    "score1": s.get("score1"),
                    "score2": s.get("score2"),
                    "status": s.get("status") or _status_from_serp_v6(data),
                    "minute": s.get("minute") or "",
                    "scorers": _v10_collect_goal_details_from_google_card(data) if '_v10_collect_goal_details_from_google_card' in globals() else [],
                    "source": source_name,
                }
                if obj.get("scorers"):
                    obj["goals_source"] = "Google Sports"
                link = _v10_extract_highlight_link(data) if '_v10_extract_highlight_link' in globals() else ""
                if link:
                    obj["highlight_url"] = link
                if _v9_valid_score_obj(obj):
                    return obj
    except Exception:
        pass
    return None


def _v10_fetch_match_from_serp_source(team1, team2, date_hint=None, source="google"):
    if not _serpapi_key():
        return None
    req1 = canonical_team_name(team1) or normalize_name(team1)
    req2 = canonical_team_name(team2) or normalize_name(team2)
    source_name = {"google": "Google Sports", "365": "365Scores عبر Google", "kooora": "Kooora عبر Google"}.get(source, "Google Sports")
    attempts = []
    for q in _v11_google_match_queries(req1, req2, date_hint, source=source):
        if source == "google":
            attempts.append((q, "ar", "sa"))
            attempts.append((q, "en", "us"))
        else:
            attempts.append((q, "ar", "sa"))
    seen = set()
    for q, hl, gl in attempts:
        key = (q, hl, gl)
        if key in seen:
            continue
        seen.add(key)
        try:
            data = serpapi_search_json(q, hl=hl, gl=gl, timeout=9)
        except Exception:
            continue
        obj = _v11_parse_google_match_data(data, req1, req2, source_name=source_name)
        if obj and _v9_valid_score_obj(obj):
            return obj
    return None


def fetch_match_from_serpapi(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")


def fetch_match_from_365(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="365")


def fetch_match_from_kooora(team1, team2, date_hint=None):
    return _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="kooora")


def fetch_live_match_data(team1, team2, mode="google", date_hint=None):
    norm = _norm_source_mode(mode)
    requested_label = mode_label_ar(norm)
    if norm in ["365", "kooora"]:
        seq = [norm, "google", "espn", "api"]
    elif norm == "latest":
        seq = ["google", "espn", "365", "kooora", "api"]
    elif norm == "official":
        seq = ["espn", "google", "api"]
    elif norm == "google":
        seq = ["google", "espn"]
    else:
        seq = ["google", "espn", "api"]
    for src in seq:
        try:
            obj = _v9_fetch_primary_from_source(team1, team2, src, date_hint=date_hint)
        except Exception:
            obj = None
        if not obj or not _v9_valid_score_obj(obj):
            continue
        obj["source"] = _v9_clean_source_name(obj.get("source") or ("ESPN" if src == "espn" else mode_label_ar(src)))
        obj["requested_source"] = requested_label
        obj["actual_source"] = obj["source"]
        if not obj.get("status"):
            obj["status"] = "مباشر" if obj.get("minute") else "غير محدد"
        scorers = _v9_sanitize_scorers(obj.get("scorers") or [])
        details_source = obj.get("goals_source") or (obj.get("source") if scorers else "")
        if not scorers:
            pref = norm if norm in ["365", "kooora"] else src
            scorers, details_source = _v9_fetch_goal_details(team1, team2, date_hint=date_hint, preferred_source=pref)
        obj["scorers"] = scorers
        obj["goals_source"] = details_source or "غير متوفر"
        if not obj.get("highlight_url") and src != "google" and _serpapi_key():
            try:
                gobj = _v10_fetch_match_from_serp_source(team1, team2, date_hint=date_hint, source="google")
                if gobj and gobj.get("highlight_url"):
                    obj["highlight_url"] = gobj.get("highlight_url")
            except Exception:
                pass
        return obj
    return None


def _source_help_text(kind, mode):
    if kind == "standings":
        return (
            f"تعذر جلب ترتيب مجموعات مؤكد من مصدر {mode_label_ar(mode)} ❌\n"
            "لن أعرض ترتيبًا غير موثوق.\n\n"
            "قوقل رجع بيانات لكن قد تكون البنية مختلفة؛ جرّب:\n"
            "/بحث_قوقل ترتيبات كأس العالم"
        )
    return f"تعذر جلب البيانات من مصدر {mode_label_ar(mode)} ❌"


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text)
    if mode_override:
        mode = _norm_source_mode(mode_override)
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر الولايات المتحدة * أستراليا * قوقل\nأو:\n/مباشر المكسيك * كوريا الجنوبية * 18/06/2026 * قوقل")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, date_hint), timeout=26)
    except Exception:
        data = None
    if not data:
        hint_cmd = f"/بحث_قوقل مباراة {team1} {team2}"
        await wait.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\n"
            f"مباراة: {team1} × {team2}\n" + (f"التاريخ: {date_hint}\n" if date_hint else "") +
            f"\nاختر مصدر آخر أو جرّب الأمر:\n{hint_cmd}",
            reply_markup=kb,
        )
        return
    path = render_live_match_card(data, mode_label_ar(mode))
    try:
        await wait.delete()
    except Exception:
        pass
    await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    text = update.message.text if getattr(update, 'message', None) else ""
    mode = mode_override
    if not mode:
        m = re.search(r"\*\s*(رسمي|سريع|الأحدث|الاحدث|official|fast|latest|google|قوقل|365|٣٦٥|كورة|كوره|kooora)\s*$", text or "", re.I)
        mode = m.group(1) if m else "latest"
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text("⏳ جاري جلب ترتيب المجموعات من قوقل...")
    try:
        groups, source_label = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_current_groups, mode), timeout=28)
    except Exception:
        groups, source_label = [], ""
    if not groups:
        await wait.edit_text(_source_help_text("standings", mode) + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    path = create_all_groups_newlook_image(groups) if 'create_all_groups_newlook_image' in globals() else create_all_groups_image(groups)
    caption = f"ترتيب المجموعات الآن ✅\nالمصدر: {source_label}\nالاستعلام المعتمد: ترتيبات كأس العالم"
    try:
        await wait.delete()
    except Exception:
        pass
    await send_photo_path_markup(update.message, path, caption, kb)
    await update.message.reply_text(build_groups_text(groups, source_label))


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    body = parse_command_body_lines(update.message.text)
    query = " ".join(body).strip() or "مباراة المكسيك كوريا الجنوبية"
    msg = await update.message.reply_text(f"⏳ أفحص قوقل: {query}")
    try:
        data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(serpapi_search_json, query, "ar", "sa", 10), timeout=14)
        sr = data.get("sports_results") if isinstance(data, dict) else None
        lines = ["نتيجة فحص قوقل:", f"query: {query}"]
        if isinstance(sr, dict):
            lines.append("✅ sports_results موجود")
            lines.append("مفاتيح: " + "، ".join(list(sr.keys())[:12]))
            if "rankings" in sr:
                groups = _v11_extract_groups_from_google_json(data)
                if groups:
                    lines.append(f"✅ قرأت ترتيب المجموعات: {len(groups)} مجموعة")
                    for title, rows in groups[:3]:
                        lines.append(f"{title}: " + "، ".join([f"{r[0]} {r[3]}" for r in rows[:4]]))
                else:
                    lines.append("⚠️ rankings موجود لكن لم أقرأ صفوف ترتيب مؤكدة")
            else:
                req1, req2 = "المكسيك", "كوريا الجنوبية"
                txt = normalize_name(query)
                found = []
                for tm in WORLD_CUP_TEAMS:
                    if simple_key(tm) in simple_key(txt):
                        found.append(tm)
                    if len(found) >= 2:
                        break
                if len(found) >= 2:
                    req1, req2 = found[0], found[1]
                obj = _v11_parse_google_match_data(data, req1, req2, source_name="Google Sports")
                if obj:
                    lines.append(f"✅ قرأت المباراة: {obj['team1']} {obj['score1']} - {obj['score2']} {obj['team2']}")
                    lines.append(f"الحالة: {obj.get('status','')}")
                    sc = _v9_sanitize_scorers(obj.get("scorers") or [])
                    if sc:
                        lines.append("الهدافون: " + "، ".join(sc[:4]))
                else:
                    lines.append("⚠️ وصلنا لقوقل لكن لم أقرأ نتيجة مؤكدة من البنية الحالية")
        else:
            lines.append("⚠️ لم يظهر sports_results")
        await msg.edit_text("\n".join(lines[:16]))
    except Exception as e:
        await msg.edit_text(f"❌ فشل فحص قوقل: {str(e)[:220]}")


async def sports_source_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    await query.answer()
    if not is_admin_user(update):
        await query.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (query.data or "").split("|")
    if len(parts) != 3:
        await query.message.reply_text("تعذر قراءة الخيار.")
        return
    _tag, token, mode = parts
    payload = context.bot_data.get("sports_source_requests", {}).get(token)
    if not payload:
        await query.message.reply_text("انتهت صلاحية الخيار، أعد تنفيذ الأمر من جديد.")
        return
    kind = payload.get("kind")
    kb = source_keyboard(context, payload)
    if kind == "standings":
        fake_update = Update(update.update_id, message=query.message)
        # بدل fake معقد: نفذ نفس المنطق وأرسل في نفس المحادثة
        msg = await query.message.reply_text(f"⏳ جاري جلب ترتيب المجموعات من {mode_label_ar(mode)}...")
        try:
            groups, src = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_current_groups, mode), timeout=28)
        except Exception:
            groups, src = [], ""
        if not groups:
            await msg.edit_text(_source_help_text("standings", mode), reply_markup=kb)
            return
        path = create_all_groups_newlook_image(groups) if 'create_all_groups_newlook_image' in globals() else create_all_groups_image(groups)
        try:
            await msg.delete()
        except Exception:
            pass
        await send_photo_path_markup(query.message, path, f"ترتيب المجموعات الآن ✅\nالمصدر: {src}", kb)
        await query.message.reply_text(build_groups_text(groups, src))
        return
    if kind == "live":
        team1 = payload.get("team1")
        team2 = payload.get("team2")
        date_hint = payload.get("date_hint")
        msg = await query.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
        try:
            data = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_live_match_data, team1, team2, mode, date_hint), timeout=26)
        except Exception:
            data = None
        if not data:
            await msg.edit_text(f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\nمباراة: {team1} × {team2}\n\nجرّب:\n/بحث_قوقل مباراة {team1} {team2}", reply_markup=kb)
            return
        path = render_live_match_card(data, mode_label_ar(mode))
        try:
            await msg.delete()
        except Exception:
            pass
        await send_photo_path_markup(query.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        return
    await query.message.reply_text("تعذر تحديد نوع الطلب.")

# ==================== END V11 FINAL PATCH ====================


# ==================== V14 SAFE START PATCH ====================
# نسخة آمنة مبنية على آخر نسخة كانت تعمل. الهدف الأساسي: تشغيل البوت بدون كسر /start.
# - يقبل أكثر من اسم لمتغير التوكن.
# - يمنع دخول أسماء ملفات الأعلام في أسماء المنتخبات قدر الإمكان.
# - يحافظ على أن المصدر في كابشن /مباشر وليس داخل الصورة.

try:
    _V14_PREV_PARSE_LIVE_COMMAND_TEXT = parse_live_command_text
except Exception:
    _V14_PREV_PARSE_LIVE_COMMAND_TEXT = None

_V14_FLAG_TO_TEAM = {
    "scotland.png": "اسكتلندا",
    "morocco.png": "المغرب",
    "usa.png": "الولايات المتحدة",
    "united_states.png": "الولايات المتحدة",
    "australia.png": "أستراليا",
    "mexico.png": "المكسيك",
    "south_korea.png": "كوريا الجنوبية",
    "korea_republic.png": "كوريا الجنوبية",
}

def _v14_clean_team_value(x):
    try:
        s = normalize_name(x)
    except Exception:
        s = str(x or "").strip()
    low = s.lower().strip()
    if low in _V14_FLAG_TO_TEAM:
        return _V14_FLAG_TO_TEAM[low]
    # لو جاء مسار كامل للعلم خذ اسم الملف فقط ثم حوله
    base = os.path.basename(low)
    if base in _V14_FLAG_TO_TEAM:
        return _V14_FLAG_TO_TEAM[base]
    if low.endswith(('.png', '.jpg', '.jpeg', '.webp')):
        # لا نسمح لاسم ملف الصورة يدخل البحث؛ نحاول نرجع اسم نظيف أو نتركه فارغ
        name = re.sub(r'\.(png|jpg|jpeg|webp)$', '', base, flags=re.I)
        name = name.replace('_', ' ').replace('-', ' ')
        can = canonical_team_name(name) if 'canonical_team_name' in globals() else None
        return can or name
    can = canonical_team_name(s) if 'canonical_team_name' in globals() else None
    return can or s

if _V14_PREV_PARSE_LIVE_COMMAND_TEXT:
    def parse_live_command_text(text):
        t1, t2, mode, date_hint = _V14_PREV_PARSE_LIVE_COMMAND_TEXT(text)
        return _v14_clean_team_value(t1), _v14_clean_team_value(t2), mode, date_hint

try:
    _V14_PREV_RENDER_QUALIFIED32_BOARD = render_qualified32_board
    def render_qualified32_board(teams):
        path = _V14_PREV_RENDER_QUALIFIED32_BOARD(teams)
        try:
            img = Image.open(path).convert('RGB')
            d = ImageDraw.Draw(img)
            w, h = img.size
            rounded_rect(d, (w//2-245, h-53, w//2+245, h-12), radius=16, fill="#061633D8", outline="#FFFFFF55", width=1)
            draw_text(d, (w//2, h-32), "المصيف يضعكم بالحدث", get_font(22), fill="#FDE68A", max_width=440)
            img.save(path, quality=96)
        except Exception:
            pass
        return path
except Exception:
    pass

# ==================== END V14 SAFE START PATCH ====================



# ==================== V22 HARD STABLE PATCH: ESPN RESULTS -> GROUP STANDINGS ====================
# الهدف: إيقاف دوامة Google/Playwright/Docker نهائيًا.
# ترتيب المجموعات هنا يسحب نتائج المباريات من ESPN JSON المباشر، ثم يحسب الترتيب فورًا.
# لا يحتاج Docker ولا Playwright ولا Chromium. requests فقط.

V22_GROUPS_SOURCE_LABEL = "ESPN Live Results"


def _v22_event_is_countable(event, obj):
    """نحسب المباراة إذا بدأت أو انتهت. نستبعد المجدولة فقط."""
    try:
        comp = (event.get("competitions") or [event])[0]
        st = comp.get("status") or event.get("status") or {}
        typ = st.get("type") or {}
        state = normalize_name(typ.get("state") or typ.get("name") or typ.get("description") or typ.get("detail") or "").lower()
        detail = normalize_name(typ.get("detail") or typ.get("shortDetail") or st.get("displayClock") or obj.get("status") or obj.get("minute") or "").lower()
        if any(x in state + " " + detail for x in ["pre", "scheduled", "not started", "لم تبدأ"]):
            return False
        # إذا فيه نتيجة رقمية نعدها، حتى لو مباشر.
        int(float(obj.get("score1", 0)))
        int(float(obj.get("score2", 0)))
        return True
    except Exception:
        return False


def _v22_date_range_for_worldcup():
    """من بداية البطولة إلى اليوم + يومين. ثابت وخفيف."""
    start = datetime(2026, 6, 11).date()
    # لا نخلي التاريخ قبل بداية البطولة حتى لو ساعة السيرفر مختلفة
    today = max(datetime.utcnow().date(), start)
    end = min(today + timedelta(days=2), datetime(2026, 7, 1).date())
    dates = []
    d = start
    while d <= end:
        dates.append(d.strftime("%Y%m%d"))
        d += timedelta(days=1)
    return dates


def fetch_standings_from_espn_live_results():
    """يسحب نتائج ESPN لكل مباريات المجموعات، ثم يحسب الترتيب.
    يرجع نفس شكل التصميم المعتمد: [(المجموعة أ, [(team, played, gd, pts), ...]), ...]
    """
    # جدول كامل لكل منتخب داخل مجموعته
    group_by_team = {}
    for g, teams in WORLD_CUP_GROUPS:
        for t in teams:
            group_by_team[t] = g
    table = {t: {"played": 0, "wins": 0, "draws": 0, "losses": 0, "gf": 0, "ga": 0, "gd": 0, "pts": 0} for t in WORLD_CUP_TEAMS}

    found_any = False
    seen = set()
    for dv in _v22_date_range_for_worldcup():
        events = []
        try:
            events = _fetch_espn_events_by_date(dv) or []
        except Exception:
            events = []
        for event in events:
            try:
                eid = str(event.get("id") or event.get("uid") or "")
                if eid and eid in seen:
                    continue
                if eid:
                    seen.add(eid)
                obj = _parse_espn_match_from_event(event)
                if not obj:
                    continue
                t1 = canonical_team_name(obj.get("team1")) or normalize_name(obj.get("team1"))
                t2 = canonical_team_name(obj.get("team2")) or normalize_name(obj.get("team2"))
                if t1 not in table or t2 not in table:
                    continue
                # لا نحسب مباريات من مجموعات مختلفة أو مباريات ودية/خارج البطولة
                if group_by_team.get(t1) != group_by_team.get(t2):
                    continue
                if not _v22_event_is_countable(event, obj):
                    continue
                s1 = int(float(obj.get("score1", 0)))
                s2 = int(float(obj.get("score2", 0)))
                found_any = True
                table[t1]["played"] += 1
                table[t2]["played"] += 1
                table[t1]["gf"] += s1
                table[t1]["ga"] += s2
                table[t2]["gf"] += s2
                table[t2]["ga"] += s1
                table[t1]["gd"] = table[t1]["gf"] - table[t1]["ga"]
                table[t2]["gd"] = table[t2]["gf"] - table[t2]["ga"]
                if s1 > s2:
                    table[t1]["wins"] += 1; table[t2]["losses"] += 1
                    table[t1]["pts"] += 3
                elif s2 > s1:
                    table[t2]["wins"] += 1; table[t1]["losses"] += 1
                    table[t2]["pts"] += 3
                else:
                    table[t1]["draws"] += 1; table[t2]["draws"] += 1
                    table[t1]["pts"] += 1; table[t2]["pts"] += 1
            except Exception:
                continue
    if not found_any:
        return []

    groups = []
    ar_letters = {"A":"أ", "B":"ب", "C":"ج", "D":"د", "E":"هـ", "F":"و", "G":"ز", "H":"ح", "I":"ط", "J":"ي", "K":"ك", "L":"ل"}
    for g, teams in WORLD_CUP_GROUPS:
        rows = []
        for t in teams:
            r = table[t]
            rows.append((t, int(r["played"]), int(r["gd"]), int(r["pts"])))
        # ترتيب قياسي: نقاط، فارق، أهداف له، الاسم لضمان ثبات الترتيب
        rows.sort(key=lambda row: (row[3], row[2], table[row[0]]["gf"], row[0]), reverse=True)
        groups.append((f"المجموعة {ar_letters.get(g, g)}", rows))
    return groups


def fetch_current_groups(mode="latest"):
    """V22: المصدر الثابت لترتيب المجموعات = نتائج ESPN المباشرة فقط."""
    groups = fetch_standings_from_espn_live_results()
    if groups:
        return groups, V22_GROUPS_SOURCE_LABEL
    return [], ""


def _source_help_text(kind, mode):
    if kind == "standings":
        return (
            "تعذر جلب نتائج المجموعات من ESPN حاليًا ❌\n"
            "البوت شغال، لكن مصدر النتائج لم يرجع مباريات كافية الآن.\n\n"
            "جرّب بعد دقيقة أو اختر مصدر آخر."
        )
    return f"تعذر جلب البيانات من مصدر {mode_label_ar(mode)} ❌"


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    """V22: أمر ثابت لا يستخدم Google/Playwright، فقط ESPN JSON."""
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text("⏳ أسحب نتائج المجموعات من ESPN وأحسب الترتيب...")
    try:
        groups, source_label = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_current_groups, "espn"), timeout=35)
    except Exception as e:
        groups, source_label = [], ""
        try:
            await wait.edit_text(f"تعذر جلب ترتيب المجموعات ❌\n{str(e)[:250]}", reply_markup=kb)
            return
        except Exception:
            pass
    if not groups:
        await wait.edit_text(_source_help_text("standings", "espn") + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    try:
        path = create_all_groups_newlook_image(groups) if 'create_all_groups_newlook_image' in globals() else create_all_groups_image(groups)
    except Exception:
        path = None
    caption = "ترتيب المجموعات الآن ✅\nالمصدر: ESPN — محسوب من نتائج المباريات المباشرة"
    try:
        await wait.delete()
    except Exception:
        pass
    if path and os.path.exists(path):
        await send_photo_path_markup(update.message, path, caption, kb)
    else:
        await update.message.reply_text(caption, reply_markup=kb)
    await update.message.reply_text(build_groups_text(groups, "ESPN Live Results"))


async def google_search_debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """نترك اسم الأمر كما هو، لكن للترتيب نعرض تشخيص ESPN الثابت بدل دوامة قوقل."""
    body = parse_command_body_lines(update.message.text)
    query = " ".join(body).strip() or "مباراة المكسيك كوريا الجنوبية"
    if any(x in normalize_name(query) for x in ["ترتيب", "ترتيبات", "standings", "المجموعات"]):
        msg = await update.message.reply_text("⏳ أفحص ESPN لترتيب المجموعات...")
        try:
            groups = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_standings_from_espn_live_results), timeout=35)
        except Exception as e:
            await msg.edit_text("نتيجة فحص ESPN:\n" + f"error: {str(e)[:500]}")
            return
        if not groups:
            await msg.edit_text("نتيجة فحص ESPN:\nلم يرجع مباريات كافية لحساب الترتيب الآن.")
            return
        lines = ["نتيجة فحص ESPN:", f"groups_read: {len(groups)}"]
        for title, rows in groups[:4]:
            lines.append(f"{title}: " + "، ".join([f"{r[0]} {r[3]}" for r in rows[:4]]))
        await msg.edit_text("\n".join(lines[:25]))
        return
    # للبحث العادي نستخدم الأمر السابق إن وجد، وإلا رد مختصر
    try:
        if '_V11_PREV_GOOGLE_SEARCH_DEBUG_COMMAND' in globals() and _V11_PREV_GOOGLE_SEARCH_DEBUG_COMMAND:
            return await _V11_PREV_GOOGLE_SEARCH_DEBUG_COMMAND(update, context)
    except Exception:
        pass
    await update.message.reply_text("استخدم /ترتيب_المجموعات_الان لترتيب المجموعات أو /مباشر للمباريات.")


# ==================== V23 CLEAN FIXTURES + STANDINGS PAGES + ESPN GOALS PATCH ====================
# - حذف تفعيل أوامر التصميم القديمة من main.
# - جدول المباريات من ملف PDF يبدأ من السبت 20 يونيو 2026.
# - /مباريات: أزرار الأيام + تحديث مباريات تحدد لاحقًا بدون تصميم مباشر.
# - /مباريات_مجمعة: أكثر من يوم في تصميم واحد، وإذا كثرت الأيام يقسم تلقائيًا.
# - /ترتيب_المجموعات_الان: صورة شاملة + 3 صور واضحة.
# - تحسين محاولة سحب الهدافين من تفاصيل ESPN Summary.

FIXTURES_UPDATES_FILE = os.path.join("data", "fixture_updates.json")

_GROUP_AR = {
    1: "المجموعة أ", 2: "المجموعة ب", 3: "المجموعة ج", 4: "المجموعة د",
    5: "المجموعة هـ", 6: "المجموعة و", 7: "المجموعة ز", 8: "المجموعة ح",
    9: "المجموعة ط", 10: "المجموعة ي", 11: "المجموعة ك", 12: "المجموعة ل",
}

# جدول ثابت من ملف PDF — البداية المعتمدة: السبت 20 يونيو 2026.
# الوقت كما في الملف: توقيت السعودية.
TOURNAMENT_FIXTURES = [
    # دور المجموعات — من 20 يونيو
    {"id":"G20-1","date":"20/06/2026","day":"السبت","time":"8:00 م","team1":"هولندا","team2":"السويد","stage":"دور المجموعات","group":"المجموعة و"},
    {"id":"G20-2","date":"20/06/2026","day":"السبت","time":"11:00 م","team1":"ألمانيا","team2":"ساحل العاج","stage":"دور المجموعات","group":"المجموعة هـ"},
    {"id":"G20-3","date":"20/06/2026","day":"السبت","time":"3:00 فجراً","team1":"الإكوادور","team2":"كوراساو","stage":"دور المجموعات","group":"المجموعة هـ"},
    {"id":"G20-4","date":"20/06/2026","day":"السبت","time":"7:00 صباحاً","team1":"تونس","team2":"اليابان","stage":"دور المجموعات","group":"المجموعة و"},

    {"id":"G21-1","date":"21/06/2026","day":"الأحد","time":"7:00 م","team1":"إسبانيا","team2":"السعودية","stage":"دور المجموعات","group":"المجموعة ح"},
    {"id":"G21-2","date":"21/06/2026","day":"الأحد","time":"10:00 م","team1":"بلجيكا","team2":"إيران","stage":"دور المجموعات","group":"المجموعة ز"},
    {"id":"G21-3","date":"21/06/2026","day":"الأحد","time":"1:00 فجراً","team1":"الأوروغواي","team2":"الرأس الأخضر","stage":"دور المجموعات","group":"المجموعة ح"},
    {"id":"G21-4","date":"21/06/2026","day":"الأحد","time":"4:00 فجراً","team1":"نيوزيلندا","team2":"مصر","stage":"دور المجموعات","group":"المجموعة ز"},

    {"id":"G22-1","date":"22/06/2026","day":"الإثنين","time":"8:00 م","team1":"الأرجنتين","team2":"النمسا","stage":"دور المجموعات","group":"المجموعة ي"},
    {"id":"G22-2","date":"22/06/2026","day":"الإثنين","time":"12:00 منتصف الليل","team1":"فرنسا","team2":"العراق","stage":"دور المجموعات","group":"المجموعة ط"},
    {"id":"G22-3","date":"22/06/2026","day":"الإثنين","time":"3:00 فجراً","team1":"النرويج","team2":"السنغال","stage":"دور المجموعات","group":"المجموعة ط"},
    {"id":"G22-4","date":"22/06/2026","day":"الإثنين","time":"6:00 صباحاً","team1":"الأردن","team2":"الجزائر","stage":"دور المجموعات","group":"المجموعة ي"},

    {"id":"G23-1","date":"23/06/2026","day":"الثلاثاء","time":"8:00 م","team1":"البرتغال","team2":"أوزبكستان","stage":"دور المجموعات","group":"المجموعة ك"},
    {"id":"G23-2","date":"23/06/2026","day":"الثلاثاء","time":"11:00 م","team1":"إنجلترا","team2":"غانا","stage":"دور المجموعات","group":"المجموعة ل"},
    {"id":"G23-3","date":"23/06/2026","day":"الثلاثاء","time":"2:00 فجراً","team1":"بنما","team2":"كرواتيا","stage":"دور المجموعات","group":"المجموعة ل"},
    {"id":"G23-4","date":"23/06/2026","day":"الثلاثاء","time":"5:00 فجراً","team1":"كولومبيا","team2":"الكونغو الديمقراطية","stage":"دور المجموعات","group":"المجموعة ك"},

    {"id":"G24-1","date":"24/06/2026","day":"الأربعاء","time":"10:00 م","team1":"سويسرا","team2":"كندا","stage":"دور المجموعات","group":"المجموعة ب"},
    {"id":"G24-2","date":"24/06/2026","day":"الأربعاء","time":"10:00 م","team1":"البوسنة والهرسك","team2":"قطر","stage":"دور المجموعات","group":"المجموعة ب"},
    {"id":"G24-3","date":"24/06/2026","day":"الأربعاء","time":"1:00 فجراً","team1":"اسكتلندا","team2":"البرازيل","stage":"دور المجموعات","group":"المجموعة ج"},
    {"id":"G24-4","date":"24/06/2026","day":"الأربعاء","time":"1:00 فجراً","team1":"المغرب","team2":"هايتي","stage":"دور المجموعات","group":"المجموعة ج"},
    {"id":"G24-5","date":"24/06/2026","day":"الأربعاء","time":"4:00 فجراً","team1":"التشيك","team2":"المكسيك","stage":"دور المجموعات","group":"المجموعة أ"},
    {"id":"G24-6","date":"24/06/2026","day":"الأربعاء","time":"4:00 فجراً","team1":"جنوب أفريقيا","team2":"كوريا الجنوبية","stage":"دور المجموعات","group":"المجموعة أ"},

    {"id":"G25-1","date":"25/06/2026","day":"الخميس","time":"11:00 م","team1":"كوراساو","team2":"ساحل العاج","stage":"دور المجموعات","group":"المجموعة هـ"},
    {"id":"G25-2","date":"25/06/2026","day":"الخميس","time":"11:00 م","team1":"الإكوادور","team2":"ألمانيا","stage":"دور المجموعات","group":"المجموعة هـ"},
    {"id":"G25-3","date":"25/06/2026","day":"الخميس","time":"2:00 فجراً","team1":"اليابان","team2":"السويد","stage":"دور المجموعات","group":"المجموعة و"},
    {"id":"G25-4","date":"25/06/2026","day":"الخميس","time":"2:00 فجراً","team1":"تونس","team2":"هولندا","stage":"دور المجموعات","group":"المجموعة و"},
    {"id":"G25-5","date":"25/06/2026","day":"الخميس","time":"5:00 فجراً","team1":"تركيا","team2":"الولايات المتحدة","stage":"دور المجموعات","group":"المجموعة د"},
    {"id":"G25-6","date":"25/06/2026","day":"الخميس","time":"5:00 فجراً","team1":"باراغواي","team2":"أستراليا","stage":"دور المجموعات","group":"المجموعة د"},

    {"id":"G26-1","date":"26/06/2026","day":"الجمعة","time":"10:00 م","team1":"النرويج","team2":"فرنسا","stage":"دور المجموعات","group":"المجموعة ط"},
    {"id":"G26-2","date":"26/06/2026","day":"الجمعة","time":"10:00 م","team1":"السنغال","team2":"العراق","stage":"دور المجموعات","group":"المجموعة ط"},
    {"id":"G26-3","date":"26/06/2026","day":"الجمعة","time":"3:00 فجراً","team1":"الرأس الأخضر","team2":"السعودية","stage":"دور المجموعات","group":"المجموعة ح"},
    {"id":"G26-4","date":"26/06/2026","day":"الجمعة","time":"3:00 فجراً","team1":"الأوروغواي","team2":"إسبانيا","stage":"دور المجموعات","group":"المجموعة ح"},
    {"id":"G26-5","date":"26/06/2026","day":"الجمعة","time":"6:00 صباحاً","team1":"مصر","team2":"إيران","stage":"دور المجموعات","group":"المجموعة ز"},
    {"id":"G26-6","date":"26/06/2026","day":"الجمعة","time":"6:00 صباحاً","team1":"نيوزيلندا","team2":"بلجيكا","stage":"دور المجموعات","group":"المجموعة ز"},

    {"id":"G27-1","date":"27/06/2026","day":"السبت","time":"12:00 منتصف الليل","team1":"بنما","team2":"إنجلترا","stage":"دور المجموعات","group":"المجموعة ل"},
    {"id":"G27-2","date":"27/06/2026","day":"السبت","time":"12:00 منتصف الليل","team1":"كرواتيا","team2":"غانا","stage":"دور المجموعات","group":"المجموعة ل"},
    {"id":"G27-3","date":"27/06/2026","day":"السبت","time":"2:30 فجراً","team1":"كولومبيا","team2":"البرتغال","stage":"دور المجموعات","group":"المجموعة ك"},
    {"id":"G27-4","date":"27/06/2026","day":"السبت","time":"2:30 فجراً","team1":"الكونغو الديمقراطية","team2":"أوزبكستان","stage":"دور المجموعات","group":"المجموعة ك"},
    {"id":"G27-5","date":"27/06/2026","day":"السبت","time":"5:00 فجراً","team1":"الجزائر","team2":"النمسا","stage":"دور المجموعات","group":"المجموعة ي"},
    {"id":"G27-6","date":"27/06/2026","day":"السبت","time":"5:00 فجراً","team1":"الأردن","team2":"الأرجنتين","stage":"دور المجموعات","group":"المجموعة ي"},

    # دور الـ32 — أطراف قابلة للتحديث من داخل البوت
    {"id":"R32-1","date":"28/06/2026","day":"الأحد","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"ثاني المجموعة أ × ثاني المجموعة ب"},
    {"id":"R32-2","date":"29/06/2026","day":"الإثنين","time":"8:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ج × ثاني المجموعة و"},
    {"id":"R32-3","date":"29/06/2026","day":"الإثنين","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة هـ × ثالث مؤهل"},
    {"id":"R32-4","date":"30/06/2026","day":"الثلاثاء","time":"4:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة و × ثاني المجموعة ج"},
    {"id":"R32-5","date":"30/06/2026","day":"الثلاثاء","time":"8:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"ثاني المجموعة هـ × ثاني المجموعة ط"},
    {"id":"R32-6","date":"01/07/2026","day":"الأربعاء","time":"12:00 منتصف الليل","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ط × ثالث مؤهل"},
    {"id":"R32-7","date":"01/07/2026","day":"الأربعاء","time":"4:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة أ × ثالث مؤهل"},
    {"id":"R32-8","date":"01/07/2026","day":"الأربعاء","time":"7:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ل × ثالث مؤهل"},
    {"id":"R32-9","date":"01/07/2026","day":"الأربعاء","time":"11:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ز × ثالث مؤهل"},
    {"id":"R32-10","date":"02/07/2026","day":"الخميس","time":"3:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة د × ثالث مؤهل"},
    {"id":"R32-11","date":"02/07/2026","day":"الخميس","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ح × ثاني المجموعة ي"},
    {"id":"R32-12","date":"03/07/2026","day":"الجمعة","time":"2:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"ثاني المجموعة ك × ثاني المجموعة ل"},
    {"id":"R32-13","date":"03/07/2026","day":"الجمعة","time":"6:00 صباحاً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ب × ثالث مؤهل"},
    {"id":"R32-14","date":"03/07/2026","day":"الجمعة","time":"9:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"ثاني المجموعة د × ثاني المجموعة ز"},
    {"id":"R32-15","date":"04/07/2026","day":"السبت","time":"1:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ي × ثاني المجموعة ح"},
    {"id":"R32-16","date":"04/07/2026","day":"السبت","time":"4:30 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ32","note":"أول المجموعة ك × ثالث مؤهل"},

    # دور الـ16 وما بعده — يتم تحديثها لاحقًا من داخل البوت
    {"id":"R16-1","date":"04/07/2026","day":"السبت","time":"8:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 1 × الفائز من مباراة 3"},
    {"id":"R16-2","date":"05/07/2026","day":"الأحد","time":"12:00 منتصف الليل","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 2 × الفائز من مباراة 5"},
    {"id":"R16-3","date":"05/07/2026","day":"الأحد","time":"11:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 4 × الفائز من مباراة 6"},
    {"id":"R16-4","date":"06/07/2026","day":"الإثنين","time":"3:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 7 × الفائز من مباراة 8"},
    {"id":"R16-5","date":"06/07/2026","day":"الإثنين","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 11 × الفائز من مباراة 12"},
    {"id":"R16-6","date":"07/07/2026","day":"الثلاثاء","time":"3:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 9 × الفائز من مباراة 10"},
    {"id":"R16-7","date":"07/07/2026","day":"الثلاثاء","time":"7:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 13 × الفائز من مباراة 15"},
    {"id":"R16-8","date":"07/07/2026","day":"الثلاثاء","time":"11:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"دور الـ16","note":"الفائز من مباراة 14 × الفائز من مباراة 16"},
    {"id":"QF-1","date":"09/07/2026","day":"الخميس","time":"11:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور ربع النهائي","note":"الفائز من ثمن النهائي الأول × الفائز من ثمن النهائي الثاني"},
    {"id":"QF-2","date":"10/07/2026","day":"الجمعة","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور ربع النهائي","note":"الفائز من ثمن النهائي الخامس × الفائز من ثمن النهائي السادس"},
    {"id":"QF-3","date":"12/07/2026","day":"الأحد","time":"12:00 منتصف الليل","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور ربع النهائي","note":"الفائز من ثمن النهائي الثالث × الفائز من ثمن النهائي الرابع"},
    {"id":"QF-4","date":"12/07/2026","day":"الأحد","time":"4:00 فجراً","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور ربع النهائي","note":"الفائز من ثمن النهائي السابع × الفائز من ثمن النهائي الثامن"},
    {"id":"SF-1","date":"14/07/2026","day":"الثلاثاء","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور نصف النهائي","note":"الفائز من ربع النهائي الأول × الفائز من ربع النهائي الثاني"},
    {"id":"SF-2","date":"15/07/2026","day":"الأربعاء","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"الدور نصف النهائي","note":"الفائز من ربع النهائي الثالث × الفائز من ربع النهائي الرابع"},
    {"id":"3RD","date":"19/07/2026","day":"الأحد","time":"12:00 منتصف الليل","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"مباراة تحديد المركز الثالث","note":"الخاسر من نصف النهائي الأول × الخاسر من نصف النهائي الثاني"},
    {"id":"FINAL","date":"19/07/2026","day":"الأحد","time":"10:00 م","team1":"تحدد لاحقًا","team2":"تحدد لاحقًا","stage":"المباراة النهائية","note":"الفائز من نصف النهائي الأول × الفائز من نصف النهائي الثاني"},
]


def _ensure_data_dir():
    os.makedirs("data", exist_ok=True)
    os.makedirs(GENERATED_DIR, exist_ok=True)


def _load_fixture_updates():
    _ensure_data_dir()
    try:
        if os.path.exists(FIXTURES_UPDATES_FILE):
            with open(FIXTURES_UPDATES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}


def _save_fixture_updates(data):
    _ensure_data_dir()
    tmp = FIXTURES_UPDATES_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data or {}, f, ensure_ascii=False, indent=2)
    os.replace(tmp, FIXTURES_UPDATES_FILE)


def _fixture_by_id(match_id):
    for m in TOURNAMENT_FIXTURES:
        if str(m.get("id")) == str(match_id):
            return dict(m)
    return None


def _apply_fixture_updates(match):
    m = dict(match)
    upd = _load_fixture_updates().get(m.get("id"), {})
    if isinstance(upd, dict):
        for k in ["team1", "team2", "time", "note"]:
            if upd.get(k):
                m[k] = upd[k]
    return m


def _all_fixtures():
    return [_apply_fixture_updates(m) for m in TOURNAMENT_FIXTURES]


def _date_key(d):
    try:
        a,b,c = str(d).split("/")
        return (int(c), int(b), int(a))
    except Exception:
        return (9999, 99, 99)


def _fixture_dates():
    seen = {}
    for m in TOURNAMENT_FIXTURES:
        d = m["date"]
        if d not in seen:
            seen[d] = m.get("day", "")
    return sorted(seen.items(), key=lambda x: _date_key(x[0]))


def _normalize_date_arg(x):
    s = str(x or "").strip()
    s = s.replace("-", "/")
    m = re.search(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", s)
    if not m:
        return ""
    d, mo, y = m.groups()
    y = y or "2026"
    if len(y) == 2:
        y = "20" + y
    return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"


def _extract_fixture_dates_from_text(text):
    body = re.sub(r"^/\S+", "", text or "").strip()
    found = []
    for m in re.finditer(r"\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?", body):
        d = _normalize_date_arg(m.group(0))
        if d and d not in found:
            found.append(d)
    # مدى: /مباريات 20/06 إلى 24/06
    if any(w in body for w in ["إلى", "الى", "لحد", "حتى"] ) and len(found) >= 2:
        start, end = found[0], found[1]
        dates = [d for d,_day in _fixture_dates() if _date_key(start) <= _date_key(d) <= _date_key(end)]
        return dates
    return found


def _fixtures_for_date(date):
    d = _normalize_date_arg(date)
    return [m for m in _all_fixtures() if m.get("date") == d]


def _has_unknown(m):
    return "تحدد" in str(m.get("team1")) or "تحدد" in str(m.get("team2"))


def _fixtures_caption(date_or_title, source="PDF جدول البطولة"):
    return f"{date_or_title}\nالمصدر: {source}\nالمصيف يضعكم بالحدث"


def _fixture_title(date):
    d = _normalize_date_arg(date)
    day = ""
    for x, dy in _fixture_dates():
        if x == d:
            day = dy
            break
    return f"{day} {d[:5]}".strip()


def _draw_fixture_card(draw, img, box, m, idx=1):
    x0,y0,x1,y1 = box
    try:
        draw.rounded_rectangle(box, radius=24, fill=(5,24,58,228), outline=(38,151,255,220), width=3)
    except Exception:
        draw.rectangle(box, fill=(5,24,58), outline=(38,151,255), width=3)
    header = f"{m.get('stage','')}" + (f" | {m.get('group')}" if m.get('group') else "")
    draw_text(draw, ((x0+x1)//2, y0+28), header, get_font(24), fill="#FBBF24", max_width=x1-x0-40)
    # flags
    fw, fh = 82, 58
    try:
        paste_flag(img, m.get("team1"), (x0+34, y0+56, x0+34+fw, y0+56+fh))
        paste_flag(img, m.get("team2"), (x1-34-fw, y0+56, x1-34, y0+56+fh))
    except Exception:
        pass
    draw_text(draw, (x0+150, y0+88), m.get("team1"), get_font(30), fill="#FFFFFF", anchor="lm", max_width=230)
    draw_text(draw, (x1-150, y0+88), m.get("team2"), get_font(30), fill="#FFFFFF", anchor="rm", max_width=230)
    draw_text(draw, ((x0+x1)//2, y0+88), "×", get_font(42), fill="#FBBF24")
    draw_text(draw, ((x0+x1)//2, y0+138), m.get("time", ""), get_font(28), fill="#FFFFFF")
    if _has_unknown(m) and m.get("note"):
        draw_text(draw, ((x0+x1)//2, y0+176), m.get("note"), get_font(22), fill="#CBD5E1", max_width=x1-x0-50)


def _fixture_bg(w=1080, h=1350):
    # نستخدم خلفية مباريات اليوم V31 إن وجدت
    for p in ["games_v31_full_bg.png", "games_v31_clean_bg.png", os.path.join("assets", "templates", "games_v31_full_bg.png")]:
        if os.path.exists(p):
            try:
                bg = Image.open(p).convert("RGB").resize((w,h))
                ov = Image.new("RGBA", (w,h), (0,10,30,92))
                bg = Image.alpha_composite(bg.convert("RGBA"), ov).convert("RGB")
                return bg
            except Exception:
                pass
    bg = Image.new("RGB", (w,h), "#071329")
    return bg


def render_fixtures_day_images(date, matches=None):
    matches = matches if matches is not None else _fixtures_for_date(date)
    if not matches:
        return []
    chunks = [matches[i:i+6] for i in range(0, len(matches), 6)]
    paths = []
    for page, chunk in enumerate(chunks, 1):
        img = _fixture_bg(1080, 1350)
        draw = ImageDraw.Draw(img, "RGBA")
        draw_text(draw, (540, 86), "MONDIAL AL MASEEF 2026", get_font(34), fill="#FFFFFF")
        draw_text(draw, (540, 132), "GAMES OF THE DAY", get_font(56), fill="#FFFFFF")
        title = _fixture_title(date)
        if len(chunks) > 1:
            title += f"  |  {page}/{len(chunks)}"
        draw_text(draw, (540, 188), title, get_font(36), fill="#FBBF24")
        y = 245
        card_h = 165 if len(chunk) > 4 else 190
        for idx, m in enumerate(chunk, 1):
            _draw_fixture_card(draw, img, (70, y, 1010, y+card_h-12), m, idx)
            y += card_h
        draw.line((250, 1242, 830, 1242), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(28), fill="#FBBF24")
        path = os.path.join(GENERATED_DIR, f"fixtures_{date.replace('/','_')}_{page}.png")
        img.save(path, quality=95)
        paths.append(path)
    return paths


def render_fixtures_combined_images(dates):
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    dates = [d for d in dates if _fixtures_for_date(d)]
    if not dates:
        return []
    # كل صورة تجمع حتى 3 أيام لتبقى مقروءة
    date_chunks = [dates[i:i+3] for i in range(0, len(dates), 3)]
    paths = []
    for page, dchunk in enumerate(date_chunks, 1):
        rows = []
        for d in dchunk:
            rows.append(("date", d))
            for m in _fixtures_for_date(d):
                rows.append(("match", m))
        h = max(1350, 250 + len(rows)*118 + 120)
        h = min(h, 2400)
        img = _fixture_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")
        draw_text(draw, (540, 74), "MONDIAL AL MASEEF 2026", get_font(32), fill="#FFFFFF")
        draw_text(draw, (540, 120), "GAMES SCHEDULE", get_font(54), fill="#FFFFFF")
        if len(date_chunks) > 1:
            draw_text(draw, (540, 172), f"صفحة {page}/{len(date_chunks)}", get_font(28), fill="#FBBF24")
        y = 220
        for kind, val in rows:
            if y > h - 140:
                break
            if kind == "date":
                draw.rounded_rectangle((90,y,990,y+50), radius=18, fill=(251,191,36,235))
                draw_text(draw, (540,y+25), _fixture_title(val), get_font(30), fill="#061329")
                y += 68
            else:
                m = val
                draw.rounded_rectangle((80,y,1000,y+92), radius=18, fill=(5,24,58,225), outline=(38,151,255,180), width=2)
                draw_text(draw, (910,y+27), m.get("time"), get_font(24), fill="#FBBF24", anchor="rm")
                draw_text(draw, (540,y+28), f"{m.get('team1')}  ×  {m.get('team2')}", get_font(28), fill="#FFFFFF", max_width=620)
                draw_text(draw, (540,y+66), m.get("stage") + (f" | {m.get('group')}" if m.get("group") else ""), get_font(20), fill="#CBD5E1", max_width=760)
                y += 108
        draw.line((250, h-88, 830, h-88), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, h-48), "المصيف يضعكم بالحدث", get_font(28), fill="#FBBF24")
        path = os.path.join(GENERATED_DIR, f"fixtures_combined_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=95)
        paths.append(path)
    return paths


def _fixtures_dates_keyboard(mode="single", selected=None):
    selected = set(selected or [])
    rows = []
    row = []
    for d, day in _fixture_dates():
        label = f"{'✅ ' if d in selected else ''}{day} {d[:5]}"
        data = f"fx|toggle|{d}" if mode == "multi" else f"fx|day|{d}"
        row.append(InlineKeyboardButton(label, callback_data=data))
        if len(row) == 2:
            rows.append(row); row = []
    if row:
        rows.append(row)
    if mode == "multi":
        rows.append([InlineKeyboardButton("تصميم كل يوم", callback_data="fx|render_each"), InlineKeyboardButton("تصميم واحد", callback_data="fx|render_combo")])
        rows.append([InlineKeyboardButton("تصفير الاختيار", callback_data="fx|clear"), InlineKeyboardButton("رجوع", callback_data="fx|menu")])
    else:
        rows.append([InlineKeyboardButton("اختيار أكثر من يوم", callback_data="fx|multi")])
    return InlineKeyboardMarkup(rows)


def _fixtures_day_keyboard(date):
    rows = [[InlineKeyboardButton("تصميم اليوم", callback_data=f"fx|render|{date}")]]
    miss = [m for m in _fixtures_for_date(date) if _has_unknown(m)]
    for i, m in enumerate(miss, 1):
        rows.append([InlineKeyboardButton(f"تحديث مباراة {i} — {m.get('time')}", callback_data=f"fx|upd|{m.get('id')}")])
    rows.append([InlineKeyboardButton("رجوع للأيام", callback_data="fx|menu")])
    return InlineKeyboardMarkup(rows)


def _fixtures_day_text(date):
    matches = _fixtures_for_date(date)
    if not matches:
        return "ما فيه مباريات لهذا التاريخ."
    lines = [f"{_fixture_title(date)}", ""]
    for i,m in enumerate(matches,1):
        lines.append(f"{i}) {m.get('team1')} × {m.get('team2')} — {m.get('time')}")
        lines.append(f"   {m.get('stage')}" + (f" | {m.get('group')}" if m.get('group') else ""))
        if _has_unknown(m) and m.get('note'):
            lines.append(f"   {m.get('note')}")
    return "\n".join(lines)


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if dates:
        for d in dates:
            paths = render_fixtures_day_images(d)
            if not paths:
                await update.message.reply_text(f"ما فيه مباريات بتاريخ {d}")
                continue
            for p in paths:
                await send_photo_path(update.message, p, _fixtures_caption(_fixture_title(d)))
        return
    await update.message.reply_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return
    paths = render_fixtures_combined_images(dates)
    if not paths:
        await update.message.reply_text("ما لقيت مباريات للتواريخ المطلوبة.")
        return
    for p in paths:
        await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_missing_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    missing = [m for m in _all_fixtures() if _has_unknown(m)]
    if not missing:
        await update.message.reply_text("كل المباريات محدثة ✅")
        return
    lines = ["المباريات التي تحتاج تحديث:", ""]
    rows = []
    for m in missing[:40]:
        lines.append(f"{m.get('id')} | {m.get('date')} | {m.get('time')} | {m.get('stage')} | {m.get('note','')}")
        rows.append([InlineKeyboardButton(f"تحديث {m.get('id')} — {m.get('date')} {m.get('time')}", callback_data=f"fx|upd|{m.get('id')}")])
    await update.message.reply_text("\n".join(lines[:45]), reply_markup=InlineKeyboardMarkup(rows[:25]))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return
    await q.answer()
    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""
    if action == "menu":
        await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
        return
    if action == "multi":
        context.user_data["fx_selected_dates"] = []
        await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
        return
    if action == "toggle" and len(parts) >= 3:
        d = parts[2]
        sel = list(context.user_data.get("fx_selected_dates") or [])
        if d in sel:
            sel.remove(d)
        else:
            sel.append(d)
        context.user_data["fx_selected_dates"] = sel
        await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", sel))
        return
    if action == "clear":
        context.user_data["fx_selected_dates"] = []
        await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
        return
    if action in ["render_each", "render_combo"]:
        sel = list(context.user_data.get("fx_selected_dates") or [])
        if not sel:
            await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
            return
        if action == "render_combo":
            paths = render_fixtures_combined_images(sel)
            for p in paths:
                await send_photo_path(q.message, p, _fixtures_caption("مباريات مجمعة"))
        else:
            for d in sel:
                for p in render_fixtures_day_images(d):
                    await send_photo_path(q.message, p, _fixtures_caption(_fixture_title(d)))
        return
    if action == "day" and len(parts) >= 3:
        d = parts[2]
        await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
        return
    if action == "render" and len(parts) >= 3:
        d = parts[2]
        paths = render_fixtures_day_images(d)
        if not paths:
            await q.message.reply_text("ما فيه مباريات لهذا اليوم.")
            return
        for p in paths:
            await send_photo_path(q.message, p, _fixtures_caption(_fixture_title(d)))
        return
    if action == "upd" and len(parts) >= 3:
        mid = parts[2]
        m = _fixture_by_id(mid)
        if not m:
            await q.message.reply_text("لم أجد المباراة.")
            return
        context.user_data["fixture_update_match_id"] = mid
        await q.message.reply_text(
            f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {m.get('time')}) كذا:\n"
            "الفريق الأول * الفريق الثاني\n\n"
            "مثال: المكسيك * أستراليا\n"
            "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
        )
        return


async def fixtures_update_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mid = context.user_data.get("fixture_update_match_id")
    if not mid:
        return
    text = (update.message.text or "").strip()
    if "*" in text:
        a,b = [x.strip() for x in text.split("*",1)]
    elif "×" in text:
        a,b = [x.strip() for x in text.split("×",1)]
    elif "-" in text:
        a,b = [x.strip() for x in text.split("-",1)]
    else:
        await update.message.reply_text("اكتبها كذا: الفريق الأول * الفريق الثاني")
        return
    if not a or not b:
        await update.message.reply_text("اكتب اسم الفريقين كاملين.")
        return
    data = _load_fixture_updates()
    data.setdefault(mid, {})
    data[mid]["team1"] = canonical_team_name(a) or normalize_name(a)
    data[mid]["team2"] = canonical_team_name(b) or normalize_name(b)
    _save_fixture_updates(data)
    context.user_data.pop("fixture_update_match_id", None)
    m = _apply_fixture_updates(_fixture_by_id(mid) or {"id":mid})
    await update.message.reply_text(
        f"✅ تم حفظ تحديث المباراة\n"
        f"{m.get('team1')} × {m.get('team2')} — {m.get('time','')}\n\n"
        f"لن أصمم الآن. وقت ما تبيها اكتب:\n/مباريات {m.get('date','')}"
    )


# تحسين ESPN: نسحب الهدافين من summary/event details إذا المصدر وفرها.
def _v23_extract_goal_strings_from_node(obj):
    out = []
    for node in _walk_json(obj) if '_walk_json' in globals() else []:
        if not isinstance(node, dict):
            continue
        blob = json.dumps(node, ensure_ascii=False).lower()
        if not any(k in blob for k in ["goal", "scored", "scores", "هدف"]):
            continue
        txt = node.get("text") or node.get("description") or node.get("headline") or node.get("displayText") or node.get("detail") or ""
        athlete = node.get("athlete") or node.get("player") or node.get("competitor") or {}
        name = ""
        if isinstance(athlete, dict):
            name = athlete.get("displayName") or athlete.get("shortName") or athlete.get("name") or ""
        clock = node.get("clock") or node.get("time") or node.get("minute") or {}
        minute = ""
        if isinstance(clock, dict):
            minute = clock.get("displayValue") or clock.get("value") or ""
        else:
            minute = str(clock or "")
        txt = normalize_name(txt)
        name = normalize_name(name)
        if not name:
            # أحيانًا النص نفسه: Alex Freeman Goal - 43'
            name = txt
        if name and len(name) < 80:
            item = name
            if minute and "'" not in item:
                item += f" {minute}"
            out.append(item)
    return _v9_sanitize_scorers(out) if '_v9_sanitize_scorers' in globals() else out[:8]


def _fetch_espn_summary_goals(event_id):
    if not event_id or not requests:
        return []
    urls = [
        f"https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/summary?event={event_id}",
        f"https://site.web.api.espn.com/apis/site/v2/sports/soccer/fifa.world/summary?event={event_id}",
        f"https://site.api.espn.com/apis/site/v2/sports/soccer/all/summary?event={event_id}",
    ]
    for url in urls:
        try:
            r = _requests_get(url, timeout=12)
            if int(getattr(r, "status_code", 200) or 200) >= 400:
                continue
            data = r.json()
            goals = _v23_extract_goal_strings_from_node(data)
            if goals:
                return goals
        except Exception:
            continue
    return []


def _parse_espn_match_from_event(event):
    try:
        comp = (event.get("competitions") or [event])[0]
        competitors = comp.get("competitors") or []
        if len(competitors) < 2:
            return None
        a = competitors[0]
        b = competitors[1]
        team1 = canonical_team_name(a.get("team", {}).get("displayName")) or normalize_name(a.get("team", {}).get("displayName"))
        team2 = canonical_team_name(b.get("team", {}).get("displayName")) or normalize_name(b.get("team", {}).get("displayName"))
        score1 = str(a.get("score", 0))
        score2 = str(b.get("score", 0))
        status = comp.get("status") or event.get("status") or {}
        detail = normalize_name(((status.get("type") or {}).get("detail")) or status.get("displayClock") or "")
        short_detail = normalize_name(((status.get("type") or {}).get("shortDetail")) or "")
        state = normalize_name(((status.get("type") or {}).get("name")) or ((status.get("type") or {}).get("state")) or "")
        status_ar = _normalize_status_text(detail or short_detail or state)
        scorers = []
        for d in comp.get("details", []) or []:
            if isinstance(d, dict):
                txt = normalize_name(d.get("text") or d.get("detail") or d.get("description") or "")
                if txt:
                    scorers.append(txt)
        event_id = str(event.get("id") or comp.get("id") or "")
        if not scorers and event_id:
            scorers = _fetch_espn_summary_goals(event_id)
        return {
            "team1": team1, "team2": team2,
            "score1": score1, "score2": score2,
            "status": status_ar,
            "minute": detail or short_detail,
            "scorers": scorers[:8],
            "goals_source": "ESPN" if scorers else "",
            "event_id": event_id,
            "source": "ESPN",
        }
    except Exception:
        return None


# ترتيب المجموعات: نرسل صورة شاملة + 3 صور مقروءة، بدون النص الطويل.
def _standings_page_image(groups, page_no=1, total_pages=3):
    chunk = groups[(page_no-1)*4:page_no*4]
    img = _fixture_bg(1080, 1350)
    draw = ImageDraw.Draw(img, "RGBA")
    draw_text(draw, (540, 76), "MONDIAL AL MASEEF 2026", get_font(32), fill="#FFFFFF")
    draw_text(draw, (540, 122), "GROUP STANDINGS", get_font(54), fill="#FFFFFF")
    draw_text(draw, (540, 174), f"ترتيب المجموعات | {page_no}/{total_pages}", get_font(30), fill="#FBBF24")
    positions = [(60,230,510,680),(570,230,1020,680),(60,715,510,1165),(570,715,1020,1165)]
    for (title, rows), box in zip(chunk, positions):
        x0,y0,x1,y1 = box
        draw.rounded_rectangle(box, radius=26, fill=(5,24,58,230), outline=(38,151,255,210), width=3)
        draw.rounded_rectangle((x0+16,y0+14,x1-16,y0+58), radius=16, fill=(251,191,36,240))
        draw_text(draw, ((x0+x1)//2,y0+36), title, get_font(28), fill="#061329")
        yy = y0 + 85
        draw_text(draw, (x1-32, yy), "نقاط", get_font(20), fill="#FBBF24", anchor="rm")
        draw_text(draw, (x1-110, yy), "فارق", get_font(20), fill="#FBBF24", anchor="rm")
        draw_text(draw, (x1-185, yy), "لعب", get_font(20), fill="#FBBF24", anchor="rm")
        draw_text(draw, (x0+34, yy), "المنتخب", get_font(20), fill="#FBBF24", anchor="lm")
        yy += 36
        for i,r in enumerate(rows[:4],1):
            name, played, gd, pts = r[0], r[1], r[2], r[3]
            draw.line((x0+20, yy-16, x1-20, yy-16), fill=(255,255,255,45), width=1)
            draw_text(draw, (x1-34, yy), str(pts), get_font(24), fill="#FFFFFF", anchor="rm")
            draw_text(draw, (x1-112, yy), f"{gd:+d}" if isinstance(gd,int) else str(gd), get_font(22), fill="#CBD5E1", anchor="rm")
            draw_text(draw, (x1-188, yy), str(played), get_font(22), fill="#CBD5E1", anchor="rm")
            draw_text(draw, (x0+34, yy), f"{i}. {name}", get_font(23), fill="#FFFFFF", anchor="lm", max_width=245)
            yy += 67
    draw.line((250, 1242, 830, 1242), fill=(255,255,255,180), width=2)
    draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(28), fill="#FBBF24")
    path = os.path.join(GENERATED_DIR, f"group_standings_page_{page_no}.png")
    img.save(path, quality=95)
    return path


def create_all_groups_three_page_images(groups):
    return [_standings_page_image(groups, i, 3) for i in range(1,4)]


async def current_groups_now_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    payload = {"kind": "standings"}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text("⏳ أسحب نتائج المجموعات من ESPN وأحسب الترتيب...")
    try:
        groups, source_label = await _asyncio_v6.wait_for(_asyncio_v6.to_thread(fetch_current_groups, "espn"), timeout=35)
    except Exception as e:
        groups, source_label = [], ""
        await wait.edit_text(f"تعذر جلب ترتيب المجموعات ❌\n{str(e)[:250]}", reply_markup=kb)
        return
    if not groups:
        await wait.edit_text(_source_help_text("standings", "espn") + "\n\nاختر مصدر آخر:", reply_markup=kb)
        return
    try:
        await wait.delete()
    except Exception:
        pass
    caption_all = "ترتيب المجموعات الآن ✅\nالمصدر: ESPN — محسوب من نتائج المباريات المباشرة"
    try:
        all_path = create_all_groups_newlook_image(groups) if 'create_all_groups_newlook_image' in globals() else create_all_groups_image(groups)
        if all_path and os.path.exists(all_path):
            await send_photo_path_markup(update.message, all_path, caption_all, kb)
    except Exception:
        pass
    try:
        for i, p in enumerate(create_all_groups_three_page_images(groups), 1):
            await send_photo_path(update.message, p, f"ترتيب المجموعات الآن — صورة {i}/3\nالمصدر: ESPN Live Results")
    except Exception:
        # إذا فشل تقسيم الصور نكتفي بالصورة الشاملة.
        pass

# ==================== END V23 CLEAN FIXTURES PATCH ====================

# ==================== END V22 HARD STABLE PATCH ====================



# ==================== V24 STABLE PATCH: fixtures design + safe commands ====================
# - إصلاح إرسال الصور مع update/message/callback message.
# - ربط /مباريات و/مباريات_مجمعة وزر تصميم اليوم بدالة التصميم فعليًا.
# - عدم السكوت عند الخطأ: كل أمر يرسل انتظار وخطأ واضح.
# - فحص API بمهلات قصيرة.
# - استيراد ملف Excel برسالة انتظار وخطأ واضح.

async def send_photo_path(target, path, caption=None):
    """يرسل صورة سواء كان الهدف Update أو Message."""
    msg = getattr(target, "message", None) or target
    if not msg:
        raise RuntimeError("لا يوجد هدف صالح لإرسال الصورة")
    if not path or not os.path.exists(path):
        raise FileNotFoundError(str(path))
    with open(path, "rb") as f:
        await msg.reply_photo(photo=f, caption=caption or "")


async def send_photo_path_markup(target, path, caption=None, reply_markup=None):
    """يرسل صورة مع أزرار سواء كان الهدف Update أو Message."""
    msg = getattr(target, "message", None) or target
    if not msg:
        raise RuntimeError("لا يوجد هدف صالح لإرسال الصورة")
    if not path or not os.path.exists(path):
        raise FileNotFoundError(str(path))
    with open(path, "rb") as f:
        await msg.reply_photo(photo=f, caption=caption or "", reply_markup=reply_markup)


def _v24_dates_from_text(text):
    dates = _extract_fixture_dates_from_text(text or "")
    # إزالة التكرار مع الحفاظ على الترتيب
    out = []
    for d in dates:
        if d not in out:
            out.append(d)
    return out


async def _v24_send_fixture_day(target, date):
    msg = getattr(target, "message", None) or target
    paths = render_fixtures_day_images(date)
    if not paths:
        await msg.reply_text(f"ما فيه مباريات بتاريخ {date}")
        return 0
    sent = 0
    for p in paths:
        await send_photo_path(msg, p, _fixtures_caption(_fixture_title(date)))
        sent += 1
    return sent


async def _v24_send_fixture_combo(target, dates):
    msg = getattr(target, "message", None) or target
    paths = render_fixtures_combined_images(dates)
    if not paths:
        await msg.reply_text("ما لقيت مباريات للتواريخ المطلوبة.")
        return 0
    sent = 0
    for p in paths:
        await send_photo_path(msg, p, _fixtures_caption("مباريات مجمعة"))
        sent += 1
    return sent


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _v24_dates_from_text(update.message.text)
    if dates:
        wait = await update.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
        try:
            total = 0
            for d in dates:
                total += await _v24_send_fixture_day(update.message, d)
            try:
                await wait.delete()
            except Exception:
                pass
            if total == 0:
                await update.message.reply_text("ما لقيت مباريات للتواريخ المطلوبة.")
        except Exception as e:
            await wait.edit_text(f"تعذر تصميم مباريات اليوم ❌\nالسبب: {str(e)[:300]}")
        return
    await update.message.reply_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _v24_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return
    wait = await update.message.reply_text("⏳ جاري تصميم المباريات المجمعة...")
    try:
        total = await _v24_send_fixture_combo(update.message, dates)
        try:
            await wait.delete()
        except Exception:
            pass
        if total == 0:
            await update.message.reply_text("ما لقيت مباريات للتواريخ المطلوبة.")
    except Exception as e:
        await wait.edit_text(f"تعذر تصميم المباريات المجمعة ❌\nالسبب: {str(e)[:300]}")


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _v24_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return
    await q.answer()
    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return
    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""
    try:
        if action == "menu":
            await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
            return
        if action == "multi":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return
        if action == "toggle" and len(parts) >= 3:
            d = parts[2]
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if d in sel:
                sel.remove(d)
            else:
                sel.append(d)
            context.user_data["fx_selected_dates"] = sel
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", sel))
            return
        if action == "clear":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return
        if action == "day" and len(parts) >= 3:
            d = parts[2]
            await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
            return
        if action == "render" and len(parts) >= 3:
            d = parts[2]
            wait = await q.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                await _v24_send_fixture_day(q.message, d)
                try:
                    await wait.delete()
                except Exception:
                    pass
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:300]}")
            return
        if action in ["render_each", "render_combo"]:
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري التصميم...")
            try:
                if action == "render_combo":
                    await _v24_send_fixture_combo(q.message, sel)
                else:
                    for d in sel:
                        await _v24_send_fixture_day(q.message, d)
                try:
                    await wait.delete()
                except Exception:
                    pass
            except Exception as e:
                await wait.edit_text(f"تعذر التصميم ❌\nالسبب: {str(e)[:300]}")
            return
        if action == "upd" and len(parts) >= 3:
            mid = parts[2]
            m = _fixture_by_id(mid)
            if not m:
                await q.message.reply_text("لم أجد المباراة.")
                return
            context.user_data["fixture_update_match_id"] = mid
            await q.message.reply_text(
                f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {m.get('time')}) كذا:\n"
                "الفريق الأول * الفريق الثاني\n\n"
                "مثال: المكسيك * أستراليا\n"
                "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
            )
            return
        await q.message.reply_text("تعذر قراءة الخيار.")
    except Exception as e:
        await q.message.reply_text(f"تعذر تنفيذ خيار المباريات ❌\n{str(e)[:300]}")



# ==================== V25 FIXTURES DESIGN OVERRIDE ====================
# هذا القسم يتعمد تعريف دوال /مباريات مرة أخيرة قبل التشغيل.
# الهدف:
# - /مباريات 20/06 = نفس تصميم /مباريات_اليوم الرئيسي (تصميم 2)
# - اختيار يوم من الأزرار = يظهر زرين تصميم 1 وتصميم 2
# - تصميم 1 = مضغوط بخلفية التمثال بدون الكلام الكبير فوق
# - تصميم 2 = التصميم الرئيسي المعتمد V31 /مباريات_اليوم
# - /مباريات_مجمعة = تصميم مضغوط لعدة أيام
# - حذف تكرار المباريات داخل نفس اليوم

def _v25_safe_txt(x):
    return str(x or "").strip()


def _v25_dedupe_fixture_matches(matches):
    seen = set()
    out = []
    for m in matches or []:
        t1 = _v25_safe_txt(m.get("team1"))
        t2 = _v25_safe_txt(m.get("team2"))
        tm = _v25_safe_txt(m.get("time"))
        dt = _v25_safe_txt(m.get("date"))
        key = (
            re.sub(r"\s+", " ", t1.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", t2.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", tm).strip(),
            dt,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(dict(m))
    return out


def _v25_fixture_simple_matches(date):
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    simple = []
    for m in rows:
        simple.append((
            _v25_safe_txt(m.get("team1")),
            _v25_safe_txt(m.get("team2")),
            _v25_safe_txt(m.get("time")),
        ))
    return rows, simple


def _v25_fixture_title(date):
    d = _normalize_date_arg(date)
    day = ""
    for x, dy in _fixture_dates():
        if x == d:
            day = dy
            break
    return f"{day} {d[:5]}".strip()


def _fixtures_caption(date_or_title, source="PDF جدول البطولة"):
    return f"{date_or_title}\nالمصدر: {source}\nالمصيف يضعكم بالحدث"


def _v25_compact_bg(w=1080, h=1350):
    """
    خلفية التمثال النظيفة للمضغوط بدون عنوان GAMES الكبير.
    """
    candidates = [
        "games_v31_clean_bg.png",
        os.path.join("assets", "templates", "games_v31_clean_bg.png"),
        "games_v31_full_bg.png",
        os.path.join("assets", "templates", "games_v31_full_bg.png"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                bg = Image.open(p).convert("RGB").resize((w, h))
                ov = Image.new("RGBA", (w, h), (0, 10, 30, 120))
                return Image.alpha_composite(bg.convert("RGBA"), ov).convert("RGB")
            except Exception:
                pass
    return Image.new("RGB", (w, h), "#071329")


def _v25_draw_compact_match(draw, box, m):
    x0, y0, x1, y1 = box
    try:
        draw.rounded_rectangle(box, radius=22, fill=(5,24,58,225), outline=(38,151,255,190), width=2)
    except Exception:
        draw.rectangle(box, fill=(5,24,58), outline=(38,151,255), width=2)

    # الوقت يمين
    draw_text(draw, (x1-35, y0+38), _v25_safe_txt(m.get("time")), get_font(27), fill="#FBBF24", anchor="rm", max_width=190)

    # المباراة بالنص
    text = f"{_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))}"
    draw_text(draw, ((x0+x1)//2, y0+36), text, get_font(30), fill="#FFFFFF", max_width=(x1-x0)-260)

    sub = _v25_safe_txt(m.get("stage"))
    if m.get("group"):
        sub += f" - {_v25_safe_txt(m.get('group'))}"
    draw_text(draw, ((x0+x1)//2, y0+75), sub, get_font(21), fill="#CBD5E1", max_width=(x1-x0)-160)


def _render_fixture_day_compact(date):
    """
    تصميم 1: مضغوط، خلفية التمثال، بدون الكلام الكبير فوق.
    """
    rows = _v25_dedupe_fixture_matches(_fixtures_for_date(date))
    if not rows:
        return []

    chunks = [rows[i:i+8] for i in range(0, len(rows), 8)]
    paths = []

    for page, chunk in enumerate(chunks, 1):
        h = 1350
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 105
        # شريط التاريخ فقط بدون عنوان كبير
        try:
            draw.rounded_rectangle((120, y, 960, y+62), radius=22, fill=(251,191,36,235))
        except Exception:
            draw.rectangle((120, y, 960, y+62), fill=(251,191,36))
        title = _v25_fixture_title(date)
        if len(chunks) > 1:
            title += f" | {page}/{len(chunks)}"
        draw_text(draw, (540, y+31), title, get_font(34), fill="#061329", max_width=780)

        y += 92
        for m in chunk:
            _v25_draw_compact_match(draw, (70, y, 1010, y+104), m)
            y += 122

        draw.line((250, 1238, 830, 1238), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(30), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_compact_{date.replace('/','_')}_{page}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _render_fixture_day_by_design(date, design=2):
    """
    design=1 => مضغوط
    design=2 => نفس تصميم /مباريات_اليوم الرئيسي
    """
    rows, simple_matches = _v25_fixture_simple_matches(date)
    if not simple_matches:
        return []

    if int(design) == 1:
        return _render_fixture_day_compact(date)

    # تصميم 2: نفس تصميم /مباريات_اليوم
    chunks = [simple_matches[i:i+7] for i in range(0, len(simple_matches), 7)]
    paths = []

    for page_idx, chunk in enumerate(chunks, start=1):
        page_title = _v25_fixture_title(date)
        if len(chunks) > 1:
            page_title = f"{page_title} | {page_idx}/{len(chunks)}"

        # هذا هو التصميم المعتمد نفسه حق /مباريات_اليوم
        path = create_matches_today_v31_full_image(page_title, chunk)
        final_path = os.path.join(
            GENERATED_DIR,
            f"fixtures_day_design2_{date.replace('/','_')}_{page_idx}.png"
        )
        try:
            Image.open(path).save(final_path, quality=96)
            paths.append(final_path)
        except Exception:
            paths.append(path)

    return paths


def render_fixtures_combined_images(dates):
    """
    تصميم مجمع: خلفية التمثال بدون الكلام الكبير فوق.
    إذا كثرت الأيام يقسم كل 3 أيام في صورة.
    """
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    dates = [d for d in dates if _fixtures_for_date(d)]
    # إزالة تكرار التواريخ مع الحفاظ على الترتيب
    clean_dates = []
    for d in dates:
        if d not in clean_dates:
            clean_dates.append(d)
    dates = clean_dates

    if not dates:
        return []

    date_chunks = [dates[i:i+3] for i in range(0, len(dates), 3)]
    paths = []

    for page, dchunk in enumerate(date_chunks, 1):
        rows = []
        for d in dchunk:
            rows.append(("date", d))
            for m in _v25_dedupe_fixture_matches(_fixtures_for_date(d)):
                rows.append(("match", m))

        h = max(1350, 170 + len(rows)*112 + 140)
        h = min(h, 2400)
        img = _v25_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 90
        if len(date_chunks) > 1:
            draw_text(draw, (540, 42), f"صفحة {page}/{len(date_chunks)}", get_font(24), fill="#FBBF24")

        for kind, val in rows:
            if y > h - 135:
                break

            if kind == "date":
                try:
                    draw.rounded_rectangle((90, y, 990, y+54), radius=20, fill=(251,191,36,235))
                except Exception:
                    draw.rectangle((90, y, 990, y+54), fill=(251,191,36))
                draw_text(draw, (540, y+27), _v25_fixture_title(val), get_font(31), fill="#061329", max_width=830)
                y += 72
            else:
                _v25_draw_compact_match(draw, (75, y, 1005, y+96), val)
                y += 112

        draw.line((250, h-88, 830, h-88), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, h-48), "المصيف يضعكم بالحدث", get_font(29), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_combined_v25_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _fixtures_day_keyboard(date):
    rows = [
        [
            InlineKeyboardButton("تصميم 1", callback_data=f"fx|render1|{date}"),
            InlineKeyboardButton("تصميم 2", callback_data=f"fx|render2|{date}")
        ]
    ]

    miss = [m for m in _fixtures_for_date(date) if _has_unknown(m)]
    for i, m in enumerate(miss, 1):
        rows.append([InlineKeyboardButton(f"تحديث مباراة {i} — {m.get('time')}", callback_data=f"fx|upd|{m.get('id')}")])

    rows.append([InlineKeyboardButton("رجوع للأيام", callback_data="fx|menu")])
    return InlineKeyboardMarkup(rows)


def _fixtures_dates_keyboard(mode="single", selected=None):
    selected = set(selected or [])
    rows = []
    row = []

    for d, day in _fixture_dates():
        label = f"{'✅ ' if d in selected else ''}{day} {d[:5]}"
        data = f"fx|toggle|{d}" if mode == "multi" else f"fx|day|{d}"
        row.append(InlineKeyboardButton(label, callback_data=data))
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    if mode == "multi":
        rows.append([
            InlineKeyboardButton("تصميم كل يوم", callback_data="fx|render_each"),
            InlineKeyboardButton("تصميم واحد", callback_data="fx|render_combo"),
        ])
        rows.append([
            InlineKeyboardButton("تصفير الاختيار", callback_data="fx|clear"),
            InlineKeyboardButton("رجوع", callback_data="fx|menu"),
        ])
    else:
        rows.append([InlineKeyboardButton("اختيار أكثر من يوم", callback_data="fx|multi")])

    return InlineKeyboardMarkup(rows)


def _fixtures_day_text(date):
    rows, matches = _v25_fixture_simple_matches(date)
    if not matches:
        return "ما فيه مباريات لهذا التاريخ."

    lines = [f"{_v25_fixture_title(date)}", ""]
    for i, m in enumerate(rows, 1):
        lines.append(f"{i}) {_v25_safe_txt(m.get('team1'))} × {_v25_safe_txt(m.get('team2'))} — {_v25_safe_txt(m.get('time'))}")
        extra = []
        if m.get("stage"):
            extra.append(_v25_safe_txt(m.get("stage")))
        if m.get("group"):
            extra.append(_v25_safe_txt(m.get("group")))
        if extra:
            lines.append("   " + " | ".join(extra))
        if _has_unknown(m) and m.get("note"):
            lines.append(f"   {_v25_safe_txt(m.get('note'))}")

    return "\n".join(lines)


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    dates = _extract_fixture_dates_from_text(text)

    if dates:
        # تاريخ واحد: يصمم مباشرة بتصميم 2، نفس /مباريات_اليوم
        if len(dates) == 1:
            d = dates[0]
            wait = await update.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=2)
                if not paths:
                    await wait.edit_text(f"ما فيه مباريات بتاريخ {d}")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(update.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        # أكثر من تاريخ: مجمع
        wait = await update.message.reply_text("⏳ جاري تصميم الأيام المجمعة...")
        try:
            paths = render_fixtures_combined_images(dates)
            if not paths:
                await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
                return
            try:
                await wait.delete()
            except Exception:
                pass
            for p in paths:
                await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
        except Exception as e:
            await wait.edit_text(f"تعذر تصميم الأيام ❌\nالسبب: {str(e)[:400]}")
        return

    await update.message.reply_text("اختر اليوم أو اكتب:\n/مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return

    wait = await update.message.reply_text("⏳ جاري تصميم المباريات المجمعة...")
    try:
        paths = render_fixtures_combined_images(dates)
        if not paths:
            await wait.edit_text("ما لقيت مباريات للتواريخ المطلوبة.")
            return
        try:
            await wait.delete()
        except Exception:
            pass
        for p in paths:
            await send_photo_path(update.message, p, _fixtures_caption("مباريات مجمعة"))
    except Exception as e:
        await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return

    await q.answer()

    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return

    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""

    try:
        if action == "menu":
            await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
            return

        if action == "multi":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "toggle" and len(parts) >= 3:
            d = parts[2]
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if d in sel:
                sel.remove(d)
            else:
                sel.append(d)
            context.user_data["fx_selected_dates"] = sel
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", sel))
            return

        if action == "clear":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اضغط التصميم المناسب:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "day" and len(parts) >= 3:
            d = parts[2]
            await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
            return

        if action in ["render1", "render2"] and len(parts) >= 3:
            d = parts[2]
            design = 1 if action == "render1" else 2
            wait = await q.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                paths = _render_fixture_day_by_design(d, design=design)
                if not paths:
                    await wait.edit_text("ما فيه مباريات لهذا اليوم.")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_each":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري تصميم الأيام...")
            try:
                try:
                    await wait.delete()
                except Exception:
                    pass
                for d in sel:
                    for p in _render_fixture_day_by_design(d, design=2):
                        await send_photo_path(q.message, p, _fixtures_caption(_v25_fixture_title(d)))
            except Exception as e:
                await q.message.reply_text(f"تعذر التصميم ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "render_combo":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            wait = await q.message.reply_text("⏳ جاري التصميم المجمع...")
            try:
                paths = render_fixtures_combined_images(sel)
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _fixtures_caption("مباريات مجمعة"))
            except Exception as e:
                await wait.edit_text(f"تعذر التصميم المجمع ❌\nالسبب: {str(e)[:400]}")
            return

        if action == "upd" and len(parts) >= 3:
            mid = parts[2]
            m = _fixture_by_id(mid)
            if not m:
                await q.message.reply_text("لم أجد المباراة.")
                return

            context.user_data["fixture_update_match_id"] = mid
            await q.message.reply_text(
                f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {m.get('time')}) كذا:\n"
                "الفريق الأول * الفريق الثاني\n\n"
                "مثال: المكسيك * أستراليا\n"
                "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
            )
            return

        await q.message.reply_text("تعذر قراءة الخيار.")
    except Exception as e:
        await q.message.reply_text(f"تعذر تنفيذ خيار المباريات ❌\n{str(e)[:400]}")


async def fixtures_update_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mid = context.user_data.get("fixture_update_match_id")
    if not mid:
        return

    text = (update.message.text or "").strip()

    if "*" in text:
        a, b = [x.strip() for x in text.split("*", 1)]
    elif "×" in text:
        a, b = [x.strip() for x in text.split("×", 1)]
    elif "-" in text:
        a, b = [x.strip() for x in text.split("-", 1)]
    else:
        await update.message.reply_text("اكتبها كذا: الفريق الأول * الفريق الثاني")
        return

    if not a or not b:
        await update.message.reply_text("اكتب اسم الفريقين كاملين.")
        return

    data = _load_fixture_updates()
    data.setdefault(mid, {})
    data[mid]["team1"] = canonical_team_name(a) or normalize_name(a)
    data[mid]["team2"] = canonical_team_name(b) or normalize_name(b)
    _save_fixture_updates(data)

    context.user_data.pop("fixture_update_match_id", None)

    m = _apply_fixture_updates(_fixture_by_id(mid) or {"id": mid})
    await update.message.reply_text(
        f"✅ تم حفظ تحديث المباراة\n"
        f"{m.get('team1')} × {m.get('team2')} — {m.get('time', '')}\n\n"
        f"لن أصمم الآن. وقت ما تبيها اكتب:\n/مباريات {m.get('date', '')}"
    )

# ==================== END V25 FIXTURES DESIGN OVERRIDE ====================


async def api_check_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ أفحص مصادر النتائج...")
    lines = ["فحص مصادر النتائج:", ""]
    if _serpapi_key():
        try:
            data = await asyncio.wait_for(asyncio.to_thread(serpapi_search_json, "مباراة المكسيك كوريا الجنوبية", "ar", "sa", 8), timeout=12)
            lines.append("✅ SERPAPI_KEY موجود والاتصال بقوقل شغال")
            if isinstance(data, dict) and isinstance(data.get("sports_results"), dict):
                lines.append("✅ Google Sports ظهر في نتيجة الفحص")
            else:
                lines.append("⚠️ الاتصال شغال، لكن لم يظهر كرت Google Sports في هذا الفحص")
        except Exception as e:
            lines.append(f"❌ SerpApi موجود لكن فشل الاتصال: {str(e)[:160]}")
    else:
        lines.append("❌ SERPAPI_KEY غير موجود")
    if _api_football_key():
        try:
            status = await asyncio.wait_for(asyncio.to_thread(_api_football_get, "/status", {}), timeout=8)
            lines.append("✅ API_FOOTBALL_KEY موجود والاتصال شغال")
        except Exception as e:
            lines.append(f"❌ API-Football فشل: {str(e)[:160]}")
    else:
        lines.append("❌ API_FOOTBALL_KEY غير موجود")
    await msg.edit_text("\n".join(lines))


async def import_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    document = update.message.document
    if document:
        filename = document.file_name or "import.xlsx"
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            await update.message.reply_text("الملف لازم يكون Excel بصيغة .xlsx أو .xlsm")
            return
        local_path, _ = await _download_document_to_folder(update, context, "imports")
        LAST_UPLOADED_FILES.setdefault(chat_id, {})["excel"] = local_path
    else:
        local_path = LAST_UPLOADED_FILES.get(chat_id, {}).get("excel")
        if not local_path or not os.path.exists(local_path):
            await update.message.reply_text("ما لقيت ملف Excel محفوظ. أرسل ملف الإكسل لحاله أولًا، وبعدها اكتب:\n/استيراد_ملف")
            return
    wait = await update.message.reply_text("⏳ جاري قراءة ملف الإكسل...")
    try:
        imported = await asyncio.wait_for(asyncio.to_thread(parse_import_excel, local_path), timeout=45)
    except Exception as e:
        await wait.edit_text(f"صار خطأ أثناء قراءة الإكسل ❌\n{str(e)[:400]}")
        return
    if not imported:
        await wait.edit_text("ما قدرت أستخرج أيام من الملف. تأكد أن الصفحات باسم: يوم 1، يوم 2 ...")
        return
    PENDING_IMPORTS[chat_id] = {"path": local_path, "data": imported}
    await wait.edit_text(import_summary_text(imported))


async def dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    wait = await update.message.reply_text("جاري إنشاء ملف الإحصائيات... ⏳")
    try:
        nums = get_numbers(update.message.text)
        if len(nums) >= 2:
            start_day, end_day = nums[0], nums[1]
            if start_day > end_day:
                start_day, end_day = end_day, start_day
        elif len(nums) == 1:
            start_day, end_day = 1, nums[0]
        else:
            start_day, end_day = 1, 31
        file_name, stats = await asyncio.wait_for(asyncio.to_thread(create_dashboard, start_day, end_day), timeout=70)
        if not stats.get("days"):
            await wait.edit_text("ما لقيت أيام لإحصائيات الداشبورد.")
            return
        try:
            add_matchups_sheet_to_dashboard(file_name, start_day, end_day)
        except Exception:
            pass
        caption = (
            "تم إنشاء ملف الإحصائيات الكامل ✅\n"
            f"النطاق: من اليوم {start_day} إلى اليوم {end_day}\n"
            f"الأيام المحسوبة: {', '.join(map(str, stats['days']))}"
        )
        try:
            await wait.delete()
        except Exception:
            pass
        with open(file_name, "rb") as file:
            await update.message.reply_document(document=file, filename=file_name, caption=caption)
    except Exception as e:
        await wait.edit_text(f"صار خطأ أثناء إنشاء الإحصائيات ❌\n\nالسبب:\n{str(e)[:500]}")


async def live_match_command(update: Update, context: ContextTypes.DEFAULT_TYPE, mode_override=None):
    team1, team2, mode, date_hint = parse_live_command_text(update.message.text)
    if mode_override:
        mode = _norm_source_mode(mode_override)
    if not team1 or not team2:
        await update.message.reply_text("اكتبها كذا:\n/مباشر البرازيل * هايتي * قوقل\nأو:\n/مباشر البرازيل * هايتي * رسمي")
        return
    payload = {"kind": "live", "team1": team1, "team2": team2, "date_hint": date_hint}
    kb = source_keyboard(context, payload)
    wait = await update.message.reply_text(f"⏳ جاري البحث عن مباراة {team1} × {team2}\nالمصدر: {mode_label_ar(mode)}")
    try:
        data = await asyncio.wait_for(asyncio.to_thread(fetch_live_match_data, team1, team2, mode, date_hint), timeout=22)
    except Exception as e:
        data = None
        err = str(e)[:160]
    if not data:
        await wait.edit_text(
            f"تعذر جلب المباراة من مصدر {mode_label_ar(mode)} ❌\n"
            f"مباراة: {team1} × {team2}\n"
            f"جرب مصدر ثاني من الأزرار.",
            reply_markup=kb,
        )
        return
    try:
        path = render_live_match_card(data, mode_label_ar(mode))
        await send_photo_path_markup(update.message, path, build_live_caption(data, mode_label_ar(mode)), kb)
        try:
            await wait.delete()
        except Exception:
            pass
    except Exception as e:
        await wait.edit_text(f"تم جلب النتيجة لكن تعذر تصميم الصورة ❌\n{str(e)[:180]}\n\n" + build_live_caption(data, mode_label_ar(mode)), reply_markup=kb)


# ==================== V26 FINAL OVERRIDE: fixtures + excel safety ====================
# هذا القسم آخر تعريف للدوال قبل main()، لذلك هو المعتمد وقت التشغيل.
# الإصلاحات:
# 1) اختصار الوقت: 3:00 ص / 10:00 م / 12:00 ص
# 2) التصميم المجمع الجديد: يومين جنب بعض
# 3) خيار تصميم جديد + تصميم 10 مباريات للأيام المتعددة
# 4) كابشن مباريات اليوم مع نموذج فانتزي المصيف
# 5) تثبيت استيراد الإكسل وعدم كسره
# 6) صيغ إضافة/حذف متسابق متعددة بدون تعارض مع /اضافه

FANTASY_MATCH_DAY_FORM_V26 = (
    "🏆 فانتزي المصيف 2026  🏆\n"
    "ً     🔥🔥🔥 اليوم العاشر🔥🔥  \n"
    "📋 نموذج المشاركة الرسمي المعتمد\n"
    "🏆 تشكيلة الفانتزي - اليوم (    )\n"
    "🧤 الحارس:\n"
    " اللاعب 1:\n"
    " اللاعب 2:\n"
    " اللاعب 3:\n"
    "👑 الكابتن :"
)


def _v26_safe_txt(x):
    return str(x or "").strip()


def _v26_short_time(t):
    s = _v26_safe_txt(t)
    if not s:
        return ""

    # تنظيف المسافات والرموز
    s = s.replace("صباحاً", "ص").replace("صباحًا", "ص").replace("صباحا", "ص")
    s = s.replace("مساءً", "م").replace("مساءا", "م").replace("مساء", "م")
    s = s.replace("فجراً", "ص").replace("فجرًا", "ص").replace("فجرا", "ص")
    s = s.replace("ليلاً", "م").replace("ليلًا", "م").replace("ليلا", "م")
    s = s.replace("منتصف الليل", "ص")
    s = s.replace("الظهر", "م").replace("ظهراً", "م").replace("ظهرًا", "م")
    s = s.replace("AM", "ص").replace("am", "ص").replace("A.M.", "ص").replace("a.m.", "ص")
    s = s.replace("PM", "م").replace("pm", "م").replace("P.M.", "م").replace("p.m.", "م")
    s = re.sub(r"\s+", " ", s).strip()

    # لو الصيغة صارت 3:00 ص أو 10:00 م
    m = re.search(r"(\d{1,2})(?::(\d{2}))?\s*([صم])", s)
    if m:
        hh = int(m.group(1))
        mm = m.group(2) or "00"
        ap = m.group(3)
        return f"{hh}:{mm} {ap}"

    # لو فيه كلمة ص/م قبل الرقم بطريقة غريبة
    m = re.search(r"([صم])\s*(\d{1,2})(?::(\d{2}))?", s)
    if m:
        ap = m.group(1)
        hh = int(m.group(2))
        mm = m.group(3) or "00"
        return f"{hh}:{mm} {ap}"

    return s


def _v26_dedupe_fixture_matches(matches):
    seen = set()
    out = []
    for m in matches or []:
        t1 = _v26_safe_txt(m.get("team1"))
        t2 = _v26_safe_txt(m.get("team2"))
        tm = _v26_short_time(m.get("time"))
        dt = _v26_safe_txt(m.get("date"))
        key = (
            re.sub(r"\s+", " ", t1.replace("ـ", "")).strip(),
            re.sub(r"\s+", " ", t2.replace("ـ", "")).strip(),
            tm,
            dt,
        )
        if key in seen:
            continue
        seen.add(key)
        x = dict(m)
        x["time"] = tm
        out.append(x)
    return out


def _v26_fixture_title(date):
    d = _normalize_date_arg(date)
    day = ""
    for x, dy in _fixture_dates():
        if x == d:
            day = dy
            break
    return f"{day} {d[:5]}".strip()


def _v26_fixture_simple_matches(date):
    rows = _v26_dedupe_fixture_matches(_fixtures_for_date(date))
    simple = []
    for m in rows:
        simple.append((
            _v26_safe_txt(m.get("team1")),
            _v26_safe_txt(m.get("team2")),
            _v26_short_time(m.get("time")),
        ))
    return rows, simple


def _v26_fixtures_caption(date_or_title, matches=None, source="PDF جدول البطولة", include_fantasy=True):
    lines = [
        "🏆 مونديال المصيف 2026 🏆",
        f"🔥 مباريات اليوم ( {date_or_title} ) 🔥",
        "",
    ]
    for i, m in enumerate(matches or [], start=1):
        if isinstance(m, dict):
            a = _v26_safe_txt(m.get("team1"))
            b = _v26_safe_txt(m.get("team2"))
            t = _v26_short_time(m.get("time"))
        else:
            a, b, t = m
            t = _v26_short_time(t)
        lines.append(f"{i}. {a} × {b}" + (f" — {t}" if t else ""))

    lines.extend(["", f"المصدر: {source}", "المصيف يضعكم بالحدث"])

    if include_fantasy:
        lines.extend(["", FANTASY_MATCH_DAY_FORM_V26])

    return "\n".join(lines)


def _fixtures_caption(date_or_title, source="PDF جدول البطولة"):
    # نخلي الدالة القديمة ترجع نفس الهوية الجديدة إذا استُخدمت في مسار قديم
    return _v26_fixtures_caption(date_or_title, [], source=source, include_fantasy=True)


def _v26_compact_bg(w=1080, h=1350):
    candidates = [
        "games_v31_clean_bg.png",
        os.path.join("assets", "templates", "games_v31_clean_bg.png"),
        "games_v31_full_bg.png",
        os.path.join("assets", "templates", "games_v31_full_bg.png"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                bg = Image.open(p).convert("RGB").resize((w, h))
                ov = Image.new("RGBA", (w, h), (0, 10, 30, 118))
                return Image.alpha_composite(bg.convert("RGBA"), ov).convert("RGB")
            except Exception:
                pass
    return Image.new("RGB", (w, h), "#071329")


def _v26_draw_day_block(draw, box, date, matches):
    x0, y0, x1, y1 = box
    try:
        draw.rounded_rectangle(box, radius=26, fill=(4, 18, 50, 218), outline=(38,151,255,190), width=2)
        draw.rounded_rectangle((x0+16, y0+16, x1-16, y0+70), radius=18, fill=(251,191,36,238))
    except Exception:
        draw.rectangle(box, fill=(4, 18, 50), outline=(38,151,255), width=2)
        draw.rectangle((x0+16, y0+16, x1-16, y0+70), fill=(251,191,36))

    draw_text(draw, ((x0+x1)//2, y0+43), _v26_fixture_title(date), get_font(28), fill="#061329", max_width=(x1-x0)-45)

    yy = y0 + 94
    row_h = 58
    if len(matches) >= 5:
        row_h = 50

    for idx, m in enumerate(matches[:6], start=1):
        if yy + row_h > y1 - 18:
            break
        time_txt = _v26_short_time(m.get("time"))
        teams = f"{_v26_safe_txt(m.get('team1'))} × {_v26_safe_txt(m.get('team2'))}"

        # خط فاصل خفيف
        if idx > 1:
            draw.line((x0+26, yy-9, x1-26, yy-9), fill=(255,255,255,45), width=1)

        # الوقت كبيل صغير لا ينكسر
        try:
            draw.rounded_rectangle((x0+22, yy+4, x0+115, yy+38), radius=12, fill=(2,10,27,230), outline=(251,191,36,180), width=1)
        except Exception:
            draw.rectangle((x0+22, yy+4, x0+115, yy+38), fill=(2,10,27))
        draw_text(draw, (x0+68, yy+21), time_txt, get_font(19), fill="#FBBF24", max_width=86)

        draw_text(draw, ((x0+x1)//2+35, yy+22), teams, get_font(23 if len(teams) < 28 else 20), fill="#FFFFFF", max_width=(x1-x0)-155)
        yy += row_h

    if len(matches) > 6:
        draw_text(draw, ((x0+x1)//2, y1-24), f"+ {len(matches)-6} مباراة", get_font(19), fill="#FBBF24")


def render_fixtures_combined_images(dates):
    """
    التصميم الجديد للأيام المتعددة:
    يومين جنب بعض — السبت/الأحد ثم الاثنين/الثلاثاء...
    """
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    clean_dates = []
    for d in dates:
        if d and _fixtures_for_date(d) and d not in clean_dates:
            clean_dates.append(d)
    dates = clean_dates
    if not dates:
        return []

    # 4 أيام لكل صورة = صفين، وكل صف يومين
    chunks = [dates[i:i+4] for i in range(0, len(dates), 4)]
    paths = []

    for page, dchunk in enumerate(chunks, start=1):
        w, h = 1080, 1350
        img = _v26_compact_bg(w, h)
        draw = ImageDraw.Draw(img, "RGBA")

        # عنوان صغير فقط، بدون الكلام الكبير القديم
        if len(chunks) > 1:
            draw_text(draw, (540, 44), f"صفحة {page}/{len(chunks)}", get_font(24), fill="#FBBF24")

        left_x, right_x = 55, 555
        col_w = 470
        block_h = 520
        top_y = 95
        gap_y = 42
        positions = [
            (left_x, top_y, left_x+col_w, top_y+block_h),
            (right_x, top_y, right_x+col_w, top_y+block_h),
            (left_x, top_y+block_h+gap_y, left_x+col_w, top_y+block_h+gap_y+block_h),
            (right_x, top_y+block_h+gap_y, right_x+col_w, top_y+block_h+gap_y+block_h),
        ]

        for d, box in zip(dchunk, positions):
            _v26_draw_day_block(draw, box, d, _v26_dedupe_fixture_matches(_fixtures_for_date(d)))

        draw.line((250, h-76, 830, h-76), fill=(255,255,255,160), width=2)
        draw_text(draw, (540, h-39), "المصيف يضعكم بالحدث", get_font(28), fill="#FBBF24")
        path = os.path.join(GENERATED_DIR, f"fixtures_combined_v26_new_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _v26_render_fixtures_10_images(dates):
    """
    تصميم 10 مباريات: قائمة واضحة، كل صفحة حتى 10 مباريات.
    """
    dates = [_normalize_date_arg(d) for d in dates if _normalize_date_arg(d)]
    flat = []
    for d in dates:
        for m in _v26_dedupe_fixture_matches(_fixtures_for_date(d)):
            x = dict(m)
            x["_date_title"] = _v26_fixture_title(d)
            flat.append(x)

    if not flat:
        return []

    chunks = [flat[i:i+10] for i in range(0, len(flat), 10)]
    paths = []

    for page, chunk in enumerate(chunks, start=1):
        w, h = 1080, 1350
        img = _v26_compact_bg(w, h)
        draw = ImageDraw.Draw(img, "RGBA")

        draw_text(draw, (540, 64), "جدول المباريات", get_font(42), fill="#FFFFFF", max_width=760)
        draw_text(draw, (540, 112), f"تصميم 10 مباريات" + (f" | صفحة {page}/{len(chunks)}" if len(chunks) > 1 else ""), get_font(26), fill="#FBBF24", max_width=820)

        y = 170
        for i, m in enumerate(chunk, start=1):
            try:
                draw.rounded_rectangle((70, y, 1010, y+92), radius=18, fill=(5,24,58,225), outline=(38,151,255,170), width=2)
            except Exception:
                draw.rectangle((70, y, 1010, y+92), fill=(5,24,58), outline=(38,151,255), width=2)

            draw_text(draw, (950, y+27), str(i + (page-1)*10), get_font(24), fill="#FBBF24")
            draw_text(draw, (820, y+28), _v26_safe_txt(m.get("_date_title")), get_font(20), fill="#CBD5E1", max_width=210)
            draw_text(draw, (540, y+34), f"{_v26_safe_txt(m.get('team1'))} × {_v26_safe_txt(m.get('team2'))}", get_font(28), fill="#FFFFFF", max_width=560)
            try:
                draw.rounded_rectangle((95, y+25, 205, y+62), radius=13, fill=(2,10,27,230), outline=(251,191,36,180), width=1)
            except Exception:
                draw.rectangle((95, y+25, 205, y+62), fill=(2,10,27))
            draw_text(draw, (150, y+44), _v26_short_time(m.get("time")), get_font(21), fill="#FBBF24", max_width=96)
            y += 106

        draw.line((250, h-76, 830, h-76), fill=(255,255,255,160), width=2)
        draw_text(draw, (540, h-39), "المصيف يضعكم بالحدث", get_font(28), fill="#FBBF24")
        path = os.path.join(GENERATED_DIR, f"fixtures_10_v26_{page}_{datetime.now().strftime('%H%M%S')}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _render_fixture_day_compact(date):
    rows = _v26_dedupe_fixture_matches(_fixtures_for_date(date))
    if not rows:
        return []

    chunks = [rows[i:i+8] for i in range(0, len(rows), 8)]
    paths = []

    for page, chunk in enumerate(chunks, 1):
        h = 1350
        img = _v26_compact_bg(1080, h)
        draw = ImageDraw.Draw(img, "RGBA")

        y = 105
        try:
            draw.rounded_rectangle((120, y, 960, y+62), radius=22, fill=(251,191,36,235))
        except Exception:
            draw.rectangle((120, y, 960, y+62), fill=(251,191,36))
        title = _v26_fixture_title(date)
        if len(chunks) > 1:
            title += f" | {page}/{len(chunks)}"
        draw_text(draw, (540, y+31), title, get_font(34), fill="#061329", max_width=780)

        y += 92
        for m in chunk:
            try:
                draw.rounded_rectangle((70, y, 1010, y+104), radius=22, fill=(5,24,58,225), outline=(38,151,255,190), width=2)
            except Exception:
                draw.rectangle((70, y, 1010, y+104), fill=(5,24,58), outline=(38,151,255), width=2)

            time_txt = _v26_short_time(m.get("time"))
            try:
                draw.rounded_rectangle((835, y+26, 982, y+70), radius=14, fill=(2,10,27,230), outline=(251,191,36,180), width=1)
            except Exception:
                draw.rectangle((835, y+26, 982, y+70), fill=(2,10,27))
            draw_text(draw, (908, y+48), time_txt, get_font(24), fill="#FBBF24", max_width=130)

            draw_text(draw, (485, y+39), f"{m.get('team1')} × {m.get('team2')}", get_font(30), fill="#FFFFFF", max_width=620)
            sub = _v26_safe_txt(m.get("stage"))
            if m.get("group"):
                sub += f" - {_v26_safe_txt(m.get('group'))}"
            draw_text(draw, (485, y+76), sub, get_font(21), fill="#CBD5E1", max_width=650)
            y += 122

        draw.line((250, 1238, 830, 1238), fill=(255,255,255,180), width=2)
        draw_text(draw, (540, 1284), "المصيف يضعكم بالحدث", get_font(30), fill="#FBBF24")

        path = os.path.join(GENERATED_DIR, f"fixtures_compact_v26_{date.replace('/','_')}_{page}.png")
        img.save(path, quality=96)
        paths.append(path)

    return paths


def _render_fixture_day_by_design(date, design=2):
    rows, simple_matches = _v26_fixture_simple_matches(date)
    if not simple_matches:
        return []

    if int(design) == 1:
        return _render_fixture_day_compact(date)

    chunks = [simple_matches[i:i+7] for i in range(0, len(simple_matches), 7)]
    paths = []

    for page_idx, chunk in enumerate(chunks, start=1):
        page_title = _v26_fixture_title(date)
        if len(chunks) > 1:
            page_title = f"{page_title} | {page_idx}/{len(chunks)}"

        path = create_matches_today_v31_full_image(page_title, chunk)
        final_path = os.path.join(
            GENERATED_DIR,
            f"fixtures_day_design2_v26_{date.replace('/','_')}_{page_idx}.png"
        )
        try:
            Image.open(path).save(final_path, quality=96)
            paths.append(final_path)
        except Exception:
            paths.append(path)

    return paths


def _fixtures_day_keyboard(date):
    rows = [
        [
            InlineKeyboardButton("تصميم 1", callback_data=f"fx|render1|{date}"),
            InlineKeyboardButton("تصميم 2", callback_data=f"fx|render2|{date}"),
        ]
    ]

    miss = [m for m in _fixtures_for_date(date) if _has_unknown(m)]
    for i, m in enumerate(miss, 1):
        rows.append([InlineKeyboardButton(f"تحديث مباراة {i} — {_v26_short_time(m.get('time'))}", callback_data=f"fx|upd|{m.get('id')}")])

    rows.append([InlineKeyboardButton("رجوع للأيام", callback_data="fx|menu")])
    return InlineKeyboardMarkup(rows)


def _fixtures_dates_keyboard(mode="single", selected=None):
    selected = set(selected or [])
    rows = []
    row = []

    for d, day in _fixture_dates():
        label = f"{'✅ ' if d in selected else ''}{day} {d[:5]}"
        data = f"fx|toggle|{d}" if mode == "multi" else f"fx|day|{d}"
        row.append(InlineKeyboardButton(label, callback_data=data))
        if len(row) == 2:
            rows.append(row)
            row = []

    if row:
        rows.append(row)

    if mode == "multi":
        rows.append([
            InlineKeyboardButton("تصميم جديد", callback_data="fx|render_combo_new"),
            InlineKeyboardButton("تصميم 10 مباريات", callback_data="fx|render_combo_10"),
        ])
        rows.append([
            InlineKeyboardButton("تصفير الاختيار", callback_data="fx|clear"),
            InlineKeyboardButton("رجوع", callback_data="fx|menu"),
        ])
    else:
        rows.append([InlineKeyboardButton("اختيار أكثر من يوم", callback_data="fx|multi")])

    return InlineKeyboardMarkup(rows)


def _fixtures_day_text(date):
    rows, matches = _v26_fixture_simple_matches(date)
    if not matches:
        return "ما فيه مباريات لهذا التاريخ."

    lines = [f"{_v26_fixture_title(date)}", ""]
    for i, m in enumerate(rows, 1):
        lines.append(f"{i}) {_v26_safe_txt(m.get('team1'))} × {_v26_safe_txt(m.get('team2'))} — {_v26_short_time(m.get('time'))}")
        extra = []
        if m.get("stage"):
            extra.append(_v26_safe_txt(m.get("stage")))
        if m.get("group"):
            extra.append(_v26_safe_txt(m.get("group")))
        if extra:
            lines.append("   " + " | ".join(extra))
        if _has_unknown(m) and m.get("note"):
            lines.append(f"   {_v26_safe_txt(m.get('note'))}")

    return "\n".join(lines)


def _v26_combo_choice_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("تصميم جديد", callback_data="fx|render_combo_new"),
            InlineKeyboardButton("تصميم 10 مباريات", callback_data="fx|render_combo_10"),
        ]
    ])


async def fixtures_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    dates = _extract_fixture_dates_from_text(text)

    if dates:
        if len(dates) == 1:
            d = dates[0]
            wait = await update.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                rows, simple = _v26_fixture_simple_matches(d)
                paths = _render_fixture_day_by_design(d, design=2)
                if not paths:
                    await wait.edit_text(f"ما فيه مباريات بتاريخ {d}")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(update.message, p, _v26_fixtures_caption(_v26_fixture_title(d), simple, include_fantasy=True))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        # أكثر من تاريخ: لا يصمم مباشرة، يعطي زرين
        clean_dates = []
        for d in dates:
            if d not in clean_dates:
                clean_dates.append(d)
        context.user_data["fx_selected_dates"] = clean_dates
        await update.message.reply_text("اختر التصميم للأيام المحددة:", reply_markup=_v26_combo_choice_keyboard())
        return

    await update.message.reply_text("اختر اليوم أو اكتب:\n/مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))


async def fixtures_combined_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مباريات_مجمعة 20/06 21/06 22/06")
        return

    clean_dates = []
    for d in dates:
        if d not in clean_dates:
            clean_dates.append(d)
    context.user_data["fx_selected_dates"] = clean_dates
    await update.message.reply_text("اختر التصميم للأيام المحددة:", reply_markup=_v26_combo_choice_keyboard())


async def fixtures_review_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dates = _extract_fixture_dates_from_text(update.message.text)
    if not dates:
        await update.message.reply_text("اكتبها كذا:\n/مراجعة_مباراة 20/07")
        return
    for d in dates:
        await update.message.reply_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))


async def fixtures_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if not q:
        return

    await q.answer()

    if not is_admin_user(update):
        await q.message.reply_text("هذا الخيار للمشرفين فقط 🔒")
        return

    parts = (q.data or "").split("|")
    action = parts[1] if len(parts) > 1 else ""

    try:
        if action == "menu":
            await q.message.edit_text("اختر اليوم أو استخدم /مباريات 20/06", reply_markup=_fixtures_dates_keyboard("single"))
            return

        if action == "multi":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اختر التصميم:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "toggle" and len(parts) >= 3:
            d = parts[2]
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if d in sel:
                sel.remove(d)
            else:
                sel.append(d)
            context.user_data["fx_selected_dates"] = sel
            await q.message.edit_text("اختر الأيام المطلوبة ثم اختر التصميم:", reply_markup=_fixtures_dates_keyboard("multi", sel))
            return

        if action == "clear":
            context.user_data["fx_selected_dates"] = []
            await q.message.edit_text("اختر الأيام المطلوبة ثم اختر التصميم:", reply_markup=_fixtures_dates_keyboard("multi", []))
            return

        if action == "day" and len(parts) >= 3:
            d = parts[2]
            await q.message.edit_text(_fixtures_day_text(d), reply_markup=_fixtures_day_keyboard(d))
            return

        if action in ["render1", "render2"] and len(parts) >= 3:
            d = parts[2]
            design = 1 if action == "render1" else 2
            wait = await q.message.reply_text("⏳ جاري تصميم مباريات اليوم...")
            try:
                rows, simple = _v26_fixture_simple_matches(d)
                paths = _render_fixture_day_by_design(d, design=design)
                if not paths:
                    await wait.edit_text("ما فيه مباريات لهذا اليوم.")
                    return
                try:
                    await wait.delete()
                except Exception:
                    pass
                for p in paths:
                    await send_photo_path(q.message, p, _v26_fixtures_caption(_v26_fixture_title(d), simple, include_fantasy=True))
            except Exception as e:
                await wait.edit_text(f"تعذر تصميم اليوم ❌\nالسبب: {str(e)[:400]}")
            return

        if action in ["render_combo_new", "render_combo_10", "render_combo"]:
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return

            title = "التصميم الجديد" if action != "render_combo_10" else "تصميم 10 مباريات"
            wait = await q.message.reply_text(f"⏳ جاري تصميم {title}...")
            try:
                if action == "render_combo_10":
                    paths = _v26_render_fixtures_10_images(sel)
                else:
                    paths = render_fixtures_combined_images(sel)

                if not paths:
                    await wait.edit_text("ما لقيت مباريات للتواريخ المحددة.")
                    return

                try:
                    await wait.delete()
                except Exception:
                    pass

                for p in paths:
                    await send_photo_path(q.message, p, _v26_fixtures_caption("مباريات مجمعة", [], include_fantasy=False))
            except Exception as e:
                await wait.edit_text(f"تعذر التصميم ❌\nالسبب: {str(e)[:400]}")
            return

        # توافق قديم لو فيه زر render_each
        if action == "render_each":
            sel = list(context.user_data.get("fx_selected_dates") or [])
            if not sel:
                await q.message.reply_text("اختر يومًا واحدًا على الأقل.")
                return
            for d in sel:
                rows, simple = _v26_fixture_simple_matches(d)
                for p in _render_fixture_day_by_design(d, design=2):
                    await send_photo_path(q.message, p, _v26_fixtures_caption(_v26_fixture_title(d), simple, include_fantasy=True))
            return

        if action == "upd" and len(parts) >= 3:
            mid = parts[2]
            m = _fixture_by_id(mid)
            if not m:
                await q.message.reply_text("لم أجد المباراة.")
                return

            context.user_data["fixture_update_match_id"] = mid
            await q.message.reply_text(
                f"اكتب طرفي المباراة لـ {mid} ({m.get('date')} {_v26_short_time(m.get('time'))}) كذا:\n"
                "الفريق الأول * الفريق الثاني\n\n"
                "مثال: المكسيك * أستراليا\n"
                "ملاحظة: سيتم الحفظ فقط، ولن يتم التصميم إلا عندما تطلب /مباريات التاريخ."
            )
            return

        await q.message.reply_text("تعذر قراءة الخيار.")
    except Exception as e:
        await q.message.reply_text(f"تعذر تنفيذ خيار المباريات ❌\n{str(e)[:400]}")


async def fixtures_update_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mid = context.user_data.get("fixture_update_match_id")
    if not mid:
        return

    text = (update.message.text or "").strip()

    if "*" in text:
        a, b = [x.strip() for x in text.split("*", 1)]
    elif "×" in text:
        a, b = [x.strip() for x in text.split("×", 1)]
    elif "-" in text:
        a, b = [x.strip() for x in text.split("-", 1)]
    else:
        await update.message.reply_text("اكتبها كذا: الفريق الأول * الفريق الثاني")
        return

    if not a or not b:
        await update.message.reply_text("اكتب اسم الفريقين كاملين.")
        return

    data = _load_fixture_updates()
    data.setdefault(mid, {})
    data[mid]["team1"] = canonical_team_name(a) or normalize_name(a)
    data[mid]["team2"] = canonical_team_name(b) or normalize_name(b)
    _save_fixture_updates(data)

    context.user_data.pop("fixture_update_match_id", None)

    m = _apply_fixture_updates(_fixture_by_id(mid) or {"id": mid})
    await update.message.reply_text(
        f"✅ تم حفظ تحديث المباراة\n"
        f"{m.get('team1')} × {m.get('team2')} — {_v26_short_time(m.get('time', ''))}\n\n"
        f"لن أصمم الآن. وقت ما تبيها اكتب:\n/مباريات {m.get('date', '')}"
    )


def _v26_extract_name_after_command(text, mode="add"):
    s = normalize_name(text or "")
    if mode == "add":
        pat = r"^/(?:إضافة_متسابق|اضافة_متسابق|إضافه_متسابق|اضافه_متسابق|إضافة\s+متسابق|اضافة\s+متسابق|إضافه\s+متسابق|اضافه\s+متسابق)\s*"
    else:
        pat = r"^/(?:حذف_متسابق|ازالة_متسابق|إزالة_متسابق|حذف\s+متسابق|ازالة\s+متسابق|إزالة\s+متسابق)\s*"
    s = re.sub(pat, "", s, flags=re.I).strip()
    return normalize_name(s)


async def add_participant_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = _v26_extract_name_after_command(update.message.text, "add")
    if not name:
        await update.message.reply_text("اكتبها كذا:\n/إضافة_متسابق عبدالله محمد")
        return
    if name in PARTICIPANTS:
        await update.message.reply_text(f"✅ {name} موجود مسبقًا في قائمة المتسابقين")
        return
    PARTICIPANTS.append(name)
    _save_participants_state()
    await update.message.reply_text(f"✅ تمت إضافة المتسابق: {name}\nالمجموع الحالي: {len(PARTICIPANTS)}")


async def remove_participant_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = _v26_extract_name_after_command(update.message.text, "remove")
    if not name:
        await update.message.reply_text("اكتبها كذا:\n/حذف_متسابق عبدالله محمد")
        return
    target = name if name in PARTICIPANTS else None
    if not target:
        close = difflib.get_close_matches(name, PARTICIPANTS, n=1, cutoff=0.75)
        target = close[0] if close else None
    if not target:
        await update.message.reply_text(f"ما لقيت المتسابق: {name} ❌")
        return
    PARTICIPANTS.remove(target)
    _save_participants_state()
    await update.message.reply_text(f"✅ تم حذف المتسابق: {target}\nالمجموع الحالي: {len(PARTICIPANTS)}")


def _v26_is_import_command_text(text):
    s = normalize_name(text or "")
    return bool(re.match(r"^/(?:استيراد_ملف|استيراد\s+ملف|استيراد|استيراد_اكسل|استيراد_إكسل|استيراد_excel)(?:\s|$)", s))


def _v26_find_latest_excel_for_chat(chat_id):
    candidates = []
    for folder in ["uploads", "imports"]:
        if not os.path.isdir(folder):
            continue
        for name in os.listdir(folder):
            low = name.lower()
            if not low.endswith((".xlsx", ".xlsm")):
                continue
            # الملفات المحفوظة تبدأ غالبًا برقم الشات
            p = os.path.join(folder, name)
            if str(chat_id) in name or True:
                try:
                    candidates.append((os.path.getmtime(p), p))
                except Exception:
                    pass
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


async def _v26_run_import_from_path(update: Update, context: ContextTypes.DEFAULT_TYPE, local_path):
    if not local_path or not os.path.exists(local_path):
        await update.message.reply_text("ملف الإكسل غير موجود. أرسله مرة ثانية.")
        return
    wait = await update.message.reply_text("⏳ جاري قراءة ملف الإكسل...")
    try:
        imported = await asyncio.wait_for(asyncio.to_thread(parse_import_excel, local_path), timeout=60)
    except Exception as e:
        await wait.edit_text(f"صار خطأ أثناء قراءة الإكسل ❌\n{str(e)[:500]}")
        return
    if not imported:
        await wait.edit_text("ما قدرت أستخرج أيام من الملف. تأكد أن الصفحات باسم: يوم 1، يوم 2 ...")
        return
    chat_id = update.effective_chat.id
    PENDING_IMPORTS[chat_id] = {"path": local_path, "data": imported}
    await wait.edit_text(import_summary_text(imported))


async def import_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    document = update.message.document

    if document:
        filename = document.file_name or "import.xlsx"
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            await update.message.reply_text("الملف لازم يكون Excel بصيغة .xlsx أو .xlsm")
            return
        local_path, _ = await _download_document_to_folder(update, context, "imports")
        LAST_UPLOADED_FILES.setdefault(chat_id, {})["excel"] = local_path
        await _v26_run_import_from_path(update, context, local_path)
        return

    local_path = LAST_UPLOADED_FILES.get(chat_id, {}).get("excel")
    if not local_path or not os.path.exists(local_path):
        local_path = _v26_find_latest_excel_for_chat(chat_id)

    if not local_path or not os.path.exists(local_path):
        await update.message.reply_text(
            "ما لقيت ملف Excel محفوظ.\n"
            "أرسل ملف الإكسل لحاله أولًا، وبعدها اكتب:\n"
            "/استيراد_ملف"
        )
        return

    LAST_UPLOADED_FILES.setdefault(chat_id, {})["excel"] = local_path
    await _v26_run_import_from_path(update, context, local_path)


async def remember_last_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    حفظ ملفات Excel/ZIP، مع دعم كابشن الاستيراد بكل الصيغ.
    """
    local_path, filename = await _download_document_to_folder(update, context, "uploads")
    if not local_path:
        return

    chat_id = update.effective_chat.id
    LAST_UPLOADED_FILES.setdefault(chat_id, {})
    lower = (filename or "").lower()
    caption = (update.message.caption or "").strip()

    if lower.endswith((".xlsx", ".xlsm")):
        LAST_UPLOADED_FILES[chat_id]["excel"] = local_path
        if _v26_is_import_command_text(caption):
            await _v26_run_import_from_path(update, context, local_path)
        else:
            await update.message.reply_text(
                "وصل ملف الإكسل ✅\n"
                "اكتب الآن:\n"
                "/استيراد_ملف"
            )
        return

    if lower.endswith(".zip"):
        LAST_UPLOADED_FILES[chat_id]["zip"] = local_path
        if caption.startswith("/استرجاع_نسخة"):
            await _run_restore_from_zip_path(update, context, local_path)
        else:
            await update.message.reply_text(
                "وصل ملف ZIP ✅\n"
                "للاسترجاع اكتب الآن:\n"
                "/استرجاع_نسخة"
            )
        return

    await update.message.reply_text("وصل الملف ✅")

# ==================== END V26 FINAL OVERRIDE ====================

# ==================== END V24 STABLE PATCH ====================

if __name__ == "__main__":
    main()